import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import json
import html
import hashlib

import uuid
from pathlib import Path
from datetime import date, datetime, timedelta
from io import BytesIO
import zipfile
import sqlite3
import re
import os
import base64
import random
import shutil
import threading
import time
from contextlib import contextmanager

from hidamari.auth.login_state import (
    INITIAL_ACCOUNT_PASSWORD,
    INITIAL_LOGIN_IDS,
    clear_login_failures,
    is_login_temporarily_locked,
    record_login_failure,
)
from hidamari.auth.password import (
    account_requires_password_change,
    hash_password,
    is_bcrypt_available,
    password_hash_needs_upgrade,
    uses_initial_password,
    validate_new_password,
    verify_password,
)
from hidamari.auth.permissions import is_admin_identity
from hidamari.config.columns import (
    ACCOUNT_COLUMNS,
    ALERT_CONDITION_COLUMNS,
    APP_SETTING_COLUMNS,
    ASSESSMENT_COLUMNS,
    BUSINESS_HANDOVER_COLUMNS,
    EXCRETION_COLUMNS,
    HANDOVER_KEYWORD_COLUMNS,
    HEALTH_COLUMNS,
    LIFE_ADL_COLUMNS,
    LOGIN_HISTORY_COLUMNS,
    MONITORING_DRAFT_COLUMNS,
    SHORT_GOAL_CHECK_COLUMNS,
    SHORT_GOAL_MASTER_COLUMNS,
    USER_COLUMNS,
    USER_NAME_ALIAS_COLUMNS,
)
from hidamari.config.constants import (
    ADL_LEVEL_OPTIONS,
    COGNITIVE_OPTIONS,
    DEFAULT_USERS,
    DENTURE_OPTIONS,
    EXCRETION_SHEET,
    EXCRETION_SLOTS,
    HEALTH_SHEET,
    MEAL_INTAKE_OPTIONS,
    MEAL_INTAKE_PERCENT,
    NUTRITION_RISK_OPTIONS,
    ORAL_STATUS_OPTIONS,
    STOOL_AMOUNT_CODE,
    STOOL_AMOUNT_OPTIONS,
    STOOL_TYPE_CODE,
    STOOL_TYPE_OPTIONS,
    URINE_AMOUNT_CODE,
    URINE_AMOUNT_OPTIONS,
    URINE_TYPE_CODE,
    URINE_TYPE_OPTIONS,
    USER_SHEET,
)
from hidamari.config.menu import (
    MENU_CATEGORY_LABELS,
    MENU_DISPLAY_LABELS,
    MENU_GROUPS_ADMIN,
    MENU_GROUPS_STAFF,
    menu_category_label,
    menu_display_label,
)
from hidamari.config.paths import (
    ACCOUNT_FILE,
    ALERT_CONDITION_FILE,
    BUSINESS_HANDOVER_EXCEL_DIR,
    BUSINESS_HANDOVER_PHOTO_DIR,
    DATA_DIR,
    EXCRETION_FILE,
    HANDOVER_FILE,
    HEALTH_FILE,
    HIDAMARI_DB_FILE,
    LIFE_ADL_FILE,
    LOGIN_HISTORY_FILE,
    MENU_CATEGORY_SETTINGS_FILE,
    MONITORING_DRAFT_FILE,
    REPORT_DIR,
    SHORT_GOAL_CHECK_FILE,
    SHORT_GOAL_MASTER_FILE,
    USER_FILE,
)
from hidamari.core.text_utils import (
    clean_text,
    get_life_option_index,
    get_option_index,
    html_escape_text,
    make_date_user_key,
    make_excretion_key,
    meal_option_from_percent,
    option_code,
    safe_float,
    safe_int,
    to_number,
)
from hidamari.core.time_utils import JST, format_now_jst, now_jst, now_jst_dt, today_jst
from hidamari.features.ai import (
    analyze_structured_insights,
    build_ai_structured_context,
    filter_records_by_period,
    get_openai_api_key,
    hidamari_ai_build_admin_report,
    hidamari_ai_filter_period,
)
from hidamari.features.ai_prompts import AI_STRUCTURED_ADVICE_SYSTEM_PROMPT
from hidamari.pdf.ai_report import (
    AI_INSIGHT_CONFIRMATION_NOTE,
    AI_INSIGHT_LOG_EMPTY_TEXT,
    AI_INSIGHT_LOG_SUMMARY_TITLE,
    AI_INSIGHT_REPORT_DISCLAIMER,
    AI_INSIGHT_REPORT_TITLE,
    ai_admin_report_file_name,
    append_markdown_lines_to_story,
)
from hidamari.pdf.common import (
    paragraph_lines,
    pdf_safe_text,
    register_japanese_pdf_fonts,
    register_single_japanese_pdf_font,
    short_goal_join_for_pdf,
    short_goal_pdf_text,
)
from hidamari.pdf.family_report import (
    HIDAMARI_REPORT_DISCLAIMER,
    HIDAMARI_REPORT_EXCRETION_HEADING,
    HIDAMARI_REPORT_NO_EXCRETION_TEXT,
    HIDAMARI_REPORT_SUMMARY_HEADING,
    HIDAMARI_REPORT_TITLE,
)
from hidamari.ui.common import (
    danger_note,
    os_mindset_box,
    product_ui_notice,
    safe_note,
    show_observation_perspective,
    ui_badges,
    ui_card,
    ui_section,
    warning_note,
)
from hidamari.ui.sidebar import configure_sidebar, render_sidebar_menu
from hidamari.ui.theme import (
    APP_COPY,
    APP_VERSION,
    UI_COLORS,
    apply_design,
    apply_product_ui_ux,
    configure_theme_settings,
    get_color_settings,
    get_ui_theme,
)

try:
    import requests
except Exception:
    requests = None

try:
    import matplotlib.pyplot as plt
except Exception:
    plt = None

try:
    from PIL import Image, ImageOps
except Exception:
    Image = None
    ImageOps = None

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
except Exception:
    colors = None



# =========================
# Ver4.8.3 OpenAIモデル切替対応
# Ver4.8.2 起動順安全化：キャッシュ関連の最小定義を最上部で先に用意
# =========================
# Streamlit Cloudでは、デコレータ行の評価時点で関数名が未定義だと
# NameErrorでアプリ全体が停止するため、後続の本定義より前に安全版を置く。
DEFAULT_QUERY_CACHE_TTL_SEC = 60
DEFAULT_RECENT_DAYS = 7
SAFE_READ_CACHE_TTL_SEC = 300

def cache_safe_master_read(ttl=SAFE_READ_CACHE_TTL_SEC):
    """Return a safe st.cache_data decorator, or the original function if caching is unavailable."""
    def _decorator(func):
        try:
            return st.cache_data(ttl=ttl, show_spinner=False)(func)
        except Exception:
            return func
    return _decorator


def log_perf(label, elapsed_sec, detail=""):
    try:
        entry = {
            "time": datetime.utcnow().isoformat(timespec="seconds"),
            "label": str(label),
            "elapsed_sec": round(float(elapsed_sec), 3),
            "detail": str(detail or ""),
        }
        logs = st.session_state.setdefault("hidamari_perf_logs", [])
        logs.append(entry)
        if len(logs) > 50:
            del logs[:-50]
    except Exception:
        pass
    try:
        print(f"[hidamari_perf] {label}: {elapsed_sec:.3f}s {detail}")
    except Exception:
        pass


@contextmanager
def perf_timer(label, detail=""):
    start = time.perf_counter()
    try:
        yield
    finally:
        elapsed = time.perf_counter() - start
        if elapsed >= 0.2:
            log_perf(label, elapsed, detail)


def _clear_cached_functions(function_names):
    cleared = []
    for name in function_names:
        func = globals().get(name)
        clear_func = getattr(func, "clear", None)
        if callable(clear_func):
            try:
                clear_func()
                cleared.append(name)
            except Exception:
                pass
    return cleared


def clear_hidamari_read_cache(reason=""):
    """Clear Streamlit read caches after saves, deletes, syncs, and restores."""
    reason_text = str(reason or "")
    targets = []
    if any(key in reason_text for key in ["利用者", "マスタ", "user"]):
        targets.extend(["load_users", "load_active_user_names", "load_user_name_aliases", "_supabase_read_table_cached"])
    if any(key in reason_text for key in ["健康", "排泄", "申し送り", "短期", "目標", "モニタリング", "復元", "Supabase", "保存", "削除"]):
        targets.extend(["_supabase_read_table_cached", "load_short_goal_master"])
    cleared = _clear_cached_functions(dict.fromkeys(targets))
    if cleared:
        return
    try:
        st.cache_data.clear()
    except Exception:
        pass


# =========================
# Ver4.7 表記修正：時間帯の「?」を「〜」に修正
# =========================

# =========================
# ページ設定
# =========================
st.set_page_config(
    page_title="ひだまり 健康チェック管理システム",
    page_icon="☀️",
    layout="wide",
)

def is_admin_user():
    role = st.session_state.get("role", "")
    user = (
        st.session_state.get("username", "")
        or st.session_state.get("user_id", "")
        or st.session_state.get("login_user", "")
        or st.session_state.get("user", "")
    )
    login_info = st.session_state.get("login_user_info", {})
    if isinstance(login_info, dict):
        role = role or login_info.get("role", "")
        user = user or login_info.get("username", "") or login_info.get("id", "")
    return is_admin_identity(role, user)


# 管理者専用メニュー制御
ADMIN_ONLY_MENUS = [
    "自分専用ダッシュボード",
    "データダウンロード",
    "LIFE入力標準化",
    "短期目標・モニタリング",
    "モニタリング下書き作成",
    "管理者LIFE入力",
    "LIFE不足チェック",
    "LIFE CSV出力",
    "LIFE登録一覧",
    "加算シミュレーション",
    "現場の気づき構造化・AI管理者支援",
    "AI管理者アシスタント",
    "セキュリティ・保守管理",
    "利用者ID移行チェック",
    "利用者名ゆれ紐づけマスタ",
]

# =========================
# 非表示メニュー制御
# =========================
# 「介護計画モニタリング下書き作成」は機能本体を残したまま、
# サイドバーのメニューからは表示しない。
# 既存のメニューカテゴリ設定に残っていても表示されないよう、
# サイドバー生成時と設定読込時の両方で除外する。
HIDDEN_MENUS = [
    "モニタリング下書き作成",
]

def filter_admin_menus(menu_list):
    hidden = set(HIDDEN_MENUS)
    if is_admin_user():
        return [m for m in menu_list if m not in hidden]
    return [m for m in menu_list if m not in ADMIN_ONLY_MENUS and m not in hidden]




def current_login_user():
    user = (
        st.session_state.get("username", "")
        or st.session_state.get("user_id", "")
        or st.session_state.get("login_user", "")
        or st.session_state.get("user", "")
    )
    login_info = st.session_state.get("login_user_info", {})
    if isinstance(login_info, dict):
        user = user or login_info.get("username", "") or login_info.get("id", "")
    return user or "kanri"

# =========================
# ログイン設定
# =========================
USERS = {
    "kanri": {"password": INITIAL_ACCOUNT_PASSWORD, "role": "admin", "label": "管理者"},
    "staff": {"password": INITIAL_ACCOUNT_PASSWORD, "role": "staff", "label": "職員"},
}


# =========================
# SQLite設定（Ver3.7 DB層分離）
# DB接続・WAL設定・保存/読込・整合性チェックは db/database.py に分離
# =========================
SQLITE_TABLE_HEALTH = "health_records"
SQLITE_TABLE_EXCRETION = "excretion_records"
SQLITE_TABLE_HANDOVER = "handover_logs"
SQLITE_TABLE_SHORT_GOAL_MASTER = "short_term_goals"
SQLITE_TABLE_SHORT_GOAL_CHECKS = "short_goal_checks"
SQLITE_TABLE_MONITORING_DRAFTS = "monitoring_drafts"
SQLITE_TABLE_ALERT_CONDITIONS = "alert_conditions"
SQLITE_TABLE_ALERTS = "alerts"
SQLITE_TABLE_ACCOUNTS = "login_accounts"
SQLITE_TABLE_LOGIN_HISTORY = "login_history"
SQLITE_TABLE_USERS = "users"
SQLITE_TABLE_APP_SETTINGS = "app_settings"
SQLITE_TABLE_LIFE_ADL = "life_adl_assessments"
SQLITE_TABLE_AI_INSIGHT_LOGS = "ai_insight_logs"
SQLITE_TABLE_USER_NAME_ALIASES = "user_name_aliases"
SQLITE_TABLE_HANDOVER_KEYWORDS = "handover_keywords"

from db import database as db_engine

db_engine.configure_database(DATA_DIR, HIDAMARI_DB_FILE)

DB_BUSY_TIMEOUT_MS = db_engine.DB_BUSY_TIMEOUT_MS
DB_WRITE_LOCK = db_engine.DB_WRITE_LOCK
DB_LAST_INTEGRITY_RESULT = db_engine.get_last_integrity_result()

validate_sqlite_identifier = db_engine.validate_sqlite_identifier
apply_sqlite_pragmas = db_engine.apply_sqlite_pragmas

# =========================
# SQLite PRAGMA互換ラッパー（Ver4.5.1）
# db/database.py 側の apply_sqlite_pragmas が for_write 引数に未対応の環境でも、
# LIFE用DB初期化で停止しないようにする。
# =========================
def safe_apply_sqlite_pragmas(conn, for_write=False):
    try:
        return apply_sqlite_pragmas(conn, for_write=for_write)
    except TypeError:
        # 旧版 db/database.py 互換
        try:
            result = apply_sqlite_pragmas(conn)
        except Exception:
            result = None
        # 書込用途の最低限PRAGMAをアプリ側で補完
        try:
            conn.execute("PRAGMA foreign_keys = ON;")
            conn.execute("PRAGMA busy_timeout = 30000;")
            if for_write:
                conn.execute("PRAGMA journal_mode = WAL;")
                conn.execute("PRAGMA synchronous = NORMAL;")
        except Exception:
            pass
        return result

hidamari_db_connection = db_engine.hidamari_db_connection
hidamari_write_transaction = db_engine.hidamari_write_transaction
get_hidamari_conn = db_engine.get_hidamari_conn

_original_sqlite_table_exists = db_engine.sqlite_table_exists
_original_sqlite_table_row_count = db_engine.sqlite_table_row_count
_original_db_write_dataframe = db_engine.db_write_dataframe
_original_db_read_dataframe = db_engine.db_read_dataframe

def sqlite_table_exists(table_name):
    """SQLite補助DBの存在確認。失敗時はFalseで返し、アプリを止めない。"""
    try:
        return _original_sqlite_table_exists(table_name)
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, f"sqlite_table_exists:{table_name}")
        except Exception:
            pass
        return False

def sqlite_table_row_count(table_name):
    """SQLite補助DBの件数確認。失敗時は0で返し、初期化・ログインを止めない。"""
    try:
        return _original_sqlite_table_row_count(table_name)
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, f"sqlite_table_row_count:{table_name}")
        except Exception:
            pass
        return 0

normalize_df_columns = db_engine.normalize_df_columns
prepare_sqlite_dataframe = db_engine.prepare_sqlite_dataframe

def db_write_dataframe(df, table_name, columns, *args, **kwargs):
    """低層SQLite書込の安全ラッパー。"""
    try:
        return _original_db_write_dataframe(df, table_name, columns, *args, **kwargs)
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, f"db_write_dataframe:{table_name}")
            _show_sqlite_backup_warning_once(e, table_name)
        except Exception:
            pass
        return False

def db_read_dataframe(table_name, columns, *args, **kwargs):
    """低層SQLite読込の安全ラッパー。"""
    try:
        return _original_db_read_dataframe(table_name, columns, *args, **kwargs)
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, f"db_read_dataframe:{table_name}")
            _show_sqlite_backup_warning_once(e, table_name)
        except Exception:
            pass
        return pd.DataFrame(columns=list(columns or []))

# =========================
# Supabase外部DB対応（Ver4.5）
# 正式方針：
#   Supabase正本：users／health_records／excretion_records／handover_logs／short_goal_checks
#   SQLite補助：その他テーブル＋Supabase成功後のミラー
#   保存方式：upsert（全削除→全保存は行わない）
# =========================
SUPABASE_CORE_TABLES = {
    "users": ["user_id"],
    "health_records": ["記録日", "user_id"],
    "excretion_records": ["記録日", "user_id", "時間帯"],
    "handover_logs": ["記録ID"],
    # Ver4.5完成版：短期目標・モニタリング系もSupabaseを正本にする
    "short_goal_checks": ["記録ID"],
    "short_term_goals": ["目標ID"],
    "monitoring_drafts": ["下書きID"],
}

SUPABASE_CORE_LABELS = {
    "users": "利用者マスタ",
    "health_records": "健康チェック",
    "excretion_records": "排泄チェック",
    "handover_logs": "業務全体申し送り",
    "short_goal_checks": "短期目標実施チェック",
    "short_term_goals": "短期目標マスタ",
    "monitoring_drafts": "モニタリング下書き",
}

_original_save_sqlite_table = db_engine.save_sqlite_table
_original_load_sqlite_table = db_engine.load_sqlite_table


def _sb_clean(value, default=""):
    try:
        if pd.isna(value):
            return default
    except Exception:
        pass
    text = str(value).strip()
    if text.lower() in ["nan", "none", "nat"]:
        return default
    return text


def _sb_json_safe(value):
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    try:
        if hasattr(value, "isoformat"):
            return value.isoformat()
    except Exception:
        pass
    try:
        if hasattr(value, "item"):
            return value.item()
    except Exception:
        pass
    return value


def _secret_get(container, key, default=""):
    """Streamlit Secretsのdict／AttrDict／通常属性を安全に読む。"""
    if container is None:
        return default
    try:
        value = container.get(key, default)
        if value not in [None, ""]:
            return value
    except Exception:
        pass
    try:
        value = container[key]
        if value not in [None, ""]:
            return value
    except Exception:
        pass
    try:
        value = getattr(container, key)
        if value not in [None, ""]:
            return value
    except Exception:
        pass
    return default


def _normalize_supabase_url(url: str) -> str:
    """
    Supabase URLを安全に正規化する。

    今回の事故防止ポイント：
    - /rest/v1/ まで貼っても自動で Project URL に戻す
    - .supabase.corest のような入力ミスを補正する
    - huufblmiqvloudeqctjp と huufblmiqvloudeqctjp の c/q 入れ替わりを補正する
    - 末尾スラッシュを削除する
    """
    if not url:
        return ""

    url = str(url).strip().strip('"').strip("'")
    url = url.replace(" ", "").replace("\n", "").replace("\r", "").replace("\t", "")

    # よくある誤入力: .co/rest が .corest になった場合
    url = url.replace(".supabase.corest", ".supabase.co/rest")

    # スキームがない場合の補完
    if url and not url.startswith(("http://", "https://")):
        url = "https://" + url

    # Data API URLをそのまま貼った場合でもProject URLに戻す
    if "/rest/v1" in url:
        url = url.split("/rest/v1")[0]

    url = url.rstrip("/")

    # 今回確認済みの正しいProject Ref
    correct_ref = "huufblmiqvloudeqctjp"
    known_bad_refs = {
        "huufblmiqvloudecqtjp": correct_ref,  # c と q の位置違い
    }
    for bad, good in known_bad_refs.items():
        url = url.replace(bad, good)

    return url


def _supabase_config():
    """
    Streamlit Secretsを柔軟に読む。

    推奨：
        [supabase]
        enabled = true
        url = "https://huufblmiqvloudeqctjp.supabase.co"
        key = "sb_secret_xxxxx"

    互換：
        SUPABASE_URL = "..."
        SUPABASE_KEY = "..."
    """
    secrets = None
    try:
        secrets = st.secrets if hasattr(st, "secrets") else None
    except Exception:
        secrets = None

    sb = _secret_get(secrets, "supabase", {})

    def pick(*keys):
        for key in keys:
            value = _secret_get(sb, key, "")
            if value not in [None, ""]:
                return str(value).strip().strip('"').strip("'")
        for key in keys:
            value = _secret_get(secrets, key, "")
            if value not in [None, ""]:
                return str(value).strip().strip('"').strip("'")
        return ""

    enabled_raw = pick("enabled", "SUPABASE_ENABLED")
    raw_url = pick("url", "SUPABASE_URL")
    key = pick("key", "service_role_key", "anon_key", "SUPABASE_KEY", "SUPABASE_SERVICE_ROLE_KEY", "SUPABASE_ANON_KEY")

    url = _normalize_supabase_url(raw_url)

    enabled = str(enabled_raw).lower() in ["1", "true", "yes", "on", "有効"]
    if not enabled and url and key:
        enabled = True

    return {
        "enabled": enabled,
        "url": url,
        "key": key,
        "raw_url": raw_url,
    }

def supabase_is_enabled():
    cfg = _supabase_config()
    return bool(cfg.get("enabled") and cfg.get("url") and cfg.get("key") and requests is not None)


def _supabase_headers(prefer="return=minimal"):
    cfg = _supabase_config()
    headers = {
        "apikey": cfg["key"],
        "Authorization": f"Bearer {cfg['key']}",
        "Content-Type": "application/json",
    }
    if prefer:
        headers["Prefer"] = prefer
    return headers


def _supabase_endpoint(table_name: str, query: str = ""):
    cfg = _supabase_config()
    return f"{cfg['url']}/rest/v1/{table_name}" + query


def _make_supabase_record_key(row: dict, table_name: str, unique_cols=None):
    cols = unique_cols or SUPABASE_CORE_TABLES.get(table_name, [])

    # 利用者マスタは user_id を最優先
    if table_name == "users":
        user_id = _sb_clean(row.get("user_id"))
        if user_id:
            return user_id
        user_name = _sb_clean(row.get("利用者名"))
        if user_name and "make_user_id_from_name" in globals():
            return make_user_id_from_name(user_name)

    # 健康・排泄は user_id が空なら利用者名でフォールバック
    if table_name in ["health_records", "excretion_records"]:
        if "user_id" in cols and not _sb_clean(row.get("user_id")):
            cols = ["記録日", "利用者名"] + (["時間帯"] if table_name == "excretion_records" else [])

    if not cols:
        cols = ["記録ID"] if "記録ID" in row else []

    parts = []
    for col in cols:
        value = _sb_clean(row.get(col))
        if col == "記録日":
            dt = pd.to_datetime(value, errors="coerce")
            value = dt.strftime("%Y-%m-%d") if not pd.isna(dt) else value
        parts.append(value)

    key = "__".join(parts).strip("_")
    return key or str(uuid.uuid4())


def _df_to_supabase_payload(df: pd.DataFrame, table_name: str, unique_cols=None):
    if df is None or df.empty:
        return []
    payload = []
    work = df.copy()
    for _, row in work.iterrows():
        data = {str(k): _sb_json_safe(v) for k, v in row.to_dict().items()}
        payload.append({
            "record_key": _make_supabase_record_key(data, table_name, unique_cols),
            "data": data,
            "updated_at": format_now_jst("%Y-%m-%dT%H:%M:%S+09:00") if "format_now_jst" in globals() else datetime.utcnow().isoformat(),
        })
    # 同じrecord_keyがある場合は後勝ちにしてupsert対象を整理
    dedup = {}
    for item in payload:
        dedup[item["record_key"]] = item
    return list(dedup.values())


# =========================
# Supabase読込の軽量化（Ver4.8）
# =========================
# 目的：
# - 毎回全件を読みに行かない
# - 今日〜直近7日など、画面表示に必要な期間だけ読む
# - 同じ条件の読込は30〜60秒キャッシュする
# - 保存・削除後は clear_hidamari_read_cache() でキャッシュを消す
DEFAULT_QUERY_CACHE_TTL_SEC = 60
DEFAULT_RECENT_DAYS = 7

def _date_to_iso(value):
    if value in [None, ""]:
        return ""
    try:
        dt = pd.to_datetime(value, errors="coerce")
        if pd.isna(dt):
            return ""
        return dt.strftime("%Y-%m-%d")
    except Exception:
        return ""

def recent_start_date(days=DEFAULT_RECENT_DAYS, base_date=None):
    base = base_date or today_jst()
    try:
        return base - timedelta(days=max(int(days), 1) - 1)
    except Exception:
        return today_jst() - timedelta(days=DEFAULT_RECENT_DAYS - 1)

def _normalize_supabase_df_from_rows(rows, columns=None):
    data_rows = []
    for item in rows or []:
        data = item.get("data") or {}
        if isinstance(data, dict):
            data_rows.append(data)
    df = pd.DataFrame(data_rows)
    if columns:
        for col in columns:
            if col not in df.columns:
                df[col] = ""
        df = df[list(columns)]
    return df

def _filter_df_by_date_range(df, date_col, start_date=None, end_date=None):
    if df is None or df.empty or not date_col or date_col not in df.columns:
        return df
    start_iso = _date_to_iso(start_date)
    end_iso = _date_to_iso(end_date)
    if not start_iso and not end_iso:
        return df
    work = df.copy()
    work[date_col] = pd.to_datetime(work[date_col], errors="coerce")
    if start_iso:
        start_dt = pd.to_datetime(start_iso, errors="coerce")
        if not pd.isna(start_dt):
            work = work[work[date_col].dt.date >= start_dt.date()]
    if end_iso:
        end_dt = pd.to_datetime(end_iso, errors="coerce")
        if not pd.isna(end_dt):
            work = work[work[date_col].dt.date <= end_dt.date()]
    return work

@cache_safe_master_read(ttl=DEFAULT_QUERY_CACHE_TTL_SEC)
def _supabase_read_table_cached(table_name: str, columns_tuple=(), date_field: str = "", start_iso: str = "", end_iso: str = "", limit: int = 0) -> pd.DataFrame:
    # PostgRESTのJSONBフィルタで data->>記録日 / data->>日付 を絞る。
    params = [("select", "record_key,data,updated_at")]
    if date_field and start_iso:
        params.append((f"data->>{date_field}", f"gte.{start_iso}"))
    if date_field and end_iso:
        params.append((f"data->>{date_field}", f"lte.{end_iso}"))
    params.append(("order", "updated_at.desc"))
    if limit and int(limit) > 0:
        params.append(("limit", str(int(limit))))

    with perf_timer("supabase_read", f"{table_name} {start_iso or ''}-{end_iso or ''}"):
        res = requests.get(
            _supabase_endpoint(table_name),
            headers=_supabase_headers(prefer=""),
            params=params,
            timeout=20,
        )
    res.raise_for_status()
    return _normalize_supabase_df_from_rows(res.json() or [], list(columns_tuple))

def supabase_read_table(table_name: str, columns=None, date_field: str = "", start_date=None, end_date=None, limit: int = 0) -> pd.DataFrame:
    if not supabase_is_enabled() or table_name not in SUPABASE_CORE_TABLES:
        df = _original_load_sqlite_table(table_name, columns or [])
        return _filter_df_by_date_range(df, date_field, start_date, end_date)
    try:
        columns_tuple = tuple(columns or [])
        start_iso = _date_to_iso(start_date)
        end_iso = _date_to_iso(end_date)
        df = _supabase_read_table_cached(table_name, columns_tuple, date_field or "", start_iso, end_iso, int(limit or 0))

        # Supabase正本化直後の安全措置：
        # Supabase側が空で、SQLiteミラー側に旧データがある場合はSQLiteを返す。
        # 期間指定がある場合はSQLite側も同じ期間に絞る。
        if df.empty:
            try:
                local_df = _original_load_sqlite_table(table_name, columns or [])
                local_df = _filter_df_by_date_range(local_df, date_field, start_date, end_date)
                if local_df is not None and not local_df.empty:
                    return local_df
            except Exception:
                pass
        return df
    except Exception as e:
        try:
            st.warning(f"Supabase読込に失敗しました。SQLite補助DBが使える場合のみ読み込みます：{table_name} / {e}")
        except Exception:
            pass
        try:
            df = _original_load_sqlite_table(table_name, columns or [])
            return _filter_df_by_date_range(df, date_field, start_date, end_date)
        except Exception as e2:
            try:
                _mark_sqlite_backup_error(e2, table_name)
            except Exception:
                pass
            return pd.DataFrame(columns=list(columns or []))


def supabase_upsert_table(df: pd.DataFrame, table_name: str, columns=None, unique_cols=None) -> bool:
    """
    Ver4.5正式方式：upsert。
    全削除→全保存は行わず、同じrecord_keyは更新、なければ追加する。
    """
    if not supabase_is_enabled() or table_name not in SUPABASE_CORE_TABLES:
        return False
    try:
        payload = _df_to_supabase_payload(df, table_name, unique_cols=unique_cols)
        if not payload:
            return True

        post_url = _supabase_endpoint(table_name, "?on_conflict=record_key")
        headers = _supabase_headers(prefer="resolution=merge-duplicates,return=minimal")

        for i in range(0, len(payload), 300):
            chunk = payload[i:i + 300]
            post_res = requests.post(post_url, headers=headers, json=chunk, timeout=30)
            post_res.raise_for_status()
        try:
            clear_hidamari_read_cache(f"Supabase保存:{table_name}")
        except Exception:
            pass
        return True
    except Exception as e:
        try:
            st.error(f"Supabase upsert保存に失敗しました。ローカルSQLiteへ保存します：{table_name} / {e}")
        except Exception:
            pass
        return False


# 旧関数名との互換。内部はupsert方式。
def supabase_replace_table(df: pd.DataFrame, table_name: str, columns=None, unique_cols=None) -> bool:
    return supabase_upsert_table(df, table_name, columns=columns, unique_cols=unique_cols)



# =========================
# SQLite破損DB自動隔離（Ver4.5.4）
# Streamlit Cloud上のSQLite補助DBが破損した場合、
# Supabase正本を守るため、壊れたDBを退避リネームして新規作成可能にする。
# =========================
SQLITE_CORRUPTION_KEYWORDS = [
    "database disk image is malformed",
    "file is not a database",
    "database is malformed",
    "malformed database schema",
]

def is_sqlite_corruption_error(e) -> bool:
    msg = str(e).lower()
    return any(k in msg for k in SQLITE_CORRUPTION_KEYWORDS)

def quarantine_corrupt_sqlite_db(reason=""):
    """
    壊れたSQLite補助DBを隔離する。
    Supabaseを正本にしているため、SQLite補助DBは再作成対象とする。
    """
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
    except Exception:
        pass

    try:
        if st.session_state.get("sqlite_quarantine_done"):
            return
    except Exception:
        pass

    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    targets = [
        HIDAMARI_DB_FILE,
        Path(str(HIDAMARI_DB_FILE) + "-wal"),
        Path(str(HIDAMARI_DB_FILE) + "-shm"),
    ]

    moved = []
    for src in targets:
        try:
            if src.exists():
                dst = src.with_name(f"{src.name}.corrupt_{timestamp}")
                src.replace(dst)
                moved.append(f"{src.name} -> {dst.name}")
        except Exception:
            pass

    try:
        st.session_state["sqlite_quarantine_done"] = True
        st.session_state["sqlite_backup_available"] = False
        st.session_state["sqlite_backup_last_error"] = f"SQLite補助DBを隔離しました: {reason}"
        if moved and not st.session_state.get("sqlite_quarantine_notice_shown"):
            st.warning("SQLite補助DBの破損を検出したため、壊れたDBを隔離しました。Supabase正本で処理を継続します。")
            st.session_state["sqlite_quarantine_notice_shown"] = True
    except Exception:
        pass


# =========================
# SQLiteバックアップ安全化（Ver4.5.2）
# Supabaseを正本にしているため、SQLiteミラー／バックアップ側のエラーで
# アプリ本体を停止させない。Streamlit Cloudの一時ファイル・DBロック・
# 破損・PRAGMA相性問題を吸収して、空DFまたはSupabase側で継続する。
# =========================
def _safe_empty_df(columns=None):
    try:
        return pd.DataFrame(columns=list(columns or []))
    except Exception:
        return pd.DataFrame()


def _mark_sqlite_backup_error(e, table_name=""):
    try:
        if is_sqlite_corruption_error(e):
            quarantine_corrupt_sqlite_db(f"{table_name}: {e}" if table_name else str(e))
    except Exception:
        pass
    try:
        st.session_state["sqlite_backup_available"] = False
        st.session_state["sqlite_backup_last_error"] = f"{table_name}: {e}" if table_name else str(e)
    except Exception:
        pass


def _show_sqlite_backup_warning_once(e, table_name=""):
    try:
        # 復元処理中は、監査ログ・権限・履歴テーブルの再初期化失敗を
        # 画面上の「エラー」に見せない。復元本体は別メッセージで結果表示する。
        if st.session_state.get("restore_suppress_sqlite_aux_warnings"):
            st.session_state[f"sqlite_backup_warning_suppressed_{table_name}"] = str(e)
            return
        key = f"sqlite_backup_warning_shown_{table_name}"
        if not st.session_state.get(key):
            st.warning(
                f"SQLiteバックアップDBの処理に失敗しました。Supabase正本で処理を継続します。"
                f"対象: {table_name or 'SQLite'}"
            )
            st.session_state[key] = True
    except Exception:
        pass


def save_sqlite_table(df, table_name, columns, date_cols=None, unique_cols=None, sort_cols=None):
    """
    Supabase正本対象テーブルは Supabase を正本としてupsert保存。
    SQLiteはミラー／バックアップ扱いにし、SQLite失敗ではアプリを止めない。
    """
    supabase_ok = False
    if table_name in SUPABASE_CORE_TABLES:
        supabase_ok = supabase_upsert_table(df, table_name, columns=columns, unique_cols=unique_cols)

    try:
        result = _original_save_sqlite_table(
            df,
            table_name,
            columns,
            date_cols=date_cols,
            unique_cols=unique_cols,
            sort_cols=sort_cols,
        )
        try:
            st.session_state["sqlite_backup_available"] = True
        except Exception:
            pass
        return result
    except Exception as e:
        _mark_sqlite_backup_error(e, table_name)
        _show_sqlite_backup_warning_once(e, table_name)
        # Supabase正本テーブルは、Supabase保存が成功していれば成功扱いで続行。
        if table_name in SUPABASE_CORE_TABLES and supabase_ok:
            return True
        # 非正本テーブルも、初期化・ログ・設定系のSQLite失敗で画面全体を落とさない。
        return False


def load_sqlite_table(table_name, columns, date_cols=None):
    """
    Supabase正本対象テーブルは Supabase 優先で読む。
    SQLite読込失敗時は空DFで返し、起動・ログイン画面を継続する。
    """
    if table_name in SUPABASE_CORE_TABLES and supabase_is_enabled():
        try:
            return supabase_read_table(table_name, columns=columns)
        except Exception as e:
            _mark_sqlite_backup_error(e, table_name)
            _show_sqlite_backup_warning_once(e, table_name)

    try:
        return _original_load_sqlite_table(table_name, columns, date_cols=date_cols)
    except Exception as e:
        _mark_sqlite_backup_error(e, table_name)
        _show_sqlite_backup_warning_once(e, table_name)
        return _safe_empty_df(columns)


# =========================
# 安全キャッシュ化（速度改善より、壊さない改善）
# =========================
# 方針：
# - 入力・削除・Supabase/SQLite同期・写真・AI分析には触れない
# - 変化が少ない「読むだけのマスタ系」だけを短時間キャッシュする
# - 保存・削除後はキャッシュを全消去して、古い表示を残しにくくする
SAFE_READ_CACHE_TTL_SEC = 300

def get_supabase_storage_status():
    cfg = _supabase_config()
    if requests is None:
        return "Supabase無効：requestsライブラリが利用できません。"
    if not cfg.get("url") or not cfg.get("key"):
        return "Supabase未設定：Streamlit Secrets に url と key を登録してください。"
    if not cfg.get("enabled"):
        return "Supabase未有効：enabled=true を設定してください。"
    try:
        checks = []
        for table_name in SUPABASE_CORE_TABLES.keys():
            url = _supabase_endpoint(table_name, "?select=record_key&limit=1")
            res = requests.get(url, headers=_supabase_headers(prefer=""), timeout=10)
            if res.status_code not in [200, 206]:
                return f"Supabase接続注意：{table_name} / HTTP {res.status_code} / {res.text[:160]}"
            checks.append(f"{SUPABASE_CORE_LABELS.get(table_name, table_name)}OK")
        return "Supabase接続OK：利用者マスタ・健康チェック・排泄チェック・申し送り・短期目標実施チェック・短期目標マスタ・モニタリング下書きは外部DB保存です。"
    except Exception as e:
        return f"Supabase接続エラー：{e}"


def get_supabase_diagnostic_rows():
    """管理画面表示用の簡易診断。"""
    cfg = _supabase_config()
    rows = [
        {"項目": "requests", "状態": "OK" if requests is not None else "NG", "詳細": "requests利用可能" if requests is not None else "requestsが利用できません"},
        {"項目": "enabled", "状態": "OK" if cfg.get("enabled") else "NG", "詳細": str(cfg.get("enabled"))},
        {"項目": "url", "状態": "OK" if cfg.get("url") else "NG", "詳細": cfg.get("url", "")},
        {"項目": "raw_url", "状態": "参考", "詳細": cfg.get("raw_url", "")},
        {"項目": "project_ref", "状態": "OK" if "huufblmiqvloudeqctjp" in cfg.get("url", "") else "確認", "詳細": "正：huufblmiqvloudeqctjp ／ 誤：huufblmiqvloudecqtjp"},
        {"項目": "key", "状態": "OK" if cfg.get("key") else "NG", "詳細": "設定済み" if cfg.get("key") else "未設定"},
    ]
    if supabase_is_enabled():
        for table_name in SUPABASE_CORE_TABLES.keys():
            try:
                url = _supabase_endpoint(table_name, "?select=record_key&limit=1")
                res = requests.get(url, headers=_supabase_headers(prefer=""), timeout=10)
                ok = res.status_code in [200, 206]
                rows.append({
                    "項目": table_name,
                    "状態": "OK" if ok else "NG",
                    "詳細": f"HTTP {res.status_code}" if not ok else "接続OK",
                })
            except Exception as e:
                rows.append({"項目": table_name, "状態": "NG", "詳細": str(e)})
    return pd.DataFrame(rows)


def get_supabase_create_table_sql():
    return """
-- Supabase SQL Editorで実行してください
create table if not exists public.users (
  record_key text primary key,
  data jsonb not null default '{}'::jsonb,
  updated_at timestamptz default now()
);

create table if not exists public.health_records (
  record_key text primary key,
  data jsonb not null default '{}'::jsonb,
  updated_at timestamptz default now()
);

create table if not exists public.excretion_records (
  record_key text primary key,
  data jsonb not null default '{}'::jsonb,
  updated_at timestamptz default now()
);

create table if not exists public.handover_logs (
  record_key text primary key,
  data jsonb not null default '{}'::jsonb,
  updated_at timestamptz default now()
);

create table if not exists public.short_goal_checks (
  record_key text primary key,
  data jsonb not null default '{}'::jsonb,
  updated_at timestamptz default now()
);

create table if not exists public.short_term_goals (
  record_key text primary key,
  data jsonb not null default '{}'::jsonb,
  updated_at timestamptz default now()
);

create table if not exists public.monitoring_drafts (
  record_key text primary key,
  data jsonb not null default '{}'::jsonb,
  updated_at timestamptz default now()
);
""".strip()


# =========================
# 共通削除関数（Ver4.2）
# SQLite DELETE + Supabase DELETE + 監査ログ + 画面再読み込みのための共通基盤
# upsert保存では既存行が消えないため、削除は必ずこの直接DELETE系を使う。
# =========================
def _sql_quote_identifier(name: str) -> str:
    return '"' + str(name).replace('"', '""') + '"'


def _normalize_delete_value(value, col_name=""):
    if value is None:
        return ""
    try:
        if pd.isna(value):
            return ""
    except Exception:
        pass
    if isinstance(value, (datetime, date)):
        if col_name in ["記録日", "日付", "開始日", "終了予定日", "作成日"]:
            return value.strftime("%Y-%m-%d")
        return value.isoformat()
    text = str(value).strip()
    if col_name in ["記録日", "日付", "開始日", "終了予定日", "作成日"]:
        dt = pd.to_datetime(text, errors="coerce")
        if not pd.isna(dt):
            return dt.strftime("%Y-%m-%d")
    return text


def sqlite_delete_records(table_name: str, where: dict) -> int:
    """SQLiteから条件一致する行を直接DELETEする。whereはAND条件。"""
    if not table_name or not where:
        return 0
    validate_sqlite_identifier(table_name)
    clauses = []
    params = []
    date_cols = {"記録日", "日付", "開始日", "終了予定日", "作成日"}
    for col, value in where.items():
        col_name = str(col)
        q_col = _sql_quote_identifier(col_name)
        norm_value = _normalize_delete_value(value, col_name)
        if col_name in date_cols:
            clauses.append(f"date({q_col}) = date(?)")
        else:
            clauses.append(f"trim(CAST({q_col} AS TEXT)) = trim(CAST(? AS TEXT))")
        params.append(norm_value)
    sql = f"DELETE FROM {_sql_quote_identifier(table_name)} WHERE " + " AND ".join(clauses)
    with hidamari_write_transaction() as conn:
        cur = conn.execute(sql, params)
        deleted = int(cur.rowcount or 0)
    return deleted


def supabase_delete_by_record_keys(table_name: str, record_keys) -> int:
    """SupabaseのJSONB汎用テーブルからrecord_keyで直接DELETEする。"""
    if not supabase_is_enabled() or table_name not in SUPABASE_CORE_TABLES:
        return 0
    if requests is None:
        return 0
    deleted = 0
    keys = []
    for key in record_keys or []:
        key = clean_text(key) if "clean_text" in globals() else str(key).strip()
        if key and key not in keys:
            keys.append(key)
    for key in keys:
        try:
            url = _supabase_endpoint(table_name, f"?record_key=eq.{requests.utils.quote(str(key), safe='')}")
            res = requests.delete(url, headers=_supabase_headers(prefer="return=minimal"), timeout=20)
            if res.status_code in [200, 202, 204]:
                deleted += 1
            else:
                try:
                    st.warning(f"Supabase削除注意：{table_name} / {key} / HTTP {res.status_code} / {res.text[:160]}")
                except Exception:
                    pass
        except Exception as e:
            try:
                st.warning(f"Supabase削除に失敗しました：{table_name} / {key} / {e}")
            except Exception:
                pass
    return deleted


def supabase_find_record_keys_by_data_field(table_name: str, field_name: str, value) -> list:
    """
    Supabase JSONB汎用テーブルで、data内の値からrecord_keyを探す。
    旧データや一部保存時に record_key と 記録ID がずれていても削除できるようにする保険。
    """
    if not supabase_is_enabled() or table_name not in SUPABASE_CORE_TABLES:
        return []
    if requests is None:
        return []
    target = clean_text(value) if "clean_text" in globals() else str(value or "").strip()
    if not target:
        return []
    try:
        # 件数が小規模施設想定なので、確実性優先でJSONを読み、Python側で照合する。
        url = _supabase_endpoint(table_name, "?select=record_key,data&limit=10000")
        res = requests.get(url, headers=_supabase_headers(prefer=""), timeout=30)
        res.raise_for_status()
        rows = res.json() or []
        keys = []
        for item in rows:
            data = item.get("data") or {}
            if not isinstance(data, dict):
                continue
            data_value = clean_text(data.get(field_name)) if "clean_text" in globals() else str(data.get(field_name, "")).strip()
            if data_value == target:
                record_key = clean_text(item.get("record_key")) if "clean_text" in globals() else str(item.get("record_key", "")).strip()
                if record_key and record_key not in keys:
                    keys.append(record_key)
        return keys
    except Exception as e:
        try:
            st.warning(f"Supabase削除キー検索に失敗しました：{table_name} / {field_name}={target} / {e}")
        except Exception:
            pass
        return []


def _build_supabase_delete_key(table_name: str, row: dict, unique_cols=None) -> str:
    try:
        return _make_supabase_record_key(row, table_name, unique_cols=unique_cols)
    except Exception:
        return ""


def delete_record_common(table_name: str, sqlite_where: dict, supabase_keys=None, operation="削除", target_key="", summary="") -> dict:
    """削除共通関数。SQLiteとSupabaseの両方を直接DELETEし、監査ログを残す。"""
    result = {"sqlite_deleted": 0, "supabase_deleted": 0, "ok": False, "error": ""}
    try:
        result["sqlite_deleted"] = sqlite_delete_records(table_name, sqlite_where)
        result["supabase_deleted"] = supabase_delete_by_record_keys(table_name, supabase_keys or [])
        result["ok"] = result["sqlite_deleted"] > 0 or result["supabase_deleted"] > 0
        try:
            add_audit_log(
                operation,
                table_name,
                target_key,
                f"{summary} / SQLite削除:{result['sqlite_deleted']} / Supabase削除:{result['supabase_deleted']}",
            )
        except Exception:
            pass
    except Exception as e:
        result["error"] = str(e)
        try:
            add_audit_log(f"{operation}失敗", table_name, target_key, str(e))
        except Exception:
            pass
    return result


def delete_business_handover_record(record_id: str, source="") -> dict:
    """
    業務全体申し送りを削除する。
    Supabase側は record_key=記録ID の想定だが、旧データ等でずれている可能性があるため、
    data["記録ID"] からもrecord_keyを探索して直接DELETEする。
    """
    record_id = clean_text(record_id) if "clean_text" in globals() else str(record_id).strip()
    supabase_keys = [record_id]
    try:
        for key in supabase_find_record_keys_by_data_field(SQLITE_TABLE_HANDOVER, "記録ID", record_id):
            if key and key not in supabase_keys:
                supabase_keys.append(key)
    except Exception:
        pass

    result = delete_record_common(
        SQLITE_TABLE_HANDOVER,
        {"記録ID": record_id},
        supabase_keys=supabase_keys,
        operation="申し送り削除",
        target_key=record_id,
        summary=source or "業務全体申し送りを削除",
    )

    # SQLite側の削除確認。直接DELETEで消えない場合の保険として、SQLiteだけ再保存する。
    try:
        local_df = _original_load_sqlite_table(SQLITE_TABLE_HANDOVER, BUSINESS_HANDOVER_COLUMNS)
        if not local_df.empty and "記録ID" in local_df.columns:
            before = len(local_df)
            local_df = local_df[local_df["記録ID"].astype(str).str.strip() != str(record_id).strip()].copy()
            if len(local_df) != before:
                _original_save_sqlite_table(
                    local_df,
                    SQLITE_TABLE_HANDOVER,
                    BUSINESS_HANDOVER_COLUMNS,
                    date_cols=["日付"],
                    unique_cols=["記録ID"],
                    sort_cols=["記録日時"],
                )
                result["sqlite_deleted"] = max(int(result.get("sqlite_deleted", 0) or 0), before - len(local_df))
                result["ok"] = True
    except Exception:
        pass

    return result


def delete_health_record(record_date, user_name, user_id="", source="") -> dict:
    date_text = _normalize_delete_value(record_date, "記録日")
    user_name = clean_text(user_name) if "clean_text" in globals() else str(user_name).strip()
    user_id = ensure_user_id_value(user_id, user_name) if "ensure_user_id_value" in globals() else str(user_id or "").strip()
    row = {"記録日": date_text, "利用者名": user_name, "user_id": user_id}
    keys = [
        _build_supabase_delete_key(SQLITE_TABLE_HEALTH, row, unique_cols=["記録日", "利用者名"]),
        _build_supabase_delete_key(SQLITE_TABLE_HEALTH, row, unique_cols=["記録日", "user_id"]),
    ]
    return delete_record_common(
        SQLITE_TABLE_HEALTH,
        {"記録日": date_text, "利用者名": user_name},
        supabase_keys=keys,
        operation="健康チェック削除",
        target_key=f"{date_text}_{user_name}",
        summary=source or "健康チェックを削除",
    )


def delete_excretion_record(record_date, user_name, slot, user_id="", source="") -> dict:
    date_text = _normalize_delete_value(record_date, "記録日")
    user_name = clean_text(user_name) if "clean_text" in globals() else str(user_name).strip()
    slot = clean_text(slot) if "clean_text" in globals() else str(slot).strip()
    user_id = ensure_user_id_value(user_id, user_name) if "ensure_user_id_value" in globals() else str(user_id or "").strip()
    row = {"記録日": date_text, "利用者名": user_name, "user_id": user_id, "時間帯": slot}
    keys = [
        _build_supabase_delete_key(SQLITE_TABLE_EXCRETION, row, unique_cols=["記録日", "利用者名", "時間帯"]),
        _build_supabase_delete_key(SQLITE_TABLE_EXCRETION, row, unique_cols=["記録日", "user_id", "時間帯"]),
    ]
    return delete_record_common(
        SQLITE_TABLE_EXCRETION,
        {"記録日": date_text, "利用者名": user_name, "時間帯": slot},
        supabase_keys=keys,
        operation="排泄チェック削除",
        target_key=f"{date_text}_{user_name}_{slot}",
        summary=source or "排泄チェックを削除",
    )


def delete_short_goal_check_records(record_ids, source="") -> dict:
    """短期目標実施チェックをSQLiteミラーとSupabase正本の両方から直接削除する。"""
    ids = []
    for rid in record_ids or []:
        rid = clean_text(rid) if "clean_text" in globals() else str(rid).strip()
        if rid and rid not in ids:
            ids.append(rid)

    total_sqlite = 0
    total_supabase = 0
    errors = []
    for rid in ids:
        try:
            total_sqlite += sqlite_delete_records(SQLITE_TABLE_SHORT_GOAL_CHECKS, {"記録ID": rid})
        except Exception as e:
            errors.append(f"SQLite {rid}: {e}")
        try:
            keys = [rid]
            for key in supabase_find_record_keys_by_data_field(SQLITE_TABLE_SHORT_GOAL_CHECKS, "記録ID", rid):
                if key and key not in keys:
                    keys.append(key)
            total_supabase += supabase_delete_by_record_keys(SQLITE_TABLE_SHORT_GOAL_CHECKS, keys)
        except Exception as e:
            errors.append(f"Supabase {rid}: {e}")

    try:
        add_audit_log(
            "短期目標実施チェック削除",
            SQLITE_TABLE_SHORT_GOAL_CHECKS,
            ",".join(ids[:10]),
            f"{source or '実施履歴を削除'} / SQLite削除:{total_sqlite} / Supabase削除:{total_supabase} / 対象:{len(ids)}件",
        )
    except Exception:
        pass
    return {"sqlite_deleted": total_sqlite, "supabase_deleted": total_supabase, "ok": (total_sqlite > 0 or total_supabase > 0), "error": " / ".join(errors)}




def delete_short_goal_master_records(goal_ids, source="") -> dict:
    """短期目標マスタをSQLiteミラーとSupabase正本の両方から直接削除する。"""
    ids = []
    for gid in goal_ids or []:
        gid = clean_text(gid) if "clean_text" in globals() else str(gid).strip()
        if gid and gid not in ids:
            ids.append(gid)
    total_sqlite = 0
    total_supabase = 0
    errors = []
    for gid in ids:
        try:
            total_sqlite += sqlite_delete_records(SQLITE_TABLE_SHORT_GOAL_MASTER, {"目標ID": gid})
        except Exception as e:
            errors.append(f"SQLite {gid}: {e}")
        try:
            keys = [gid]
            for key in supabase_find_record_keys_by_data_field(SQLITE_TABLE_SHORT_GOAL_MASTER, "目標ID", gid):
                if key and key not in keys:
                    keys.append(key)
            total_supabase += supabase_delete_by_record_keys(SQLITE_TABLE_SHORT_GOAL_MASTER, keys)
        except Exception as e:
            errors.append(f"Supabase {gid}: {e}")
    try:
        add_audit_log(
            "短期目標マスタ削除",
            SQLITE_TABLE_SHORT_GOAL_MASTER,
            ",".join(ids[:10]),
            f"{source or '短期目標マスタを削除'} / SQLite削除:{total_sqlite} / Supabase削除:{total_supabase} / 対象:{len(ids)}件",
        )
    except Exception:
        pass
    if total_sqlite > 0 or total_supabase > 0:
        clear_hidamari_read_cache("短期目標マスタ削除")
    return {"sqlite_deleted": total_sqlite, "supabase_deleted": total_supabase, "ok": (total_sqlite > 0 or total_supabase > 0), "error": " / ".join(errors)}


def delete_monitoring_draft_records(draft_ids, source="") -> dict:
    """モニタリング下書きをSQLiteミラーとSupabase正本の両方から直接削除する。"""
    ids = []
    for did in draft_ids or []:
        did = clean_text(did) if "clean_text" in globals() else str(did).strip()
        if did and did not in ids:
            ids.append(did)
    total_sqlite = 0
    total_supabase = 0
    errors = []
    for did in ids:
        try:
            total_sqlite += sqlite_delete_records(SQLITE_TABLE_MONITORING_DRAFTS, {"下書きID": did})
        except Exception as e:
            errors.append(f"SQLite {did}: {e}")
        try:
            keys = [did]
            for key in supabase_find_record_keys_by_data_field(SQLITE_TABLE_MONITORING_DRAFTS, "下書きID", did):
                if key and key not in keys:
                    keys.append(key)
            total_supabase += supabase_delete_by_record_keys(SQLITE_TABLE_MONITORING_DRAFTS, keys)
        except Exception as e:
            errors.append(f"Supabase {did}: {e}")
    try:
        add_audit_log(
            "モニタリング下書き削除",
            SQLITE_TABLE_MONITORING_DRAFTS,
            ",".join(ids[:10]),
            f"{source or 'モニタリング下書きを削除'} / SQLite削除:{total_sqlite} / Supabase削除:{total_supabase} / 対象:{len(ids)}件",
        )
    except Exception:
        pass
    return {"sqlite_deleted": total_sqlite, "supabase_deleted": total_supabase, "ok": (total_sqlite > 0 or total_supabase > 0), "error": " / ".join(errors)}

def show_delete_result_and_rerun(result: dict, success_message="削除しました。"):
    if result.get("error"):
        st.error(f"削除に失敗しました：{result.get('error')}")
        return
    if result.get("ok"):
        st.success(success_message)
        st.rerun()
    else:
        st.error("削除対象が見つかりません。")


def initialize_sqlite_engine():
    """SQLiteバックアップDBの初期化。失敗してもSupabase正本で起動継続する。"""
    try:
        result = db_engine.initialize_sqlite_engine()
        try:
            st.session_state["sqlite_backup_available"] = True
        except Exception:
            pass
        return result
    except Exception as e:
        _mark_sqlite_backup_error(e, "initialize_sqlite_engine")
        # 破損DBを隔離できた場合は、1回だけ新規DBとして再初期化を試す
        try:
            if is_sqlite_corruption_error(e):
                result = db_engine.initialize_sqlite_engine()
                st.session_state["sqlite_backup_available"] = True
                return result
        except Exception as e2:
            _mark_sqlite_backup_error(e2, "initialize_sqlite_engine_retry")
        _show_sqlite_backup_warning_once(e, "initialize_sqlite_engine")
        return None

def run_db_integrity_check(auto_repair: bool = True) -> dict:
    global DB_LAST_INTEGRITY_RESULT
    try:
        DB_LAST_INTEGRITY_RESULT = db_engine.run_db_integrity_check(auto_repair=auto_repair)
    except Exception as e:
        _mark_sqlite_backup_error(e, "integrity_check")
        DB_LAST_INTEGRITY_RESULT = {"ok": False, "error": str(e), "message": "SQLiteバックアップDBの整合性確認に失敗しました。Supabase正本で継続中です。"}
    return DB_LAST_INTEGRITY_RESULT

def get_db_integrity_status_text() -> str:
    return db_engine.get_db_integrity_status_text()



# =========================
# 商品化向け：設定系SQLite一元管理
# =========================
def ensure_app_settings_table():
    """JSON/Excelへ散らばりやすい設定をSQLiteへ集約するための共通テーブルを用意する。"""
    try:
        if not sqlite_table_exists(SQLITE_TABLE_APP_SETTINGS):
            db_write_dataframe(pd.DataFrame(columns=APP_SETTING_COLUMNS), SQLITE_TABLE_APP_SETTINGS, APP_SETTING_COLUMNS, unique_cols=["設定キー"])
    except Exception:
        # 初期化途中でもアプリ全体を止めない
        pass


def _json_dumps_safe(value) -> str:
    try:
        return json.dumps(value, ensure_ascii=False)
    except Exception:
        return json.dumps(str(value), ensure_ascii=False)


def _json_loads_safe(value, default=None):
    if value is None or value == "":
        return default
    try:
        return json.loads(value)
    except Exception:
        return default


def get_app_setting(setting_key, default=None):
    """SQLite app_settings から設定値を取得する。値はJSONとして保存・復元する。"""
    setting_key = clean_text(setting_key)
    if not setting_key:
        return default
    try:
        ensure_app_settings_table()
        df = load_sqlite_table(SQLITE_TABLE_APP_SETTINGS, APP_SETTING_COLUMNS)
        hit = df[df["設定キー"].astype(str) == setting_key]
        if hit.empty:
            return default
        raw = hit.iloc[-1].get("設定値", "")
        return _json_loads_safe(raw, default)
    except Exception:
        return default


def set_app_setting(setting_key, value, category="一般設定", description=""):
    """SQLite app_settings へ設定値を保存する。"""
    setting_key = clean_text(setting_key)
    if not setting_key:
        return
    try:
        ensure_app_settings_table()
        df = load_sqlite_table(SQLITE_TABLE_APP_SETTINGS, APP_SETTING_COLUMNS)
        df = df[df["設定キー"].astype(str) != setting_key].copy()
        row = {
            "設定キー": setting_key,
            "設定値": _json_dumps_safe(value),
            "分類": clean_text(category, "一般設定"),
            "説明": clean_text(description),
            "更新日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
            "更新者": current_login_user() if "current_login_user" in globals() else "",
        }
        df = pd.concat([df, pd.DataFrame([row], columns=APP_SETTING_COLUMNS)], ignore_index=True)
        save_sqlite_table(df, SQLITE_TABLE_APP_SETTINGS, APP_SETTING_COLUMNS, unique_cols=["設定キー"])
    except Exception:
        pass


def delete_app_setting(setting_key):
    setting_key = clean_text(setting_key)
    try:
        ensure_app_settings_table()
        df = load_sqlite_table(SQLITE_TABLE_APP_SETTINGS, APP_SETTING_COLUMNS)
        df = df[df["設定キー"].astype(str) != setting_key].copy()
        save_sqlite_table(df, SQLITE_TABLE_APP_SETTINGS, APP_SETTING_COLUMNS, unique_cols=["設定キー"])
    except Exception:
        pass


def migrate_json_file_setting_to_db(setting_key, json_path, category="移行設定", default=None):
    """
    旧JSONファイルからSQLiteへ初回移行する。
    DB側に既に値があれば上書きしない。
    """
    existing = get_app_setting(setting_key, None)
    if existing is not None:
        return existing
    value = default
    try:
        path = Path(json_path)
        if path.exists():
            value = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        value = default
    set_app_setting(setting_key, value, category=category, description=f"{Path(json_path).name} から移行")
    return value


def get_all_app_settings_df():
    ensure_app_settings_table()
    return load_sqlite_table(SQLITE_TABLE_APP_SETTINGS, APP_SETTING_COLUMNS)


def initialize_default_app_settings():
    """商品化前提の標準設定をSQLiteへ初期投入する。既存設定は維持する。"""
    ensure_app_settings_table()

    # メニューカテゴリ設定：旧JSONがあれば初回移行
    try:
        if "MENU_CATEGORY_SETTINGS_FILE" in globals():
            migrate_json_file_setting_to_db(
                "menu_category_settings_all",
                MENU_CATEGORY_SETTINGS_FILE,
                category="メニュー設定",
                default={},
            )
    except Exception:
        pass

    # 自分専用ダッシュボード設定：旧JSONがあれば初回移行
    try:
        if "DASHBOARD_SETTINGS_FILE" in globals():
            migrate_json_file_setting_to_db(
                "dashboard_settings_all",
                DASHBOARD_SETTINGS_FILE,
                category="ダッシュボード設定",
                default={},
            )
    except Exception:
        pass

    if get_app_setting("ui_settings", None) is None:
        set_app_setting(
            "ui_settings",
            {
                "テーマ": "ひだまり標準",
                "iPad最適化": True,
                "ボタン大型化": True,
                "カード表示": True,
                "フォント倍率": 1.0,
            },
            category="UI設定",
            description="画面表示・iPad対応の基本設定",
        )

    if get_app_setting("color_settings", None) is None:
        set_app_setting(
            "color_settings",
            {
                "staff_bg": "#FFFDF7",
                "staff_accent": "#C9705C",
                "admin_bg": "#F6F8F7",
                "admin_accent": "#2F6F5E",
                "alert": "#C9705C",
                "success": "#2F6F5E",
            },
            category="色設定",
            description="ブランドカラー・注意色・管理者色の設定",
        )

    if get_app_setting("life_settings", None) is None:
        set_app_setting(
            "life_settings",
            {
                "対象月初期値": "当月",
                "LIFE不足表示": True,
                "CSV出力前確認": True,
                "診断表現を避ける": True,
                "AIは整理係": True,
            },
            category="LIFE設定",
            description="LIFE管理・CSV出力・AI整理に関する設定",
        )

    if get_app_setting("facility_settings", None) is None:
        set_app_setting(
            "facility_settings",
            {
                "施設名": "ひだまり",
                "事業種別": "小規模介護施設",
                "定員": "",
                "所在地": "",
                "管理者名": "",
                "連絡先": "",
            },
            category="施設設定",
            description="施設名・管理者名・帳票表示用の基本設定",
        )

    if get_app_setting("photo_storage_settings", None) is None:
        set_app_setting(
            "photo_storage_settings",
            {
                "auto_compress": True,
                "max_kb": 300,
                "max_width": 800,
                "retention_days": 180,
                "backup_before_delete": True,
            },
            category="写真設定",
            description="申し送り写真の自動圧縮・保存期間・削除前バックアップ設定",
        )





def get_storage_unification_status():
    """商品版向け：保存先の簡易表示用。"""
    if "supabase_is_enabled" in globals() and supabase_is_enabled():
        return {
            "正データ": "Supabase（利用者マスタ・健康チェック・排泄チェック・申し送り・短期目標実施チェック・短期目標マスタ・モニタリング下書き）／SQLite（その他）",
            "Excel保存": "廃止（ダウンロード出力のみ）",
            "JSON保存": "廃止（app_settingsテーブルへ統合）",
            "バックアップ": "主要7機能はSupabase正本＋SQLiteミラー／その他はSQLite DB + 添付ファイル",
        }
    return {
        "正データ": "SQLite",
        "Excel保存": "廃止（ダウンロード出力のみ）",
        "JSON保存": "廃止（app_settingsテーブルへ統合）",
        "バックアップ": "SQLite DB + 添付ファイル",
    }

def migrate_excel_to_sqlite_if_needed(table_name: str, excel_path: Path, sheet_name: str, columns: list, date_cols=None, unique_cols=None):
    """初回起動時のみ、既存ExcelデータをSQLiteへ移行する。DBに1件でもあれば上書きしない。"""
    ensure_dirs()
    if sqlite_table_row_count(table_name) > 0:
        return

    if not excel_path.exists():
        save_sqlite_table(pd.DataFrame(columns=columns), table_name, columns, date_cols=date_cols, unique_cols=unique_cols)
        return

    try:
        df = pd.read_excel(excel_path, sheet_name=sheet_name)
    except Exception:
        df = pd.DataFrame(columns=columns)

    df = normalize_df_columns(df, columns)
    save_sqlite_table(df, table_name, columns, date_cols=date_cols, unique_cols=unique_cols)


def ensure_hidamari_db():
    """主要テーブルを作成し、起動時にDB整合性を確認する。

    SQLiteはバックアップ／ミラーのため、各初期化処理は個別に保護する。
    どこか1か所のSQLiteエラーでログイン画面・入力画面を止めない。
    """
    ensure_dirs()
    initialize_sqlite_engine()

    init_jobs = [
        lambda: save_sqlite_table(load_sqlite_table(SQLITE_TABLE_HEALTH, HEALTH_COLUMNS, date_cols=["記録日"]), SQLITE_TABLE_HEALTH, HEALTH_COLUMNS, date_cols=["記録日"], unique_cols=["記録日", "利用者名"]),
        lambda: save_sqlite_table(load_sqlite_table(SQLITE_TABLE_EXCRETION, EXCRETION_COLUMNS, date_cols=["記録日"]), SQLITE_TABLE_EXCRETION, EXCRETION_COLUMNS, date_cols=["記録日"], unique_cols=["記録日", "利用者名", "時間帯"]),
        lambda: save_sqlite_table(load_sqlite_table(SQLITE_TABLE_HANDOVER, BUSINESS_HANDOVER_COLUMNS, date_cols=["日付"]), SQLITE_TABLE_HANDOVER, BUSINESS_HANDOVER_COLUMNS, date_cols=["日付"], unique_cols=["記録ID"]),
        lambda: save_sqlite_table(load_sqlite_table(SQLITE_TABLE_SHORT_GOAL_MASTER, SHORT_GOAL_MASTER_COLUMNS, date_cols=["開始日", "終了予定日"]), SQLITE_TABLE_SHORT_GOAL_MASTER, SHORT_GOAL_MASTER_COLUMNS, date_cols=["開始日", "終了予定日"], unique_cols=["目標ID"]),
        lambda: save_sqlite_table(load_sqlite_table(SQLITE_TABLE_SHORT_GOAL_CHECKS, SHORT_GOAL_CHECK_COLUMNS, date_cols=["日付"]), SQLITE_TABLE_SHORT_GOAL_CHECKS, SHORT_GOAL_CHECK_COLUMNS, date_cols=["日付"], unique_cols=["記録ID"]),
        lambda: save_sqlite_table(load_sqlite_table(SQLITE_TABLE_MONITORING_DRAFTS, MONITORING_DRAFT_COLUMNS, date_cols=["作成日"]), SQLITE_TABLE_MONITORING_DRAFTS, MONITORING_DRAFT_COLUMNS, date_cols=["作成日"], unique_cols=["下書きID"]),
        lambda: save_sqlite_table(load_sqlite_table(SQLITE_TABLE_ALERT_CONDITIONS, ALERT_CONDITION_COLUMNS), SQLITE_TABLE_ALERT_CONDITIONS, ALERT_CONDITION_COLUMNS, unique_cols=["条件ID"]),
        lambda: save_sqlite_table(load_sqlite_table(SQLITE_TABLE_ALERTS, ["通知ID", "日付", "利用者名", "重要度", "分類", "通知内容", "対応状況", "作成日時"], date_cols=["日付"]), SQLITE_TABLE_ALERTS, ["通知ID", "日付", "利用者名", "重要度", "分類", "通知内容", "対応状況", "作成日時"], date_cols=["日付"], unique_cols=["通知ID"]),
        ensure_account_file,
        ensure_login_history_file,
        ensure_user_file,
        ensure_user_name_alias_table,
        ensure_handover_keyword_table,
        ensure_life_adl_file,
        ensure_ai_insight_log_file,
        initialize_default_app_settings,
        lambda: run_db_integrity_check(auto_repair=True),
    ]

    for job in init_jobs:
        try:
            job()
        except Exception as e:
            _mark_sqlite_backup_error(e, getattr(job, "__name__", "ensure_hidamari_db"))
            _show_sqlite_backup_warning_once(e, getattr(job, "__name__", "ensure_hidamari_db"))
            continue

# =========================
# セキュリティ・保守機能（Ver2.1）
# 自動バックアップ／監査ログ／権限管理／データ復元
# =========================
BACKUP_DIR = DATA_DIR / "backups"
RESTORE_DIR = DATA_DIR / "restore_uploads"

SQLITE_TABLE_AUDIT_LOGS = "audit_logs"
SQLITE_TABLE_ROLE_PERMISSIONS = "role_permissions"
SQLITE_TABLE_BACKUP_HISTORY = "backup_history"

AUDIT_LOG_COLUMNS = [
    "監査ID",
    "日時",
    "ログインID",
    "表示名",
    "権限",
    "操作種別",
    "対象テーブル",
    "対象キー",
    "概要",
    "変更前",
    "変更後",
]

ROLE_PERMISSION_COLUMNS = [
    "権限",
    "機能",
    "閲覧",
    "登録更新",
    "削除",
    "復元",
    "備考",
]

BACKUP_HISTORY_COLUMNS = [
    "バックアップID",
    "日時",
    "種類",
    "ファイル名",
    "サイズKB",
    "実行者",
    "結果",
    "メモ",
]

DEFAULT_ROLE_PERMISSIONS = [
    {"権限": "admin", "機能": "全機能", "閲覧": 1, "登録更新": 1, "削除": 1, "復元": 1, "備考": "管理者は全機能を利用可能"},
    {"権限": "staff", "機能": "健康チェック入力", "閲覧": 1, "登録更新": 1, "削除": 0, "復元": 0, "備考": "職員は日々の入力中心"},
    {"権限": "staff", "機能": "排泄チェック入力", "閲覧": 1, "登録更新": 1, "削除": 0, "復元": 0, "備考": "職員は日々の入力中心"},
    {"権限": "staff", "機能": "業務全体申し送り", "閲覧": 1, "登録更新": 1, "削除": 0, "復元": 0, "備考": "職員は申し送り入力可能"},
    {"権限": "staff", "機能": "日々の実施チェック", "閲覧": 1, "登録更新": 1, "削除": 0, "復元": 0, "備考": "職員は短期目標の実施入力可能"},
]


def ensure_security_dirs():
    ensure_dirs()
    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    RESTORE_DIR.mkdir(parents=True, exist_ok=True)


def ensure_security_tables():
    """セキュリティ関連テーブルを作成する。"""
    ensure_security_dirs()
    save_sqlite_table(
        load_sqlite_table(SQLITE_TABLE_AUDIT_LOGS, AUDIT_LOG_COLUMNS),
        SQLITE_TABLE_AUDIT_LOGS,
        AUDIT_LOG_COLUMNS,
        unique_cols=["監査ID"],
        sort_cols=["日時"],
    )
    perms = load_sqlite_table(SQLITE_TABLE_ROLE_PERMISSIONS, ROLE_PERMISSION_COLUMNS)
    if perms.empty:
        perms = pd.DataFrame(DEFAULT_ROLE_PERMISSIONS, columns=ROLE_PERMISSION_COLUMNS)
    save_sqlite_table(
        perms,
        SQLITE_TABLE_ROLE_PERMISSIONS,
        ROLE_PERMISSION_COLUMNS,
        unique_cols=["権限", "機能"],
    )
    save_sqlite_table(
        load_sqlite_table(SQLITE_TABLE_BACKUP_HISTORY, BACKUP_HISTORY_COLUMNS),
        SQLITE_TABLE_BACKUP_HISTORY,
        BACKUP_HISTORY_COLUMNS,
        unique_cols=["バックアップID"],
        sort_cols=["日時"],
    )


def get_current_user_info_for_log():
    login_id = current_login_user()
    role = clean_text(st.session_state.get("role", ""))
    label = clean_text(st.session_state.get("user_label", login_id), login_id)
    info = st.session_state.get("login_user_info", {})
    if isinstance(info, dict):
        role = role or clean_text(info.get("role", ""))
        label = label or clean_text(info.get("label", login_id), login_id)
    return login_id, label, role


def add_audit_log(operation, table_name="", target_key="", summary="", before="", after=""):
    """誰が、いつ、何をしたかをSQLiteに保存する。"""
    try:
        ensure_security_tables()
        login_id, label, role = get_current_user_info_for_log()
        df = load_sqlite_table(SQLITE_TABLE_AUDIT_LOGS, AUDIT_LOG_COLUMNS)
        row = {
            "監査ID": str(uuid.uuid4()),
            "日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
            "ログインID": login_id,
            "表示名": label,
            "権限": role,
            "操作種別": clean_text(operation),
            "対象テーブル": clean_text(table_name),
            "対象キー": clean_text(target_key),
            "概要": clean_text(summary),
            "変更前": clean_text(before),
            "変更後": clean_text(after),
        }
        df = pd.concat([df, pd.DataFrame([row], columns=AUDIT_LOG_COLUMNS)], ignore_index=True)
        # 長期運用で肥大化しすぎないよう直近5000件を保持
        if len(df) > 5000:
            df = df.tail(5000)
        save_sqlite_table(df, SQLITE_TABLE_AUDIT_LOGS, AUDIT_LOG_COLUMNS, unique_cols=["監査ID"], sort_cols=["日時"])
    except Exception:
        # 監査ログ失敗で本体入力を止めない
        pass


def has_permission(feature_name, action="閲覧"):
    """権限表に基づいて操作可否を返す。adminは常に許可。"""
    role = clean_text(st.session_state.get("role", "staff"), "staff")
    if role == "admin":
        return True
    try:
        ensure_security_tables()
        perms = load_sqlite_table(SQLITE_TABLE_ROLE_PERMISSIONS, ROLE_PERMISSION_COLUMNS)
        if perms.empty:
            return False
        # 全機能または対象機能の行を見る
        candidates = perms[
            (perms["権限"].astype(str) == role)
            & ((perms["機能"].astype(str) == feature_name) | (perms["機能"].astype(str) == "全機能"))
        ]
        if candidates.empty:
            return False
        col = action if action in ["閲覧", "登録更新", "削除", "復元"] else "閲覧"
        return any(str(v).lower() in ["1", "true", "yes", "有", "可"] for v in candidates[col].tolist())
    except Exception:
        return role == "admin"


def require_permission(feature_name, action="閲覧"):
    if not has_permission(feature_name, action):
        st.warning(f"この操作は権限がありません：{feature_name}／{action}")
        add_audit_log("権限エラー", "role_permissions", feature_name, f"{action} が拒否されました")
        return False
    return True


def record_backup_history(kind, file_path, result="成功", memo=""):
    try:
        ensure_security_tables()
        df = load_sqlite_table(SQLITE_TABLE_BACKUP_HISTORY, BACKUP_HISTORY_COLUMNS)
        size_kb = 0
        try:
            size_kb = round(Path(file_path).stat().st_size / 1024, 1)
        except Exception:
            pass
        row = {
            "バックアップID": str(uuid.uuid4()),
            "日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
            "種類": kind,
            "ファイル名": Path(file_path).name if file_path else "",
            "サイズKB": size_kb,
            "実行者": current_login_user(),
            "結果": result,
            "メモ": memo,
        }
        df = pd.concat([df, pd.DataFrame([row], columns=BACKUP_HISTORY_COLUMNS)], ignore_index=True)
        if len(df) > 1000:
            df = df.tail(1000)
        save_sqlite_table(df, SQLITE_TABLE_BACKUP_HISTORY, BACKUP_HISTORY_COLUMNS, unique_cols=["バックアップID"], sort_cols=["日時"])
    except Exception:
        pass



def _list_sqlite_tables_for_backup():
    """SQLite内のユーザーテーブル一覧を取得する。"""
    try:
        if not HIDAMARI_DB_FILE.exists():
            return []
        with get_hidamari_conn() as conn:
            rows = conn.execute(
                "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
            ).fetchall()
        return [r[0] for r in rows]
    except Exception:
        return []


def _sqlite_table_to_df_for_backup(table_name: str) -> pd.DataFrame:
    try:
        validate_sqlite_identifier(table_name)
        with get_hidamari_conn() as conn:
            return pd.read_sql_query(f'SELECT * FROM "{table_name}"', conn)
    except Exception as e:
        return pd.DataFrame([{"backup_error": str(e)}])


def _make_all_sqlite_tables_excel_bytes():
    """SQLite全テーブルを1つのExcelに退避する。復旧確認・監査用。"""
    output = BytesIO()
    tables = _list_sqlite_tables_for_backup()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if not tables:
            pd.DataFrame([{"message": "SQLiteテーブルが見つかりません"}]).to_excel(writer, sheet_name="empty", index=False)
        for table_name in tables:
            df = _sqlite_table_to_df_for_backup(table_name)
            # Excelのシート名は31文字制限
            safe_sheet = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(table_name))[:31] or "table"
            df.to_excel(writer, sheet_name=safe_sheet, index=False)
    output.seek(0)
    return output.getvalue()


def _read_supabase_core_tables_for_backup():
    """
    完全バックアップ用にSupabase主要テーブルを全件取得する。

    重要：
    - 画面表示用の supabase_read_table() はキャッシュ・期間絞り込みの影響を受ける可能性がある。
    - バックアップではキャッシュを使わず、PostgRESTから直接ページング取得する。
    - data(JSONB)内の本体データに加えて、確認用に __record_key / __updated_at も残す。
    """
    result = {}
    targets = [
        (SQLITE_TABLE_USERS, USER_COLUMNS if "USER_COLUMNS" in globals() else []),
        (SQLITE_TABLE_HEALTH, HEALTH_COLUMNS if "HEALTH_COLUMNS" in globals() else []),
        (SQLITE_TABLE_EXCRETION, EXCRETION_COLUMNS if "EXCRETION_COLUMNS" in globals() else []),
        (SQLITE_TABLE_HANDOVER, BUSINESS_HANDOVER_COLUMNS if "BUSINESS_HANDOVER_COLUMNS" in globals() else []),
        (SQLITE_TABLE_SHORT_GOAL_CHECKS, SHORT_GOAL_CHECK_COLUMNS if "SHORT_GOAL_CHECK_COLUMNS" in globals() else []),
        (SQLITE_TABLE_SHORT_GOAL_MASTER, SHORT_GOAL_MASTER_COLUMNS if "SHORT_GOAL_MASTER_COLUMNS" in globals() else []),
        (SQLITE_TABLE_MONITORING_DRAFTS, MONITORING_DRAFT_COLUMNS if "MONITORING_DRAFT_COLUMNS" in globals() else []),
    ]

    if not ("supabase_is_enabled" in globals() and supabase_is_enabled()):
        for table_name, _cols in targets:
            result[table_name] = [{"backup_error": "Supabase未設定または接続不可"}]
        return result

    for table_name, cols in targets:
        try:
            all_records = []
            page_size = 1000
            offset = 0
            while True:
                headers = _supabase_headers(prefer="")
                headers["Range-Unit"] = "items"
                headers["Range"] = f"{offset}-{offset + page_size - 1}"
                res = requests.get(
                    _supabase_endpoint(table_name),
                    headers=headers,
                    params=[("select", "record_key,data,updated_at"), ("order", "updated_at.asc")],
                    timeout=30,
                )
                if res.status_code not in [200, 206]:
                    res.raise_for_status()
                rows = res.json() or []
                for item in rows:
                    data = item.get("data") or {}
                    if isinstance(data, dict):
                        row = dict(data)
                        row["__record_key"] = item.get("record_key", "")
                        row["__updated_at"] = item.get("updated_at", "")
                        all_records.append(row)
                if len(rows) < page_size:
                    break
                offset += page_size

            if cols:
                df = pd.DataFrame(all_records)
                for col in cols:
                    if col not in df.columns:
                        df[col] = ""
                # 確認用メタ列は末尾に残す
                meta_cols = [c for c in ["__record_key", "__updated_at"] if c in df.columns]
                result[table_name] = df[list(cols) + meta_cols].to_dict(orient="records")
            else:
                result[table_name] = all_records
        except Exception as e:
            result[table_name] = [{"backup_error": str(e)}]
    return result

def _make_supabase_core_excel_bytes(supabase_data: dict):
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        if not supabase_data:
            pd.DataFrame([{"message": "Supabase未設定または主要データなし"}]).to_excel(writer, sheet_name="supabaseなし", index=False)
        for table_name, records in supabase_data.items():
            df = pd.DataFrame(records)
            safe_sheet = re.sub(r"[\[\]\:\*\?\/\\]", "_", str(table_name))[:31] or "table"
            df.to_excel(writer, sheet_name=safe_sheet, index=False)
    output.seek(0)
    return output.getvalue()


def get_backup_target_status_df():
    """バックアップ対象の検査表。管理画面に表示する。"""
    rows = []
    rows.append({
        "対象": "SQLite DB本体",
        "状態": "OK" if HIDAMARI_DB_FILE.exists() else "未作成",
        "詳細": str(HIDAMARI_DB_FILE),
    })
    rows.append({
        "対象": "SQLite全テーブルExcel退避",
        "状態": "OK" if _list_sqlite_tables_for_backup() else "確認",
        "詳細": f"{len(_list_sqlite_tables_for_backup())} テーブル",
    })
    rows.append({
        "対象": "Supabase本番データ全件退避",
        "状態": "OK" if ("supabase_is_enabled" in globals() and supabase_is_enabled()) else "未設定",
        "詳細": "キャッシュを使わず、利用者・健康・排泄・申し送り・短期目標・モニタリングを全件取得",
    })
    for folder in [BUSINESS_HANDOVER_PHOTO_DIR, BUSINESS_HANDOVER_EXCEL_DIR, REPORT_DIR]:
        count = 0
        if folder.exists():
            count = len([f for f in folder.rglob("*") if f.is_file()])
        rows.append({
            "対象": folder.name,
            "状態": "OK" if folder.exists() else "未作成",
            "詳細": f"{count} ファイル",
        })
    return pd.DataFrame(rows)


def _write_dataframe_to_zip_excel(zf, arcname, df, fallback_message="データなし"):
    """DataFrameをExcelとしてZIPへ安全に書き込む。"""
    try:
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            work = df if isinstance(df, pd.DataFrame) and not df.empty else pd.DataFrame([{"message": fallback_message}])
            work.to_excel(writer, sheet_name="data", index=False)
        output.seek(0)
        zf.writestr(arcname, output.getvalue())
    except Exception as e:
        zf.writestr(str(arcname).replace(".xlsx", "_error.txt"), str(e))


def _write_folder_to_zip(zf, folder: Path, arc_prefix: str):
    """指定フォルダをZIPへ安全に格納する。"""
    try:
        folder = Path(folder)
        if not folder.exists():
            zf.writestr(f"{arc_prefix.rstrip('/')}/_folder_not_found.txt", str(folder))
            return 0
        count = 0
        for file in folder.rglob("*"):
            if file.is_file():
                rel = file.relative_to(folder)
                zf.write(file, arcname=f"{arc_prefix.rstrip('/')}/{rel.as_posix()}")
                count += 1
        if count == 0:
            zf.writestr(f"{arc_prefix.rstrip('/')}/_empty.txt", "対象ファイルはありません。")
        return count
    except Exception as e:
        zf.writestr(f"{arc_prefix.rstrip('/')}/_backup_error.txt", str(e))
        return 0


def _read_sqlite_table_for_complete_backup(table_name, columns=None):
    """完全バックアップ用にSQLiteテーブルを安全に読む。"""
    try:
        if not sqlite_table_exists(table_name):
            return pd.DataFrame(columns=list(columns or []))
        df = _original_load_sqlite_table(table_name, columns or [])
        if columns:
            for col in columns:
                if col not in df.columns:
                    df[col] = ""
            df = df[list(columns)]
        return df
    except Exception as e:
        return pd.DataFrame([{"backup_error": str(e), "table": table_name}])


def _write_complete_system_excels(zf):
    """監査ログ・ログイン履歴・設定などをSystem配下へ退避する。"""
    system_targets = [
        ("System/audit_log.xlsx", SQLITE_TABLE_AUDIT_LOGS if "SQLITE_TABLE_AUDIT_LOGS" in globals() else "audit_logs", AUDIT_LOG_COLUMNS if "AUDIT_LOG_COLUMNS" in globals() else []),
        ("System/login_history.xlsx", SQLITE_TABLE_LOGIN_HISTORY if "SQLITE_TABLE_LOGIN_HISTORY" in globals() else "login_history", LOGIN_HISTORY_COLUMNS if "LOGIN_HISTORY_COLUMNS" in globals() else []),
        ("System/settings.xlsx", SQLITE_TABLE_APP_SETTINGS if "SQLITE_TABLE_APP_SETTINGS" in globals() else "app_settings", APP_SETTING_COLUMNS if "APP_SETTING_COLUMNS" in globals() else []),
        ("System/accounts.xlsx", SQLITE_TABLE_ACCOUNTS if "SQLITE_TABLE_ACCOUNTS" in globals() else "login_accounts", ACCOUNT_COLUMNS if "ACCOUNT_COLUMNS" in globals() else []),
        ("System/role_permissions.xlsx", SQLITE_TABLE_ROLE_PERMISSIONS if "SQLITE_TABLE_ROLE_PERMISSIONS" in globals() else "role_permissions", ROLE_PERMISSION_COLUMNS if "ROLE_PERMISSION_COLUMNS" in globals() else []),
        ("System/backup_history.xlsx", SQLITE_TABLE_BACKUP_HISTORY if "SQLITE_TABLE_BACKUP_HISTORY" in globals() else "backup_history", BACKUP_HISTORY_COLUMNS if "BACKUP_HISTORY_COLUMNS" in globals() else []),
        ("System/ai_insight_logs.xlsx", SQLITE_TABLE_AI_INSIGHT_LOGS if "SQLITE_TABLE_AI_INSIGHT_LOGS" in globals() else "ai_insight_logs", AI_INSIGHT_LOG_COLUMNS if "AI_INSIGHT_LOG_COLUMNS" in globals() else []),
        ("System/user_name_aliases.xlsx", SQLITE_TABLE_USER_NAME_ALIASES if "SQLITE_TABLE_USER_NAME_ALIASES" in globals() else "user_name_aliases", USER_NAME_ALIAS_COLUMNS if "USER_NAME_ALIAS_COLUMNS" in globals() else []),
        ("System/handover_keywords.xlsx", SQLITE_TABLE_HANDOVER_KEYWORDS if "SQLITE_TABLE_HANDOVER_KEYWORDS" in globals() else "handover_keywords", HANDOVER_KEYWORD_COLUMNS if "HANDOVER_KEYWORD_COLUMNS" in globals() else []),
    ]
    for arcname, table_name, cols in system_targets:
        try:
            _write_dataframe_to_zip_excel(zf, arcname, _read_sqlite_table_for_complete_backup(table_name, cols))
        except Exception as e:
            zf.writestr(arcname.replace(".xlsx", "_error.txt"), str(e))


def _write_supabase_complete_backup_files(zf, supabase_data: dict):
    """Supabase全件を用途別Excelと互換JSON/Excelで格納する。"""
    label_map = {
        SQLITE_TABLE_USERS: "users.xlsx",
        SQLITE_TABLE_HEALTH: "health.xlsx",
        SQLITE_TABLE_EXCRETION: "excretion.xlsx",
        SQLITE_TABLE_HANDOVER: "handover.xlsx",
        SQLITE_TABLE_SHORT_GOAL_CHECKS: "short_goal_checks.xlsx",
        SQLITE_TABLE_SHORT_GOAL_MASTER: "short_goal_master.xlsx",
        SQLITE_TABLE_MONITORING_DRAFTS: "monitoring.xlsx",
    }
    for table_name, file_name in label_map.items():
        records = supabase_data.get(table_name, [])
        df = pd.DataFrame(records if isinstance(records, list) else [])
        _write_dataframe_to_zip_excel(zf, f"Supabase/{file_name}", df, fallback_message=f"{table_name} は空です。")

    # 復元機能との互換用。既存の復元処理は exports/supabase_core_tables.json を読む。
    zf.writestr(
        "exports/supabase_core_tables.json",
        json.dumps(supabase_data, ensure_ascii=False, indent=2, default=str),
    )
    zf.writestr("exports/supabase_core_tables.xlsx", _make_supabase_core_excel_bytes(supabase_data))

def create_backup_zip(kind="手動"):
    """
    完全バックアップZIPを作成する。

    Ver4.8.5 完全バックアップ方針：
    - Supabase本番データはキャッシュを使わず全件取得
    - SQLite DB本体、WAL/SHM、全SQLiteテーブルExcelを同梱
    - 写真、添付Excel、reports配下の帳票を同梱
    - 監査ログ、ログイン履歴、設定、アカウント等をSystem配下へExcel出力
    - 既存復元処理との互換のため exports/supabase_core_tables.json も維持
    """
    ensure_security_dirs()
    timestamp = format_now_jst("%Y%m%d_%H%M%S")
    safe_kind = re.sub(r"[^\w一-龥ぁ-んァ-ンー\-]", "_", str(kind))
    zip_path = BACKUP_DIR / f"hidamari_complete_backup_{safe_kind}_{timestamp}.zip"

    try:
        sqlite_db_ok_for_file_backup = False
        sqlite_check_message = ""
        if HIDAMARI_DB_FILE.exists():
            try:
                with get_hidamari_conn() as conn:
                    conn.execute("PRAGMA quick_check;")
                    conn.execute("PRAGMA wal_checkpoint(FULL);")
                sqlite_db_ok_for_file_backup = True
            except Exception as e:
                sqlite_check_message = str(e)
                _mark_sqlite_backup_error(e, "complete_backup_sqlite_checkpoint")
                if is_sqlite_corruption_error(e):
                    quarantine_corrupt_sqlite_db(f"complete_backup_sqlite_checkpoint: {e}")
                sqlite_db_ok_for_file_backup = False

        # ここは画面表示用キャッシュを使わない全件取得
        supabase_data = _read_supabase_core_tables_for_backup()

        backup_info = {
            "app": "ひだまり 健康チェック管理システム",
            "backup_type": "complete",
            "version_note": "Ver4.8.5 完全バックアップZIP",
            "created_at": format_now_jst("%Y-%m-%d %H:%M:%S"),
            "kind": str(kind),
            "user": current_login_user() if "current_login_user" in globals() else "",
            "supabase_enabled": bool("supabase_is_enabled" in globals() and supabase_is_enabled()),
            "supabase_tables": list(supabase_data.keys()) if isinstance(supabase_data, dict) else [],
            "sqlite_db": str(HIDAMARI_DB_FILE),
            "sqlite_db_included": bool(HIDAMARI_DB_FILE.exists() and sqlite_db_ok_for_file_backup),
            "sqlite_check_message": sqlite_check_message,
            "sqlite_tables": _list_sqlite_tables_for_backup(),
            "included_folders": {
                "Photos": str(BUSINESS_HANDOVER_PHOTO_DIR),
                "Attachments": str(BUSINESS_HANDOVER_EXCEL_DIR),
                "Reports": str(REPORT_DIR),
                "SQLite": str(DATA_DIR),
                "System": "SQLite内の監査ログ・ログイン履歴・設定等をExcel出力",
            },
            "note": "Supabase本番データ、SQLite、写真、帳票、設定/ログ系を含む完全バックアップです。画面表示用キャッシュや7日表示には依存しません。",
        }

        with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
            zf.writestr("backup_info.json", json.dumps(backup_info, ensure_ascii=False, indent=2, default=str))

            # Supabase本番データ全件
            try:
                _write_supabase_complete_backup_files(zf, supabase_data)
            except Exception as e:
                zf.writestr("Supabase/_backup_error.txt", str(e))

            # SQLite DB本体と関連ファイル
            sqlite_files = [
                HIDAMARI_DB_FILE,
                Path(str(HIDAMARI_DB_FILE) + "-wal"),
                Path(str(HIDAMARI_DB_FILE) + "-shm"),
                DATA_DIR / "hidamari_life.db",
                Path("hidamari_life.db"),
            ]
            wrote_sqlite_file = False
            for db_file in sqlite_files:
                try:
                    if db_file.exists():
                        # メインDBが壊れている場合は、本体だけスキップし、WAL/SHMや他DBは可能な範囲で残す
                        if db_file == HIDAMARI_DB_FILE and not sqlite_db_ok_for_file_backup:
                            zf.writestr("SQLite/hidamari_health_db_skipped.txt", f"SQLite補助DBが破損または読込不可のためDB本体同梱をスキップしました: {sqlite_check_message}")
                            continue
                        zf.write(db_file, arcname=f"SQLite/{db_file.name}")
                        wrote_sqlite_file = True
                except Exception as e:
                    zf.writestr(f"SQLite/{db_file.name}_backup_error.txt", str(e))
            if not wrote_sqlite_file:
                zf.writestr("SQLite/_no_sqlite_file.txt", "同梱可能なSQLite DBファイルはありませんでした。")

            # SQLite全テーブル確認用Excel
            try:
                sqlite_excel = _make_all_sqlite_tables_excel_bytes()
                zf.writestr("SQLite/sqlite_all_tables.xlsx", sqlite_excel)
                zf.writestr("exports/sqlite_all_tables.xlsx", sqlite_excel)  # 旧パス互換
            except Exception as e:
                zf.writestr("SQLite/sqlite_all_tables_error.txt", str(e))

            # System：ログ・設定・アカウント等
            try:
                _write_complete_system_excels(zf)
            except Exception as e:
                zf.writestr("System/_backup_error.txt", str(e))

            # 写真・添付・帳票
            _write_folder_to_zip(zf, BUSINESS_HANDOVER_PHOTO_DIR, "Photos/business_handover_photos")
            _write_folder_to_zip(zf, BUSINESS_HANDOVER_EXCEL_DIR, "Attachments/business_handover_excels")
            _write_folder_to_zip(zf, REPORT_DIR, "Reports")

            # data内の旧Excel/JSON設定類も念のため保管
            try:
                for file in DATA_DIR.glob("*"):
                    if file.is_file() and file.suffix.lower() in [".xlsx", ".json", ".csv", ".txt"]:
                        zf.write(file, arcname=f"System/data_files/{file.name}")
            except Exception as e:
                zf.writestr("System/data_files_backup_error.txt", str(e))

        record_backup_history(kind, zip_path, "成功", "完全バックアップZIP作成")
        add_audit_log("完全バックアップ作成", "backup_history", zip_path.name, f"{kind}完全バックアップを作成")
        return zip_path, ""
    except Exception as e:
        # 完全バックアップ全体が落ちた場合も、Supabase本番データだけは緊急退避する。
        try:
            emergency_path = BACKUP_DIR / f"hidamari_complete_backup_{safe_kind}_{timestamp}_supabase_only.zip"
            supabase_data = _read_supabase_core_tables_for_backup()
            with zipfile.ZipFile(emergency_path, "w", zipfile.ZIP_DEFLATED) as zf:
                zf.writestr("backup_error.txt", str(e))
                _write_supabase_complete_backup_files(zf, supabase_data)
            try:
                record_backup_history(kind, emergency_path, "一部成功", f"完全バックアップ失敗。Supabase本番データのみ退避: {e}")
            except Exception:
                pass
            try:
                add_audit_log("完全バックアップ一部成功", "backup_history", emergency_path.name, str(e))
            except Exception:
                pass
            return emergency_path, ""
        except Exception as e2:
            try:
                record_backup_history(kind, zip_path, "失敗", str(e2))
            except Exception:
                pass
            try:
                add_audit_log("完全バックアップ失敗", "backup_history", zip_path.name, str(e2))
            except Exception:
                pass
            return None, str(e2)

def run_daily_auto_backup():
    """1日1回だけ自動バックアップを作成する。"""
    try:
        ensure_security_dirs()
        today_key = today_jst().strftime("%Y%m%d")
        marker = BACKUP_DIR / f".auto_backup_{today_key}.done"
        if marker.exists():
            return
        zip_path, err = create_backup_zip(kind="自動")
        if zip_path and not err:
            marker.write_text(format_now_jst("%Y-%m-%d %H:%M:%S"), encoding="utf-8")
            # 古い自動バックアップは30世代程度に整理
            auto_files = sorted(BACKUP_DIR.glob("hidamari_backup_自動_*.zip"), key=lambda x: x.stat().st_mtime, reverse=True)
            for old in auto_files[30:]:
                try:
                    old.unlink()
                except Exception:
                    pass
    except Exception:
        # 自動バックアップ失敗で本体起動を止めない
        pass



def _restore_supabase_core_tables_from_backup(zf, names):
    """
    バックアップZIP内の exports/supabase_core_tables.json から、
    Supabase主要7機能（利用者・健康・排泄・申し送り・短期目標実施チェック・短期目標マスタ・モニタリング下書き）へ復元する。
    安全優先のため、現在のSupabase全削除は行わず、record_key単位のupsertで戻す。
    """
    result = {
        "ok": False,
        "restored_counts": {},
        "messages": [],
        "error": "",
    }

    if not ("supabase_is_enabled" in globals() and supabase_is_enabled()):
        result["error"] = "Supabaseが未設定または接続できません。"
        return result

    if "exports/supabase_core_tables.json" not in names:
        result["error"] = "このバックアップZIPには Supabase主要7機能データ（exports/supabase_core_tables.json）が含まれていません。"
        return result

    try:
        raw = zf.read("exports/supabase_core_tables.json").decode("utf-8")
        supabase_data = json.loads(raw) if raw.strip() else {}
        if not isinstance(supabase_data, dict):
            result["error"] = "Supabase復元データの形式が正しくありません。"
            return result

        targets = [
            (SQLITE_TABLE_USERS, USER_COLUMNS if "USER_COLUMNS" in globals() else [], ["user_id"]),
            (SQLITE_TABLE_HEALTH, HEALTH_COLUMNS if "HEALTH_COLUMNS" in globals() else [], ["記録日", "user_id"]),
            (SQLITE_TABLE_EXCRETION, EXCRETION_COLUMNS if "EXCRETION_COLUMNS" in globals() else [], ["記録日", "user_id", "時間帯"]),
            (SQLITE_TABLE_HANDOVER, BUSINESS_HANDOVER_COLUMNS if "BUSINESS_HANDOVER_COLUMNS" in globals() else [], ["記録ID"]),
            (SQLITE_TABLE_SHORT_GOAL_CHECKS, SHORT_GOAL_CHECK_COLUMNS if "SHORT_GOAL_CHECK_COLUMNS" in globals() else [], ["記録ID"]),
            (SQLITE_TABLE_SHORT_GOAL_MASTER, SHORT_GOAL_MASTER_COLUMNS if "SHORT_GOAL_MASTER_COLUMNS" in globals() else [], ["目標ID"]),
            (SQLITE_TABLE_MONITORING_DRAFTS, MONITORING_DRAFT_COLUMNS if "MONITORING_DRAFT_COLUMNS" in globals() else [], ["下書きID"]),
        ]

        restored_any = False
        for table_name, cols, unique_cols in targets:
            records = supabase_data.get(table_name, [])
            if not isinstance(records, list):
                records = []

            # backup_error行は復元対象から除外
            clean_records = []
            for item in records:
                if isinstance(item, dict) and "backup_error" not in item:
                    clean_records.append(item)

            df = pd.DataFrame(clean_records)
            if cols:
                try:
                    df = normalize_df_columns(df, cols)
                except Exception:
                    for col in cols:
                        if col not in df.columns:
                            df[col] = ""
                    df = df[cols] if cols else df

            if df.empty:
                result["restored_counts"][table_name] = 0
                continue

            ok = supabase_upsert_table(df, table_name, columns=cols, unique_cols=unique_cols)
            if ok:
                restored_any = True
                result["restored_counts"][table_name] = len(df)
                result["messages"].append(f"{SUPABASE_CORE_LABELS.get(table_name, table_name)}：{len(df)}件")
                # Supabase復元後、SQLiteミラーにも同じ内容を残す
                try:
                    _original_save_sqlite_table(
                        df,
                        table_name,
                        cols,
                        date_cols=(
                            ["記録日"] if table_name in [SQLITE_TABLE_HEALTH, SQLITE_TABLE_EXCRETION]
                            else (["日付"] if table_name in [SQLITE_TABLE_HANDOVER, SQLITE_TABLE_SHORT_GOAL_CHECKS]
                                  else (["開始日", "終了予定日"] if table_name == SQLITE_TABLE_SHORT_GOAL_MASTER
                                        else (["作成日"] if table_name == SQLITE_TABLE_MONITORING_DRAFTS else None)))
                        ),
                        unique_cols=unique_cols,
                    )
                except Exception:
                    pass
            else:
                result["restored_counts"][table_name] = 0
                result["messages"].append(f"{SUPABASE_CORE_LABELS.get(table_name, table_name)}：復元失敗")

        result["ok"] = restored_any
        if not restored_any:
            result["error"] = "Supabaseへ復元できるデータがありませんでした。"
        return result

    except Exception as e:
        result["error"] = str(e)
        return result



# =========================
# 復元安全化（Ver4.8.6）
# 完全バックアップZIP／旧バックアップZIPの両方に対応する。
# =========================
def _zip_has_member(names, candidates):
    name_set = set(names or [])
    for candidate in candidates:
        if candidate in name_set:
            return candidate
    return ""

def _safe_zip_member_to_target(zf, member_name: str, target_path: Path) -> bool:
    """
    ZIP内ファイルを指定先へ安全に復元する。
    - ディレクトリトラバーサル対策
    - 親フォルダ自動作成
    """
    if not member_name or member_name.endswith("/"):
        return False
    target_path = Path(target_path)
    target_path.parent.mkdir(parents=True, exist_ok=True)
    target_path.write_bytes(zf.read(member_name))
    return True

def _restore_folder_prefix_from_zip(zf, names, zip_prefix: str, target_root: Path) -> int:
    """ZIP内の指定prefix配下を、target_rootへ相対パスを保って復元する。"""
    count = 0
    zip_prefix = str(zip_prefix).strip("/")
    target_root = Path(target_root)
    for name in names:
        if not name or name.endswith("/"):
            continue
        normalized = name.strip("/")
        if not normalized.startswith(zip_prefix + "/"):
            continue
        rel = normalized[len(zip_prefix) + 1:]
        if not rel or rel.startswith("../") or "/../" in rel or rel == "..":
            continue
        try:
            _safe_zip_member_to_target(zf, name, target_root / rel)
            count += 1
        except Exception:
            continue
    return count

def _restore_legacy_folder_from_zip(zf, names, folder: Path) -> int:
    """
    旧バックアップ形式の data/... パスをそのまま復元する。
    Streamlit上の相対Pathのみを対象にして、危険な絶対パスは除外する。
    """
    count = 0
    folder_text = str(folder).replace("\\", "/").strip("/")
    for name in names:
        normalized = str(name).replace("\\", "/").strip("/")
        if not normalized or normalized.endswith("/"):
            continue
        if normalized.startswith("../") or "/../" in normalized or Path(normalized).is_absolute():
            continue
        if normalized.startswith(folder_text + "/"):
            try:
                _safe_zip_member_to_target(zf, name, Path(normalized))
                count += 1
            except Exception:
                continue
    return count

def _restore_sqlite_files_from_zip(zf, names) -> list:
    """
    SQLite DB本体を復元する。
    Ver4.8.5完全バックアップ形式:
      SQLite/hidamari_health.db
      SQLite/hidamari_health.db-wal
      SQLite/hidamari_health.db-shm
      SQLite/hidamari_life.db
    旧形式:
      data/hidamari_health.db
      data/hidamari_life.db
    """
    restored = []

    main_db_member = _zip_has_member(names, [
        f"SQLite/{HIDAMARI_DB_FILE.name}",
        f"data/{HIDAMARI_DB_FILE.name}",
        HIDAMARI_DB_FILE.name,
    ])
    if not main_db_member:
        raise FileNotFoundError(
            "SQLite復元を選択していますが、このZIPには hidamari_health.db が含まれていません。"
            "完全バックアップZIPの場合は SQLite/hidamari_health.db、旧バックアップの場合は data/hidamari_health.db が必要です。"
        )

    # 既存WAL/SHMが残ると復元後に不整合が出ることがあるため、先に退避削除する。
    for sidecar in [Path(str(HIDAMARI_DB_FILE) + "-wal"), Path(str(HIDAMARI_DB_FILE) + "-shm")]:
        try:
            if sidecar.exists():
                sidecar.unlink()
        except Exception:
            pass

    HIDAMARI_DB_FILE.parent.mkdir(parents=True, exist_ok=True)
    HIDAMARI_DB_FILE.write_bytes(zf.read(main_db_member))
    restored.append("hidamari_health.db")

    # WAL/SHMは完全バックアップに含まれている場合のみ復元する。
    sidecar_map = [
        (f"SQLite/{HIDAMARI_DB_FILE.name}-wal", Path(str(HIDAMARI_DB_FILE) + "-wal")),
        (f"SQLite/{HIDAMARI_DB_FILE.name}-shm", Path(str(HIDAMARI_DB_FILE) + "-shm")),
        (f"data/{HIDAMARI_DB_FILE.name}-wal", Path(str(HIDAMARI_DB_FILE) + "-wal")),
        (f"data/{HIDAMARI_DB_FILE.name}-shm", Path(str(HIDAMARI_DB_FILE) + "-shm")),
    ]
    for member, target in sidecar_map:
        try:
            if member in names:
                target.write_bytes(zf.read(member))
                restored.append(target.name)
        except Exception:
            pass

    life_member = _zip_has_member(names, [
        "SQLite/hidamari_life.db",
        "data/hidamari_life.db",
        "hidamari_life.db",
    ])
    if life_member:
        try:
            (DATA_DIR / "hidamari_life.db").write_bytes(zf.read(life_member))
            restored.append("hidamari_life.db")
        except Exception:
            pass

    return restored

def _restore_files_from_complete_or_legacy_zip(zf, names) -> int:
    """
    写真・添付・帳票を復元する。
    Ver4.8.5完全バックアップ形式と旧形式の両方に対応。
    """
    file_count = 0

    # Ver4.8.5 完全バックアップ形式
    file_count += _restore_folder_prefix_from_zip(
        zf, names, "Photos/business_handover_photos", BUSINESS_HANDOVER_PHOTO_DIR
    )
    file_count += _restore_folder_prefix_from_zip(
        zf, names, "Attachments/business_handover_excels", BUSINESS_HANDOVER_EXCEL_DIR
    )
    file_count += _restore_folder_prefix_from_zip(
        zf, names, "Reports", REPORT_DIR
    )

    # 旧形式：data/... や reports/... をそのまま復元
    for folder in [BUSINESS_HANDOVER_PHOTO_DIR, BUSINESS_HANDOVER_EXCEL_DIR, REPORT_DIR]:
        file_count += _restore_legacy_folder_from_zip(zf, names, folder)

    return file_count

def get_restore_zip_diagnostic(uploaded_file) -> pd.DataFrame:
    """復元前にZIPの中身を診断するための補助。画面表示に使える。"""
    rows = []
    try:
        pos = uploaded_file.tell()
    except Exception:
        pos = None
    try:
        uploaded_file.seek(0)
        with zipfile.ZipFile(uploaded_file, "r") as zf:
            names = zf.namelist()
            rows.append({"項目": "ZIPファイル", "状態": "OK", "詳細": f"{len(names)}件のファイル"})
            rows.append({"項目": "backup_info.json", "状態": "OK" if "backup_info.json" in names else "確認", "詳細": "あり" if "backup_info.json" in names else "なし"})
            rows.append({"項目": "SQLite DB", "状態": "OK" if _zip_has_member(names, [f"SQLite/{HIDAMARI_DB_FILE.name}", f"data/{HIDAMARI_DB_FILE.name}", HIDAMARI_DB_FILE.name]) else "NG", "詳細": "復元可能" if _zip_has_member(names, [f"SQLite/{HIDAMARI_DB_FILE.name}", f"data/{HIDAMARI_DB_FILE.name}", HIDAMARI_DB_FILE.name]) else "hidamari_health.dbが見つかりません"})
            rows.append({"項目": "Supabase退避JSON", "状態": "OK" if "exports/supabase_core_tables.json" in names else "確認", "詳細": "あり" if "exports/supabase_core_tables.json" in names else "なし"})
            photos = len([n for n in names if n.startswith("Photos/") or n.startswith(str(BUSINESS_HANDOVER_PHOTO_DIR).replace("\\", "/"))])
            reports = len([n for n in names if n.startswith("Reports/") or n.startswith(str(REPORT_DIR).replace("\\", "/"))])
            rows.append({"項目": "写真", "状態": "OK" if photos else "確認", "詳細": f"{photos}件"})
            rows.append({"項目": "帳票", "状態": "OK" if reports else "確認", "詳細": f"{reports}件"})
    except Exception as e:
        rows.append({"項目": "ZIP診断", "状態": "NG", "詳細": str(e)})
    finally:
        try:
            if pos is not None:
                uploaded_file.seek(pos)
            else:
                uploaded_file.seek(0)
        except Exception:
            pass
    return pd.DataFrame(rows)


def _post_restore_initialize_aux_tables_quietly() -> list:
    """
    復元後の補助テーブル再初期化を静かに行う。
    SQLite DB本体を復元した直後は、監査ログ・権限・バックアップ履歴の保存で
    一時的に失敗することがあるため、画面には警告を出さず、内部メモだけ残す。
    """
    messages = []
    try:
        initialize_sqlite_engine()
        messages.append("SQLiteエンジン再初期化OK")
    except Exception as e:
        messages.append(f"SQLiteエンジン再初期化は保留: {e}")
    try:
        ensure_security_tables()
        messages.append("セキュリティ補助テーブル確認OK")
    except Exception as e:
        messages.append(f"セキュリティ補助テーブル確認は保留: {e}")
    try:
        initialize_default_app_settings()
        messages.append("設定テーブル確認OK")
    except Exception as e:
        messages.append(f"設定テーブル確認は保留: {e}")
    return messages


def _safe_add_restore_audit_and_history(restore_zip_path: Path, summary_text: str, pre_backup: Path = None):
    """
    復元完了後の監査ログ・履歴保存。
    ここで失敗しても復元本体を失敗扱いにしない。
    """
    notes = []
    try:
        add_audit_log("データ復元", "restore", restore_zip_path.name, f"復元対象：{summary_text}")
        notes.append("監査ログ記録OK")
    except Exception as e:
        notes.append(f"監査ログ記録は保留: {e}")
    try:
        record_backup_history("復元", restore_zip_path, "成功", f"復元対象：{summary_text}")
        notes.append("バックアップ履歴記録OK")
    except Exception as e:
        notes.append(f"バックアップ履歴記録は保留: {e}")
    return notes

def restore_from_backup_zip(uploaded_file, restore_sqlite=True, restore_files=True, restore_supabase=False):
    """
    バックアップZIPから復元する。管理者のみ。

    Ver4.8.7 復元後の警告表示安全化:
    - SQLite DB復元直後の監査ログ/権限/バックアップ履歴テーブル初期化失敗を画面上のエラー風警告にしない
    - 復元本体（SQLite/Supabase/写真/帳票）が成功していれば、結果を成功として返す
    - 復元後に可能な範囲でSQLite安全再初期化・セキュリティテーブル再作成を行う

    Ver4.8.6:
    - Ver4.8.5完全バックアップZIPの新パスに対応
      SQLite/・Photos/・Attachments/・Reports/・System/
    - 旧バックアップZIPの data/・reports/ パスにも互換対応
    - 復元前バックアップを必ず作成
    - Supabase復元は安全優先のupsert方式
    """
    if not is_admin_user():
        return False, "管理者のみ復元できます。"
    ensure_security_dirs()

    if not restore_sqlite and not restore_files and not restore_supabase:
        return False, "復元対象を1つ以上選択してください。"

    pre_backup, pre_err = create_backup_zip(kind="復元前")
    if pre_err:
        return False, f"復元前バックアップに失敗しました：{pre_err}"

    restore_summary = []
    restore_notes = []

    # 復元中はSQLite補助テーブルの警告を抑制し、復元結果メッセージへ集約する。
    previous_suppress = bool(st.session_state.get("restore_suppress_sqlite_aux_warnings", False))
    st.session_state["restore_suppress_sqlite_aux_warnings"] = True

    try:
        filename = clean_text(getattr(uploaded_file, "name", "restore.zip"), "restore.zip")
        restore_zip_path = RESTORE_DIR / f"{format_now_jst('%Y%m%d_%H%M%S')}_{filename}"
        uploaded_file.seek(0)
        restore_zip_path.write_bytes(uploaded_file.read())

        with zipfile.ZipFile(restore_zip_path, "r") as zf:
            bad_file = zf.testzip()
            if bad_file:
                return False, f"ZIPファイルに破損があります：{bad_file}"

            names = zf.namelist()

            if restore_sqlite:
                restored_sqlite = _restore_sqlite_files_from_zip(zf, names)
                restore_summary.append("SQLite DB本体（" + "、".join(restored_sqlite) + "）")

            if restore_files:
                file_count = _restore_files_from_complete_or_legacy_zip(zf, names)
                restore_summary.append(f"写真・添付・帳票ファイル {file_count}件")

            if restore_supabase:
                sb_result = _restore_supabase_core_tables_from_backup(zf, names)
                if not sb_result.get("ok"):
                    return False, f"Supabase主要7機能の復元に失敗しました：{sb_result.get('error')}"
                sb_msg = " / ".join(sb_result.get("messages", []))
                restore_summary.append(f"Supabase主要7機能（{sb_msg}）")

        try:
            clear_hidamari_read_cache("バックアップ復元")
        except Exception:
            pass

        # 復元後に、補助テーブルを可能な範囲で静かに整える。
        restore_notes.extend(_post_restore_initialize_aux_tables_quietly())

        try:
            clear_hidamari_read_cache("バックアップ復元")
        except Exception:
            pass

        summary_text = "、".join(restore_summary) if restore_summary else "復元対象なし"
        restore_notes.extend(_safe_add_restore_audit_and_history(restore_zip_path, summary_text, pre_backup))

        note_text = " / ".join([n for n in restore_notes if n])
        note_suffix = f" 補足：{note_text}" if note_text else ""
        return True, f"復元しました。対象：{summary_text}。復元前バックアップも作成済みです：{pre_backup.name if pre_backup else ''}。{note_suffix}"
    except Exception as e:
        try:
            # 失敗ログも、ここで失敗して復元エラーを上書きしない。
            add_audit_log("データ復元失敗", "restore", "", str(e))
        except Exception:
            pass
        return False, f"復元に失敗しました：{e}"
    finally:
        try:
            st.session_state["restore_suppress_sqlite_aux_warnings"] = previous_suppress
        except Exception:
            pass

# =========================
# 起動時DB安全チェック・異常時復元（Ver4.8）
# 自動復元はしない。異常を検知した場合のみ、管理者確認で最新バックアップから復元する。
# =========================
STARTUP_CRITICAL_TABLES = [
    SQLITE_TABLE_USERS,
    SQLITE_TABLE_HEALTH,
    SQLITE_TABLE_EXCRETION,
    SQLITE_TABLE_HANDOVER,
    SQLITE_TABLE_SHORT_GOAL_CHECKS,
    SQLITE_TABLE_APP_SETTINGS,
]

def get_latest_backup_zip():
    """最新のバックアップZIPを返す。なければNone。"""
    try:
        ensure_security_dirs()
        backups = sorted(
            BACKUP_DIR.glob("hidamari_backup_*.zip"),
            key=lambda x: x.stat().st_mtime,
            reverse=True,
        )
        return backups[0] if backups else None
    except Exception:
        return None


def _startup_sqlite_quick_check():
    """SQLite quick_checkを直接実行する。起動前検査なので本体初期化に依存しすぎない。"""
    try:
        if not HIDAMARI_DB_FILE.exists():
            return False, "DBファイルが存在しません。"
        conn = sqlite3.connect(HIDAMARI_DB_FILE, timeout=DB_BUSY_TIMEOUT_MS / 1000)
        try:
            cur = conn.execute("PRAGMA quick_check;")
            result = cur.fetchone()
            text = str(result[0]) if result else ""
            ok = text.lower() == "ok"
            return ok, text or "quick_check結果なし"
        finally:
            conn.close()
    except Exception as e:
        return False, str(e)


def _startup_table_counts():
    """主要テーブルの存在と件数を確認する。"""
    rows = []
    try:
        if not HIDAMARI_DB_FILE.exists():
            return rows
        conn = sqlite3.connect(HIDAMARI_DB_FILE, timeout=DB_BUSY_TIMEOUT_MS / 1000)
        try:
            existing = set(
                r[0] for r in conn.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%'"
                ).fetchall()
            )
            for table_name in STARTUP_CRITICAL_TABLES:
                if table_name not in existing:
                    rows.append({
                        "テーブル": table_name,
                        "状態": "未作成",
                        "件数": "",
                        "詳細": "主要テーブルが見つかりません",
                    })
                    continue
                try:
                    count = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
                    rows.append({
                        "テーブル": table_name,
                        "状態": "OK",
                        "件数": int(count),
                        "詳細": "",
                    })
                except Exception as e:
                    rows.append({
                        "テーブル": table_name,
                        "状態": "確認エラー",
                        "件数": "",
                        "詳細": str(e),
                    })
        finally:
            conn.close()
    except Exception as e:
        rows.append({"テーブル": "SQLite接続", "状態": "確認エラー", "件数": "", "詳細": str(e)})
    return rows


def check_startup_database_health():
    """
    起動時の安全確認。
    戻り値:
      ok=True なら通常起動。
      ok=False なら本体起動を止め、管理者だけ復元操作を出す。
    """
    latest_backup = get_latest_backup_zip()
    result = {
        "ok": True,
        "severity": "normal",
        "messages": [],
        "db_exists": HIDAMARI_DB_FILE.exists(),
        "quick_check_ok": None,
        "quick_check_detail": "",
        "table_rows": [],
        "latest_backup": latest_backup,
        "checked_at": format_now_jst("%Y-%m-%d %H:%M:%S"),
    }

    # 管理者が「新規DBとして続行」を明示した場合だけ、同一セッションでは止めない。
    if st.session_state.get("startup_db_continue_without_restore"):
        result["messages"].append("管理者確認により、このセッションでは新規DBとして続行します。")
        return result

    if not HIDAMARI_DB_FILE.exists():
        if latest_backup:
            result["ok"] = False
            result["severity"] = "missing_db_with_backup"
            result["messages"].append("SQLite DBファイルが見つかりません。最新バックアップから復元できます。")
        else:
            # 初回起動など、バックアップがない場合は通常初期化へ進める
            result["messages"].append("SQLite DBファイルは未作成です。バックアップが無いため初回起動として初期化します。")
        return result

    quick_ok, quick_detail = _startup_sqlite_quick_check()
    result["quick_check_ok"] = quick_ok
    result["quick_check_detail"] = quick_detail
    if not quick_ok:
        result["ok"] = False
        result["severity"] = "corrupt_db"
        result["messages"].append(f"SQLite整合性チェックに失敗しました：{quick_detail}")

    table_rows = _startup_table_counts()
    result["table_rows"] = table_rows

    missing_tables = [r["テーブル"] for r in table_rows if r.get("状態") == "未作成"]
    error_tables = [r["テーブル"] for r in table_rows if r.get("状態") == "確認エラー"]

    if latest_backup and missing_tables:
        result["ok"] = False
        result["severity"] = "missing_tables"
        result["messages"].append("主要テーブルが不足しています：" + "、".join(missing_tables))

    if error_tables:
        result["ok"] = False
        result["severity"] = "table_error"
        result["messages"].append("主要テーブルの確認に失敗しました：" + "、".join(error_tables))

    # バックアップがあるのに主要データがすべて0件の場合は、Streamlit Cloud再起動等による消失疑いとして止める。
    key_tables = [SQLITE_TABLE_USERS, SQLITE_TABLE_HEALTH, SQLITE_TABLE_EXCRETION, SQLITE_TABLE_HANDOVER]
    count_map = {}
    for row in table_rows:
        try:
            if row.get("状態") == "OK":
                count_map[row.get("テーブル")] = int(row.get("件数", 0))
        except Exception:
            pass
    key_total = sum(count_map.get(t, 0) for t in key_tables)
    if latest_backup and count_map and key_total == 0:
        result["ok"] = False
        result["severity"] = "empty_core_tables_with_backup"
        result["messages"].append("バックアップが存在するのに主要データが0件です。データ消失の可能性があります。")

    return result


def restore_from_backup_path(backup_path: Path):
    """サーバー内の既存バックアップZIPから復元する。管理者のみ。"""
    if not is_admin_user():
        return False, "管理者のみ復元できます。"
    try:
        backup_path = Path(backup_path)
        ensure_security_dirs()
        if not backup_path.exists():
            return False, "指定されたバックアップファイルが見つかりません。"
        # BACKUP_DIR配下のZIPだけ許可
        try:
            backup_path.resolve().relative_to(BACKUP_DIR.resolve())
        except Exception:
            return False, "安全のため、バックアップフォルダ内のZIPのみ復元できます。"

        pre_backup, pre_err = create_backup_zip(kind="起動時復元前")
        if pre_err:
            return False, f"復元前バックアップに失敗しました：{pre_err}"

        with zipfile.ZipFile(backup_path, "r") as zf:
            names = zf.namelist()
            if f"data/{HIDAMARI_DB_FILE.name}" not in names:
                return False, "このZIPには hidamari_health.db が含まれていません。"

            ensure_dirs()
            HIDAMARI_DB_FILE.write_bytes(zf.read(f"data/{HIDAMARI_DB_FILE.name}"))

            if "data/hidamari_life.db" in names:
                (DATA_DIR / "hidamari_life.db").write_bytes(zf.read("data/hidamari_life.db"))

            for folder in [BUSINESS_HANDOVER_PHOTO_DIR, BUSINESS_HANDOVER_EXCEL_DIR, REPORT_DIR]:
                for name in names:
                    if name.startswith(str(folder)) and not name.endswith("/"):
                        target = Path(name)
                        target.parent.mkdir(parents=True, exist_ok=True)
                        target.write_bytes(zf.read(name))

        st.session_state.pop("startup_db_continue_without_restore", None)
        add_audit_log("起動時データ復元", "restore", backup_path.name, "起動時異常検知後、最新バックアップZIPから復元")
        record_backup_history("起動時復元", backup_path, "成功", "起動時異常検知後、最新バックアップZIPから復元")
        return True, f"最新バックアップから復元しました：{backup_path.name} / 復元前バックアップ：{pre_backup.name if pre_backup else ''}"
    except Exception as e:
        try:
            add_audit_log("起動時データ復元失敗", "restore", "", str(e))
        except Exception:
            pass
        return False, f"復元に失敗しました：{e}"


def show_startup_recovery_panel(check_result: dict):
    """
    起動時異常画面。
    異常時は本体メニューを出さず、管理者にだけ復元ボタンを表示する。
    """
    st.error("起動時のデータ確認で異常を検知しました。")
    st.caption("古いバックアップを誤って読み込まないため、自動復元は行いません。管理者が確認して復元します。")

    messages = check_result.get("messages") or []
    for msg in messages:
        st.warning(msg)

    rows = [
        {"確認項目": "DB存在チェック", "状態": "OK" if check_result.get("db_exists") else "NG", "詳細": str(HIDAMARI_DB_FILE)},
        {"確認項目": "SQLite整合性チェック", "状態": "OK" if check_result.get("quick_check_ok") else "NG", "詳細": check_result.get("quick_check_detail", "")},
        {"確認項目": "検査日時", "状態": "情報", "詳細": check_result.get("checked_at", "")},
    ]
    st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    table_rows = check_result.get("table_rows") or []
    if table_rows:
        st.markdown("#### 主要テーブル件数チェック")
        st.dataframe(pd.DataFrame(table_rows), use_container_width=True, hide_index=True)

    latest_backup = check_result.get("latest_backup")
    st.markdown("#### 復元候補")
    if latest_backup:
        try:
            st.info(f"最新バックアップ：{latest_backup.name} / サイズ：{round(latest_backup.stat().st_size / 1024, 1)} KB")
        except Exception:
            st.info(f"最新バックアップ：{latest_backup.name}")
    else:
        st.warning("復元に使えるバックアップZIPが見つかりません。")

    if not is_admin_user():
        st.info("この画面は管理者確認が必要です。管理者でログインし直してください。")
        return False

    if latest_backup:
        confirm = st.checkbox("最新バックアップから復元することを理解しました", key="startup_restore_confirm")
        if st.button("最新バックアップから復元して起動する", type="primary", use_container_width=True):
            if not confirm:
                st.error("確認チェックを入れてください。")
            else:
                ok, msg = restore_from_backup_path(latest_backup)
                if ok:
                    st.success(msg)
                    st.info("復元後に画面を再読み込みします。")
                    st.rerun()
                else:
                    st.error(msg)

    with st.expander("管理者用：復元せず新規DBとして続行", expanded=False):
        st.warning("初回導入時など、バックアップが不要な場合だけ使います。既存データがあるはずの運用中は押さないでください。")
        confirm_new = st.checkbox("復元せずに続行することを理解しました", key="startup_continue_confirm")
        if st.button("復元せずに新規DBとして続行", use_container_width=True):
            if not confirm_new:
                st.error("確認チェックを入れてください。")
            else:
                st.session_state["startup_db_continue_without_restore"] = True
                try:
                    add_audit_log("起動時復元スキップ", "startup_check", "", "管理者が復元せず新規DBとして続行")
                except Exception:
                    pass
                st.rerun()
    return False


def run_startup_database_guard():
    """
    ログイン後・本体初期化前に実行する。
    Trueなら通常起動、Falseならst.stop()前提。
    """
    try:
        check_result = check_startup_database_health()
        st.session_state["startup_db_check_result"] = check_result
        if check_result.get("ok", True):
            return True
        show_startup_recovery_panel(check_result)
        return False
    except Exception as e:
        st.error(f"起動時データ確認に失敗しました：{e}")
        if not is_admin_user():
            st.info("管理者でログインし直してください。")
            return False
        st.warning("管理者のみ、初回起動として続行できます。")
        if st.button("管理者として初回起動を続行", use_container_width=True):
            st.session_state["startup_db_continue_without_restore"] = True
            st.rerun()
        return False


# =========================
# バックアップ整合性チェック（Ver4.8.8）
# 本番データを復元・更新せず、現在Supabase件数とバックアップZIP内件数を比較する。
# =========================
def _supabase_backup_check_targets():
    return [
        (SQLITE_TABLE_USERS, "利用者マスタ"),
        (SQLITE_TABLE_HEALTH, "健康チェック"),
        (SQLITE_TABLE_EXCRETION, "排泄チェック"),
        (SQLITE_TABLE_HANDOVER, "業務全体申し送り"),
        (SQLITE_TABLE_SHORT_GOAL_CHECKS, "短期目標実施チェック"),
        (SQLITE_TABLE_SHORT_GOAL_MASTER, "短期目標マスタ"),
        (SQLITE_TABLE_MONITORING_DRAFTS, "モニタリング下書き"),
    ]


def _count_current_supabase_records_uncached(table_name: str) -> tuple:
    """
    現在のSupabase本番テーブル件数を、画面表示キャッシュを使わずに取得する。
    戻り値：(件数, メッセージ)
    """
    if not ("supabase_is_enabled" in globals() and supabase_is_enabled()):
        return None, "Supabase未設定または接続不可"
    if requests is None:
        return None, "requestsが利用できません"
    try:
        total = 0
        page_size = 1000
        offset = 0
        while True:
            headers = _supabase_headers(prefer="")
            headers["Range-Unit"] = "items"
            headers["Range"] = f"{offset}-{offset + page_size - 1}"
            res = requests.get(
                _supabase_endpoint(table_name),
                headers=headers,
                params=[("select", "record_key"), ("order", "updated_at.asc")],
                timeout=30,
            )
            if res.status_code not in [200, 206]:
                res.raise_for_status()
            rows = res.json() or []
            total += len(rows)
            if len(rows) < page_size:
                break
            offset += page_size
        return int(total), ""
    except Exception as e:
        return None, str(e)


def _clean_backup_records_for_count(records):
    if not isinstance(records, list):
        return []
    clean = []
    for item in records:
        if isinstance(item, dict) and "backup_error" not in item:
            clean.append(item)
    return clean


def _read_backup_supabase_counts_from_zip(uploaded_or_path) -> tuple:
    """
    バックアップZIP内のSupabase主要7機能件数を読む。
    本番データには一切書き込まない。
    対応：
      - exports/supabase_core_tables.json
      - exports/supabase_core_tables.xlsx
      - Supabase/*.xlsx
    """
    counts = {}
    messages = []

    try:
        if hasattr(uploaded_or_path, "getvalue"):
            zip_bytes = uploaded_or_path.getvalue()
            zip_source = BytesIO(zip_bytes)
        else:
            zip_source = uploaded_or_path

        with zipfile.ZipFile(zip_source, "r") as zf:
            names = zf.namelist()

            # 最優先：復元互換JSON
            if "exports/supabase_core_tables.json" in names:
                raw = zf.read("exports/supabase_core_tables.json").decode("utf-8")
                data = json.loads(raw) if raw.strip() else {}
                if isinstance(data, dict):
                    for table_name, _label in _supabase_backup_check_targets():
                        counts[table_name] = len(_clean_backup_records_for_count(data.get(table_name, [])))
                    return counts, messages
                messages.append("exports/supabase_core_tables.json の形式が正しくありません。")

            # 次点：互換Excel
            if "exports/supabase_core_tables.xlsx" in names:
                xls_bytes = BytesIO(zf.read("exports/supabase_core_tables.xlsx"))
                for table_name, _label in _supabase_backup_check_targets():
                    try:
                        df = pd.read_excel(xls_bytes, sheet_name=table_name)
                        # 次のシート読込のためにBytesIO位置を戻す
                        xls_bytes.seek(0)
                        if "backup_error" in df.columns:
                            df = df[df["backup_error"].isna()]
                        counts[table_name] = int(len(df))
                    except Exception as e:
                        xls_bytes.seek(0)
                        counts[table_name] = None
                        messages.append(f"{table_name}: Excel読込不可 / {e}")
                return counts, messages

            # 完全バックアップの用途別Excel
            file_map = {
                SQLITE_TABLE_USERS: "Supabase/users.xlsx",
                SQLITE_TABLE_HEALTH: "Supabase/health.xlsx",
                SQLITE_TABLE_EXCRETION: "Supabase/excretion.xlsx",
                SQLITE_TABLE_HANDOVER: "Supabase/handover.xlsx",
                SQLITE_TABLE_SHORT_GOAL_CHECKS: "Supabase/short_goal_checks.xlsx",
                SQLITE_TABLE_SHORT_GOAL_MASTER: "Supabase/short_goal_master.xlsx",
                SQLITE_TABLE_MONITORING_DRAFTS: "Supabase/monitoring.xlsx",
            }
            found_any = False
            for table_name, member in file_map.items():
                if member in names:
                    found_any = True
                    try:
                        df = pd.read_excel(BytesIO(zf.read(member)))
                        # 空テーブル用のmessage行だけの場合は0件扱い
                        if list(df.columns) == ["message"]:
                            counts[table_name] = 0
                        elif "backup_error" in df.columns:
                            counts[table_name] = int(len(df[df["backup_error"].isna()]))
                        else:
                            counts[table_name] = int(len(df))
                    except Exception as e:
                        counts[table_name] = None
                        messages.append(f"{member}: 読込不可 / {e}")
            if found_any:
                return counts, messages

            messages.append("SupabaseバックアップデータがZIP内に見つかりません。")
            return counts, messages
    except Exception as e:
        messages.append(f"ZIP読込エラー: {e}")
        return counts, messages


def check_backup_integrity_against_current_supabase(uploaded_or_path) -> dict:
    """
    バックアップ整合性チェック本体。
    復元・保存・削除は一切しない。
    """
    backup_counts, messages = _read_backup_supabase_counts_from_zip(uploaded_or_path)

    rows = []
    matched = 0
    checked = 0

    for table_name, label in _supabase_backup_check_targets():
        current_count, current_msg = _count_current_supabase_records_uncached(table_name)
        backup_count = backup_counts.get(table_name, None)

        if current_count is None:
            status = "現在件数取得不可"
            detail = current_msg
        elif backup_count is None:
            status = "バックアップ件数取得不可"
            detail = "ZIP内に対象データがない、または読込できません。"
        else:
            checked += 1
            diff = int(backup_count) - int(current_count)
            if diff == 0:
                matched += 1
                status = "一致"
                detail = "OK"
            else:
                status = "差異あり"
                detail = f"バックアップ - 現在 = {diff:+d} 件"

        rows.append({
            "対象": label,
            "テーブル": table_name,
            "現在Supabase件数": "" if current_count is None else int(current_count),
            "バックアップ件数": "" if backup_count is None else int(backup_count),
            "判定": status,
            "詳細": detail,
        })

    match_rate = round((matched / checked * 100), 1) if checked else 0
    overall = "一致" if checked and matched == checked else ("確認が必要" if checked else "判定不可")

    return {
        "overall": overall,
        "match_rate": match_rate,
        "checked": checked,
        "matched": matched,
        "messages": messages,
        "df": pd.DataFrame(rows),
    }

def show_security_maintenance_menu():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    ensure_security_tables()
    st.header("セキュリティ・保守管理")
    st.caption("DBバックアップ・復元・監査ログ・権限管理を、この画面にまとめています。")

    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs(["バックアップ", "監査ログ", "権限管理", "データ復元", "DB整合性", "Supabase設定", "写真設定"])

    with tab1:
        st.subheader("バックアップ")
        st.write("SQLite DB、SQLite全テーブルExcel、Supabase主要7機能、写真・添付ファイルをZIPで保存します。")
        st.markdown("#### バックアップ対象の検査")
        try:
            st.dataframe(get_backup_target_status_df(), use_container_width=True, hide_index=True)
        except Exception as e:
            st.warning(f"バックアップ対象の検査表示に失敗しました：{e}")

        st.divider()
        st.markdown("#### バックアップ整合性チェック")
        st.caption("本番Supabaseには一切書き込まず、現在件数とバックアップZIP内件数を比較します。復元テストの代わりに安全確認できます。")

        check_source = st.radio(
            "チェックするバックアップの選び方",
            ["保存済みバックアップから選ぶ", "ZIPをアップロードする"],
            horizontal=True,
            key="backup_integrity_source",
        )

        integrity_target = None
        if check_source == "保存済みバックアップから選ぶ":
            try:
                saved_backups = sorted(
                    list(BACKUP_DIR.glob("hidamari_backup_*.zip")) + list(BACKUP_DIR.glob("hidamari_complete_backup_*.zip")),
                    key=lambda x: x.stat().st_mtime,
                    reverse=True,
                )
            except Exception:
                saved_backups = []
            if saved_backups:
                selected_integrity = st.selectbox(
                    "整合性チェックする保存済みバックアップ",
                    [b.name for b in saved_backups],
                    key="selected_integrity_backup",
                )
                integrity_target = BACKUP_DIR / selected_integrity
            else:
                st.info("保存済みバックアップが見つかりません。ZIPをアップロードして確認してください。")
        else:
            integrity_upload = st.file_uploader(
                "整合性チェックするバックアップZIPを選択",
                type=["zip"],
                key="backup_integrity_upload",
            )
            integrity_target = integrity_upload

        if st.button("バックアップ整合性チェックを実行", use_container_width=True):
            if integrity_target is None:
                st.error("チェックするバックアップZIPを選択してください。")
            elif not ("supabase_is_enabled" in globals() and supabase_is_enabled()):
                st.error("Supabaseが未設定または接続できません。")
            else:
                with st.spinner("バックアップZIPと現在Supabase件数を照合しています..."):
                    result = check_backup_integrity_against_current_supabase(integrity_target)
                if result.get("overall") == "一致":
                    st.success(f"整合性OK：{result.get('matched')}/{result.get('checked')} テーブル一致（一致率 {result.get('match_rate')}%）")
                elif result.get("overall") == "確認が必要":
                    st.warning(f"確認が必要：{result.get('matched')}/{result.get('checked')} テーブル一致（一致率 {result.get('match_rate')}%）")
                else:
                    st.error("判定できませんでした。バックアップZIPの形式を確認してください。")
                for msg in result.get("messages", []):
                    st.info(msg)
                check_df = result.get("df", pd.DataFrame())
                if isinstance(check_df, pd.DataFrame) and not check_df.empty:
                    st.dataframe(check_df, use_container_width=True, hide_index=True)
                    st.download_button(
                        "整合性チェック結果をExcelでダウンロード",
                        data=to_excel_download(check_df),
                        file_name=f"backup_integrity_check_{today_jst().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )
                try:
                    add_audit_log(
                        "バックアップ整合性チェック",
                        "backup",
                        "",
                        f"結果:{result.get('overall')} / 一致率:{result.get('match_rate')}%",
                    )
                except Exception:
                    pass

        st.divider()

        if st.button("今すぐ手動バックアップを作成", type="primary", use_container_width=True):
            zip_path, err = create_backup_zip(kind="手動")
            if err:
                st.error(err)
            else:
                st.success(f"バックアップを作成しました：{zip_path.name}")

        backups = sorted(
            list(BACKUP_DIR.glob("hidamari_backup_*.zip")) + list(BACKUP_DIR.glob("hidamari_complete_backup_*.zip")),
            key=lambda x: x.stat().st_mtime,
            reverse=True,
        )
        if backups:
            selected = st.selectbox("ダウンロードするバックアップ", [b.name for b in backups])
            target = BACKUP_DIR / selected
            with open(target, "rb") as f:
                st.download_button(
                    "選択したバックアップをダウンロード",
                    data=f.read(),
                    file_name=target.name,
                    mime="application/zip",
                    use_container_width=True,
                )
        else:
            st.info("バックアップファイルはまだありません。")

        st.divider()
        st.subheader("バックアップ履歴")
        history = load_sqlite_table(SQLITE_TABLE_BACKUP_HISTORY, BACKUP_HISTORY_COLUMNS)
        if history.empty:
            st.info("履歴はまだありません。")
        else:
            st.dataframe(history.sort_values("日時", ascending=False).head(200), use_container_width=True, hide_index=True)

    with tab2:
        st.subheader("監査ログ")
        logs = load_sqlite_table(SQLITE_TABLE_AUDIT_LOGS, AUDIT_LOG_COLUMNS)
        c1, c2, c3 = st.columns(3)
        with c1:
            op_filter = st.text_input("操作種別で検索", key="audit_op_filter")
        with c2:
            user_filter = st.text_input("ログインIDで検索", key="audit_user_filter")
        with c3:
            limit = st.number_input("表示件数", min_value=50, max_value=1000, value=200, step=50)
        view = logs.copy()
        if not view.empty:
            if clean_text(op_filter):
                view = view[view["操作種別"].astype(str).str.contains(clean_text(op_filter), case=False, na=False)]
            if clean_text(user_filter):
                view = view[view["ログインID"].astype(str).str.contains(clean_text(user_filter), case=False, na=False)]
            view = view.sort_values("日時", ascending=False).head(int(limit))
            st.dataframe(view, use_container_width=True, hide_index=True)
            st.download_button(
                "監査ログをExcelでダウンロード",
                data=to_excel_download(view),
                file_name=f"audit_logs_{today_jst().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        else:
            st.info("監査ログはまだありません。")

    with tab3:
        st.subheader("権限管理")
        st.caption("職員権限で閲覧・登録更新・削除・復元をどこまで許可するかを管理します。")
        perms = load_sqlite_table(SQLITE_TABLE_ROLE_PERMISSIONS, ROLE_PERMISSION_COLUMNS)
        edited = st.data_editor(
            perms,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "閲覧": st.column_config.CheckboxColumn("閲覧"),
                "登録更新": st.column_config.CheckboxColumn("登録更新"),
                "削除": st.column_config.CheckboxColumn("削除"),
                "復元": st.column_config.CheckboxColumn("復元"),
            },
        )
        if st.button("権限設定を保存", type="primary", use_container_width=True):
            save_sqlite_table(edited, SQLITE_TABLE_ROLE_PERMISSIONS, ROLE_PERMISSION_COLUMNS, unique_cols=["権限", "機能"])
            add_audit_log("権限設定更新", "role_permissions", "", "権限設定を更新")
            st.success("権限設定を保存しました。")
            st.rerun()

    with tab4:
        st.subheader("データ復元")
        st.warning("復元すると選択した対象がバックアップ時点の内容で復元されます。実行前に自動で『復元前バックアップ』を作成します。")

        uploaded = st.file_uploader("復元するバックアップZIPを選択", type=["zip"])

        st.markdown("#### 復元対象を選択")
        restore_sqlite = st.checkbox(
            "SQLiteだけ復元（設定・監査ログ・バックアップ履歴など）",
            value=True,
            help="ローカルSQLite DB本体を復元します。Supabase正本のデータ表示には直接反映されない場合があります。",
        )
        restore_files = st.checkbox(
            "写真・添付ファイルを復元",
            value=True,
            help="申し送り写真・添付Excel等を復元します。",
        )
        restore_supabase = st.checkbox(
            "Supabase主要7機能も復元（利用者・健康・排泄・申し送り・短期目標チェック・短期目標マスタ・モニタリング下書き）",
            value=False,
            help="バックアップZIP内の Supabase退避データを使い、主要7機能をSupabaseへ戻します。安全優先のupsert方式です。",
        )

        if restore_supabase:
            st.info("Supabase復元は、バックアップに含まれるデータを上書き・追加します。安全のため、現在Supabaseにある別データの全削除は行いません。")
            supabase_confirm_text = st.text_input(
                "Supabase復元を実行する場合は「SUPABASE復元」と入力",
                key="supabase_restore_confirm_text",
            )
        else:
            supabase_confirm_text = ""

        confirm = st.checkbox("現在のデータが上書き・追加復元されることを理解しました")

        if st.button("選択したバックアップから復元", type="primary", use_container_width=True):
            if not uploaded:
                st.error("復元するZIPを選択してください。")
            elif not confirm:
                st.error("確認チェックを入れてください。")
            elif restore_supabase and supabase_confirm_text.strip() != "SUPABASE復元":
                st.error("Supabase主要7機能を復元する場合は、確認欄に「SUPABASE復元」と入力してください。")
            elif not restore_sqlite and not restore_files and not restore_supabase:
                st.error("復元対象を1つ以上選択してください。")
            else:
                ok, msg = restore_from_backup_zip(
                    uploaded,
                    restore_sqlite=restore_sqlite,
                    restore_files=restore_files,
                    restore_supabase=restore_supabase,
                )
                if ok:
                    st.success(msg)
                    st.info("復元後はアプリを再読み込みしてください。Supabase復元を行った場合は、画面表示もSupabase側の復元内容に戻ります。")
                else:
                    st.error(msg)

    with tab5:
        st.subheader("DB整合性チェック")
        st.caption("SQLiteのWALモード、quick_check、チェックポイント状態を確認します。")
        status_text = get_db_integrity_status_text() if "get_db_integrity_status_text" in globals() else "DB整合性: 未確認"
        if DB_LAST_INTEGRITY_RESULT.get("ok", True):
            st.success(status_text)
        else:
            st.error(status_text)
        if st.button("DB整合性を今すぐ再チェック", type="primary", use_container_width=True):
            result = run_db_integrity_check(auto_repair=True)
            if result.get("ok"):
                st.success(get_db_integrity_status_text())
                add_audit_log("DB整合性チェック", "sqlite", "", "quick_check ok")
            else:
                st.error(get_db_integrity_status_text())
                add_audit_log("DB整合性チェックエラー", "sqlite", "", " / ".join(result.get("messages", [])))
            st.rerun()

    with tab6:
        st.subheader("Supabase設定")
        st.caption("利用者マスタ・健康チェック・排泄チェック・業務全体申し送り・短期目標実施チェック・短期目標マスタ・モニタリング下書きの7機能をSupabase正本として保存します。その他の補助機能は従来通りSQLite＋バックアップ方式です。")
        status = get_supabase_storage_status() if "get_supabase_storage_status" in globals() else "Supabase設定関数が見つかりません。"
        if "接続OK" in status:
            st.success(status)
        elif "未設定" in status or "未有効" in status:
            st.info(status)
        else:
            st.warning(status)

        st.markdown("#### Supabase接続診断")
        try:
            diag_df = get_supabase_diagnostic_rows() if "get_supabase_diagnostic_rows" in globals() else pd.DataFrame()
            if not diag_df.empty:
                st.dataframe(diag_df, use_container_width=True, hide_index=True)
        except Exception as e:
            st.warning(f"診断表示に失敗しました：{e}")

        st.markdown("#### Streamlit Secrets 設定例")
        st.code('''[supabase]
enabled = true
url = "https://huufblmiqvloudeqctjp.supabase.co"
key = "sb_secret_xxxxxxxxxxxxxxxxx"''', language="toml")

        st.markdown("#### Supabase SQL Editorで実行するSQL")
        st.code(get_supabase_create_table_sql() if "get_supabase_create_table_sql" in globals() else "", language="sql")

        st.markdown("#### 保存方針")
        st.write(get_storage_unification_status() if "get_storage_unification_status" in globals() else {})

        st.markdown("#### 既存SQLiteデータの移行")
        st.caption("手元や旧環境のSQLiteに残っている主要7機能データを、Supabaseへ初回移行するためのボタンです。")
        if st.button("ローカルSQLiteから主要7機能をSupabaseへ移行", use_container_width=True):
            if not supabase_is_enabled():
                st.error("Supabaseが未設定または接続できません。Secretsとテーブル作成を確認してください。")
            else:
                migrated = []
                targets = [
                    (SQLITE_TABLE_USERS, USER_COLUMNS, ["user_id"]),
                    (SQLITE_TABLE_HEALTH, HEALTH_COLUMNS, ["記録日", "user_id"]),
                    (SQLITE_TABLE_EXCRETION, EXCRETION_COLUMNS, ["記録日", "user_id", "時間帯"]),
                    (SQLITE_TABLE_HANDOVER, BUSINESS_HANDOVER_COLUMNS, ["記録ID"]),
                    (SQLITE_TABLE_SHORT_GOAL_CHECKS, SHORT_GOAL_CHECK_COLUMNS, ["記録ID"]),
                    (SQLITE_TABLE_SHORT_GOAL_MASTER, SHORT_GOAL_MASTER_COLUMNS, ["目標ID"]),
                    (SQLITE_TABLE_MONITORING_DRAFTS, MONITORING_DRAFT_COLUMNS, ["下書きID"]),
                ]
                for table_name, cols, keys in targets:
                    local_df = _original_load_sqlite_table(table_name, cols)
                    ok = supabase_replace_table(local_df, table_name, columns=cols, unique_cols=keys)
                    migrated.append({"テーブル": table_name, "件数": len(local_df), "結果": "移行OK" if ok else "移行失敗"})
                st.dataframe(pd.DataFrame(migrated), use_container_width=True, hide_index=True)

    with tab7:
        st.subheader("写真設定")
        st.caption("申し送り写真の容量を抑え、Supabase無料枠やSQLite DBの肥大化を防ぎます。")
        settings = get_photo_settings()

        c1, c2, c3 = st.columns(3)
        with c1:
            auto_compress = st.checkbox("自動圧縮を有効にする", value=bool(settings.get("auto_compress", True)))
        with c2:
            max_kb = st.number_input("圧縮後の最大サイズ（KB）", min_value=100, max_value=1024, value=int(settings.get("max_kb", PHOTO_MAX_DISPLAY_KB)), step=50)
        with c3:
            max_width = st.number_input("画像の最大長辺（px）", min_value=480, max_value=1600, value=int(settings.get("max_width", PHOTO_MAX_WIDTH)), step=80)

        c4, c5 = st.columns(2)
        with c4:
            retention_days = st.number_input("写真保存期間（日）", min_value=30, max_value=3650, value=int(settings.get("retention_days", PHOTO_RETENTION_DAYS)), step=30)
        with c5:
            backup_before_delete = st.checkbox("削除前にバックアップを作成する", value=bool(settings.get("backup_before_delete", True)))

        st.info(f"現在の標準：写真1枚／自動圧縮／最大{int(max_kb)}KB／保存期間{int(retention_days)}日／削除前バックアップ")

        if st.button("写真設定を保存", type="primary", use_container_width=True):
            set_app_setting(
                "photo_storage_settings",
                {
                    "auto_compress": bool(auto_compress),
                    "max_kb": int(max_kb),
                    "max_width": int(max_width),
                    "retention_days": int(retention_days),
                    "backup_before_delete": bool(backup_before_delete),
                },
                category="写真設定",
                description="申し送り写真の圧縮・保存期間設定",
            )
            add_audit_log("写真設定更新", "app_settings", "photo_storage_settings", "写真軽量化・保存期間設定を更新")
            st.success("写真設定を保存しました。")
            st.rerun()

        st.divider()
        st.subheader("写真保存期間の手動整理")
        st.caption("保存期間を超えた写真を、削除前バックアップ後に整理します。DB内のbase64写真も保存期間終了マーカーに置き換えます。")
        if st.button("期限切れ写真を今すぐ整理", use_container_width=True):
            result = cleanup_expired_handover_photos(
                retention_days=int(retention_days),
                backup_before_delete=bool(backup_before_delete),
            )
            if result.get("error"):
                st.error(result.get("error"))
            else:
                st.success(f"整理完了：対象確認 {result.get('checked', 0)}件／整理 {result.get('updated', 0)}件／削除ファイル {result.get('deleted', 0)}件")
                if result.get("backup"):
                    st.info(f"削除前バックアップ：{result.get('backup')}")




DEFAULT_ALERT_CONDITIONS = [
    {"条件ID": "C001", "使用": True, "条件名": "未排便3日", "重要度": "注意", "分類": "排泄", "条件種別": "未排便", "閾値1": 3, "閾値2": "", "日数": 3, "キーワード": "", "表示メッセージ": "直近{日数}日間、排便記録がありません。水分・食事量・腹部症状を確認してください。", "並び順": 10},
    {"条件ID": "C002", "使用": True, "条件名": "水様便・下痢便あり", "重要度": "注意", "分類": "排泄", "条件種別": "便性状", "閾値1": "", "閾値2": "", "日数": 1, "キーワード": "水様便,下痢便", "表示メッセージ": "水様便・下痢便の記録があります。回数・腹部症状・感染症状を確認してください。", "並び順": 20},
    {"条件ID": "C003", "使用": True, "条件名": "濃縮尿あり", "重要度": "観察", "分類": "排泄", "条件種別": "尿性状", "閾値1": "", "閾値2": "", "日数": 1, "キーワード": "濃縮尿", "表示メッセージ": "濃縮尿の記録があります。水分摂取量や発熱の有無を確認してください。", "並び順": 30},
    {"条件ID": "C004", "使用": True, "条件名": "食事量50%以下", "重要度": "観察", "分類": "食事", "条件種別": "食事低下", "閾値1": 50, "閾値2": "", "日数": 1, "キーワード": "", "表示メッセージ": "食事摂取率が{閾値1}%以下です。食欲・口腔状態・体調変化を確認してください。", "並び順": 40},
    {"条件ID": "C005", "使用": True, "条件名": "食事量50%以下が2日続く", "重要度": "注意", "分類": "食事", "条件種別": "食事低下連続", "閾値1": 50, "閾値2": "", "日数": 2, "キーワード": "", "表示メッセージ": "食事摂取率{閾値1}%以下が{日数}日続いています。継続観察と共有が必要です。", "並び順": 50},
    {"条件ID": "C006", "使用": True, "条件名": "発熱37.5℃以上", "重要度": "注意", "分類": "バイタル", "条件種別": "発熱", "閾値1": 37.5, "閾値2": "", "日数": 1, "キーワード": "", "表示メッセージ": "体温が{閾値1}℃以上です。再検・水分・食事・普段との違いを確認してください。", "並び順": 60},
    {"条件ID": "C007", "使用": True, "条件名": "SpO2 93%以下", "重要度": "注意", "分類": "バイタル", "条件種別": "SpO2低下", "閾値1": 93, "閾値2": "", "日数": 1, "キーワード": "", "表示メッセージ": "SpO2が{閾値1}%以下です。再測定・呼吸状態・顔色・傾眠の有無を確認してください。", "並び順": 70},
    {"条件ID": "C008", "使用": True, "条件名": "血圧上160以上", "重要度": "観察", "分類": "バイタル", "条件種別": "血圧高値", "閾値1": 160, "閾値2": "", "日数": 1, "キーワード": "", "表示メッセージ": "血圧上が{閾値1}以上です。再測定と普段との差を確認してください。", "並び順": 80},
    {"条件ID": "C009", "使用": True, "条件名": "1週間で体重1kg以上減少", "重要度": "観察", "分類": "体重", "条件種別": "体重減少", "閾値1": 1.0, "閾値2": "", "日数": 7, "キーワード": "", "表示メッセージ": "直近{日数}日で体重が{閾値1}kg以上減少しています。食事量・水分・むくみ等を確認してください。", "並び順": 90},
    {"条件ID": "C010", "使用": True, "条件名": "気になる変化キーワード", "重要度": "注意", "分類": "変化", "条件種別": "キーワード", "閾値1": "", "閾値2": "", "日数": 1, "キーワード": "不穏,傾眠,ふらつき,転倒,食欲なし,いつもと違う,拒否,痛み,息苦しい", "表示メッセージ": "気になる変化に注意キーワードがあります：{該当内容}", "並び順": 100},
    {"条件ID": "C011", "使用": True, "条件名": "発熱＋食事低下", "重要度": "至急", "分類": "複合", "条件種別": "複合:発熱+食事低下", "閾値1": 37.5, "閾値2": 50, "日数": 1, "キーワード": "", "表示メッセージ": "発熱と食事量低下が重なっています。体調変化として優先的に共有してください。", "並び順": 110},
    {"条件ID": "C012", "使用": True, "条件名": "濃縮尿＋食事水分低下", "重要度": "注意", "分類": "複合", "条件種別": "複合:濃縮尿+食事低下", "閾値1": 50, "閾値2": "", "日数": 1, "キーワード": "", "表示メッセージ": "濃縮尿と食事量低下が重なっています。脱水傾向に注意して水分摂取を確認してください。", "並び順": 120},
    {"条件ID": "C013", "使用": True, "条件名": "SpO2低下＋傾眠等", "重要度": "至急", "分類": "複合", "条件種別": "複合:SpO2低下+キーワード", "閾値1": 93, "閾値2": "", "日数": 1, "キーワード": "傾眠,息苦しい,顔色,呼吸,ぐったり", "表示メッセージ": "SpO2低下と気になる変化が重なっています。呼吸状態を優先確認してください。", "並び順": 130},
]

# =========================
# 共通関数
# =========================
def ensure_dirs():
    DATA_DIR.mkdir(exist_ok=True)
    REPORT_DIR.mkdir(exist_ok=True)
    BUSINESS_HANDOVER_PHOTO_DIR.mkdir(parents=True, exist_ok=True)
    BUSINESS_HANDOVER_EXCEL_DIR.mkdir(parents=True, exist_ok=True)



# =========================
# Ver4.0 利用者ID移行準備
# =========================
def normalize_user_name_for_match(name: str) -> str:
    """利用者名の表記ゆれ確認用。保存値は変えず、照合だけに使う。"""
    text = clean_text(name)
    text = re.sub(r"\s+", "", text)
    text = text.replace("　", "")
    text = text.replace("様", "").replace("さん", "").replace("殿", "")
    return text.lower()


def make_user_id_from_name(user_name: str) -> str:
    """既存利用者名から安定したuser_idを生成する。"""
    base = normalize_user_name_for_match(user_name) or clean_text(user_name)
    digest = hashlib.sha1(base.encode("utf-8")).hexdigest()[:12]
    return f"usr_{digest}"


def ensure_user_id_value(user_id, user_name) -> str:
    user_id = clean_text(user_id)
    return user_id if user_id else make_user_id_from_name(user_name)


def ensure_user_name_alias_table():
    """利用者名ゆれ紐づけマスタをSQLiteに用意する。"""
    try:
        if not sqlite_table_exists(SQLITE_TABLE_USER_NAME_ALIASES):
            save_sqlite_table(
                pd.DataFrame(columns=USER_NAME_ALIAS_COLUMNS),
                SQLITE_TABLE_USER_NAME_ALIASES,
                USER_NAME_ALIAS_COLUMNS,
                unique_cols=["alias_id"],
            )
    except Exception:
        pass


def normalize_user_name_alias_df(df: pd.DataFrame) -> pd.DataFrame:
    """表記ゆれマスタの列・値を整える。"""
    if df is None:
        df = pd.DataFrame(columns=USER_NAME_ALIAS_COLUMNS)
    work = df.copy()
    for col in USER_NAME_ALIAS_COLUMNS:
        if col not in work.columns:
            work[col] = ""
    work = work[USER_NAME_ALIAS_COLUMNS].copy()
    work["alias_id"] = work["alias_id"].map(lambda x: clean_text(x))
    work["表記ゆれ名"] = work["表記ゆれ名"].map(lambda x: clean_text(x))
    work["紐づけ先 user_id"] = work["紐づけ先 user_id"].map(lambda x: clean_text(x))
    work["正式利用者名"] = work["正式利用者名"].map(lambda x: clean_text(x))
    work["有効/無効"] = work["有効/無効"].map(lambda x: clean_text(x, "有効"))
    work.loc[~work["有効/無効"].isin(["有効", "無効"]), "有効/無効"] = "有効"
    work["備考"] = work["備考"].map(lambda x: clean_text(x))
    work["更新日時"] = work["更新日時"].map(lambda x: clean_text(x))
    work["更新者"] = work["更新者"].map(lambda x: clean_text(x))
    # 空の表記ゆれ名は除外。同じ表記ゆれ名は最後の設定を優先。
    work = work[work["表記ゆれ名"] != ""].copy()
    for idx, row in work.iterrows():
        if not clean_text(row.get("alias_id")):
            source = f"{clean_text(row.get('表記ゆれ名'))}__{clean_text(row.get('紐づけ先 user_id'))}"
            work.at[idx, "alias_id"] = "alias_" + hashlib.sha1(source.encode("utf-8")).hexdigest()[:12]
    work = work.drop_duplicates(subset=["表記ゆれ名"], keep="last")
    return work.reset_index(drop=True)


def load_user_name_aliases(include_disabled=False) -> pd.DataFrame:
    ensure_user_name_alias_table()
    df = load_sqlite_table(SQLITE_TABLE_USER_NAME_ALIASES, USER_NAME_ALIAS_COLUMNS)
    df = normalize_user_name_alias_df(df)
    if not include_disabled:
        df = df[df["有効/無効"] == "有効"].copy()
    return df.reset_index(drop=True)


def save_user_name_aliases(df: pd.DataFrame):
    work = normalize_user_name_alias_df(df)
    save_sqlite_table(work, SQLITE_TABLE_USER_NAME_ALIASES, USER_NAME_ALIAS_COLUMNS, unique_cols=["alias_id"])


def build_user_name_to_id_map(include_hidden=True, include_aliases=True) -> dict:
    """
    利用者名→user_idの対応表を作る。
    安全のため、正式利用者名は原則「完全一致」のみ。
    表記ゆれは、管理者が利用者名ゆれ紐づけマスタで有効登録したものだけ自動補完する。
    """
    try:
        users = load_users(include_hidden=include_hidden)
    except Exception:
        users = pd.DataFrame(columns=USER_COLUMNS)
    mapping = {}
    official_by_id = {}
    for _, row in users.iterrows():
        name = clean_text(row.get("利用者名"))
        uid = ensure_user_id_value(row.get("user_id", ""), name)
        if name and uid:
            mapping[name] = uid
            official_by_id[uid] = name

    if include_aliases:
        try:
            aliases = load_user_name_aliases(include_disabled=False)
            for _, row in aliases.iterrows():
                alias_name = clean_text(row.get("表記ゆれ名"))
                uid = clean_text(row.get("紐づけ先 user_id"))
                if alias_name and uid and uid in official_by_id:
                    # 表記ゆれ名は完全一致・照合キー一致の両方を登録するが、管理者登録済みのものに限る。
                    mapping[alias_name] = uid
                    mapping[normalize_user_name_for_match(alias_name)] = uid
        except Exception:
            pass
    return mapping


def attach_user_ids(df: pd.DataFrame, name_col="利用者名", id_col="user_id") -> pd.DataFrame:
    """
    既存データへuser_id列を安全に補完する。
    正式利用者名の完全一致、または管理者が登録した表記ゆれマスタに一致した場合のみ補完する。
    未確認の名称は空欄のまま残し、管理者の確認対象にする。
    """
    if df is None:
        return df
    work = df.copy()
    if id_col not in work.columns:
        work[id_col] = ""
    if name_col not in work.columns:
        return work
    mapping = build_user_name_to_id_map(include_hidden=True, include_aliases=True)

    def resolve(row):
        existing = clean_text(row.get(id_col, ""))
        if existing:
            return existing
        name = clean_text(row.get(name_col, ""))
        return mapping.get(name) or mapping.get(normalize_user_name_for_match(name)) or ""

    if not work.empty:
        work[id_col] = work.apply(resolve, axis=1)
    return work


def get_user_id_by_name(user_name: str) -> str:
    mapping = build_user_name_to_id_map(include_hidden=True, include_aliases=True)
    name = clean_text(user_name)
    return mapping.get(name) or mapping.get(normalize_user_name_for_match(name)) or ""


def get_user_name_by_id(user_id: str) -> str:
    user_id = clean_text(user_id)
    if not user_id:
        return ""
    try:
        users = load_users(include_hidden=True)
        hit = users[users["user_id"].astype(str) == user_id]
        if not hit.empty:
            return clean_text(hit.iloc[0].get("利用者名"))
    except Exception:
        pass
    return ""

def build_handover_target_options():
    """業務全体申し送り用の対象選択肢を作る。先頭は施設全体の申し送り。"""
    options = ["業務全般"]
    try:
        users = load_users(include_hidden=False)
        if not users.empty:
            for _, row in users.iterrows():
                name = clean_text(row.get("利用者名"))
                uid = clean_text(row.get("user_id"))
                if name:
                    options.append(f"{name}（ID:{uid}）" if uid else name)
    except Exception:
        pass
    return options


def resolve_handover_target(selection_text: str) -> tuple[str, str, str]:
    """対象選択肢から 対象区分・user_id・利用者名 を返す。"""
    value = clean_text(selection_text, "業務全般")
    if value == "業務全般":
        return "業務全般", "", "業務全般"

    # 「山田様（ID:usr_xxx）」形式からIDを拾う
    m = re.search(r"（ID:(.*?)）", value)
    user_id = clean_text(m.group(1)) if m else ""
    user_name = re.sub(r"（ID:.*?）", "", value).strip()

    if not user_id and user_name:
        user_id = get_user_id_by_name(user_name)
    if not user_name and user_id:
        user_name = get_user_name_by_id(user_id)

    return "利用者", user_id, user_name


def make_handover_target_label(user_id: str, user_name: str, target_type: str = "") -> str:
    """保存済みの対象情報から画面用ラベルを作る。"""
    target_type = clean_text(target_type)
    user_id = clean_text(user_id)
    user_name = clean_text(user_name)
    if target_type == "業務全般" or (not user_id and (not user_name or user_name == "業務全般")):
        return "業務全般"
    if user_name and user_id:
        return f"{user_name}（ID:{user_id}）"
    return user_name or "業務全般"



def apply_user_id_migration_preview():
    """主要テーブルの利用者名→user_id補完状況を確認する。保存はしない。"""
    targets = [
        ("健康チェック", SQLITE_TABLE_HEALTH, HEALTH_COLUMNS, ["記録日", "利用者名"]),
        ("排泄チェック", SQLITE_TABLE_EXCRETION, EXCRETION_COLUMNS, ["記録日", "利用者名", "時間帯"]),
        ("短期目標マスタ", SQLITE_TABLE_SHORT_GOAL_MASTER, SHORT_GOAL_MASTER_COLUMNS, ["目標ID"]),
        ("短期目標実施", SQLITE_TABLE_SHORT_GOAL_CHECKS, SHORT_GOAL_CHECK_COLUMNS, ["記録ID"]),
        ("モニタリング下書き", SQLITE_TABLE_MONITORING_DRAFTS, MONITORING_DRAFT_COLUMNS, ["下書きID"]),
        ("LIFE ADL評価", SQLITE_TABLE_LIFE_ADL, LIFE_ADL_COLUMNS, ["評価ID"]),
    ]
    rows = []
    for label, table, columns, key_cols in targets:
        try:
            df = load_sqlite_table(table, columns)
            before_missing = 0 if "user_id" not in df.columns else int((df["user_id"].astype(str).str.strip() == "").sum())
            after_df = attach_user_ids(df)
            after_missing = int((after_df["user_id"].astype(str).str.strip() == "").sum()) if "user_id" in after_df.columns else len(after_df)
            rows.append({
                "対象": label,
                "テーブル": table,
                "件数": len(df),
                "移行前 user_id空欄": before_missing,
                "移行後 user_id空欄見込み": after_missing,
                "状態": "OK" if after_missing == 0 else "要確認",
            })
        except Exception as e:
            rows.append({"対象": label, "テーブル": table, "件数": 0, "移行前 user_id空欄": "", "移行後 user_id空欄見込み": "", "状態": f"エラー: {e}"})
    return pd.DataFrame(rows)


def run_user_id_migration_apply():
    """主要テーブルにuser_id列を追加・補完する。画面表示の利用者名は残す。"""
    # 利用者マスタを先に正規化保存
    users = load_users(include_hidden=True)
    save_users(users)

    targets = [
        (SQLITE_TABLE_HEALTH, HEALTH_COLUMNS, ["記録日"], ["記録日", "利用者名"]),
        (SQLITE_TABLE_EXCRETION, EXCRETION_COLUMNS, ["記録日"], ["記録日", "利用者名", "時間帯"]),
        (SQLITE_TABLE_SHORT_GOAL_MASTER, SHORT_GOAL_MASTER_COLUMNS, ["開始日", "終了予定日"], ["目標ID"]),
        (SQLITE_TABLE_SHORT_GOAL_CHECKS, SHORT_GOAL_CHECK_COLUMNS, ["日付"], ["記録ID"]),
        (SQLITE_TABLE_MONITORING_DRAFTS, MONITORING_DRAFT_COLUMNS, ["作成日"], ["下書きID"]),
        (SQLITE_TABLE_LIFE_ADL, LIFE_ADL_COLUMNS, ["評価日"], ["評価ID"]),
    ]
    results = []
    for table, columns, date_cols, unique_cols in targets:
        df = load_sqlite_table(table, columns, date_cols=date_cols)
        df = attach_user_ids(df)
        save_sqlite_table(df, table, columns, date_cols=date_cols, unique_cols=unique_cols)
        results.append({"テーブル": table, "件数": len(df), "結果": "user_id補完済み"})
    try:
        add_audit_log("利用者ID移行準備", "users", "", "主要テーブルへuser_idを補完しました")
    except Exception:
        pass
    return pd.DataFrame(results)


def build_user_name_variation_df():
    """表記ゆれ候補を確認するための一覧を作る。"""
    users = load_users(include_hidden=True)
    if users.empty:
        return pd.DataFrame(columns=["照合キー", "利用者名一覧", "件数", "状態"])
    tmp = users.copy()
    tmp["照合キー"] = tmp["利用者名"].map(normalize_user_name_for_match)
    grouped = tmp.groupby("照合キー")["利用者名"].apply(lambda s: " / ".join(sorted(set([clean_text(x) for x in s if clean_text(x)])))).reset_index()
    grouped["件数"] = grouped["利用者名"].map(lambda x: len([v for v in x.split(" / ") if v]))
    grouped["状態"] = grouped["件数"].map(lambda n: "表記ゆれ候補" if n >= 2 else "OK")
    return grouped.sort_values(["状態", "照合キー"], ascending=[False, True])


def show_user_id_migration_check():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return
    st.header("利用者ID移行チェック")
    st.caption("職員画面では今まで通り利用者名を表示し、内部保存だけ user_id で安定化するための準備画面です。")

    st.subheader("1. 利用者マスタの user_id")
    users = load_users(include_hidden=True)
    st.dataframe(users[["user_id", "利用者名", "表示"]], use_container_width=True, hide_index=True)

    st.subheader("2. 表記ゆれ候補")
    variations = build_user_name_variation_df()
    st.dataframe(variations, use_container_width=True, hide_index=True)
    if not variations.empty and (variations["状態"] == "表記ゆれ候補").any():
        st.warning("表記ゆれ候補があります。必要に応じて「利用者名ゆれ紐づけマスタ」で正式利用者に紐づけてから補完してください。")
    else:
        st.success("大きな表記ゆれ候補は見つかっていません。")

    st.subheader("3. 主要テーブルの移行見込み")
    preview = apply_user_id_migration_preview()
    st.dataframe(preview, use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("4. user_id補完を実行")
    st.info("この処理は、既存の利用者名を残したまま user_id 列を補完します。画面表示や職員の入力方法は変わりません。")
    confirm = st.checkbox("既存データへ user_id を補完することを理解しました", key="confirm_user_id_migration")
    if st.button("利用者ID移行準備を実行", type="primary", use_container_width=True):
        if not confirm:
            st.error("確認チェックを入れてください。")
        else:
            result = run_user_id_migration_apply()
            st.success("user_id補完を実行しました。")
            st.dataframe(result, use_container_width=True, hide_index=True)
            st.rerun()




def build_unmatched_user_names_df():
    """主要テーブルから、user_id未補完の利用者名を集める。"""
    targets = [
        ("健康チェック", SQLITE_TABLE_HEALTH, HEALTH_COLUMNS),
        ("排泄チェック", SQLITE_TABLE_EXCRETION, EXCRETION_COLUMNS),
        ("短期目標マスタ", SQLITE_TABLE_SHORT_GOAL_MASTER, SHORT_GOAL_MASTER_COLUMNS),
        ("短期目標実施", SQLITE_TABLE_SHORT_GOAL_CHECKS, SHORT_GOAL_CHECK_COLUMNS),
        ("モニタリング下書き", SQLITE_TABLE_MONITORING_DRAFTS, MONITORING_DRAFT_COLUMNS),
        ("LIFE ADL評価", SQLITE_TABLE_LIFE_ADL, LIFE_ADL_COLUMNS),
    ]
    rows = []
    for label, table, columns in targets:
        try:
            df = load_sqlite_table(table, columns)
            if df.empty or "利用者名" not in df.columns:
                continue
            df = attach_user_ids(df)
            if "user_id" not in df.columns:
                df["user_id"] = ""
            missing = df[df["user_id"].astype(str).str.strip() == ""]
            for name, g in missing.groupby("利用者名", dropna=False):
                name = clean_text(name)
                if name:
                    rows.append({
                        "対象": label,
                        "表記ゆれ名": name,
                        "照合キー": normalize_user_name_for_match(name),
                        "件数": len(g),
                    })
        except Exception:
            pass
    if not rows:
        return pd.DataFrame(columns=["対象", "表記ゆれ名", "照合キー", "件数"])
    out = pd.DataFrame(rows)
    out = out.groupby(["表記ゆれ名", "照合キー"], as_index=False).agg({"対象": lambda s: " / ".join(sorted(set(s))), "件数": "sum"})
    return out.sort_values(["件数", "表記ゆれ名"], ascending=[False, True]).reset_index(drop=True)


def add_user_name_alias(alias_name, target_user_id, memo=""):
    """管理者確認済みの表記ゆれ→user_id紐づけを追加する。"""
    alias_name = clean_text(alias_name)
    target_user_id = clean_text(target_user_id)
    if not alias_name:
        return False, "表記ゆれ名を入力してください。"
    if not target_user_id:
        return False, "紐づけ先利用者を選択してください。"
    official_name = get_user_name_by_id(target_user_id)
    if not official_name:
        return False, "紐づけ先の正式利用者が見つかりません。"
    if alias_name == official_name:
        return False, "正式利用者名と同じ名前は登録不要です。"

    df = load_user_name_aliases(include_disabled=True)
    now_text = format_now_jst("%Y-%m-%d %H:%M:%S")
    alias_id = "alias_" + hashlib.sha1(f"{alias_name}__{target_user_id}".encode("utf-8")).hexdigest()[:12]
    # 同じ表記ゆれ名は最後の設定で上書きする。
    df = df[df["表記ゆれ名"].astype(str) != alias_name].copy()
    row = {
        "alias_id": alias_id,
        "表記ゆれ名": alias_name,
        "紐づけ先 user_id": target_user_id,
        "正式利用者名": official_name,
        "有効/無効": "有効",
        "備考": clean_text(memo),
        "更新日時": now_text,
        "更新者": current_login_user(),
    }
    df = pd.concat([df, pd.DataFrame([row], columns=USER_NAME_ALIAS_COLUMNS)], ignore_index=True)
    save_user_name_aliases(df)
    try:
        add_audit_log("利用者名ゆれ紐づけ登録", SQLITE_TABLE_USER_NAME_ALIASES, alias_name, f"{alias_name} → {target_user_id} {official_name}")
    except Exception:
        pass
    return True, f"{alias_name} → {official_name} として登録しました。"


def apply_user_name_aliases_to_records():
    """登録済みの表記ゆれマスタに基づき、未補完user_idだけを補完する。"""
    targets = [
        (SQLITE_TABLE_HEALTH, HEALTH_COLUMNS, ["記録日"], ["記録日", "利用者名"]),
        (SQLITE_TABLE_EXCRETION, EXCRETION_COLUMNS, ["記録日"], ["記録日", "利用者名", "時間帯"]),
        (SQLITE_TABLE_SHORT_GOAL_MASTER, SHORT_GOAL_MASTER_COLUMNS, ["開始日", "終了予定日"], ["目標ID"]),
        (SQLITE_TABLE_SHORT_GOAL_CHECKS, SHORT_GOAL_CHECK_COLUMNS, ["日付"], ["記録ID"]),
        (SQLITE_TABLE_MONITORING_DRAFTS, MONITORING_DRAFT_COLUMNS, ["作成日"], ["下書きID"]),
        (SQLITE_TABLE_LIFE_ADL, LIFE_ADL_COLUMNS, ["評価日"], ["評価ID"]),
    ]
    rows = []
    for table, columns, date_cols, unique_cols in targets:
        df = load_sqlite_table(table, columns, date_cols=date_cols)
        before = 0 if "user_id" not in df.columns else int((df["user_id"].astype(str).str.strip() == "").sum())
        df2 = attach_user_ids(df)
        after = int((df2["user_id"].astype(str).str.strip() == "").sum()) if "user_id" in df2.columns else len(df2)
        save_sqlite_table(df2, table, columns, date_cols=date_cols, unique_cols=unique_cols)
        rows.append({"テーブル": table, "件数": len(df2), "補完前未紐づけ": before, "補完後未紐づけ": after, "今回補完": max(before - after, 0)})
    try:
        add_audit_log("利用者名ゆれ紐づけ適用", SQLITE_TABLE_USER_NAME_ALIASES, "", "表記ゆれマスタに基づいて未補完user_idを補完")
    except Exception:
        pass
    return pd.DataFrame(rows)


def show_user_name_alias_master_menu():
    """管理者確認済みの利用者名ゆれ紐づけマスタ。"""
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    st.header("利用者名ゆれ紐づけマスタ")
    st.caption("完全自動ではなく、管理者が確認した表記ゆれだけを user_id に紐づけます。正式な利用者名は画面表示に残し、内部だけ安全に統一します。")

    tab1, tab2, tab3 = st.tabs(["未紐づけ候補", "マスタ編集", "補完実行"])

    users = load_users(include_hidden=True)
    user_options = []
    user_label_to_id = {}
    for _, row in users.iterrows():
        uid = ensure_user_id_value(row.get("user_id", ""), row.get("利用者名", ""))
        name = clean_text(row.get("利用者名"))
        if uid and name:
            label = f"{name}（{uid}）"
            user_options.append(label)
            user_label_to_id[label] = uid

    with tab1:
        st.subheader("未紐づけ候補")
        st.info("ここに出る名前は、まだ正式利用者または登録済み表記ゆれに紐づいていない名称です。内容を確認して、必要なものだけマスタ登録してください。")
        unmatched = build_unmatched_user_names_df()
        if unmatched.empty:
            st.success("未紐づけ候補はありません。")
        else:
            st.dataframe(unmatched, use_container_width=True, hide_index=True)
            selected_alias = st.selectbox("登録する表記ゆれ名", unmatched["表記ゆれ名"].tolist(), key="alias_candidate_select")
            selected_user = st.selectbox("紐づけ先の正式利用者", user_options, key="alias_candidate_user_select") if user_options else ""
            memo = st.text_input("備考", value="未紐づけ候補から登録", key="alias_candidate_memo")
            if st.button("この候補を紐づけマスタへ登録", type="primary", use_container_width=True):
                ok, msg = add_user_name_alias(selected_alias, user_label_to_id.get(selected_user, ""), memo)
                if ok:
                    st.success(msg)
                    st.rerun()
                else:
                    st.error(msg)

    with tab2:
        st.subheader("マスタ編集")
        aliases = load_user_name_aliases(include_disabled=True)
        st.caption("表記ゆれ名を直接追加・無効化できます。紐づけ先 user_id は上の候補登録を使うと安全です。")
        edited = st.data_editor(
            aliases,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "有効/無効": st.column_config.SelectboxColumn("有効/無効", options=["有効", "無効"]),
                "紐づけ先 user_id": st.column_config.TextColumn("紐づけ先 user_id"),
                "正式利用者名": st.column_config.TextColumn("正式利用者名"),
            },
            key="user_name_alias_editor",
        )
        if st.button("マスタを保存", type="primary", use_container_width=True):
            work = normalize_user_name_alias_df(edited)
            # user_idから正式利用者名を補完する。存在しないuser_idは保存前に警告。
            valid_ids = set(users["user_id"].astype(str).tolist()) if "user_id" in users.columns else set()
            invalid = work[(work["紐づけ先 user_id"].astype(str) != "") & (~work["紐づけ先 user_id"].astype(str).isin(valid_ids))]
            if not invalid.empty:
                st.error("存在しない user_id が含まれています。候補登録から選ぶか、正式な user_id に修正してください。")
                st.dataframe(invalid, use_container_width=True, hide_index=True)
            else:
                for idx, row in work.iterrows():
                    uid = clean_text(row.get("紐づけ先 user_id"))
                    name = get_user_name_by_id(uid)
                    if name:
                        work.at[idx, "正式利用者名"] = name
                    work.at[idx, "更新日時"] = format_now_jst("%Y-%m-%d %H:%M:%S")
                    work.at[idx, "更新者"] = current_login_user()
                save_user_name_aliases(work)
                add_audit_log("利用者名ゆれ紐づけマスタ保存", SQLITE_TABLE_USER_NAME_ALIASES, "", "マスタを保存")
                st.success("利用者名ゆれ紐づけマスタを保存しました。")
                st.rerun()

    with tab3:
        st.subheader("登録済みマスタを既存データへ適用")
        st.warning("この処理は、登録済みの表記ゆれマスタに一致した未補完データだけ user_id を入れます。名称自体は書き換えません。")
        preview = apply_user_id_migration_preview()
        st.dataframe(preview, use_container_width=True, hide_index=True)
        confirm = st.checkbox("管理者が確認した紐づけマスタだけを既存データへ適用する", key="confirm_apply_aliases")
        if st.button("利用者名ゆれ紐づけを適用", type="primary", use_container_width=True):
            if not confirm:
                st.error("確認チェックを入れてください。")
            else:
                result = apply_user_name_aliases_to_records()
                st.success("登録済みの表記ゆれマスタに基づいて user_id を補完しました。")
                st.dataframe(result, use_container_width=True, hide_index=True)
                st.rerun()


# =========================
# Ver4.2 体重は「測定した日だけ入力」
# =========================
def parse_optional_weight(value):
    """体重は毎日必須にせず、未測定なら空欄で保存する。"""
    text = clean_text(value)
    if text == "":
        return "", ""
    text = text.replace("kg", "").replace("ＫＧ", "").replace("ｋｇ", "").strip()
    try:
        weight = float(text)
    except Exception:
        return "", "体重は数値で入力してください。未測定の場合は空欄でOKです。"
    if weight <= 0:
        return "", "体重は0より大きい数値で入力してください。未測定の場合は空欄でOKです。"
    if weight > 200:
        return "", "体重が200kgを超えています。入力値を確認してください。"
    return round(weight, 1), ""


def format_weight_value(value):
    w = safe_float(value, 0)
    if w <= 0:
        return ""
    return f"{w:.1f}"


def build_latest_weight_summary(health_df, users=None, target_date=None):
    """利用者ごとの最新体重を返す。体重0・空欄は未測定として扱う。"""
    if target_date is None:
        target_date = today_jst()
    try:
        target_date = pd.to_datetime(target_date).date()
    except Exception:
        target_date = today_jst()

    users = users or []
    rows = []
    work = health_df.copy() if health_df is not None else pd.DataFrame(columns=HEALTH_COLUMNS)
    if not work.empty:
        work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
        work["体重_num"] = pd.to_numeric(work.get("体重", ""), errors="coerce")
        work = work[(work["記録日"].notna()) & (work["体重_num"].notna()) & (work["体重_num"] > 0)].copy()

    for user in users:
        user = clean_text(user)
        if not user:
            continue
        latest = pd.DataFrame()
        if not work.empty and "利用者名" in work.columns:
            latest = work[work["利用者名"].astype(str).str.strip() == user].sort_values("記録日")
        if latest.empty:
            rows.append({
                "利用者名": user,
                "最新体重": "未測定",
                "測定日": "",
                "経過日数": "",
                "状態": "体重記録なし",
            })
        else:
            r = latest.iloc[-1]
            measured_date = pd.to_datetime(r.get("記録日"), errors="coerce").date()
            days = (target_date - measured_date).days
            rows.append({
                "利用者名": user,
                "最新体重": f"{float(r.get('体重_num')):.1f}kg",
                "測定日": measured_date.strftime("%Y/%m/%d"),
                "経過日数": f"{days}日前" if days >= 0 else "確認日より後",
                "状態": "14日以上未測定" if days >= 14 else "OK",
            })
    return pd.DataFrame(rows, columns=["利用者名", "最新体重", "測定日", "経過日数", "状態"])


def build_weight_not_measured_users(health_df, users=None, target_date=None, threshold_days=14):
    """14日以上体重未測定、または体重記録なしの利用者を返す。"""
    summary = build_latest_weight_summary(health_df, users, target_date)
    if summary.empty:
        return summary

    def is_overdue(row):
        status = clean_text(row.get("状態"))
        if status == "体重記録なし":
            return True
        days_text = clean_text(row.get("経過日数"))
        m = re.search(r"(\d+)", days_text)
        return bool(m and int(m.group(1)) >= threshold_days)

    out = summary[summary.apply(is_overdue, axis=1)].copy()
    if not out.empty:
        out["確認すること"] = "体重測定の予定を確認してください。未測定には理由がある場合があります。"
    return out


def show_latest_weight_block(health_df, users=None, target_date=None):
    st.subheader("最新体重")
    st.caption("体重は毎日入力ではなく、測定した日だけ入力します。最新の測定値を確認します。")
    summary = build_latest_weight_summary(health_df, users, target_date)
    if summary.empty:
        st.info("利用者または健康チェックデータがありません。")
    else:
        st.dataframe(summary, use_container_width=True, hide_index=True)


def show_weight_overdue_block(health_df, users=None, target_date=None, threshold_days=14):
    st.subheader(f"{threshold_days}日以上体重未測定")
    st.caption("未入力を責めるためではなく、測定予定・拒否・体調などを確認するための表示です。")
    overdue = build_weight_not_measured_users(health_df, users, target_date, threshold_days=threshold_days)
    if overdue.empty:
        st.success(f"{threshold_days}日以上体重未測定の利用者はいません。")
    else:
        st.warning("体重測定の予定を確認したい利用者がいます。")
        st.dataframe(overdue, use_container_width=True, hide_index=True)


def ensure_excel_file(path, sheet_name, columns):
    """
    Ver3.4以降の互換用。
    商品版ではExcelファイルを正データとして作成しません。
    既存Excelからの初回移行だけは各ensure_*関数側で行います。
    """
    ensure_dirs()
    return

def upgrade_account_password_hash(login_id: str, password: str):
    """旧SHA256／平文パスワードをbcryptへ自動移行する。"""
    if not is_bcrypt_available():
        return
    try:
        accounts = load_accounts()
        login_id = clean_text(login_id).lower()
        matches = accounts[accounts["ログインID"] == login_id].index.tolist()
        if not matches:
            return
        idx = matches[-1]
        current_hash = clean_text(accounts.at[idx, "パスワードハッシュ"])
        if password_hash_needs_upgrade(current_hash):
            accounts.at[idx, "パスワードハッシュ"] = hash_password(password)
            accounts.at[idx, "更新日時"] = format_now_jst("%Y-%m-%d %H:%M:%S")
            save_accounts(accounts)
            try:
                add_audit_log("パスワードハッシュ自動移行", "login_account_master", login_id, "旧形式からbcryptへ自動移行")
            except Exception:
                pass
    except Exception:
        pass


def default_account_rows():
    """初期アカウントを返す。DBが空の場合のみ使用する。"""
    now_text = format_now_jst("%Y-%m-%d %H:%M:%S")
    return [
        {
            "ログインID": "kanri",
            "表示名": "管理者",
            "パスワードハッシュ": hash_password(INITIAL_ACCOUNT_PASSWORD),
            "権限": "admin",
            "状態": "有効",
            "備考": "初期管理者。削除・無効化するとログインできなくなるため注意してください。",
            "作成日時": now_text,
            "更新日時": now_text,
            "初回パスワード変更必須": "はい",
            "最終パスワード変更日時": "",
        },
        {
            "ログインID": "staff",
            "表示名": "職員",
            "パスワードハッシュ": hash_password(INITIAL_ACCOUNT_PASSWORD),
            "権限": "staff",
            "状態": "有効",
            "備考": "初期職員アカウント",
            "作成日時": now_text,
            "更新日時": now_text,
            "初回パスワード変更必須": "はい",
            "最終パスワード変更日時": "",
        },
    ]


def normalize_accounts_df(df):
    """ログインアカウントDataFrameを正規化する。"""
    if df is None:
        df = pd.DataFrame(columns=ACCOUNT_COLUMNS)
    work = df.copy()
    for col in ACCOUNT_COLUMNS:
        if col not in work.columns:
            work[col] = ""
    work = work[ACCOUNT_COLUMNS].copy()
    work["ログインID"] = work["ログインID"].fillna("").astype(str).str.strip().str.lower()
    work["表示名"] = work["表示名"].fillna("").astype(str).str.strip()
    work["権限"] = work["権限"].fillna("staff").astype(str).str.strip()
    work["状態"] = work["状態"].fillna("有効").astype(str).str.strip()
    work = work[work["ログインID"] != ""].drop_duplicates(subset=["ログインID"], keep="last")

    # Ver3.9：初回パスワード変更必須化。
    # 既存DBに列がない場合はここで安全に補完する。
    # 初期ID（kanri/staff）で、まだ既定パスワードのままなら必ず変更対象にする。
    if "初回パスワード変更必須" in work.columns:
        work["初回パスワード変更必須"] = work["初回パスワード変更必須"].map(lambda x: clean_text(x))
        for idx, row in work.iterrows():
            current_value = clean_text(row.get("初回パスワード変更必須"))
            login_id = clean_text(row.get("ログインID")).lower()
            password_hash = clean_text(row.get("パスワードハッシュ"))
            if current_value == "":
                try:
                    default_pw = verify_password(INITIAL_ACCOUNT_PASSWORD, password_hash)
                except Exception:
                    default_pw = False
                work.at[idx, "初回パスワード変更必須"] = "はい" if (login_id in INITIAL_LOGIN_IDS and default_pw) else "いいえ"
            elif current_value.lower() in ["true", "1", "yes", "有", "必須", "on"]:
                work.at[idx, "初回パスワード変更必須"] = "はい"
            else:
                work.at[idx, "初回パスワード変更必須"] = "いいえ"

    if "最終パスワード変更日時" in work.columns:
        work["最終パスワード変更日時"] = work["最終パスワード変更日時"].map(lambda x: clean_text(x))

    return work.reset_index(drop=True)


def ensure_account_file():
    """
    ログインアカウントを用意する。
    SQLite補助DBが壊れている／ロックしている場合でも、既定アカウントでログイン画面を継続する。
    """
    try:
        ensure_dirs()
    except Exception:
        pass

    try:
        if sqlite_table_row_count(SQLITE_TABLE_ACCOUNTS) > 0:
            return
    except Exception:
        pass

    # 旧Excelがあれば移行を試みる。失敗しても止めない。
    try:
        if ACCOUNT_FILE.exists():
            try:
                df = pd.read_excel(ACCOUNT_FILE, sheet_name="ログインアカウント")
            except Exception:
                df = pd.DataFrame(columns=ACCOUNT_COLUMNS)
            df = normalize_accounts_df(df)
            if not df.empty:
                save_sqlite_table(df, SQLITE_TABLE_ACCOUNTS, ACCOUNT_COLUMNS, unique_cols=["ログインID"])
                return
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, "ensure_account_file_excel")
        except Exception:
            pass

    # 何もなければ初期アカウントを作成。SQLite保存に失敗してもログイン自体は継続可能。
    try:
        df = pd.DataFrame(default_account_rows(), columns=ACCOUNT_COLUMNS)
        save_sqlite_table(df, SQLITE_TABLE_ACCOUNTS, ACCOUNT_COLUMNS, unique_cols=["ログインID"])
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, "ensure_account_file_default")
        except Exception:
            pass


def load_accounts():
    """
    アカウント読込。
    SQLite補助DBのエラーや空データでログイン不能にならないよう、最後はdefault_account_rowsへフォールバックする。
    """
    try:
        ensure_account_file()
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, "load_accounts.ensure")
        except Exception:
            pass

    try:
        df = load_sqlite_table(SQLITE_TABLE_ACCOUNTS, ACCOUNT_COLUMNS)
        df = normalize_accounts_df(df)
        if df is not None and not df.empty:
            return df
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, "load_accounts")
            _show_sqlite_backup_warning_once(e, "login_accounts")
        except Exception:
            pass

    # 最終フォールバック：初期ログインを必ず維持する
    try:
        return normalize_accounts_df(pd.DataFrame(default_account_rows(), columns=ACCOUNT_COLUMNS))
    except Exception:
        return pd.DataFrame(default_account_rows())


def save_accounts(df):
    try:
        work = normalize_accounts_df(df)
        return save_sqlite_table(work, SQLITE_TABLE_ACCOUNTS, ACCOUNT_COLUMNS, unique_cols=["ログインID"])
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, "save_accounts")
        except Exception:
            pass
        return False


def ensure_login_history_file():
    """
    ログイン履歴をSQLiteで管理する。
    SQLite補助DBの不調でログイン処理を止めない。
    """
    try:
        ensure_dirs()
    except Exception:
        pass

    try:
        if sqlite_table_row_count(SQLITE_TABLE_LOGIN_HISTORY) > 0:
            return
    except Exception:
        pass

    try:
        if LOGIN_HISTORY_FILE.exists():
            try:
                df = pd.read_excel(LOGIN_HISTORY_FILE, sheet_name="ログイン履歴")
            except Exception:
                df = pd.DataFrame(columns=LOGIN_HISTORY_COLUMNS)
            for col in LOGIN_HISTORY_COLUMNS:
                if col not in df.columns:
                    df[col] = ""
            df = df[LOGIN_HISTORY_COLUMNS].copy()
            if not df.empty:
                save_sqlite_table(df, SQLITE_TABLE_LOGIN_HISTORY, LOGIN_HISTORY_COLUMNS, sort_cols=["日時"])
                return
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, "ensure_login_history_file_excel")
        except Exception:
            pass

    try:
        save_sqlite_table(pd.DataFrame(columns=LOGIN_HISTORY_COLUMNS), SQLITE_TABLE_LOGIN_HISTORY, LOGIN_HISTORY_COLUMNS, sort_cols=["日時"])
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, "ensure_login_history_file_empty")
        except Exception:
            pass


def load_login_history():
    try:
        ensure_login_history_file()
        df = load_sqlite_table(SQLITE_TABLE_LOGIN_HISTORY, LOGIN_HISTORY_COLUMNS)
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, "load_login_history")
        except Exception:
            pass
        df = pd.DataFrame(columns=LOGIN_HISTORY_COLUMNS)

    for col in LOGIN_HISTORY_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    return df[LOGIN_HISTORY_COLUMNS].copy()


def save_login_history(df):
    try:
        work = df.copy()
        for col in LOGIN_HISTORY_COLUMNS:
            if col not in work.columns:
                work[col] = ""
        work = work[LOGIN_HISTORY_COLUMNS]
        return save_sqlite_table(work, SQLITE_TABLE_LOGIN_HISTORY, LOGIN_HISTORY_COLUMNS, sort_cols=["日時"])
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, "save_login_history")
        except Exception:
            pass
        return False


def add_login_history(login_id, label, role, result, memo=""):
    """ログイン履歴追加。履歴保存に失敗してもログイン処理を止めない。"""
    try:
        df = load_login_history()
        row = {
            "日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
            "ログインID": clean_text(login_id).lower(),
            "表示名": clean_text(label),
            "権限": clean_text(role),
            "結果": clean_text(result),
            "メモ": clean_text(memo),
        }
        df = pd.concat([df, pd.DataFrame([row], columns=LOGIN_HISTORY_COLUMNS)], ignore_index=True)
        if len(df) > 1000:
            df = df.tail(1000)
        save_login_history(df)
    except Exception as e:
        try:
            _mark_sqlite_backup_error(e, "add_login_history")
        except Exception:
            pass


def update_account_password(login_id, new_password, force_change="いいえ"):
    """パスワードを更新し、初回変更必須フラグを更新する。"""
    accounts = load_accounts()
    login_id = clean_text(login_id).lower()
    matches = accounts[accounts["ログインID"] == login_id].index.tolist()
    if not matches:
        return False, "アカウントが見つかりません。"
    idx = matches[-1]
    now_text = format_now_jst("%Y-%m-%d %H:%M:%S")
    accounts.at[idx, "パスワードハッシュ"] = hash_password(new_password)
    accounts.at[idx, "初回パスワード変更必須"] = "はい" if force_change in [True, "はい", "1", "true"] else "いいえ"
    accounts.at[idx, "最終パスワード変更日時"] = now_text
    accounts.at[idx, "更新日時"] = now_text
    save_accounts(accounts)
    try:
        add_audit_log("パスワード変更", "login_accounts", login_id, "パスワードを更新し、初回変更必須フラグを解除/設定")
    except Exception:
        pass
    return True, "パスワードを更新しました。"


def authenticate_user(login_id, password):
    """ログインID・パスワードを認証し、アカウント情報dictを返す。"""
    login_id = clean_text(login_id).lower()
    password = clean_text(password)
    accounts = load_accounts()
    hit = accounts[accounts["ログインID"] == login_id]
    if hit.empty:
        add_login_history(login_id, "", "", "失敗", "IDなし")
        return None, "IDまたはパスワードが違います。"
    row = hit.iloc[-1].to_dict()
    if clean_text(row.get("状態")) != "有効":
        add_login_history(login_id, row.get("表示名", ""), row.get("権限", ""), "失敗", "無効アカウント")
        return None, "このアカウントは無効です。管理者へ確認してください。"
    if not verify_password(password, row.get("パスワードハッシュ", "")):
        add_login_history(login_id, row.get("表示名", ""), row.get("権限", ""), "失敗", "パスワード違い")
        return None, "IDまたはパスワードが違います。"

    uses_initial_password_value = uses_initial_password(login_id, password)
    if uses_initial_password_value:
        row["初回パスワード変更必須"] = "はい"
        try:
            idx = hit.index[-1]
            accounts.at[idx, "初回パスワード変更必須"] = "はい"
            save_accounts(accounts)
        except Exception:
            pass

    # 旧SHA256／平文形式だった場合、ログイン成功時にbcryptへ自動移行する
    if password_hash_needs_upgrade(row.get("パスワードハッシュ", "")):
        upgrade_account_password_hash(login_id, password)
        row["パスワードハッシュ"] = "********"

    add_login_history(login_id, row.get("表示名", ""), row.get("権限", ""), "成功", "")
    return row, ""


def show_login_user_management_menu():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    st.header("ログイン・職員ID管理")
    st.caption("職員のログインID、パスワード、権限、状態を管理します。ログイン履歴も確認できます。")
    if is_bcrypt_available():
        st.success("パスワード保存方式：bcrypt（安全性の高いハッシュ保存）")
    else:
        st.warning("bcryptライブラリが未導入です。requirements.txt に bcrypt を追加すると、パスワードがbcrypt形式で保存されます。")

    tab1, tab2, tab3 = st.tabs(["職員ID管理", "新規ID追加", "ログイン履歴"])

    with tab1:
        st.subheader("登録済みアカウント")
        accounts = load_accounts()
        if accounts.empty:
            st.info("アカウントがありません。")
        else:
            display_df = accounts.copy()
            display_df["パスワードハッシュ"] = "********"
            st.dataframe(display_df, use_container_width=True, hide_index=True)

        st.divider()
        st.subheader("アカウント編集")
        accounts = load_accounts()
        id_list = accounts["ログインID"].tolist()
        if not id_list:
            st.info("編集できるアカウントがありません。")
        else:
            selected_id = st.selectbox("編集するログインID", id_list, key="account_edit_select")
            row = accounts[accounts["ログインID"] == selected_id].iloc[-1]
            with st.form("account_edit_form", clear_on_submit=False):
                c1, c2, c3 = st.columns(3)
                with c1:
                    label = st.text_input("表示名", value=clean_text(row.get("表示名")))
                with c2:
                    role = st.selectbox("権限", ["admin", "staff"], index=0 if clean_text(row.get("権限")) == "admin" else 1)
                with c3:
                    status = st.selectbox("状態", ["有効", "無効"], index=0 if clean_text(row.get("状態"), "有効") == "有効" else 1)
                new_password = st.text_input("新しいパスワード（変更しない場合は空欄）", type="password")
                force_change_next = st.checkbox(
                    "次回ログイン時にパスワード変更を求める",
                    value=account_requires_password_change(row.to_dict()),
                    help="仮パスワードを管理者が設定した場合はONにしてください。",
                )
                memo = st.text_area("備考", value=clean_text(row.get("備考")), height=80)
                submitted = st.form_submit_button("この内容で更新", type="primary", use_container_width=True)

            if submitted:
                if selected_id == "kanri" and status != "有効":
                    st.error("初期管理者 kanri は無効化できません。")
                elif selected_id == "kanri" and role != "admin":
                    st.error("初期管理者 kanri の権限は admin のままにしてください。")
                else:
                    idx = accounts[accounts["ログインID"] == selected_id].index[-1]
                    accounts.at[idx, "表示名"] = clean_text(label, selected_id)
                    accounts.at[idx, "権限"] = role
                    accounts.at[idx, "状態"] = status
                    if clean_text(new_password):
                        ok_pw, pw_msg = validate_new_password(selected_id, new_password, new_password, clean_text(accounts.at[idx, "パスワードハッシュ"]))
                        if not ok_pw:
                            st.error(pw_msg)
                            st.stop()
                        accounts.at[idx, "パスワードハッシュ"] = hash_password(new_password)
                        accounts.at[idx, "最終パスワード変更日時"] = format_now_jst("%Y-%m-%d %H:%M:%S")
                    accounts.at[idx, "初回パスワード変更必須"] = "はい" if force_change_next else "いいえ"
                    accounts.at[idx, "備考"] = clean_text(memo)
                    accounts.at[idx, "更新日時"] = format_now_jst("%Y-%m-%d %H:%M:%S")
                    save_accounts(accounts)
                    add_audit_log("アカウント更新", "login_accounts", selected_id, "アカウント情報を更新")
                    st.success("アカウントを更新しました。")
                    st.rerun()

        st.divider()
        st.subheader("アカウント削除")
        st.caption("退職者など、不要になったIDを削除できます。通常は削除より『無効』がおすすめです。")
        accounts = load_accounts()
        deletable = [x for x in accounts["ログインID"].tolist() if x != "kanri"]
        if deletable:
            del_id = st.selectbox("削除するログインID", deletable, key="account_delete_select")
            if st.button("選択したIDを削除", type="secondary"):
                accounts = accounts[accounts["ログインID"] != del_id]
                save_accounts(accounts)
                st.success(f"{del_id} を削除しました。")
                st.rerun()
        else:
            st.info("削除可能なアカウントはありません。")

    with tab2:
        st.subheader("新規ID追加")
        with st.form("account_add_form", clear_on_submit=False):
            c1, c2 = st.columns(2)
            with c1:
                new_id = st.text_input("ログインID", placeholder="例：tanaka")
                new_label = st.text_input("表示名", placeholder="例：田中")
            with c2:
                new_role = st.selectbox("権限", ["staff", "admin"], index=0)
                new_status = st.selectbox("状態", ["有効", "無効"], index=0)
            new_force_change = st.checkbox("初回ログイン時にパスワード変更を必須にする", value=True)
            pw1 = st.text_input("パスワード", type="password")
            pw2 = st.text_input("パスワード確認", type="password")
            new_memo = st.text_area("備考", height=80)
            add_submitted = st.form_submit_button("新規IDを追加", type="primary", use_container_width=True)

        if add_submitted:
            login_id = clean_text(new_id).lower()
            if not login_id:
                st.error("ログインIDを入力してください。")
            elif not re.match(r"^[a-zA-Z0-9_\-\.]+$", login_id):
                st.error("ログインIDは半角英数字、_、-、. のみ使用できます。")
            elif not clean_text(pw1):
                st.error("パスワードを入力してください。")
            elif pw1 != pw2:
                st.error("パスワード確認が一致しません。")
            else:
                ok_pw, pw_msg = validate_new_password(login_id, pw1, pw2, "")
                if not ok_pw:
                    st.error(pw_msg)
                    st.stop()
                accounts = load_accounts()
                if login_id in accounts["ログインID"].tolist():
                    st.error("同じログインIDが既に存在します。")
                else:
                    now_text = format_now_jst("%Y-%m-%d %H:%M:%S")
                    row = {
                        "ログインID": login_id,
                        "表示名": clean_text(new_label, login_id),
                        "パスワードハッシュ": hash_password(pw1),
                        "権限": new_role,
                        "状態": new_status,
                        "備考": clean_text(new_memo),
                        "作成日時": now_text,
                        "更新日時": now_text,
                        "初回パスワード変更必須": "はい" if new_force_change else "いいえ",
                        "最終パスワード変更日時": "",
                    }
                    accounts = pd.concat([accounts, pd.DataFrame([row], columns=ACCOUNT_COLUMNS)], ignore_index=True)
                    save_accounts(accounts)
                    st.success("新規IDを追加しました。")
                    st.rerun()

    with tab3:
        st.subheader("ログイン履歴")
        logs = load_login_history()
        if logs.empty:
            st.info("ログイン履歴はまだありません。")
        else:
            c1, c2, c3 = st.columns(3)
            with c1:
                result_filter = st.selectbox("結果", ["すべて", "成功", "失敗"], key="login_result_filter")
            with c2:
                id_filter = st.text_input("ログインIDで検索", key="login_id_filter")
            with c3:
                show_count = st.number_input("表示件数", min_value=10, max_value=500, value=100, step=10)
            view = logs.copy()
            if result_filter != "すべて":
                view = view[view["結果"].astype(str) == result_filter]
            if clean_text(id_filter):
                view = view[view["ログインID"].astype(str).str.contains(clean_text(id_filter).lower(), case=False, na=False)]
            view = view.tail(int(show_count)).sort_index(ascending=False)
            st.dataframe(view, use_container_width=True, hide_index=True)

            output = BytesIO()
            view.to_excel(output, index=False, sheet_name="ログイン履歴")
            st.download_button(
                "表示中のログイン履歴をExcelでダウンロード",
                data=output.getvalue(),
                file_name=f"login_history_{today_jst().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

# =========================
# 利用者マスタ
# =========================
def default_user_rows():
    """初期利用者マスタを返す。DBが空の場合のみ使用する。"""
    rows = []
    for name in DEFAULT_USERS:
        row = {"user_id": make_user_id_from_name(name), "利用者名": name, "表示": "表示"}
        for col in ASSESSMENT_COLUMNS:
            row[col] = ""
        rows.append(row)
    return rows


def normalize_users_df(df):
    """利用者マスタDataFrameを正規化する。"""
    if df is None:
        df = pd.DataFrame(columns=USER_COLUMNS)
    work = df.copy()

    for col in USER_COLUMNS:
        if col not in work.columns:
            work[col] = ""

    work = work[USER_COLUMNS].copy()
    work["利用者名"] = work["利用者名"].fillna("").astype(str).str.strip()
    work["user_id"] = work.apply(lambda row: ensure_user_id_value(row.get("user_id", ""), row.get("利用者名", "")), axis=1)
    work["表示"] = work["表示"].fillna("表示").astype(str).str.strip()
    work.loc[~work["表示"].isin(["表示", "非表示"]), "表示"] = "表示"
    work = work[work["利用者名"] != ""].drop_duplicates(subset=["user_id"], keep="first")
    return work.reset_index(drop=True)



def ensure_user_file():
    """
    利用者マスタを用意する。
    Ver4.5ではSupabase対象テーブルだが、SQLiteにもミラー保存する。
    Supabaseに既存利用者がある場合は、それを優先してローカルへミラーする。
    """
    ensure_dirs()

    # Supabaseに既存データがあれば、それをローカルSQLiteへミラーして終了
    if "supabase_is_enabled" in globals() and supabase_is_enabled():
        try:
            remote_df = supabase_read_table(SQLITE_TABLE_USERS, USER_COLUMNS)
            remote_df = normalize_users_df(remote_df)
            if not remote_df.empty:
                _original_save_sqlite_table(remote_df, SQLITE_TABLE_USERS, USER_COLUMNS, unique_cols=["user_id"])
                return
        except Exception:
            pass

    # 既にSQLiteにデータがあれば何もしない
    if sqlite_table_row_count(SQLITE_TABLE_USERS) > 0:
        return

    # 旧Excelがあれば移行
    if USER_FILE.exists():
        try:
            df = pd.read_excel(USER_FILE, sheet_name=USER_SHEET)
        except Exception:
            try:
                df = pd.read_excel(USER_FILE)
            except Exception:
                df = pd.DataFrame(columns=USER_COLUMNS)
        df = normalize_users_df(df)
        if not df.empty:
            save_sqlite_table(df, SQLITE_TABLE_USERS, USER_COLUMNS, unique_cols=["user_id"])
            return

    # 何もなければ初期利用者マスタを作成
    df = pd.DataFrame(default_user_rows(), columns=USER_COLUMNS)
    save_sqlite_table(df, SQLITE_TABLE_USERS, USER_COLUMNS, unique_cols=["user_id"])

USER_LIST_COLUMNS = ["user_id", "利用者名", "表示"]


@cache_safe_master_read(ttl=SAFE_READ_CACHE_TTL_SEC)
def load_active_user_names(include_hidden=False):
    with perf_timer("load_active_user_names", f"include_hidden={include_hidden}"):
        ensure_user_file()
        df = load_sqlite_table(SQLITE_TABLE_USERS, USER_LIST_COLUMNS)
        for col in USER_LIST_COLUMNS:
            if col not in df.columns:
                df[col] = ""
        if not include_hidden:
            df = df[df["表示"].fillna("表示") == "表示"]
        return df["利用者名"].fillna("").astype(str).str.strip().replace("", pd.NA).dropna().tolist()


@cache_safe_master_read(ttl=SAFE_READ_CACHE_TTL_SEC)
def load_users(include_hidden=False):
    with perf_timer("load_users", f"include_hidden={include_hidden}"):
        ensure_user_file()
        df = load_sqlite_table(SQLITE_TABLE_USERS, USER_COLUMNS)
        df = normalize_users_df(df)

        if not include_hidden:
            df = df[df["表示"].fillna("表示") == "表示"]

        return df.reset_index(drop=True)


def save_users(df):
    work = normalize_users_df(df)
    save_sqlite_table(work, SQLITE_TABLE_USERS, USER_COLUMNS, unique_cols=["user_id"])
    clear_hidamari_read_cache("利用者マスタ保存")


def export_user_master_excel_bytes():
    """SQLite上の利用者マスタをExcel形式で出力する。"""
    return to_excel_download(load_users(include_hidden=True))


def add_user(user_name):
    user_name = clean_text(user_name)

    if not user_name:
        return False, "利用者名を入力してください。"

    df = load_users(include_hidden=True)

    if user_name in df["利用者名"].tolist():
        df.loc[df["利用者名"] == user_name, "表示"] = "表示"
        save_users(df)
        return True, f"{user_name}を表示に戻しました。"

    row = {"user_id": make_user_id_from_name(user_name), "利用者名": user_name, "表示": "表示"}
    for col in ASSESSMENT_COLUMNS:
        row[col] = ""

    df = pd.concat([df, pd.DataFrame([row])], ignore_index=True)
    save_users(df)

    return True, f"{user_name}を追加しました。"


def hide_user(user_name):
    df = load_users(include_hidden=True)

    if user_name not in df["利用者名"].tolist():
        return False, "対象の利用者が見つかりません。"

    df.loc[df["利用者名"] == user_name, "表示"] = "非表示"
    save_users(df)

    return True, f"{user_name}を入力候補から外しました。"


def get_user_assessment(user_name):
    df = load_users(include_hidden=True)
    row = df[df["利用者名"] == user_name]

    if row.empty:
        return {}

    row = row.iloc[0]

    return {
        col: clean_text(row.get(col, ""))
        for col in ASSESSMENT_COLUMNS
        if clean_text(row.get(col, ""))
    }


def build_assessment_context_text(user_name):
    data = get_user_assessment(user_name)

    if not data:
        return ""

    order = ["主訴", "生活状況", "ADL", "IADL", "認知機能", "健康状態", "課題", "支援内容"]
    lines = []

    for col in order:
        if data.get(col):
            lines.append(f"{col}：{data[col]}")

    return "\n".join(lines)



# =========================
# 健康チェックデータ
# =========================
def ensure_health_file():
    """互換用。実データはSQLite（hidamari_health.db / health_records）へ保存します。"""
    migrate_excel_to_sqlite_if_needed(
        SQLITE_TABLE_HEALTH,
        HEALTH_FILE,
        HEALTH_SHEET,
        HEALTH_COLUMNS,
        date_cols=["記録日"],
        unique_cols=["記録日", "利用者名"],
    )


def load_health_data(start_date=None, end_date=None, recent_days=None):
    """健康チェックを読み込む。start_date/end_date指定時はSupabase側で期間を絞って高速化する。"""
    with perf_timer("load_health_data", f"{start_date or ''}-{end_date or ''} recent={recent_days or ''}"):
        ensure_health_file()
        if recent_days and start_date is None and end_date is None:
            start_date = recent_start_date(recent_days)
            end_date = today_jst()
        if supabase_is_enabled():
            df = supabase_read_table(SQLITE_TABLE_HEALTH, HEALTH_COLUMNS, date_field="記録日", start_date=start_date, end_date=end_date)
        else:
            df = load_sqlite_table(SQLITE_TABLE_HEALTH, HEALTH_COLUMNS, date_cols=["記録日"])
            df = _filter_df_by_date_range(df, "記録日", start_date, end_date)
        df = attach_user_ids(df)

        if not df.empty:
            df["記録日"] = pd.to_datetime(df["記録日"], errors="coerce")
            df["利用者名"] = df["利用者名"].astype(str).str.strip()

        return df.astype("object")


def save_health_data(df):
    """健康チェックをSQLiteへ保存する。"""
    ensure_dirs()
    df = normalize_df_columns(df, HEALTH_COLUMNS)
    df = attach_user_ids(df)

    if not df.empty:
        df["記録日"] = pd.to_datetime(df["記録日"], errors="coerce")
        df["利用者名"] = df["利用者名"].astype(str).str.strip()
        df["_key"] = df.apply(lambda row: make_date_user_key(row["記録日"], row["利用者名"]), axis=1)
        df = df[df["_key"] != ""]
        df = df.drop_duplicates(subset=["_key"], keep="last")
        df = df.drop(columns=["_key"])

    save_sqlite_table(
        df,
        SQLITE_TABLE_HEALTH,
        HEALTH_COLUMNS,
        date_cols=["記録日"],
        unique_cols=["記録日", "利用者名"],
    )
    clear_hidamari_read_cache("健康チェック保存")


def find_health_index(df, record_date, user_name):
    if df.empty:
        return None

    work = df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
    work["利用者名"] = work["利用者名"].astype(str).str.strip()

    target_date = pd.to_datetime(record_date, errors="coerce")
    if pd.isna(target_date):
        return None

    mask = (work["記録日"].dt.date == target_date.date()) & (work["利用者名"] == clean_text(user_name))
    matches = work.index[mask].tolist()

    if not matches:
        return None

    return matches[0]


def upsert_health_record(record):
    record["user_id"] = ensure_user_id_value(record.get("user_id", ""), record.get("利用者名", ""))
    df = load_health_data()
    df = df.astype("object")

    idx = find_health_index(df, record["記録日"], record["利用者名"])

    if idx is None:
        new_df = pd.DataFrame([record], columns=HEALTH_COLUMNS).astype("object")
        df = pd.concat([df, new_df], ignore_index=True)
        action = "登録"
    else:
        for col in HEALTH_COLUMNS:
            df.at[idx, col] = record.get(col, "")
        action = "更新"

    save_health_data(df)
    add_audit_log(action, SQLITE_TABLE_HEALTH, make_date_user_key(record["記録日"], record["利用者名"]), "健康チェックを保存しました")

    return action


def get_month_health_data(df, user_name, year, month):
    if df.empty:
        return df

    work = df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")

    return work[
        (work["利用者名"] == user_name)
        & (work["記録日"].dt.year == int(year))
        & (work["記録日"].dt.month == int(month))
    ].sort_values("記録日")


# =========================
# 排泄チェックデータ
# =========================
def ensure_excretion_file():
    """互換用。実データはSQLite（hidamari_health.db / excretion_records）へ保存します。"""
    migrate_excel_to_sqlite_if_needed(
        SQLITE_TABLE_EXCRETION,
        EXCRETION_FILE,
        EXCRETION_SHEET,
        EXCRETION_COLUMNS,
        date_cols=["記録日"],
        unique_cols=["記録日", "利用者名", "時間帯"],
    )


def normalize_excretion_record(record):
    urine_amount = clean_text(record.get("尿量", "なし"), "なし")
    urine_type = clean_text(record.get("尿性状", "なし"), "なし")
    stool_amount = clean_text(record.get("便量", "なし"), "なし")
    stool_type = clean_text(record.get("便性状", "なし"), "なし")

    if urine_amount == "":
        urine_amount = "なし"
    if urine_type == "":
        urine_type = "なし"
    if stool_amount == "":
        stool_amount = "なし"
    if stool_type == "":
        stool_type = "なし"

    if urine_amount == "なし":
        urine_type = "なし"

    if stool_amount == "なし":
        stool_type = "なし"

    record["尿量"] = urine_amount
    record["尿量コード"] = URINE_AMOUNT_CODE.get(urine_amount, "")
    record["尿性状"] = urine_type
    record["尿性状コード"] = URINE_TYPE_CODE.get(urine_type, "")
    record["便量"] = stool_amount
    record["便量コード"] = STOOL_AMOUNT_CODE.get(stool_amount, "")
    record["便性状"] = stool_type
    record["便性状コード"] = STOOL_TYPE_CODE.get(stool_type, "")

    return record


def load_excretion_data(start_date=None, end_date=None, recent_days=None):
    """排泄チェックを読み込む。start_date/end_date指定時はSupabase側で期間を絞って高速化する。"""
    with perf_timer("load_excretion_data", f"{start_date or ''}-{end_date or ''} recent={recent_days or ''}"):
        ensure_excretion_file()
        if recent_days and start_date is None and end_date is None:
            start_date = recent_start_date(recent_days)
            end_date = today_jst()
        if supabase_is_enabled():
            df = supabase_read_table(SQLITE_TABLE_EXCRETION, EXCRETION_COLUMNS, date_field="記録日", start_date=start_date, end_date=end_date)
        else:
            df = load_sqlite_table(SQLITE_TABLE_EXCRETION, EXCRETION_COLUMNS, date_cols=["記録日"])
            df = _filter_df_by_date_range(df, "記録日", start_date, end_date)

        if not df.empty:
            df["記録日"] = pd.to_datetime(df["記録日"], errors="coerce")
            for col in ["利用者名", "時間帯", "時間帯目安", "尿量", "尿性状", "便量", "便性状", "排泄メモ", "入力者", "登録日時"]:
                df[col] = df[col].fillna("").astype(str)

        return df.astype("object")


def save_excretion_data(df):
    """排泄チェックをSQLiteへ保存する。"""
    ensure_dirs()
    df = normalize_df_columns(df, EXCRETION_COLUMNS)
    df = attach_user_ids(df)

    records = []
    for _, row in df.iterrows():
        rec = row.to_dict()
        rec = normalize_excretion_record(rec)
        records.append(rec)

    df = pd.DataFrame(records, columns=EXCRETION_COLUMNS).astype("object")

    if not df.empty:
        df["記録日"] = pd.to_datetime(df["記録日"], errors="coerce")
        df["利用者名"] = df["利用者名"].astype(str).str.strip()
        df["時間帯"] = df["時間帯"].astype(str).str.strip()
        df["_key"] = df.apply(lambda row: make_excretion_key(row["記録日"], row["利用者名"], row["時間帯"]), axis=1)
        df = df[df["_key"] != ""]
        df = df.drop_duplicates(subset=["_key"], keep="last")
        df = df.drop(columns=["_key"])

    df = df[EXCRETION_COLUMNS]
    save_sqlite_table(
        df,
        SQLITE_TABLE_EXCRETION,
        EXCRETION_COLUMNS,
        date_cols=["記録日"],
        unique_cols=["記録日", "利用者名", "時間帯"],
    )
    clear_hidamari_read_cache("排泄チェック保存")


def find_excretion_index(df, record_date, user_name, slot):
    if df.empty:
        return None

    work = df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
    work["利用者名"] = work["利用者名"].astype(str).str.strip()
    work["時間帯"] = work["時間帯"].astype(str).str.strip()

    target_date = pd.to_datetime(record_date, errors="coerce")
    if pd.isna(target_date):
        return None

    mask = (
        (work["記録日"].dt.date == target_date.date())
        & (work["利用者名"] == clean_text(user_name))
        & (work["時間帯"] == clean_text(slot))
    )

    matches = work.index[mask].tolist()

    if not matches:
        return None

    return matches[0]


def get_excretion_row(df, record_date, user_name, slot):
    idx = find_excretion_index(df, record_date, user_name, slot)

    if idx is None:
        return None

    return df.loc[idx]


def upsert_excretion_record(record):
    record["user_id"] = ensure_user_id_value(record.get("user_id", ""), record.get("利用者名", ""))
    record = normalize_excretion_record(record)

    df = load_excretion_data()
    idx = find_excretion_index(
        df,
        record["記録日"],
        record["利用者名"],
        record["時間帯"],
    )

    if idx is None:
        df = pd.concat(
            [df, pd.DataFrame([record], columns=EXCRETION_COLUMNS).astype("object")],
            ignore_index=True,
        )
        action = "登録"
    else:
        for col in EXCRETION_COLUMNS:
            df.at[idx, col] = record.get(col, "")
        action = "更新"

    save_excretion_data(df)
    add_audit_log(action, SQLITE_TABLE_EXCRETION, make_excretion_key(record["記録日"], record["利用者名"], record["時間帯"]), "排泄チェックを保存しました")

    return action


def get_day_excretion_data(df, record_date, user_name=None):
    if df.empty:
        return df

    work = df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")

    target_date = pd.to_datetime(record_date, errors="coerce")
    if pd.isna(target_date):
        return pd.DataFrame(columns=EXCRETION_COLUMNS)

    work = work[work["記録日"].dt.date == target_date.date()]

    if user_name and user_name != "全員":
        work = work[work["利用者名"] == user_name]

    slot_order = {slot: i for i, (slot, _) in enumerate(EXCRETION_SLOTS)}
    work["_slot_order"] = work["時間帯"].map(slot_order).fillna(99)
    work = work.sort_values(["利用者名", "_slot_order"]).drop(columns=["_slot_order"])

    return work


def get_month_excretion_data(df, user_name, year, month):
    if df.empty:
        return df

    work = df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")

    return work[
        (work["利用者名"] == user_name)
        & (work["記録日"].dt.year == int(year))
        & (work["記録日"].dt.month == int(month))
    ].sort_values(["記録日", "時間帯"])


def is_present_excretion_value(value, none_words=None):
    """排尿・排便の有無判定を共通化する。

    空欄や「なし」「無」「0」などは未実施・記録なしとして扱う。
    これにより、便量が空欄の行を「排便あり」と誤判定しない。
    """
    none_words = none_words or {
        "", "なし", "無し", "無", "ない", "ナシ", "未", "未記録", "未入力",
        "0", "０", "nan", "none", "nat", "null", "-", "ー", "―", "－",
    }
    try:
        if pd.isna(value):
            return False
    except Exception:
        pass
    text = str(value).strip()
    return text.lower() not in {str(v).lower() for v in none_words}


def is_stool_present_row(row):
    """1行の排泄データから、排便ありかを判定する。

    Ver4.5.3修正：
    - 排便3日なし抽出で誤判定しないため、原則として「便量」を主判定にする。
    - 便性状が「普通便」の初期値だけ入っている行は、排便ありにしない。
    - 便量コードが1以上、または便量が「少・中・大・多・普・普通・あり・有」なら排便あり。
    - 便量が空欄の場合のみ、下痢便・水様便・硬便など明確な便性状を補助判定に使う。
    """
    if row is None:
        return False
    try:
        getv = row.get
    except Exception:
        return False

    stool_amount = clean_text(getv("便量", "")) if "clean_text" in globals() else str(getv("便量", "") or "").strip()
    stool_amount_code = clean_text(getv("便量コード", "")) if "clean_text" in globals() else str(getv("便量コード", "") or "").strip()
    stool_type = clean_text(getv("便性状", "")) if "clean_text" in globals() else str(getv("便性状", "") or "").strip()
    stool_type_code = clean_text(getv("便性状コード", "")) if "clean_text" in globals() else str(getv("便性状コード", "") or "").strip()

    none_values = {"", "なし", "無し", "無", "ない", "ナシ", "未", "未記録", "未入力", "0", "０", "nan", "none", "nat", "null", "-", "ー", "―", "－"}
    positive_amount_values = {"少", "中", "大", "多", "普", "普通", "あり", "有", "有り", "排便あり"}
    abnormal_type_values = {"硬便", "下痢便", "水様便", "軟便", "泥状便"}

    amount_norm = str(stool_amount).strip()
    amount_code_norm = str(stool_amount_code).strip()
    type_norm = str(stool_type).strip()
    type_code_norm = str(stool_type_code).strip()

    # 便量が明確に「なし」の場合は、便性状に初期値が残っていても排便なし。
    if amount_norm.lower() in {v.lower() for v in none_values}:
        return False

    # 便量コードが1以上なら排便あり。
    if amount_code_norm not in ["", "0", "０"]:
        try:
            if int(float(amount_code_norm)) > 0:
                return True
        except Exception:
            return True

    # 便量の日本語値で判定。
    if amount_norm in positive_amount_values:
        return True

    # 便量が空欄・旧データの場合だけ、便性状を補助判定に使う。
    # ただし「普通便」単独は、初期値の可能性が高いため排便ありにしない。
    if amount_norm == "":
        if type_code_norm not in ["", "0", "０", "1"]:
            return True
        if type_norm in abnormal_type_values:
            return True

    return False

def count_stool_records(df):
    """実際に排便ありとみなせる行数を数える。"""
    if df is None or df.empty:
        return 0
    return int(df.apply(is_stool_present_row, axis=1).sum())


def count_urine_records(df):
    """尿量をもとに、実際に排尿ありとみなせる行数を数える。"""
    if df is None or df.empty or "尿量" not in df.columns:
        return 0
    return int(df["尿量"].apply(is_present_excretion_value).sum())


def summarize_excretion(df):
    if df.empty:
        return {
            "排尿回数": 0,
            "排便回数": 0,
            "濃縮尿": 0,
            "下痢便": 0,
            "水様便": 0,
            "排便なし枠": 0,
        }

    stool_count = count_stool_records(df)
    urine_count = count_urine_records(df)

    return {
        "排尿回数": urine_count,
        "排便回数": stool_count,
        "濃縮尿": int((df["尿性状"].fillna("") == "濃縮尿").sum()) if "尿性状" in df.columns else 0,
        "下痢便": int((df["便性状"].fillna("") == "下痢便").sum()) if "便性状" in df.columns else 0,
        "水様便": int((df["便性状"].fillna("") == "水様便").sum()) if "便性状" in df.columns else 0,
        "排便なし枠": int(len(df) - stool_count),
    }


def build_excretion_text(df):
    if df.empty:
        return "排泄記録はありません。"

    lines = []

    for _, row in df.iterrows():
        lines.append(
            f"{row['記録日'].strftime('%m/%d') if pd.notna(row['記録日']) else ''} "
            f"{row['時間帯']}：尿 {row['尿量']}・{row['尿性状']} ／ 便 {row['便量']}・{row['便性状']}"
        )

    return "\\n".join(lines)




# =========================
# 入力チェック・注意通知・差分検知
# =========================
def validate_health_record(record):
    """健康チェック入力の整合性を確認する。"""
    warnings = []
    errors = []

    temp = safe_float(record.get("体温"), 0)
    spo2 = safe_int(record.get("SpO2"), 0)
    bp_high = safe_int(record.get("血圧上"), 0)
    bp_low = safe_int(record.get("血圧下"), 0)
    pulse = safe_int(record.get("脈拍"), 0)
    weight = safe_float(record.get("体重"), 0)

    if temp == 0:
        warnings.append("体温が0です。未測定の場合はそのままでもよいですが、入力漏れでないか確認してください。")
    elif temp < 34.0 or temp > 42.0:
        errors.append("体温が通常の入力範囲から外れています。入力値を確認してください。")
    elif temp >= 37.5:
        warnings.append("体温が37.5℃以上です。発熱傾向として申し送り対象になります。")

    if spo2 == 0:
        warnings.append("SpO2が0です。未測定か入力漏れか確認してください。")
    elif spo2 < 80:
        errors.append("SpO2が80未満です。入力ミスの可能性があります。")
    elif spo2 <= 93:
        warnings.append("SpO2が93％以下です。注意して確認してください。")

    if bp_high == 0 or bp_low == 0:
        warnings.append("血圧が0です。未測定か入力漏れか確認してください。")
    elif bp_low > bp_high:
        errors.append("血圧下が血圧上を上回っています。入力値を確認してください。")
    elif bp_high >= 160:
        warnings.append("血圧上が160以上です。注意して確認してください。")

    if pulse == 0:
        warnings.append("脈拍が0です。未測定か入力漏れか確認してください。")
    elif pulse < 40 or pulse > 130:
        warnings.append("脈拍が通常範囲から外れています。入力値と状態を確認してください。")

    if weight < 0:
        errors.append("体重がマイナスです。入力値を確認してください。")

    for meal in ["朝食摂取率", "昼食摂取率", "夕食摂取率"]:
        value = safe_int(record.get(meal), 0)
        if value < 0 or value > 100:
            errors.append(f"{meal}は0?100％で入力してください。")
        elif value <= 50:
            warnings.append(f"{meal}が50％以下です。食事量低下として確認してください。")

    return errors, warnings


def validate_excretion_record(record):
    """排泄チェック入力の整合性を確認する。"""
    warnings = []
    errors = []

    urine_amount = clean_text(record.get("尿量", "なし"), "なし")
    urine_type = clean_text(record.get("尿性状", "なし"), "なし")
    stool_amount = clean_text(record.get("便量", "なし"), "なし")
    stool_type = clean_text(record.get("便性状", "なし"), "なし")

    if urine_amount == "なし" and urine_type != "なし":
        errors.append("尿量が「なし」の場合、尿性状も「なし」にしてください。")
    if stool_amount == "なし" and stool_type != "なし":
        errors.append("便量が「なし」の場合、便性状も「なし」にしてください。")

    if urine_amount != "なし" and urine_type == "なし":
        warnings.append("尿量がありますが、尿性状が「なし」です。普通尿・濃縮尿の確認をおすすめします。")
    if stool_amount != "なし" and stool_type == "なし":
        warnings.append("便量がありますが、便性状が「なし」です。普通便・下痢便・水様便の確認をおすすめします。")

    if urine_type == "濃縮尿":
        warnings.append("濃縮尿の記録があります。水分摂取や体調変化の確認対象です。")
    if stool_type in ["下痢便", "水様便"]:
        warnings.append(f"{stool_type}の記録があります。体調変化として確認対象です。")

    return errors, warnings


def get_previous_health_record(health_df, record_date, user_name):
    """指定日より前の直近健康記録を取得する。"""
    if health_df.empty:
        return None

    work = health_df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
    target_date = pd.to_datetime(record_date, errors="coerce")

    if pd.isna(target_date):
        return None

    work = work[
        (work["利用者名"] == user_name)
        & (work["記録日"].dt.date < target_date.date())
    ].sort_values("記録日")

    if work.empty:
        return None

    return work.iloc[-1]


def build_health_diff_text(health_df, record_date, user_name, current_record=None):
    """前回記録との差分を文章化する。"""
    prev = get_previous_health_record(health_df, record_date, user_name)

    if prev is None:
        return "前回比較：比較できる過去記録はありません。"

    if current_record is None:
        idx = find_health_index(health_df, record_date, user_name)
        if idx is None:
            return "前回比較：本日の健康記録がありません。"
        current_record = health_df.loc[idx].to_dict()

    lines = []

    checks = [
        ("体温", 0.5, "℃"),
        ("SpO2", 3, "％"),
        ("体重", 1.0, "kg"),
        ("朝食摂取率", 30, "％"),
        ("昼食摂取率", 30, "％"),
        ("夕食摂取率", 30, "％"),
    ]

    for col, threshold, unit in checks:
        now = safe_float(current_record.get(col), 0)
        before = safe_float(prev.get(col), 0)

        if now == 0 or before == 0:
            continue

        diff = now - before

        if abs(diff) >= threshold:
            direction = "上昇" if diff > 0 else "低下"
            lines.append(f"{col}が前回より{abs(round(diff, 1))}{unit}{direction}")

    if not lines:
        return "前回比較：大きな差分は目立ちません。"

    return "前回比較：" + "、".join(lines)


def build_excretion_diff_text(ex_df, record_date, user_name):
    """前回排泄記録との差分を文章化する。"""
    if ex_df.empty:
        return "排泄差分：比較できる過去記録はありません。"

    work = ex_df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
    target_date = pd.to_datetime(record_date, errors="coerce")

    if pd.isna(target_date):
        return "排泄差分：日付を確認できません。"

    today_df = get_day_excretion_data(work, target_date.date(), user_name)

    prev_dates = work[
        (work["利用者名"] == user_name)
        & (work["記録日"].dt.date < target_date.date())
    ]["記録日"].dt.date.dropna().unique()

    if len(prev_dates) == 0:
        return "排泄差分：比較できる過去排泄記録はありません。"

    prev_date = sorted(prev_dates)[-1]
    prev_df = get_day_excretion_data(work, prev_date, user_name)

    now_sum = summarize_excretion(today_df)
    prev_sum = summarize_excretion(prev_df)

    lines = []

    if now_sum["排尿回数"] > prev_sum["排尿回数"] + 2:
        lines.append("排尿回数が前回より増えています")
    if now_sum["排便回数"] == 0 and prev_sum["排便回数"] > 0:
        lines.append("前回は排便記録がありましたが、本日は排便記録がありません")
    if now_sum["濃縮尿"] > prev_sum["濃縮尿"]:
        lines.append("濃縮尿の記録が前回より増えています")
    if now_sum["下痢便"] + now_sum["水様便"] > prev_sum["下痢便"] + prev_sum["水様便"]:
        lines.append("下痢便・水様便の記録が前回より増えています")

    if not lines:
        return "排泄差分：前回と比べて大きな変化は目立ちません。"

    return "排泄差分：" + "、".join(lines)


def build_attention_users(health_df, ex_df, target_date):
    """今日の注意利用者一覧を作成する。

    自分専用ダッシュボード・管理者ダッシュボードで使う注意一覧。
    体温・SpO2・食事量などの注意項目に加えて、健康チェック入力の
    「気になる変化」の具体内容も同じ一覧に表示する。
    """
    rows = []

    for user in active_users:
        notes = []
        change_text = ""
        family_memo = ""

        # 健康記録
        if not health_df.empty:
            idx = find_health_index(health_df, target_date, user)
            if idx is not None:
                h = health_df.loc[idx]
                if safe_float(h.get("体温"), 0) >= 37.5:
                    notes.append("発熱傾向")
                if safe_int(h.get("SpO2"), 100) <= 93 and safe_int(h.get("SpO2"), 100) != 0:
                    notes.append("SpO2低下")
                for meal in ["朝食摂取率", "昼食摂取率", "夕食摂取率"]:
                    if safe_int(h.get(meal), 100) <= 50:
                        notes.append(f"{meal.replace('摂取率','')}50％以下")

                change_text = clean_text(h.get("気になる変化", ""))
                family_memo = clean_text(h.get("家族共有メモ", ""))
                if change_text:
                    notes.append("気になる変化あり")

        # 排泄記録
        user_ex = get_day_excretion_data(ex_df, target_date, user)
        if not user_ex.empty:
            ex_sum = summarize_excretion(user_ex)
            if ex_sum["水様便"] > 0:
                notes.append("水様便")
            if ex_sum["下痢便"] > 0:
                notes.append("下痢便")
            if ex_sum["濃縮尿"] > 0:
                notes.append("濃縮尿")
            if ex_sum["排便回数"] == 0:
                notes.append("確認日排便記録なし")

        # 未排便3日
        if not ex_df.empty:
            work = ex_df.copy()
            work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
            recent_dates = sorted([
                d for d in work[work["利用者名"] == user]["記録日"].dt.date.dropna().unique()
                if d <= target_date
            ])[-3:]

            if len(recent_dates) >= 3:
                no_stool_all = True
                for d in recent_dates:
                    ddf = get_day_excretion_data(work, d, user)
                    if summarize_excretion(ddf)["排便回数"] > 0:
                        no_stool_all = False
                        break
                if no_stool_all:
                    notes.append("未排便3日")

        if notes:
            rows.append({
                "利用者名": user,
                "注意項目": "、".join(sorted(set(notes))),
                "気になる変化": change_text,
                "家族共有メモ": family_memo,
            })

    columns = ["利用者名", "注意項目", "気になる変化", "家族共有メモ"]
    return pd.DataFrame(rows, columns=columns)



def get_active_user_names_for_dashboard():
    """ダッシュボード用の利用者一覧を安全に取得する。

    active_users の初期化タイミングに左右されないよう、
    必要に応じて利用者マスタを直接読み直す。
    """
    users = []
    try:
        if "active_users" in globals() and active_users:
            users = [clean_text(u) for u in active_users if clean_text(u)]
    except Exception:
        users = []

    if not users:
        try:
            df_users = load_users(include_hidden=False)
            if df_users is not None and not df_users.empty and "利用者名" in df_users.columns:
                users = [clean_text(u) for u in df_users["利用者名"].tolist() if clean_text(u)]
        except Exception:
            users = []

    return list(dict.fromkeys(users))


def build_no_stool_3days_users(ex_df, target_date):
    """確認日時点で「3日以上排便がない」利用者を一覧化する。

    Ver4.5.2修正：
    - 直近3日間の行だけでなく「最終排便日」から未排便日数を計算する。
    - 最終排便日が確認日の3日前以前なら抽出する。
      例：確認日6/7、最終排便6/4 → 6/5・6/6・6/7の3日間なしとして抽出。
    - 排泄データがない利用者も「排便記録なし」として抽出する。
    - 便量・便量コード・便性状・便性状コードを総合して排便ありを判定する。
    """
    columns = ["利用者名", "未排便日数", "対象期間", "最終排便記録", "確認メモ"]
    rows = []

    target = pd.to_datetime(target_date, errors="coerce")
    if pd.isna(target):
        return pd.DataFrame(columns=columns)

    target_day = target.date()
    check_start = target_day - timedelta(days=2)
    period_text = f"{check_start.strftime('%m/%d')}〜{target_day.strftime('%m/%d')}"

    work = ex_df.copy() if ex_df is not None else pd.DataFrame()
    if not work.empty:
        if "記録日" not in work.columns:
            work = pd.DataFrame()
        else:
            work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
            work = work[work["記録日"].notna()].copy()
            if "利用者名" in work.columns:
                work["利用者名"] = work["利用者名"].astype(str).str.strip()

    target_users = get_active_user_names_for_dashboard()

    for user in target_users:
        if work.empty or "利用者名" not in work.columns:
            user_df = pd.DataFrame()
        else:
            user_df = work[
                (work["利用者名"].astype(str).str.strip() == str(user).strip())
                & (work["記録日"].dt.date <= target_day)
            ].copy()

        last_stool_text = "確認できません"
        no_stool_days = "3日以上"

        if user_df.empty:
            # 排泄データ自体がない場合も、確認対象として出す
            rows.append({
                "利用者名": user,
                "未排便日数": no_stool_days,
                "対象期間": period_text,
                "最終排便記録": last_stool_text,
                "確認メモ": "排便記録が確認できません。水分・食事量・腹部症状・普段の排便間隔を確認してください。",
            })
            continue

        stool_df = user_df[user_df.apply(is_stool_present_row, axis=1)].copy()

        if stool_df.empty:
            rows.append({
                "利用者名": user,
                "未排便日数": no_stool_days,
                "対象期間": period_text,
                "最終排便記録": last_stool_text,
                "確認メモ": "確認日までの排便記録がありません。水分・食事量・腹部症状・普段の排便間隔を確認してください。",
            })
            continue

        last_stool_date = stool_df["記録日"].max().date()
        days_since = (target_day - last_stool_date).days
        last_stool_text = last_stool_date.strftime("%Y/%m/%d")

        if days_since >= 3:
            rows.append({
                "利用者名": user,
                "未排便日数": f"{days_since}日",
                "対象期間": period_text,
                "最終排便記録": last_stool_text,
                "確認メモ": "3日以上排便記録がありません。水分・食事量・腹部症状・普段の排便間隔を確認してください。",
            })

    return pd.DataFrame(rows, columns=columns)





# =========================
# 写真から半自動入力（OCR補助）
# =========================
PHOTO_IMPORT_COLUMNS = [
    "取り込む",
    "記録日",
    "利用者名",
    "体温",
    "血圧上",
    "血圧下",
    "脈拍",
    "SpO2",
    "体重",
    "朝食摂取率",
    "昼食摂取率",
    "夕食摂取率",
    "家族共有メモ",
    "気になる変化",
]

# =========================
# OpenAIモデル切替設定（Ver4.8.3）
# =========================
# Streamlit Secrets / 環境変数 / app_settings の順でモデル名を取得する。
# 例:
# OPENAI_MODEL = "gpt-4o-mini"
# OPENAI_MODEL_VISION = "gpt-4o-mini"
# OPENAI_MODEL_SHORT_GOAL = "gpt-4o-mini"
# OPENAI_MODEL_MONITORING = "gpt-4o-mini"
# OPENAI_MODEL_ADMIN = "gpt-4.1-mini"
OPENAI_MODEL_DEFAULTS = {
    "default": "gpt-4o-mini",
    "vision": "gpt-4o-mini",
    "short_goal": "gpt-4o-mini",
    "monitoring": "gpt-4o-mini",
    "admin": "gpt-4.1-mini",
}

OPENAI_MODEL_LABELS = {
    "default": "共通",
    "vision": "写真AI取込",
    "short_goal": "短期目標AI要約",
    "monitoring": "モニタリングAI整形",
    "admin": "AI管理者アドバイス",
}


def _read_openai_model_from_secrets(key_name):
    """OPENAI_MODEL系の値をStreamlit Secrets / [openai] / 環境変数から読む。"""
    try:
        value = st.secrets.get(key_name, "")
        if value:
            return str(value).strip()
    except Exception:
        pass
    try:
        openai_section = st.secrets.get("openai", {})
        short = key_name.replace("OPENAI_MODEL_", "").lower()
        candidates = []
        if short and short != key_name.lower():
            candidates.extend([f"model_{short}", f"{short}_model", short])
        if key_name == "OPENAI_MODEL":
            candidates.extend(["model", "default_model"])
        for k in candidates:
            value = _secret_get(openai_section, k, "")
            if value:
                return str(value).strip()
    except Exception:
        pass
    value = os.environ.get(key_name, "")
    if value:
        return str(value).strip()
    return ""


def get_openai_model(purpose="default", default=None):
    """
    用途別にOpenAIモデル名を取得する。
    優先順位:
    1. OPENAI_MODEL_<PURPOSE>
    2. OPENAI_MODEL
    3. app_settings の openai_model_<purpose>
    4. 既定値
    """
    purpose = clean_text(purpose, "default").lower()
    default_model = default or OPENAI_MODEL_DEFAULTS.get(purpose, OPENAI_MODEL_DEFAULTS["default"])

    env_key = f"OPENAI_MODEL_{purpose.upper()}"
    if purpose != "default":
        value = _read_openai_model_from_secrets(env_key)
        if value:
            return value

    value = _read_openai_model_from_secrets("OPENAI_MODEL")
    if value:
        return value

    try:
        setting_value = get_app_setting(f"openai_model_{purpose}", "")
        if setting_value:
            return clean_text(setting_value)
    except Exception:
        pass

    try:
        setting_value = get_app_setting("openai_model_default", "")
        if setting_value:
            return clean_text(setting_value)
    except Exception:
        pass

    return default_model


def get_openai_model_status_df():
    """用途別モデル設定を表にする。"""
    rows = []
    for purpose, label in OPENAI_MODEL_LABELS.items():
        rows.append({
            "用途": label,
            "purpose": purpose,
            "使用モデル": get_openai_model(purpose),
            "既定モデル": OPENAI_MODEL_DEFAULTS.get(purpose, OPENAI_MODEL_DEFAULTS["default"]),
            "Secretsキー例": "OPENAI_MODEL" if purpose == "default" else f"OPENAI_MODEL_{purpose.upper()}",
        })
    return pd.DataFrame(rows)


def try_ocr_image(uploaded_file):
    """
    旧方式OCR。Streamlit CloudではTesseract本体が入っていないことが多いため、
    使えない場合は空文字を返します。
    """
    try:
        from PIL import Image
        import pytesseract
        uploaded_file.seek(0)
        image = Image.open(uploaded_file)
        text = pytesseract.image_to_string(image, lang="jpn+eng")
        return clean_text(text)
    except Exception:
        return ""


def try_openai_vision_photo_import(uploaded_file, default_user, year, month, api_key=""):
    """
    OpenAI Visionで写真を読み取り、健康チェック候補表を返します。
    AIは下書き作成のみ。保存前に必ず職員が確認します。
    """
    api_key = get_openai_api_key(api_key)
    if not api_key:
        return pd.DataFrame(columns=PHOTO_IMPORT_COLUMNS), "OpenAI APIキーが未設定です。"

    try:
        from openai import OpenAI
    except Exception:
        return pd.DataFrame(columns=PHOTO_IMPORT_COLUMNS), "openaiライブラリが未インストールです。requirements.txtに openai を追加してください。"

    try:
        uploaded_file.seek(0)
        img_bytes = uploaded_file.read()
        b64 = base64.b64encode(img_bytes).decode("utf-8")
        mime = "image/png"
        name = clean_text(getattr(uploaded_file, "name", "")).lower()
        if name.endswith(".jpg") or name.endswith(".jpeg"):
            mime = "image/jpeg"

        client = OpenAI(api_key=api_key)
        prompt = f"""
あなたは介護施設の健康チェック表を読み取る補助者です。
画像内の手書き表から、健康チェックデータの候補を抽出してください。
対象年月は {int(year)}年{int(month)}月 です。

重要ルール：
- 推測で埋めない。読めない値は空欄または0にする。
- 日付、体温、血圧上、血圧下、脈拍を優先する。
- SpO2、体重が読める場合のみ入れる。
- 排便の丸やメモは「気になる変化」に短く入れてよい。
- 出力はJSONのみ。説明文は不要。

JSON形式：
{{
  "rows": [
    {{"day": 1, "temp": 36.5, "bp_high": 128, "bp_low": 70, "pulse": 82, "spo2": 0, "weight": 0, "memo": ""}}
  ]
}}
"""
        resp = client.chat.completions.create(
            model=get_openai_model("vision", "gpt-4o-mini"),
            messages=[
                {"role": "system", "content": "画像から健康チェック表の候補データをJSONで抽出します。医療判断はしません。"},
                {"role": "user", "content": [
                    {"type": "text", "text": prompt},
                    {"type": "image_url", "image_url": {"url": f"data:{mime};base64,{b64}"}},
                ]},
            ],
            temperature=0,
            response_format={"type": "json_object"},
        )
        content = resp.choices[0].message.content or "{}"
        data = json.loads(content)
        rows = []
        for item in data.get("rows", []):
            day = safe_int(item.get("day"), 0)
            if day < 1 or day > 31:
                continue
            try:
                record_date = date(int(year), int(month), day)
            except Exception:
                continue
            rows.append({
                "取り込む": True,
                "記録日": record_date,
                "利用者名": default_user,
                "体温": safe_float(item.get("temp"), 0),
                "血圧上": safe_int(item.get("bp_high"), 0),
                "血圧下": safe_int(item.get("bp_low"), 0),
                "脈拍": safe_int(item.get("pulse"), 0),
                "SpO2": safe_int(item.get("spo2"), 0),
                "体重": safe_float(item.get("weight"), 0),
                "朝食摂取率": 100,
                "昼食摂取率": 100,
                "夕食摂取率": 100,
                "家族共有メモ": "",
                "気になる変化": clean_text(item.get("memo", "写真AI取込候補"), "写真AI取込候補"),
            })
        return pd.DataFrame(rows, columns=PHOTO_IMPORT_COLUMNS), ""
    except Exception as e:
        return pd.DataFrame(columns=PHOTO_IMPORT_COLUMNS), f"OpenAI画像読み取りでエラーが出ました：{e}"


def parse_photo_ocr_text(raw_text, default_user, year, month, input_staff=""):
    """
    OCRテキストから候補データを作成します。
    手書き帳票は誤読が起きるため、ここでは「候補作成」に留め、
    必ず st.data_editor で職員が確認してから保存します。
    """
    rows = []
    raw_text = clean_text(raw_text)

    if not raw_text:
        return pd.DataFrame(columns=PHOTO_IMPORT_COLUMNS)

    lines = [line.strip() for line in raw_text.splitlines() if line.strip()]
    last_day = None

    for line in lines:
        normalized = line.replace("／", "/").replace("｜", " ").replace("|", " ").replace("　", " ")
        normalized = re.sub(r"\s+", " ", normalized)

        # 日付候補：単独の1〜31、または 5/12 のような表記
        day = None
        m_md = re.search(r"(?:\d{1,2}/)?([1-9]|[12]\d|3[01])(?:日)?", normalized)
        if m_md:
            try:
                candidate = int(m_md.group(1))
                if 1 <= candidate <= 31:
                    day = candidate
                    last_day = day
            except Exception:
                day = None

        if day is None:
            day = last_day

        # 体温候補：35.0〜42.0程度
        temp = ""
        temp_matches = re.findall(r"(3[5-9]\.\d|4[0-2]\.\d|3[5-9]|4[0-2])", normalized)
        if temp_matches:
            temp = temp_matches[0]

        # 血圧候補：120/70 など
        bp_high = ""
        bp_low = ""
        bp = re.search(r"(\d{2,3})\s*/\s*(\d{2,3})", normalized)
        if bp:
            bp_high = bp.group(1)
            bp_low = bp.group(2)

        # 脈拍候補：血圧の後ろに出る2〜3桁を優先
        pulse = ""
        if bp:
            after_bp = normalized[bp.end():]
            nums = re.findall(r"\b([4-9]\d|1[0-4]\d)\b", after_bp)
            if nums:
                pulse = nums[0]

        # 体温または血圧がある行だけ候補化
        if temp or bp_high or bp_low or pulse:
            try:
                record_date = date(int(year), int(month), int(day)) if day else today_jst()
            except Exception:
                record_date = today_jst()

            rows.append({
                "取り込む": True,
                "記録日": record_date,
                "利用者名": default_user,
                "体温": safe_float(temp, 0) if temp != "" else 0,
                "血圧上": safe_int(bp_high, 0),
                "血圧下": safe_int(bp_low, 0),
                "脈拍": safe_int(pulse, 0),
                "SpO2": 0,
                "体重": 0.0,
                "朝食摂取率": 100,
                "昼食摂取率": 100,
                "夕食摂取率": 100,
                "家族共有メモ": "",
                "気になる変化": "写真取込候補",
            })

    return pd.DataFrame(rows, columns=PHOTO_IMPORT_COLUMNS)


def make_blank_photo_import_rows(default_user, year, month):
    """OCRが難しい時用に、選択月の1か月分の確認入力表を作成します。"""
    rows = []
    try:
        year = int(year)
        month = int(month)
        if month == 12:
            next_month = date(year + 1, 1, 1)
        else:
            next_month = date(year, month + 1, 1)
        last_day = (next_month - timedelta(days=1)).day
    except Exception:
        year = today_jst().year
        month = today_jst().month
        last_day = 31

    for d in range(1, last_day + 1):
        rows.append({
            "取り込む": False,
            "記録日": date(year, month, d),
            "利用者名": default_user,
            "体温": 0.0,
            "血圧上": 0,
            "血圧下": 0,
            "脈拍": 0,
            "SpO2": 0,
            "体重": 0.0,
            "朝食摂取率": 100,
            "昼食摂取率": 100,
            "夕食摂取率": 100,
            "家族共有メモ": "",
            "気になる変化": "",
        })
    return pd.DataFrame(rows, columns=PHOTO_IMPORT_COLUMNS)


def photo_import_rows_to_health_records(df, input_staff):
    records = []
    now_text = format_now_jst("%Y-%m-%d %H:%M:%S")

    for _, row in df.iterrows():
        if not bool(row.get("取り込む", False)):
            continue

        user_name = clean_text(row.get("利用者名"))
        record_date = pd.to_datetime(row.get("記録日"), errors="coerce")
        if not user_name or pd.isna(record_date):
            continue

        breakfast = safe_int(row.get("朝食摂取率"), 100)
        lunch = safe_int(row.get("昼食摂取率"), 100)
        dinner = safe_int(row.get("夕食摂取率"), 100)

        record = {
            "記録日": record_date.date(),
            "利用者名": user_name,
            "体温": safe_float(row.get("体温"), 0),
            "血圧上": safe_int(row.get("血圧上"), 0),
            "血圧下": safe_int(row.get("血圧下"), 0),
            "脈拍": safe_int(row.get("脈拍"), 0),
            "SpO2": safe_int(row.get("SpO2"), 0),
            "体重": safe_float(row.get("体重"), 0),
            "朝食摂取率": breakfast,
            "昼食摂取率": lunch,
            "夕食摂取率": dinner,
            "朝食摂取区分": meal_option_from_percent(breakfast),
            "昼食摂取区分": meal_option_from_percent(lunch),
            "夕食摂取区分": meal_option_from_percent(dinner),
            "水分摂取量ml": 0,
            "栄養リスク": "0: 通常",
            "口腔状態": "9: 未確認",
            "義歯使用": "9: 未確認",
            "LIFE補助メモ": "写真から半自動入力",
            "家族共有メモ": clean_text(row.get("家族共有メモ")),
            "気になる変化": clean_text(row.get("気になる変化")),
            "登録日時": now_text,
            "入力者": clean_text(input_staff, current_login_user()),
        }
        records.append(record)

    return records


def show_photo_import_menu():
    st.header("写真から半自動入力")
    st.caption("紙の健康チェック表の写真をもとに、候補データを作成します。必ず職員が確認してから保存します。")

    if not active_users:
        st.warning("利用者マスタに表示中の利用者がいません。")
        return

    st.info("手書き帳票は誤読が起きます。ここではAI/OCRを『下書き』として使い、保存前に人が確認する設計です。")

    c1, c2, c3, c4 = st.columns(4)
    with c1:
        target_year = st.number_input("対象年", min_value=2024, max_value=2035, value=today_jst().year, step=1, key="photo_import_year")
    with c2:
        target_month = st.number_input("対象月", min_value=1, max_value=12, value=today_jst().month, step=1, key="photo_import_month")
    with c3:
        default_user = st.selectbox("主な利用者名", active_users, key="photo_import_default_user")
    with c4:
        input_staff = st.text_input("入力者", value=current_login_user(), key="photo_import_staff")

    uploaded_file = st.file_uploader(
        "健康チェック表の写真をアップロード",
        type=["jpg", "jpeg", "png"],
        key="photo_import_uploader",
    )

    if uploaded_file is not None:
        st.image(uploaded_file, caption="アップロード画像", use_container_width=True)

    st.subheader("1. 読み取り・候補作成")
    st.caption("Streamlit Cloudでは通常のOCRが使えないことがあります。その場合はOpenAI画像読み取りで候補表を作ります。")

    with st.expander("OpenAI APIキー設定（Streamlit Secretsに設定済みなら入力不要）", expanded=False):
        openai_api_key_input = st.text_input(
            "OPENAI_API_KEY",
            type="password",
            placeholder="sk-...（この画面入力は保存されません）",
            key="photo_import_openai_api_key",
        )
        st.caption('Streamlit Cloudでは Settings → Secrets に OPENAI_API_KEY = "sk-..." と登録すると毎回入力不要です。')

    if uploaded_file is not None:
        c_ai, c_old = st.columns(2)
        with c_ai:
            if st.button("AIで写真から候補表を作成", use_container_width=True):
                with st.spinner("画像を読み取り中です。読めない値は空欄になります。"):
                    candidate_df, err = try_openai_vision_photo_import(
                        uploaded_file,
                        default_user,
                        target_year,
                        target_month,
                        openai_api_key_input,
                    )
                if err:
                    st.warning(err)
                elif candidate_df.empty:
                    st.warning("AI読み取り候補が作成できませんでした。空の確認表を作成して入力してください。")
                else:
                    st.session_state["photo_import_candidate_df"] = candidate_df
                    st.success(f"AI読み取り候補を{len(candidate_df)}件作成しました。下の確認表で修正してください。")
        with c_old:
            if st.button("旧OCRで文字だけ読む", use_container_width=True):
                ocr_text = try_ocr_image(uploaded_file)
                if ocr_text:
                    st.session_state["photo_import_ocr_text"] = ocr_text
                    st.success("旧OCR読み取りが完了しました。下の欄で確認してください。")
                else:
                    st.session_state["photo_import_ocr_text"] = ""
                    st.warning("この環境では旧OCRを利用できません。AI読み取り、または空の確認表を使ってください。")

    raw_text = st.text_area(
        "OCR結果・手入力テキスト",
        value=st.session_state.get("photo_import_ocr_text", ""),
        height=180,
        placeholder="例：\n1 36.5 128/70 82\n2 36.4 120/68 78\n※読み取れない場合は空の確認表を作成して直接入力できます。",
        key="photo_import_raw_text",
    )

    b1, b2 = st.columns(2)
    with b1:
        if st.button("テキストから候補表を作成", use_container_width=True):
            candidate_df = parse_photo_ocr_text(raw_text, default_user, target_year, target_month, input_staff)
            if candidate_df.empty:
                st.warning("候補データを作成できませんでした。空の確認表を作成して入力してください。")
            st.session_state["photo_import_candidate_df"] = candidate_df
    with b2:
        if st.button("空の確認表を作成", use_container_width=True):
            st.session_state["photo_import_candidate_df"] = make_blank_photo_import_rows(default_user, target_year, target_month)

    st.subheader("2. 職員確認・修正")
    candidate_df = st.session_state.get("photo_import_candidate_df")
    if candidate_df is None or len(candidate_df) == 0:
        st.info("候補表はまだありません。写真を読み取るか、空の確認表を作成してください。")
        return

    edited_df = st.data_editor(
        candidate_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "取り込む": st.column_config.CheckboxColumn("取り込む"),
            "記録日": st.column_config.DateColumn("記録日", format="YYYY-MM-DD"),
            "利用者名": st.column_config.SelectboxColumn("利用者名", options=active_users),
            "体温": st.column_config.NumberColumn("体温", min_value=0.0, max_value=42.0, step=0.1),
            "血圧上": st.column_config.NumberColumn("血圧上", min_value=0, max_value=250, step=1),
            "血圧下": st.column_config.NumberColumn("血圧下", min_value=0, max_value=200, step=1),
            "脈拍": st.column_config.NumberColumn("脈拍", min_value=0, max_value=200, step=1),
            "SpO2": st.column_config.NumberColumn("SpO2", min_value=0, max_value=100, step=1),
            "体重": st.column_config.NumberColumn("体重", min_value=0.0, max_value=200.0, step=0.1),
            "朝食摂取率": st.column_config.NumberColumn("朝食摂取率", min_value=0, max_value=100, step=10),
            "昼食摂取率": st.column_config.NumberColumn("昼食摂取率", min_value=0, max_value=100, step=10),
            "夕食摂取率": st.column_config.NumberColumn("夕食摂取率", min_value=0, max_value=100, step=10),
        },
        key="photo_import_editor",
    )

    st.subheader("3. 確認して保存")
    selected_count = int(pd.Series(edited_df.get("取り込む", [])).fillna(False).astype(bool).sum())
    st.caption(f"保存対象：{selected_count}件")

    if st.button("確認済みデータを健康チェックへ保存", type="primary", use_container_width=True):
        records = photo_import_rows_to_health_records(edited_df, input_staff)
        if not records:
            st.warning("保存対象がありません。取り込む行にチェックを入れてください。")
            return

        saved = 0
        warning_rows = []
        for record in records:
            errors, warnings = validate_health_record(record)
            if errors:
                warning_rows.append({
                    "利用者名": record["利用者名"],
                    "記録日": record["記録日"],
                    "内容": "／".join(errors),
                })
                continue
            upsert_health_record(record)
            saved += 1
            if warnings:
                warning_rows.append({
                    "利用者名": record["利用者名"],
                    "記録日": record["記録日"],
                    "内容": "／".join(warnings),
                })

        st.success(f"{saved}件を健康チェックへ保存しました。")
        if warning_rows:
            st.warning("確認が必要な行があります。")
            st.dataframe(pd.DataFrame(warning_rows), use_container_width=True, hide_index=True)

        st.session_state["photo_import_candidate_df"] = edited_df



# =========================
# LIFE入力標準化・ADL評価データ
# =========================

def ensure_life_adl_file():
    """
    LIFE ADL評価をSQLiteで管理する。
    旧Excelがある場合のみ初回移行し、以後はSQLiteを正とする。
    """
    ensure_dirs()
    if sqlite_table_row_count(SQLITE_TABLE_LIFE_ADL) > 0:
        return

    df = pd.DataFrame(columns=LIFE_ADL_COLUMNS)
    if LIFE_ADL_FILE.exists():
        try:
            df = pd.read_excel(LIFE_ADL_FILE, sheet_name="ADL評価")
        except Exception:
            try:
                df = pd.read_excel(LIFE_ADL_FILE)
            except Exception:
                df = pd.DataFrame(columns=LIFE_ADL_COLUMNS)

    df = normalize_df_columns(df, LIFE_ADL_COLUMNS)
    df = attach_user_ids(df)
    if not df.empty:
        df["評価日"] = pd.to_datetime(df["評価日"], errors="coerce")
        df["対象月"] = df["対象月"].astype(str)
        df["利用者名"] = df["利用者名"].astype(str).str.strip()
        if "評価ID" in df.columns:
            df["評価ID"] = df["評価ID"].astype(str)
            missing = df["評価ID"].astype(str).str.strip() == ""
            df.loc[missing, "評価ID"] = [str(uuid.uuid4()) for _ in range(int(missing.sum()))]
        df = df.drop_duplicates(subset=["対象月", "利用者名"], keep="last")

    save_sqlite_table(
        df,
        SQLITE_TABLE_LIFE_ADL,
        LIFE_ADL_COLUMNS,
        date_cols=["評価日"],
        unique_cols=["対象月", "利用者名"],
    )


def load_life_adl_data():
    """LIFE ADL評価をSQLiteから読み込む。"""
    ensure_life_adl_file()
    df = load_sqlite_table(SQLITE_TABLE_LIFE_ADL, LIFE_ADL_COLUMNS, date_cols=["評価日"])
    df = attach_user_ids(df)

    if not df.empty:
        df["評価日"] = pd.to_datetime(df["評価日"], errors="coerce")
        for col in LIFE_ADL_COLUMNS:
            if col != "評価日":
                df[col] = df[col].fillna("").astype(str)
    return df.astype("object")


def save_life_adl_data(df):
    """LIFE ADL評価をSQLiteへ保存する。Excelには保存しない。"""
    ensure_dirs()
    df = normalize_df_columns(df, LIFE_ADL_COLUMNS)
    df = attach_user_ids(df)
    if not df.empty:
        df["評価日"] = pd.to_datetime(df["評価日"], errors="coerce")
        df["対象月"] = df["対象月"].astype(str)
        df["利用者名"] = df["利用者名"].astype(str).str.strip()
        if "評価ID" in df.columns:
            df["評価ID"] = df["評価ID"].astype(str)
            missing = df["評価ID"].astype(str).str.strip() == ""
            df.loc[missing, "評価ID"] = [str(uuid.uuid4()) for _ in range(int(missing.sum()))]
        df = df.drop_duplicates(subset=["対象月", "利用者名"], keep="last")

    save_sqlite_table(
        df,
        SQLITE_TABLE_LIFE_ADL,
        LIFE_ADL_COLUMNS,
        date_cols=["評価日"],
        unique_cols=["対象月", "利用者名"],
    )

def upsert_life_adl_record(record):
    df = load_life_adl_data()
    target_month = clean_text(record.get("対象月"))
    user_name = clean_text(record.get("利用者名"))
    if df.empty:
        idx = None
    else:
        mask = (df["対象月"].astype(str) == target_month) & (df["利用者名"].astype(str) == user_name)
        matches = df.index[mask].tolist()
        idx = matches[0] if matches else None

    if idx is None:
        df = pd.concat([df, pd.DataFrame([record], columns=LIFE_ADL_COLUMNS)], ignore_index=True)
        action = "登録"
    else:
        for col in LIFE_ADL_COLUMNS:
            df.at[idx, col] = record.get(col, "")
        action = "更新"
    save_life_adl_data(df)
    return action


def build_life_month_summary(user_name, target_year, target_month):
    month_start = date(int(target_year), int(target_month), 1)
    month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    health_df = get_month_health_data(load_health_data(start_date=month_start, end_date=month_end), user_name, target_year, target_month)
    ex_df = get_month_excretion_data(load_excretion_data(start_date=month_start, end_date=month_end), user_name, target_year, target_month)

    result = {
        "利用者名": user_name,
        "対象月": f"{int(target_year):04d}-{int(target_month):02d}",
        "健康記録日数": len(health_df),
        "平均体温": "",
        "平均SpO2": "",
        "平均体重": "",
        "平均食事摂取率": "",
        "排尿回数": 0,
        "排便回数": 0,
        "水様便・下痢便回数": 0,
        "気になる変化件数": 0,
        "不足項目": "",
    }

    missing = []
    if health_df.empty:
        missing.append("健康チェック未入力")
    else:
        for col, label in [("体温", "体温"), ("SpO2", "SpO2"), ("体重", "体重")]:
            vals = to_number(health_df[col]) if col in health_df.columns else pd.Series(dtype=float)
            vals = vals[vals > 0]
            if vals.empty:
                missing.append(f"{label}未入力")
            else:
                result[f"平均{label}"] = round(float(vals.mean()), 1)

        meal_cols = ["朝食摂取率", "昼食摂取率", "夕食摂取率"]
        meal_vals = []
        for col in meal_cols:
            if col in health_df.columns:
                meal_vals.extend(to_number(health_df[col]).dropna().tolist())
        if meal_vals:
            result["平均食事摂取率"] = round(float(pd.Series(meal_vals).mean()), 1)
        else:
            missing.append("食事摂取率未入力")

        if "気になる変化" in health_df.columns:
            result["気になる変化件数"] = int((health_df["気になる変化"].fillna("").astype(str).str.strip() != "").sum())

    if ex_df.empty:
        missing.append("排泄チェック未入力")
    else:
        ex_sum = summarize_excretion(ex_df)
        result["排尿回数"] = ex_sum["排尿回数"]
        result["排便回数"] = ex_sum["排便回数"]
        result["水様便・下痢便回数"] = ex_sum["水様便"] + ex_sum["下痢便"]

    adl_df = load_life_adl_data()
    target_ym = f"{int(target_year):04d}-{int(target_month):02d}"
    if adl_df.empty or adl_df[(adl_df["利用者名"] == user_name) & (adl_df["対象月"] == target_ym)].empty:
        missing.append("ADL月次評価未入力")

    result["不足項目"] = "、".join(missing) if missing else "不足なし"
    return result


def show_life_standardization_menu():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return
    st.header("LIFE入力標準化")
    st.caption("日々の記録をLIFE提出補助へつなげるため、自由入力ではなく選択式・コード化した項目を増やします。")

    if not active_users:
        st.warning("利用者マスタに表示中の利用者がいません。")
        return

    tab1, tab2, tab3 = st.tabs(["ADL月次評価", "月次不足チェック", "入力基準表"])

    with tab1:
        st.subheader("ADL月次評価")
        st.info("ADLは毎日ではなく、月1回または状態変化時に評価する想定です。")
        adl_df = load_life_adl_data()
        c1, c2, c3 = st.columns(3)
        with c1:
            eval_date = st.date_input("評価日", value=today_jst(), key="life_adl_eval_date")
        with c2:
            user_name = st.selectbox("利用者名", active_users, key="life_adl_user")
        with c3:
            input_staff = st.text_input("入力者", placeholder="例：藤野", key="life_adl_staff")

        target_month = eval_date.strftime("%Y-%m")
        existing = adl_df[(adl_df["対象月"].astype(str) == target_month) & (adl_df["利用者名"].astype(str) == user_name)]
        row = existing.iloc[-1] if not existing.empty else None

        def adl_default(col):
            if row is None:
                return 4  # 9: 未確認
            return get_life_option_index(ADL_LEVEL_OPTIONS, row.get(col, "9: 未確認"), 4)

        with st.form("life_adl_form", clear_on_submit=False):
            a1, a2, a3 = st.columns(3)
            with a1:
                walk = st.selectbox("歩行", ADL_LEVEL_OPTIONS, index=adl_default("歩行"))
            with a2:
                transfer = st.selectbox("移乗", ADL_LEVEL_OPTIONS, index=adl_default("移乗"))
            with a3:
                meal_adl = st.selectbox("食事動作", ADL_LEVEL_OPTIONS, index=adl_default("食事動作"))

            a4, a5, a6 = st.columns(3)
            with a4:
                toilet_adl = st.selectbox("排泄動作", ADL_LEVEL_OPTIONS, index=adl_default("排泄動作"))
            with a5:
                dressing = st.selectbox("更衣", ADL_LEVEL_OPTIONS, index=adl_default("更衣"))
            with a6:
                cognitive = st.selectbox("認知・行動", COGNITIVE_OPTIONS, index=get_life_option_index(COGNITIVE_OPTIONS, row.get("認知・行動", "9: 未確認") if row is not None else "9: 未確認", 4))

            memo = st.text_area("評価メモ", value=clean_text(row.get("評価メモ", "")) if row is not None else "", placeholder="状態変化や判断理由を記録")
            submitted = st.form_submit_button("ADL評価を保存", use_container_width=True)

        if submitted:
            record = {
                "評価ID": clean_text(row.get("評価ID", "")) if row is not None else str(uuid.uuid4()),
                "評価日": eval_date,
                "対象月": target_month,
                "利用者名": user_name,
                "歩行": walk,
                "移乗": transfer,
                "食事動作": meal_adl,
                "排泄動作": toilet_adl,
                "更衣": dressing,
                "認知・行動": cognitive,
                "評価メモ": clean_text(memo),
                "入力者": clean_text(input_staff),
                "登録日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
            }
            action = upsert_life_adl_record(record)
            st.success(f"ADL月次評価を{action}しました。")
            st.rerun()

        st.subheader("ADL評価一覧")
        if adl_df.empty:
            st.info("ADL評価はまだ登録されていません。")
        else:
            st.dataframe(adl_df.drop(columns=["評価ID"], errors="ignore"), use_container_width=True, hide_index=True)

    with tab2:
        st.subheader("月次不足チェック")
        today = today_jst()
        c1, c2 = st.columns(2)
        with c1:
            target_year = st.number_input("年", min_value=2024, max_value=2035, value=today.year, step=1)
        with c2:
            target_month = st.number_input("月", min_value=1, max_value=12, value=today.month, step=1)

        if st.button("月次チェックを表示", use_container_width=True):
            rows = [build_life_month_summary(user, target_year, target_month) for user in active_users]
            summary_df = pd.DataFrame(rows)
            st.dataframe(summary_df, use_container_width=True, hide_index=True)
            output = BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                summary_df.to_excel(writer, index=False, sheet_name="LIFE月次不足チェック")
            st.download_button(
                "月次不足チェックをExcelでダウンロード",
                data=output.getvalue(),
                file_name=f"LIFE月次不足チェック_{int(target_year)}-{int(target_month):02d}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    with tab3:
        st.subheader("入力基準表")
        rows = []
        for label, opts in [
            ("食事摂取区分", MEAL_INTAKE_OPTIONS),
            ("栄養リスク", NUTRITION_RISK_OPTIONS),
            ("口腔状態", ORAL_STATUS_OPTIONS),
            ("義歯使用", DENTURE_OPTIONS),
            ("ADL", ADL_LEVEL_OPTIONS),
            ("認知・行動", COGNITIVE_OPTIONS),
            ("尿量", [f"{v}: {k}" for k, v in URINE_AMOUNT_CODE.items()]),
            ("尿性状", [f"{v}: {k}" for k, v in URINE_TYPE_CODE.items()]),
            ("便量", [f"{v}: {k}" for k, v in STOOL_AMOUNT_CODE.items()]),
            ("便性状", [f"{v}: {k}" for k, v in STOOL_TYPE_CODE.items()]),
        ]:
            for opt in opts:
                rows.append({"項目": label, "入力基準": opt})
        st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)



# =========================
# 加算シミュレーション
# =========================
def get_yamato_gh_addon_candidates(resident_count=9, days_per_month=30):
    """神奈川県大和市・認知症対応型共同生活介護（GH）向けの加算候補マスタ。
    すべての加算を網羅するものではなく、9名GHで検討頻度が高い候補を画面から追加できるようにするための初期マスタ。
    単位数・要件は制度改定や施設状況で変わるため、請求前に必ず最新資料・請求ソフトで確認する。
    """
    rc = int(resident_count) if resident_count else 9
    dm = int(days_per_month) if days_per_month else 30
    return pd.DataFrame([
        {
            "追加": False,
            "サービス": "GH",
            "地域": "神奈川県大和市",
            "地域区分": "5級地",
            "1単位単価": 10.80,
            "加算カテゴリ": "LIFE・記録",
            "加算名": "科学的介護推進体制加算（LIFE）",
            "単位": 40,
            "算定単位": "人/月",
            "対象人数": rc,
            "月回数": 1,
            "取得": False,
            "確認ポイント": "LIFE提出、フィードバック活用、記録・計画への反映。",
            "メモ": "9名満床なら 40単位×9名×10.80円＝月3,888円。",
        },
        {
            "追加": False,
            "サービス": "GH",
            "地域": "神奈川県大和市",
            "地域区分": "5級地",
            "1単位単価": 10.80,
            "加算カテゴリ": "職員体制",
            "加算名": "サービス提供体制強化加算Ⅰ（候補）",
            "単位": 22,
            "算定単位": "人/日",
            "対象人数": rc,
            "月回数": dm,
            "取得": False,
            "確認ポイント": "介護福祉士割合、勤続年数、常勤割合等の体制要件。",
            "メモ": "候補値。Ⅰ・Ⅱ・Ⅲのどれに該当するか確認が必要。",
        },
        {
            "追加": False,
            "サービス": "GH",
            "地域": "神奈川県大和市",
            "地域区分": "5級地",
            "1単位単価": 10.80,
            "加算カテゴリ": "医療連携",
            "加算名": "医療連携体制加算Ⅰ（候補）",
            "単位": 39,
            "算定単位": "人/日",
            "対象人数": rc,
            "月回数": dm,
            "取得": False,
            "確認ポイント": "看護師配置・連携体制、重度化対応、健康管理体制。",
            "メモ": "施設の届出区分により単位が変わる可能性あり。",
        },
        {
            "追加": False,
            "サービス": "GH",
            "地域": "神奈川県大和市",
            "地域区分": "5級地",
            "1単位単価": 10.80,
            "加算カテゴリ": "医療連携",
            "加算名": "協力医療機関連携加算（候補）",
            "単位": 100,
            "算定単位": "事業所/月",
            "対象人数": 1,
            "月回数": 1,
            "取得": False,
            "確認ポイント": "協力医療機関との実効性ある連携、会議・情報共有、届出。",
            "メモ": "令和6年度改定で重要度が上がった候補。区分・経過措置の確認が必要。",
        },
        {
            "追加": False,
            "サービス": "GH",
            "地域": "神奈川県大和市",
            "地域区分": "5級地",
            "1単位単価": 10.80,
            "加算カテゴリ": "認知症ケア",
            "加算名": "認知症チームケア推進加算（候補）",
            "単位": 150,
            "算定単位": "人/月",
            "対象人数": rc,
            "月回数": 1,
            "取得": False,
            "確認ポイント": "BPSD等の評価、チームケア体制、計画・会議・記録。",
            "メモ": "Ⅰ・Ⅱ等の区分や対象者要件の確認が必要。",
        },
        {
            "追加": False,
            "サービス": "GH",
            "地域": "神奈川県大和市",
            "地域区分": "5級地",
            "1単位単価": 10.80,
            "加算カテゴリ": "夜間体制",
            "加算名": "夜間支援体制加算（候補）",
            "単位": 50,
            "算定単位": "人/日",
            "対象人数": rc,
            "月回数": dm,
            "取得": False,
            "確認ポイント": "夜勤・宿直等の配置、ユニット数、届出区分。",
            "メモ": "区分により単位差あり。施設の夜間配置で確認。",
        },
        {
            "追加": False,
            "サービス": "GH",
            "地域": "神奈川県大和市",
            "地域区分": "5級地",
            "1単位単価": 10.80,
            "加算カテゴリ": "口腔・栄養",
            "加算名": "口腔・栄養スクリーニング加算（候補）",
            "単位": 20,
            "算定単位": "人/6か月",
            "対象人数": rc,
            "月回数": 1,
            "取得": False,
            "確認ポイント": "6か月ごとのスクリーニング、結果共有、記録保存。",
            "メモ": "算定月だけ発生する加算として扱う。",
        },
        {
            "追加": False,
            "サービス": "GH",
            "地域": "神奈川県大和市",
            "地域区分": "5級地",
            "1単位単価": 10.80,
            "加算カテゴリ": "入退居支援",
            "加算名": "退居時情報提供加算（候補）",
            "単位": 250,
            "算定単位": "人/回",
            "対象人数": 0,
            "月回数": 1,
            "取得": False,
            "確認ポイント": "医療機関等への情報提供、退居時の記録・様式。",
            "メモ": "該当者がいる月だけ人数を入力。",
        },
        {
            "追加": False,
            "サービス": "GH",
            "地域": "神奈川県大和市",
            "地域区分": "5級地",
            "1単位単価": 10.80,
            "加算カテゴリ": "個別要件",
            "加算名": "若年性認知症利用者受入加算（該当者のみ）",
            "単位": 120,
            "算定単位": "人/日",
            "対象人数": 0,
            "月回数": dm,
            "取得": False,
            "確認ポイント": "若年性認知症の該当者、個別担当者・支援内容の確認。",
            "メモ": "該当利用者がいる場合のみ対象人数を入力。",
        },
        {
            "追加": False,
            "サービス": "GH",
            "地域": "神奈川県大和市",
            "地域区分": "5級地",
            "1単位単価": 10.80,
            "加算カテゴリ": "自由入力",
            "加算名": "独自入力欄",
            "単位": 0,
            "算定単位": "人/月",
            "対象人数": rc,
            "月回数": 1,
            "取得": False,
            "確認ポイント": "請求ソフト・指定権者資料で確認した加算を入力。",
            "メモ": "マスタにない加算を追加するための行。",
        },
    ])



def should_sync_resident_count(row):
    """基本設定の利用者数を反映する対象かを判定する。
    個別該当者だけの加算や退居時などは、0人のまま手入力できるように残す。
    """
    unit = str(row.get("算定単位", ""))
    name = str(row.get("加算名", ""))
    category = str(row.get("加算カテゴリ", ""))

    if not unit.startswith("人/"):
        return False

    # 該当者だけ入力する加算は自動上書きしない
    exclusion_words = ["該当者", "若年性", "退居時", "退居", "個別要件"]
    if any(w in name for w in exclusion_words) or any(w in category for w in ["個別要件", "入退居支援"]):
        return False

    return True


def sync_addon_basic_settings(df, resident_count, days_per_month, unit_price, region_name, region_class):
    """基本設定（人数・日数・単価・地域）を加算表へ反映する。
    st.session_state に古い人数が残っていても、通常の人/月・人/日加算は現在の利用者数へ合わせる。
    """
    work = df.copy()
    if work.empty:
        return work

    if "対象人数" not in work.columns:
        work["対象人数"] = 0
    if "月回数" not in work.columns:
        work["月回数"] = 1
    if "1単位単価" not in work.columns:
        work["1単位単価"] = unit_price
    if "地域" not in work.columns:
        work["地域"] = region_name
    if "地域区分" not in work.columns:
        work["地域区分"] = region_class

    work["対象人数"] = work.apply(
        lambda r: int(resident_count) if should_sync_resident_count(r) else int(safe_int(r.get("対象人数"), 0)),
        axis=1,
    )
    work["月回数"] = work.apply(
        lambda r: int(days_per_month) if str(r.get("算定単位", "")).endswith("/日") else int(safe_int(r.get("月回数"), 1)),
        axis=1,
    )
    work["1単位単価"] = float(unit_price)
    work["地域"] = region_name
    work["地域区分"] = region_class or "5級地"
    return work

def get_gh_addon_master_default(resident_count=9, days_per_month=30):
    """初期表示用マスタ。大和市候補のうち、基本的な行だけを表示する。"""
    candidates = get_yamato_gh_addon_candidates(resident_count, days_per_month)
    base_names = [
        "科学的介護推進体制加算（LIFE）",
        "サービス提供体制強化加算Ⅰ（候補）",
        "医療連携体制加算Ⅰ（候補）",
        "口腔・栄養スクリーニング加算（候補）",
        "若年性認知症利用者受入加算（該当者のみ）",
        "独自入力欄",
    ]
    df = candidates[candidates["加算名"].isin(base_names)].copy()
    df = df.drop(columns=["追加"], errors="ignore")
    return df.reset_index(drop=True)


def calc_addon_simulation(df, unit_price):
    work = df.copy()
    for col in ["単位", "対象人数", "月回数"]:
        work[col] = pd.to_numeric(work[col], errors="coerce").fillna(0)
    if "取得" not in work.columns:
        work["取得"] = False
    if "1単位単価" not in work.columns:
        work["1単位単価"] = float(unit_price)
    work["1単位単価"] = pd.to_numeric(work["1単位単価"], errors="coerce").fillna(float(unit_price))
    work["月間単位"] = work.apply(
        lambda r: int(r["単位"] * r["対象人数"] * r["月回数"]) if bool(r.get("取得", False)) else 0,
        axis=1,
    )
    work["月額目安"] = (work["月間単位"] * work["1単位単価"]).round(0).astype(int)
    work["年間目安"] = (work["月額目安"] * 12).astype(int)
    return work


def show_addon_simulation_menu():
    st.header("加算シミュレーション")
    st.caption("神奈川県大和市のグループホームを前提に、取得候補の加算をマスタから追加して概算します。")

    st.info(
        "この画面は概算用です。大和市は5級地として1単位10.80円を初期値にしています。"
        "ただし、単位数・要件・届出区分は改定や施設状況で変わります。請求前には必ず最新資料・指定権者・請求ソフトで確認してください。"
    )

    with st.expander("基本設定", expanded=True):
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            facility_type = st.selectbox(
                "サービス種別",
                ["認知症対応型共同生活介護（グループホーム）", "通所介護", "小規模多機能", "その他"],
                index=0,
            )
        with c2:
            resident_count = st.number_input("利用者数", min_value=0, max_value=99, value=len(active_users) if active_users else 9, step=1)
        with c3:
            days_per_month = st.number_input("月の日数", min_value=1, max_value=31, value=30, step=1)
        with c4:
            unit_price = st.number_input("1単位単価（円）", min_value=1.0, max_value=20.0, value=10.80, step=0.01)

        c5, c6 = st.columns(2)
        with c5:
            region_name = st.selectbox("地域", ["神奈川県大和市", "神奈川県綾瀬市", "神奈川県藤沢市", "神奈川県横浜市", "その他"], index=0)
        with c6:
            region_class = st.text_input("地域区分", value="5級地" if region_name == "神奈川県大和市" else "")

    if "addon_editor_base" not in st.session_state:
        st.session_state["addon_editor_base"] = get_gh_addon_master_default(resident_count, days_per_month)

    st.subheader("1. 大和市で検討できる加算候補を追加")
    st.caption("追加したい加算にチェックを入れて、下のボタンを押すとシミュレーション表へ行が追加されます。")

    candidate_df = get_yamato_gh_addon_candidates(resident_count, days_per_month)
    candidate_df["1単位単価"] = unit_price
    candidate_df["地域"] = region_name
    candidate_df["地域区分"] = region_class or candidate_df["地域区分"]

    candidate_edited = st.data_editor(
        candidate_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "追加": st.column_config.CheckboxColumn("追加", default=False),
            "取得": st.column_config.CheckboxColumn("取得", default=False),
            "単位": st.column_config.NumberColumn("単位", min_value=0, step=1),
            "対象人数": st.column_config.NumberColumn("対象人数", min_value=0, step=1),
            "月回数": st.column_config.NumberColumn("月回数", min_value=0, step=1),
            "1単位単価": st.column_config.NumberColumn("1単位単価", min_value=1.0, max_value=20.0, step=0.01),
            "算定単位": st.column_config.SelectboxColumn("算定単位", options=["人/月", "人/日", "人/回", "人/6か月", "事業所/月", "その他"]),
        },
        key="yamato_addon_candidate_editor",
    )

    b1, b2 = st.columns(2)
    with b1:
        if st.button("チェックした加算をシミュレーション表へ追加", use_container_width=True):
            add_df = candidate_edited[candidate_edited.get("追加", False) == True].copy()
            if add_df.empty:
                st.warning("追加する加算にチェックを入れてください。")
            else:
                add_df = add_df.drop(columns=["追加"], errors="ignore")
                add_df = sync_addon_basic_settings(add_df, resident_count, days_per_month, unit_price, region_name, region_class)
                base_df = st.session_state.get("addon_editor_base", pd.DataFrame())
                base_df = sync_addon_basic_settings(base_df, resident_count, days_per_month, unit_price, region_name, region_class)
                merged = pd.concat([base_df, add_df], ignore_index=True)
                if "加算名" in merged.columns:
                    merged = merged.drop_duplicates(subset=["加算名"], keep="last")
                st.session_state["addon_editor_base"] = merged.reset_index(drop=True)
                st.success(f"{len(add_df)}件の加算候補を追加しました。")
                st.rerun()
    with b2:
        if st.button("大和市GH候補マスタに戻す", use_container_width=True):
            st.session_state["addon_editor_base"] = get_gh_addon_master_default(resident_count, days_per_month)
            st.success("初期候補マスタに戻しました。")
            st.rerun()

    st.subheader("2. 加算マスタ入力・編集")
    st.caption("単位数・人数・日数・取得チェックはここで編集できます。")

    default_df = st.session_state.get("addon_editor_base", get_gh_addon_master_default(resident_count, days_per_month)).copy()
    default_df = sync_addon_basic_settings(default_df, resident_count, days_per_month, unit_price, region_name, region_class)

    edited_df = st.data_editor(
        default_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "取得": st.column_config.CheckboxColumn("取得", default=False),
            "単位": st.column_config.NumberColumn("単位", min_value=0, step=1),
            "対象人数": st.column_config.NumberColumn("対象人数", min_value=0, step=1),
            "月回数": st.column_config.NumberColumn("月回数", min_value=0, step=1),
            "1単位単価": st.column_config.NumberColumn("1単位単価", min_value=1.0, max_value=20.0, step=0.01),
            "算定単位": st.column_config.SelectboxColumn("算定単位", options=["人/月", "人/日", "人/回", "人/6か月", "事業所/月", "その他"]),
        },
        key="addon_simulation_editor",
    )

    st.session_state["addon_editor_base"] = edited_df.copy()

    result_df = calc_addon_simulation(edited_df, unit_price)

    st.subheader("3. 概算結果")
    st.caption("ここでも取得チェックを直接変更できます。チェック後、月間単位・月額目安・年間目安を再計算します。")

    display_cols = [
        "取得", "地域", "地域区分", "加算カテゴリ", "加算名", "単位", "算定単位", "対象人数", "月回数", "1単位単価",
        "月間単位", "月額目安", "年間目安", "確認ポイント", "メモ"
    ]
    for col in display_cols:
        if col not in result_df.columns:
            result_df[col] = ""

    # 概算結果でもチェック操作できるように st.data_editor を使う
    result_input_df = result_df[display_cols].copy()
    if "取得" in result_input_df.columns:
        result_input_df["取得"] = result_input_df["取得"].fillna(False).astype(bool)

    result_edited_df = st.data_editor(
        result_input_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "取得": st.column_config.CheckboxColumn("取得", default=False),
            "単位": st.column_config.NumberColumn("単位", min_value=0, step=1),
            "対象人数": st.column_config.NumberColumn("対象人数", min_value=0, step=1),
            "月回数": st.column_config.NumberColumn("月回数", min_value=0, step=1),
            "1単位単価": st.column_config.NumberColumn("1単位単価", min_value=1.0, max_value=20.0, step=0.01),
            "算定単位": st.column_config.SelectboxColumn("算定単位", options=["人/月", "人/日", "人/回", "人/6か月", "事業所/月", "その他"]),
            "月間単位": st.column_config.NumberColumn("月間単位", disabled=True),
            "月額目安": st.column_config.NumberColumn("月額目安", disabled=True),
            "年間目安": st.column_config.NumberColumn("年間目安", disabled=True),
        },
        key="addon_result_checkbox_editor",
    )

    # 結果欄で変更した内容をもとに再計算し、次回表示にも反映する
    result_df = calc_addon_simulation(result_edited_df, unit_price)
    st.session_state["addon_editor_base"] = result_df.drop(columns=["月間単位", "月額目安", "年間目安"], errors="ignore").copy()

    total_units = int(result_df["月間単位"].sum()) if not result_df.empty else 0
    total_month_yen = int(result_df["月額目安"].sum()) if not result_df.empty else 0
    total_year_yen = int(result_df["年間目安"].sum()) if not result_df.empty else 0

    m1, m2, m3, m4 = st.columns(4)
    m1.metric("月間単位", f"{total_units:,} 単位")
    m2.metric("月額目安", f"{total_month_yen:,} 円")
    m3.metric("年間目安", f"{total_year_yen:,} 円")
    per_user = int(total_month_yen / resident_count) if resident_count else 0
    m4.metric("1人あたり月額", f"{per_user:,} 円")

    with st.expander("再計算後の明細を確認", expanded=False):
        st.dataframe(result_df[display_cols], use_container_width=True, hide_index=True)

    st.subheader("4. 取り漏れ確認メモ")
    st.markdown(
        """
- LIFE系：科学的介護推進体制加算、口腔・栄養関連、ADL・モニタリング系の記録連携を確認
- 医療連携系：医療連携体制加算、協力医療機関連携加算、退居時情報提供加算を確認
- 体制系：サービス提供体制強化加算、夜間支援体制、認知症チームケア推進加算を確認
- 書類系：計画書、同意、モニタリング、会議録、LIFE提出履歴、職員配置根拠を確認
- 注意：加算は「取れるか」より「継続して根拠を残せるか」が重要
        """
    )

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame([{
            "サービス種別": facility_type,
            "地域": region_name,
            "地域区分": region_class,
            "利用者数": resident_count,
            "月の日数": days_per_month,
            "1単位単価": unit_price,
            "月間単位合計": total_units,
            "月額目安合計": total_month_yen,
            "年間目安合計": total_year_yen,
            "注意": "概算用。請求前に最新の介護報酬・指定権者資料・請求ソフトで確認。",
        }]).to_excel(writer, index=False, sheet_name="概算サマリー")
        result_df.to_excel(writer, index=False, sheet_name="加算シミュレーション")
        candidate_edited.to_excel(writer, index=False, sheet_name="大和市GH候補マスタ")

    st.download_button(
        "加算シミュレーション結果をExcelでダウンロード",
        data=output.getvalue(),
        file_name=f"大和市_GH_加算シミュレーション_{format_now_jst('%Y%m%d_%H%M')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )




# =========================
# 業務全体申し送りデータ
# =========================
def ensure_business_handover_file():
    """互換用。実データはSQLite（hidamari_health.db / handover_logs）へ保存します。"""
    migrate_excel_to_sqlite_if_needed(
        SQLITE_TABLE_HANDOVER,
        HANDOVER_FILE,
        "業務全体申し送り",
        BUSINESS_HANDOVER_COLUMNS,
        date_cols=["日付"],
        unique_cols=["記録ID"],
    )


def load_business_handover_data(start_date=None, end_date=None, recent_days=None):
    """業務全体申し送りを読み込む。期間指定時はSupabase側で絞って高速化する。"""
    with perf_timer("load_business_handover_data", f"{start_date or ''}-{end_date or ''} recent={recent_days or ''}"):
        ensure_business_handover_file()
        if recent_days and start_date is None and end_date is None:
            start_date = recent_start_date(recent_days)
            end_date = today_jst()
        if supabase_is_enabled():
            df = supabase_read_table(SQLITE_TABLE_HANDOVER, BUSINESS_HANDOVER_COLUMNS, date_field="日付", start_date=start_date, end_date=end_date)
        else:
            df = load_sqlite_table(SQLITE_TABLE_HANDOVER, BUSINESS_HANDOVER_COLUMNS, date_cols=["日付"])
            df = _filter_df_by_date_range(df, "日付", start_date, end_date)

        if not df.empty:
            df["日付"] = pd.to_datetime(df["日付"], errors="coerce")
            for col in ["記録ID", "勤務帯", "記入者", "対象区分", "user_id", "利用者名", "全体申し送り", "要確認事項", "優先度", "対応状況", "写真1", "写真2", "Excel自動抽出情報", "入力Excelファイル", "入力Excel表示情報", "記録日時"]:
                if col not in df.columns:
                    df[col] = ""
                df[col] = df[col].fillna("").astype(str)

        return df.astype("object")


def save_business_handover_data(df):
    """業務全体申し送りをSQLiteへ保存する。"""
    ensure_dirs()
    df = normalize_df_columns(df, BUSINESS_HANDOVER_COLUMNS)

    if not df.empty:
        df["日付"] = pd.to_datetime(df["日付"], errors="coerce")
        df["_sort_dt"] = pd.to_datetime(df["記録日時"], errors="coerce")
        df = df.sort_values(["日付", "_sort_dt"], ascending=[False, False]).drop(columns=["_sort_dt"])

    save_sqlite_table(
        df,
        SQLITE_TABLE_HANDOVER,
        BUSINESS_HANDOVER_COLUMNS,
        date_cols=["日付"],
        unique_cols=["記録ID"],
        sort_cols=["記録日時"],
    )
    clear_hidamari_read_cache("申し送り保存")
    add_audit_log("保存", SQLITE_TABLE_HANDOVER, "", "業務全体申し送りを保存しました")


def make_business_handover_id(record_date, shift_type, staff_name):
    d = pd.to_datetime(record_date, errors="coerce")
    date_text = d.strftime("%Y%m%d") if not pd.isna(d) else format_now_jst("%Y%m%d")
    staff_text = clean_text(staff_name, "未入力").replace(" ", "").replace("　", "")
    shift_text = clean_text(shift_type, "勤務")
    now_text = format_now_jst("%H%M%S")
    return f"BH-{date_text}-{shift_text}-{staff_text}-{now_text}"


def get_business_handover_by_date(df, target_date):
    if df.empty:
        return pd.DataFrame(columns=BUSINESS_HANDOVER_COLUMNS)

    work = df.copy()
    work["日付"] = pd.to_datetime(work["日付"], errors="coerce")
    target = pd.to_datetime(target_date, errors="coerce")

    if pd.isna(target):
        return pd.DataFrame(columns=BUSINESS_HANDOVER_COLUMNS)

    work = work[work["日付"].dt.date == target.date()].copy()
    if not work.empty:
        work["_sort_dt"] = pd.to_datetime(work["記録日時"], errors="coerce")
        work = work.sort_values("_sort_dt", ascending=False).drop(columns=["_sort_dt"])

    return work



def get_business_handover_in_progress(df, exclude_today=False):
    """対応状況が「対応中」の申し送りを一覧化する。"""
    if df.empty:
        return pd.DataFrame(columns=BUSINESS_HANDOVER_COLUMNS)

    work = df.copy()
    work["日付"] = pd.to_datetime(work["日付"], errors="coerce")
    work["対応状況"] = work["対応状況"].fillna("").astype(str).str.strip()
    work = work[work["対応状況"] == "対応中"].copy()

    if exclude_today:
        today_value = today_jst()
        work = work[work["日付"].dt.date != today_value].copy()

    if not work.empty:
        work["_sort_date"] = pd.to_datetime(work["日付"], errors="coerce")
        work["_sort_dt"] = pd.to_datetime(work["記録日時"], errors="coerce")
        work = work.sort_values(["_sort_date", "_sort_dt"], ascending=[False, False]).drop(columns=["_sort_date", "_sort_dt"])

    return work


def show_business_handover_in_progress_section(df):
    """本日の申し送りの下に、未完了の対応中案件を表示する。"""
    st.subheader("対応中案件")
    in_progress_df = get_business_handover_in_progress(df, exclude_today=False)

    if in_progress_df.empty:
        st.info("現在、対応中の業務全体申し送りはありません。")
        return

    st.warning(f"対応中の案件が {len(in_progress_df)} 件あります。対応状況を確認してください。")
    for _, row in in_progress_df.iterrows():
        render_business_handover_card(row)


def get_business_handover_alerts(df):
    if df.empty:
        return pd.DataFrame(columns=BUSINESS_HANDOVER_COLUMNS)

    work = df.copy()
    work["日付"] = pd.to_datetime(work["日付"], errors="coerce")
    alert_df = work[
        (work["対応状況"].isin(["未対応", "対応中"]))
        | (work["優先度"] == "至急")
    ].copy()

    if not alert_df.empty:
        alert_df["_sort_dt"] = pd.to_datetime(alert_df["記録日時"], errors="coerce")
        alert_df = alert_df.sort_values(["日付", "_sort_dt"], ascending=[False, False]).drop(columns=["_sort_dt"])

    return alert_df



# =========================
# 申し送り写真 軽量化・保存期間管理（Ver4.1）
# =========================
PHOTO_MAX_BYTES = 300 * 1024  # 商品版標準：DB/Supabase容量保護のため圧縮後300KB以下
PHOTO_MAX_DISPLAY_KB = 300
PHOTO_MAX_WIDTH = 800
PHOTO_JPEG_MIN_QUALITY = 35
PHOTO_RETENTION_DAYS = 180


def format_file_size(size_bytes) -> str:
    try:
        size_bytes = int(size_bytes or 0)
    except Exception:
        size_bytes = 0
    if size_bytes >= 1024 * 1024:
        return f"{size_bytes / (1024 * 1024):.2f}MB"
    return f"{size_bytes / 1024:.1f}KB"


def get_photo_settings():
    default = {
        "auto_compress": True,
        "max_kb": PHOTO_MAX_DISPLAY_KB,
        "max_width": PHOTO_MAX_WIDTH,
        "retention_days": PHOTO_RETENTION_DAYS,
        "backup_before_delete": True,
    }
    try:
        saved = get_app_setting("photo_storage_settings", default)
        if isinstance(saved, dict):
            default.update(saved)
    except Exception:
        pass

    # 安全側に丸める
    try:
        default["max_kb"] = int(default.get("max_kb") or PHOTO_MAX_DISPLAY_KB)
    except Exception:
        default["max_kb"] = PHOTO_MAX_DISPLAY_KB
    try:
        default["max_width"] = int(default.get("max_width") or PHOTO_MAX_WIDTH)
    except Exception:
        default["max_width"] = PHOTO_MAX_WIDTH
    try:
        default["retention_days"] = int(default.get("retention_days") or PHOTO_RETENTION_DAYS)
    except Exception:
        default["retention_days"] = PHOTO_RETENTION_DAYS

    default["max_kb"] = max(100, min(default["max_kb"], 1024))
    default["max_width"] = max(480, min(default["max_width"], 1600))
    default["retention_days"] = max(30, min(default["retention_days"], 3650))
    default["auto_compress"] = bool(default.get("auto_compress", True))
    default["backup_before_delete"] = bool(default.get("backup_before_delete", True))
    return default


def _open_uploaded_image(uploaded_file):
    if uploaded_file is None or Image is None:
        return None, b""
    uploaded_file.seek(0)
    raw = uploaded_file.read()
    uploaded_file.seek(0)
    if not raw:
        return None, b""
    image = Image.open(BytesIO(raw))
    if ImageOps is not None:
        image = ImageOps.exif_transpose(image)
    return image, raw


def compress_handover_photo_upload(uploaded_file, max_bytes=None, max_width=None):
    """アップロード写真を300KB目安まで自動圧縮する。失敗時は安全にエラー情報を返す。"""
    settings = get_photo_settings()
    max_bytes = int(max_bytes or settings.get("max_kb", PHOTO_MAX_DISPLAY_KB) * 1024)
    max_width = int(max_width or settings.get("max_width", PHOTO_MAX_WIDTH))

    result = {
        "ok": False,
        "bytes": b"",
        "mime": "image/jpeg",
        "suffix": ".jpg",
        "original_size": 0,
        "compressed_size": 0,
        "compressed": False,
        "message": "",
    }

    if uploaded_file is None:
        result["message"] = "写真が選択されていません。"
        return result

    original_name = clean_text(getattr(uploaded_file, "name", ""), "photo.jpg")
    original_suffix = Path(original_name).suffix.lower()
    if original_suffix in [".jpg", ".jpeg"]:
        original_mime = "image/jpeg"
    elif original_suffix == ".png":
        original_mime = "image/png"
    elif original_suffix == ".webp":
        original_mime = "image/webp"
    else:
        original_mime = "image/jpeg"

    try:
        image, raw = _open_uploaded_image(uploaded_file)
        result["original_size"] = len(raw)

        # Pillowが使えない場合は、300KB以下の画像だけ許可
        if Image is None or image is None:
            if len(raw) <= max_bytes:
                result.update({
                    "ok": True,
                    "bytes": raw,
                    "mime": original_mime,
                    "suffix": original_suffix if original_suffix in [".jpg", ".jpeg", ".png", ".webp"] else ".jpg",
                    "compressed_size": len(raw),
                    "compressed": False,
                    "message": "Pillow未導入のため元画像を保存しました。",
                })
            else:
                result["message"] = "画像圧縮ライブラリ（Pillow）がないため、300KBを超える写真は保存できません。"
            return result

        # 透過PNG等も安全にJPEG化するためRGBへ変換
        if image.mode not in ("RGB", "L"):
            bg = Image.new("RGB", image.size, (255, 255, 255))
            if image.mode in ("RGBA", "LA"):
                bg.paste(image, mask=image.split()[-1])
                image = bg
            else:
                image = image.convert("RGB")
        elif image.mode == "L":
            image = image.convert("RGB")

        # 長辺をmax_width以内に縮小
        width, height = image.size
        long_side = max(width, height)
        if long_side > max_width:
            ratio = max_width / long_side
            new_size = (max(1, int(width * ratio)), max(1, int(height * ratio)))
            image = image.resize(new_size, Image.LANCZOS)

        # JPEG品質を段階的に下げて300KB以下へ
        best_bytes = b""
        best_quality = None
        for quality in [85, 78, 72, 68, 64, 60, 55, 50, 45, 40, PHOTO_JPEG_MIN_QUALITY]:
            out = BytesIO()
            image.save(out, format="JPEG", quality=quality, optimize=True)
            data = out.getvalue()
            best_bytes = data
            best_quality = quality
            if len(data) <= max_bytes:
                break

        # まだ大きい場合は幅をさらに小さくして再試行
        current_width = max(image.size)
        while len(best_bytes) > max_bytes and current_width > 480:
            current_width = int(current_width * 0.85)
            width, height = image.size
            long_side = max(width, height)
            ratio = current_width / long_side
            resized = image.resize((max(1, int(width * ratio)), max(1, int(height * ratio))), Image.LANCZOS)
            for quality in [60, 50, 45, 40, PHOTO_JPEG_MIN_QUALITY]:
                out = BytesIO()
                resized.save(out, format="JPEG", quality=quality, optimize=True)
                data = out.getvalue()
                best_bytes = data
                best_quality = quality
                if len(data) <= max_bytes:
                    image = resized
                    break
            if len(best_bytes) <= max_bytes:
                break
            image = resized

        if len(best_bytes) > max_bytes:
            result["message"] = f"写真を圧縮しましたが、上限{int(max_bytes/1024)}KB以下にできませんでした。別の写真で再試行してください。"
            result["compressed_size"] = len(best_bytes)
            return result

        result.update({
            "ok": True,
            "bytes": best_bytes,
            "mime": "image/jpeg",
            "suffix": ".jpg",
            "compressed_size": len(best_bytes),
            "compressed": len(best_bytes) < len(raw) or original_mime != "image/jpeg",
            "message": f"圧縮OK：{format_file_size(len(raw))} → {format_file_size(len(best_bytes))}（JPEG品質{best_quality}）",
        })
        return result
    except Exception as e:
        result["message"] = f"写真の圧縮に失敗しました：{e}"
        return result


def render_photo_compression_preview(uploaded_file):
    """フォーム内で圧縮前後サイズを表示する。"""
    if uploaded_file is None:
        return
    info = compress_handover_photo_upload(uploaded_file)
    if info.get("ok"):
        st.success(f"写真軽量化：{format_file_size(info.get('original_size'))} → {format_file_size(info.get('compressed_size'))}（上限{get_photo_settings().get('max_kb')}KB）")
        try:
            st.image(info.get("bytes"), caption="圧縮後プレビュー", use_container_width=True)
        except Exception:
            st.image(uploaded_file, caption="添付予定の写真", use_container_width=True)
    else:
        st.warning(info.get("message", "写真を圧縮できませんでした。"))


def build_expired_photo_marker(info, reason="保存期間終了"):
    marker = {
        "mode": "expired",
        "filename": clean_text(info.get("filename")),
        "original_name": clean_text(info.get("original_name")),
        "local_path": clean_text(info.get("local_path")),
        "saved_at": clean_text(info.get("saved_at")),
        "expired_at": format_now_jst("%Y-%m-%d %H:%M:%S"),
        "reason": reason,
        "message": "写真は保存期間終了のため削除しました。記録本文は保持されています。",
    }
    return json.dumps(marker, ensure_ascii=False)


def cleanup_expired_handover_photos(retention_days=None, backup_before_delete=True):
    """180日を超えた申し送り写真を、削除前バックアップ後に削除する。"""
    settings = get_photo_settings()
    retention_days = int(retention_days or settings.get("retention_days", PHOTO_RETENTION_DAYS))
    backup_before_delete = bool(backup_before_delete and settings.get("backup_before_delete", True))

    try:
        df = load_business_handover_data()
        if df.empty:
            return {"checked": 0, "deleted": 0, "updated": 0, "backup": ""}

        cutoff = now_jst_dt() - timedelta(days=retention_days)
        updated = False
        deleted_count = 0
        checked = 0
        backup_name = ""

        # 先に対象有無だけ確認
        targets = []
        for idx, row in df.iterrows():
            for col in ["写真1", "写真2"]:
                info = _parse_handover_photo_value(row.get(col, ""))
                if not info or info.get("mode") == "expired":
                    continue
                checked += 1
                saved_at = clean_text(info.get("saved_at")) or clean_text(row.get("記録日時"))
                dt = pd.to_datetime(saved_at, errors="coerce")
                if pd.isna(dt):
                    continue
                try:
                    dt_py = dt.to_pydatetime()
                    if dt_py.tzinfo is None and JST:
                        dt_py = dt_py.replace(tzinfo=JST)
                except Exception:
                    dt_py = None
                if dt_py and dt_py < cutoff:
                    targets.append((idx, col, info))

        if not targets:
            return {"checked": checked, "deleted": 0, "updated": 0, "backup": ""}

        if backup_before_delete and "create_backup_zip" in globals():
            zip_path, err = create_backup_zip(kind="写真削除前")
            if zip_path and not err:
                backup_name = zip_path.name
            elif err:
                try:
                    st.warning(f"写真削除前バックアップに失敗したため、自動削除を中止しました：{err}")
                except Exception:
                    pass
                return {"checked": checked, "deleted": 0, "updated": 0, "backup": "", "error": err}

        for idx, col, info in targets:
            local_path = clean_text(info.get("local_path"))
            if local_path:
                try:
                    path = Path(local_path)
                    if path.exists() and path.is_file():
                        path.unlink()
                        deleted_count += 1
                except Exception:
                    pass
            # DB内base64も消すため、写真欄を期限切れマーカーへ置換
            df.at[idx, col] = build_expired_photo_marker(info)
            updated = True

        if updated:
            save_business_handover_data(df)
            try:
                add_audit_log("写真保存期間整理", SQLITE_TABLE_HANDOVER, "", f"{retention_days}日超の写真を{len(targets)}件整理。削除前バックアップ：{backup_name}")
            except Exception:
                pass

        return {"checked": checked, "deleted": deleted_count, "updated": len(targets), "backup": backup_name}
    except Exception as e:
        return {"checked": 0, "deleted": 0, "updated": 0, "backup": "", "error": str(e)}


def run_daily_photo_retention_cleanup():
    """1日1回だけ写真保存期間整理を実行する。"""
    try:
        ensure_security_dirs()
        today_key = today_jst().strftime("%Y%m%d")
        marker = BACKUP_DIR / f".photo_cleanup_{today_key}.done"
        if marker.exists():
            return
        result = cleanup_expired_handover_photos()
        marker.write_text(json.dumps(result, ensure_ascii=False), encoding="utf-8")
    except Exception:
        pass


def _guess_image_mime(suffix: str) -> str:
    suffix = (suffix or "").lower()
    if suffix in [".jpg", ".jpeg"]:
        return "image/jpeg"
    if suffix == ".png":
        return "image/png"
    if suffix == ".webp":
        return "image/webp"
    return "image/jpeg"


def _parse_handover_photo_value(photo_value):
    """写真保存値を後方互換で読む。旧：ファイルパス／新：JSON(base64埋込)。"""
    text = clean_text(photo_value)
    if not text:
        return {}
    if text.startswith("data:image/"):
        return {"mode": "data_url", "data_url": text, "filename": "photo"}
    if text.startswith("{"):
        try:
            data = json.loads(text)
            if isinstance(data, dict):
                return data
        except Exception:
            pass
    return {"mode": "local_path", "local_path": text, "filename": Path(text).name}


def save_business_handover_photo(uploaded_file, record_id, photo_no=1):
    """
    業務全体申し送りに添付された写真を保存し、保存値を返す。

    Ver4.1 写真軽量化・保存期間管理版：
    - 写真は1件1枚運用
    - 保存前に自動圧縮し、300KB以下を標準にする
    - 圧縮前後サイズを保存値に保持する
    - ローカル保存＋DB内data_url保持（Supabase/Streamlit Cloudでも表示可能）
    - 180日超の写真は削除前バックアップ後、写真欄を保存期間終了マーカーへ置換
    """
    if uploaded_file is None:
        return ""

    ensure_dirs()
    BUSINESS_HANDOVER_PHOTO_DIR.mkdir(parents=True, exist_ok=True)
    original_name = clean_text(getattr(uploaded_file, "name", ""), "photo.jpg")

    safe_record_id = re.sub(r"[^A-Za-z0-9_-]", "_", clean_text(record_id, "handover"))
    file_name = f"{safe_record_id}_photo{photo_no}.jpg"
    save_path = BUSINESS_HANDOVER_PHOTO_DIR / file_name

    try:
        compressed = compress_handover_photo_upload(uploaded_file)
        if not compressed.get("ok"):
            try:
                st.warning(compressed.get("message", "写真を保存できませんでした。"))
            except Exception:
                pass
            return ""

        photo_bytes = compressed["bytes"]
        if not photo_bytes:
            return ""

        # ローカルにも保存（バックアップZIP用）
        try:
            save_path.write_bytes(photo_bytes)
        except Exception:
            pass

        mime = compressed.get("mime", "image/jpeg")
        data_url = f"data:{mime};base64,{base64.b64encode(photo_bytes).decode('ascii')}"

        value = {
            "mode": "embedded_base64",
            "filename": file_name,
            "original_name": original_name,
            "mime": mime,
            "size": len(photo_bytes),
            "original_size": compressed.get("original_size", 0),
            "compressed_size": compressed.get("compressed_size", len(photo_bytes)),
            "compressed": compressed.get("compressed", False),
            "max_kb": get_photo_settings().get("max_kb", PHOTO_MAX_DISPLAY_KB),
            "retention_days": get_photo_settings().get("retention_days", PHOTO_RETENTION_DAYS),
            "local_path": str(save_path),
            "data_url": data_url,
            "saved_at": format_now_jst("%Y-%m-%d %H:%M:%S") if "format_now_jst" in globals() else datetime.utcnow().isoformat(),
        }
        try:
            st.success(f"写真を軽量化して保存しました：{format_file_size(value['original_size'])} → {format_file_size(value['compressed_size'])}")
        except Exception:
            pass
        return json.dumps(value, ensure_ascii=False)
    except Exception as e:
        try:
            st.warning(f"写真の保存に失敗しました：{e}")
        except Exception:
            pass
        return ""

def show_business_handover_photo(photo_path, caption="添付写真"):
    """保存済み写真を画面表示する。旧ファイルパス／新base64埋込の両方に対応。"""
    info = _parse_handover_photo_value(photo_path)
    if not info:
        return
    try:
        if info.get("mode") == "expired":
            st.caption(clean_text(info.get("message"), "写真は保存期間終了のため削除済みです。"))
            return

        data_url = clean_text(info.get("data_url"))
        if data_url.startswith("data:image/") and ";base64," in data_url:
            b64 = data_url.split(";base64,", 1)[1]
            st.image(base64.b64decode(b64), caption=caption, use_container_width=True)
            return

        local_path = clean_text(info.get("local_path"))
        if local_path:
            path = Path(local_path)
            if path.exists():
                st.image(str(path), caption=caption, use_container_width=True)
                return

        st.caption(f"{caption}：保存ファイルが見つかりません。")
    except Exception:
        st.caption(f"{caption}：画像を表示できませんでした。")


def save_business_handover_excel(uploaded_file, record_id):
    """業務全体申し送りに添付されたExcel/CSVを保存し、保存パスを返す。"""
    if uploaded_file is None:
        return ""

    ensure_dirs()
    original_name = clean_text(getattr(uploaded_file, "name", ""), "handover.xlsx")
    suffix = Path(original_name).suffix.lower()
    if suffix not in [".xlsx", ".xls", ".csv"]:
        suffix = ".xlsx"

    safe_record_id = re.sub(r"[^A-Za-z0-9_-]", "_", clean_text(record_id, "handover"))
    file_name = f"{safe_record_id}_input_excel{suffix}"
    save_path = BUSINESS_HANDOVER_EXCEL_DIR / file_name

    try:
        uploaded_file.seek(0)
        save_path.write_bytes(uploaded_file.read())
        return str(save_path)
    except Exception:
        return ""


def read_uploaded_excel_preview(uploaded_file_or_path, max_rows=20):
    """アップロードExcel/CSVを読み込み、表示用の概要テキストとプレビューDataFrameを返す。"""
    try:
        if uploaded_file_or_path is None:
            return "入力Excelデータはありません。", pd.DataFrame()

        name = clean_text(getattr(uploaded_file_or_path, "name", "")) if not isinstance(uploaded_file_or_path, (str, Path)) else str(uploaded_file_or_path)
        suffix = Path(name).suffix.lower()

        if hasattr(uploaded_file_or_path, "seek"):
            uploaded_file_or_path.seek(0)

        if suffix == ".csv":
            df = pd.read_csv(uploaded_file_or_path)
        else:
            df = pd.read_excel(uploaded_file_or_path)

        if df is None or df.empty:
            return "入力Excelデータは読み込めましたが、表示できる行がありません。", pd.DataFrame()

        df = df.copy()
        # 列名と値を安全に文字列化し、空白列を整理
        df.columns = [clean_text(c, f"列{i+1}") for i, c in enumerate(df.columns)]
        df = df.dropna(how="all")
        preview_df = df.head(max_rows).copy()

        summary = f"ファイル名：{Path(name).name}\n行数：{len(df)}行／列数：{len(df.columns)}列\n表示：先頭{min(len(df), max_rows)}行"
        return summary, preview_df
    except Exception as e:
        return f"入力Excelデータを読み込めませんでした：{e}", pd.DataFrame()


def build_uploaded_excel_display_text(uploaded_file):
    """保存用に、入力Excelの概要と先頭行をテキスト化する。"""
    summary, preview_df = read_uploaded_excel_preview(uploaded_file, max_rows=10)
    if preview_df.empty:
        return summary

    lines = [summary, "", "【先頭データ】"]
    for idx, row in preview_df.iterrows():
        values = []
        for col in preview_df.columns[:8]:
            val = clean_text(row.get(col, ""))
            if val:
                values.append(f"{col}:{val}")
        if values:
            lines.append("・" + "／".join(values))
    return "\n".join(lines)


def show_business_handover_excel_preview(excel_path, display_text=""):
    """保存済みの入力Excelを画面表示する。"""
    excel_path = clean_text(excel_path)
    display_text = clean_text(display_text)

    if display_text:
        st.markdown("**入力Excel表示情報**")
        st.info(display_text)

    if not excel_path:
        return

    path = Path(excel_path)
    if not path.exists():
        st.caption("入力Excel：保存ファイルが見つかりません。")
        return

    summary, preview_df = read_uploaded_excel_preview(path, max_rows=20)
    st.markdown("**入力Excelプレビュー**")
    st.caption(summary)
    if not preview_df.empty:
        st.dataframe(preview_df, use_container_width=True, hide_index=True)
    try:
        st.download_button(
            "入力Excelをダウンロード",
            data=path.read_bytes(),
            file_name=path.name,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet" if path.suffix.lower() != ".csv" else "text/csv",
            use_container_width=True,
            key=f"download_{path.name}_{uuid.uuid4().hex[:6]}",
        )
    except Exception:
        pass



def ensure_alert_condition_file():
    """申し送り自動抽出の条件マスタをSQLiteへ作成する。"""
    ensure_dirs()
    if sqlite_table_row_count(SQLITE_TABLE_ALERT_CONDITIONS) == 0:
        # 既存Excelがあれば初回移行、なければ標準条件を登録
        if ALERT_CONDITION_FILE.exists():
            try:
                df = pd.read_excel(ALERT_CONDITION_FILE, sheet_name="条件マスタ")
            except Exception:
                df = pd.DataFrame(DEFAULT_ALERT_CONDITIONS, columns=ALERT_CONDITION_COLUMNS)
        else:
            df = pd.DataFrame(DEFAULT_ALERT_CONDITIONS, columns=ALERT_CONDITION_COLUMNS)
        df = normalize_alert_condition_master_df(df)
        save_sqlite_table(df, SQLITE_TABLE_ALERT_CONDITIONS, ALERT_CONDITION_COLUMNS, unique_cols=["条件ID"])


def normalize_alert_condition_master_df(df):
    """条件マスタを st.data_editor で安全に編集できる型へ整える。
    ※ mixed/object 型のまま CheckboxColumn や NumberColumn に渡すと
      StreamlitAPIException になるため、ここで明示的に型をそろえる。
    """
    if df is None or len(df) == 0:
        df = pd.DataFrame(DEFAULT_ALERT_CONDITIONS, columns=ALERT_CONDITION_COLUMNS)
    else:
        df = df.copy()

    for col in ALERT_CONDITION_COLUMNS:
        if col not in df.columns:
            df[col] = ""

    df = df[ALERT_CONDITION_COLUMNS].copy()

    def to_bool(v):
        if isinstance(v, bool):
            return v
        text = str(v).strip().lower()
        return text in ["true", "1", "yes", "on", "使用", "表示", "有効", "checked"]

    df["使用"] = df["使用"].apply(to_bool).astype(bool)

    text_cols = [
        "条件ID", "条件名", "重要度", "分類", "条件種別",
        "閾値1", "閾値2", "キーワード", "表示メッセージ"
    ]
    for col in text_cols:
        df[col] = df[col].fillna("").astype(str)
        df[col] = df[col].replace(["nan", "None", "NaT"], "")

    df["日数"] = pd.to_numeric(df["日数"], errors="coerce").fillna(1).astype(int)
    df["並び順"] = pd.to_numeric(df["並び順"], errors="coerce").fillna(999).astype(int)

    allowed_severity = ["至急", "注意", "観察", "通常"]
    df.loc[~df["重要度"].isin(allowed_severity), "重要度"] = "観察"

    allowed_category = ["排泄", "食事", "バイタル", "体重", "変化", "複合", "その他"]
    df.loc[~df["分類"].isin(allowed_category), "分類"] = "その他"

    allowed_kind = [
        "未排便", "便性状", "尿性状", "食事低下", "食事低下連続",
        "発熱", "SpO2低下", "血圧高値", "体重減少", "キーワード",
        "複合:発熱+食事低下", "複合:濃縮尿+食事低下", "複合:SpO2低下+キーワード",
    ]
    df.loc[~df["条件種別"].isin(allowed_kind), "条件種別"] = "キーワード"

    df = df.sort_values("並び順").reset_index(drop=True)
    return df


def load_alert_condition_master():
    ensure_alert_condition_file()
    df = load_sqlite_table(SQLITE_TABLE_ALERT_CONDITIONS, ALERT_CONDITION_COLUMNS)
    if df.empty:
        df = pd.DataFrame(DEFAULT_ALERT_CONDITIONS, columns=ALERT_CONDITION_COLUMNS)
    return normalize_alert_condition_master_df(df)



def save_alert_condition_master(df):
    """
    条件マスタをSQLiteへ保存する。
    商品版ではExcelファイルを正データとして書き出しません。
    """
    ensure_dirs()
    df = normalize_alert_condition_master_df(df)

    existing_ids = set()
    for i, row in df.iterrows():
        cid = clean_text(row.get("条件ID"))
        if not cid:
            cid = f"C{(i + 1):03d}"
            while cid in existing_ids:
                cid = f"C{random.randint(100, 999)}"
            df.at[i, "条件ID"] = cid
        existing_ids.add(cid)

    save_sqlite_table(df, SQLITE_TABLE_ALERT_CONDITIONS, ALERT_CONDITION_COLUMNS, unique_cols=["条件ID"])
    return df

def parse_keywords(value):
    text = clean_text(value)
    if not text:
        return []
    return [x.strip() for x in re.split(r"[,、\n\s]+", text) if x.strip()]


def get_health_row_for_day(health_df, target_day, user_name):
    if health_df.empty:
        return None
    h = health_df.copy()
    h["記録日"] = pd.to_datetime(h["記録日"], errors="coerce")
    h = h[(h["記録日"].dt.date == target_day) & (h["利用者名"].astype(str) == user_name)]
    if h.empty:
        return None
    return h.iloc[-1]


def get_health_days_for_user(health_df, target_day, user_name, days):
    if health_df.empty:
        return pd.DataFrame(columns=HEALTH_COLUMNS)
    start_day = target_day - timedelta(days=max(int(days), 1) - 1)
    h = health_df.copy()
    h["記録日"] = pd.to_datetime(h["記録日"], errors="coerce")
    return h[
        (h["利用者名"].astype(str) == user_name)
        & (h["記録日"].dt.date >= start_day)
        & (h["記録日"].dt.date <= target_day)
    ].sort_values("記録日")


def day_meal_min(row):
    if row is None:
        return None
    vals = []
    for col in ["朝食摂取率", "昼食摂取率", "夕食摂取率"]:
        v = safe_float(row.get(col), -1)
        if v >= 0:
            vals.append(v)
    return min(vals) if vals else None


def check_no_stool_days(ex_df, target_day, user_name, days):
    days = max(int(safe_int(days, 3)), 1)
    check_dates = [target_day - timedelta(days=i) for i in range(days - 1, -1, -1)]
    stool_counts = []
    for d in check_dates:
        ddf = get_day_excretion_data(ex_df, d, user_name)
        stool_counts.append(summarize_excretion(ddf)["排便回数"])
    return sum(stool_counts) == 0, f"{check_dates[0].strftime('%m/%d')}〜{check_dates[-1].strftime('%m/%d')}"


def check_alert_condition(rule, health_df, ex_df, target_day, user_name):
    """条件マスタ1行を判定し、該当すれば辞書を返す。診断ではなく申し送り候補。"""
    kind = clean_text(rule.get("条件種別"))
    name = clean_text(rule.get("条件名"), kind)
    severity = clean_text(rule.get("重要度"), "観察")
    category = clean_text(rule.get("分類"), "その他")
    threshold1 = safe_float(rule.get("閾値1"), 0)
    threshold2 = safe_float(rule.get("閾値2"), 0)
    days = max(safe_int(rule.get("日数"), 1), 1)
    message = clean_text(rule.get("表示メッセージ"), name)
    hrow = get_health_row_for_day(health_df, target_day, user_name)
    ex_day = get_day_excretion_data(ex_df, target_day, user_name)
    detail = ""
    hit = False
    matched_text = ""

    if kind == "未排便":
        hit, period = check_no_stool_days(ex_df, target_day, user_name, days)
        detail = f"対象期間：{period}"

    elif kind == "便性状":
        keywords = parse_keywords(rule.get("キーワード")) or ["下痢便", "水様便"]
        if not ex_day.empty:
            warn = ex_day[ex_day["便性状"].fillna("").astype(str).isin(keywords)]
            hit = not warn.empty
            if hit:
                detail = "、".join([f"{clean_text(r.get('時間帯'))}:{clean_text(r.get('便性状'))}" for _, r in warn.iterrows()])

    elif kind == "尿性状":
        keywords = parse_keywords(rule.get("キーワード")) or ["濃縮尿"]
        if not ex_day.empty:
            warn = ex_day[ex_day["尿性状"].fillna("").astype(str).isin(keywords)]
            hit = not warn.empty
            if hit:
                detail = "、".join([f"{clean_text(r.get('時間帯'))}:{clean_text(r.get('尿性状'))}" for _, r in warn.iterrows()])

    elif kind == "食事低下":
        meal_min = day_meal_min(hrow)
        hit = meal_min is not None and meal_min <= threshold1
        if hit:
            detail = f"最小食事摂取率 {meal_min:.0f}%"

    elif kind == "食事低下連続":
        hdays = get_health_days_for_user(health_df, target_day, user_name, days)
        low_days = 0
        for _, r in hdays.iterrows():
            meal_min = day_meal_min(r)
            if meal_min is not None and meal_min <= threshold1:
                low_days += 1
        hit = low_days >= days
        if hit:
            detail = f"{days}日中{low_days}日が{threshold1:.0f}%以下"

    elif kind == "発熱":
        temp = safe_float(hrow.get("体温"), 0) if hrow is not None else 0
        hit = temp >= threshold1 and temp > 0
        if hit:
            detail = f"体温 {temp:.1f}℃"

    elif kind == "SpO2低下":
        spo2 = safe_float(hrow.get("SpO2"), 0) if hrow is not None else 0
        hit = 0 < spo2 <= threshold1
        if hit:
            detail = f"SpO2 {spo2:.0f}%"

    elif kind == "血圧高値":
        bp_high = safe_float(hrow.get("血圧上"), 0) if hrow is not None else 0
        hit = bp_high >= threshold1 and bp_high > 0
        if hit:
            detail = f"血圧上 {bp_high:.0f}"

    elif kind == "体重減少":
        hdays = get_health_days_for_user(health_df, target_day, user_name, days + 1)
        weights = []
        for _, r in hdays.iterrows():
            w = safe_float(r.get("体重"), 0)
            if w > 0:
                weights.append((pd.to_datetime(r.get("記録日"), errors="coerce"), w))
        if len(weights) >= 2:
            before = weights[0][1]
            now = weights[-1][1]
            diff = before - now
            hit = diff >= threshold1
            if hit:
                detail = f"{before:.1f}kg → {now:.1f}kg（-{diff:.1f}kg）"

    elif kind == "キーワード":
        keywords = parse_keywords(rule.get("キーワード"))
        notes = []
        if hrow is not None:
            notes.append(clean_text(hrow.get("気になる変化")))
            notes.append(clean_text(hrow.get("家族共有メモ")))
        note_text = " ".join([n for n in notes if n])
        matched = [kw for kw in keywords if kw and kw in note_text]
        hit = bool(matched)
        if hit:
            matched_text = "、".join(matched)
            detail = clean_text(note_text[:80])

    elif kind == "複合:発熱+食事低下":
        temp = safe_float(hrow.get("体温"), 0) if hrow is not None else 0
        meal_min = day_meal_min(hrow)
        hit = temp >= threshold1 and meal_min is not None and meal_min <= threshold2
        if hit:
            detail = f"体温 {temp:.1f}℃／最小食事摂取率 {meal_min:.0f}%"

    elif kind == "複合:濃縮尿+食事低下":
        urine_hit = False
        if not ex_day.empty:
            urine_hit = (ex_day["尿性状"].fillna("").astype(str) == "濃縮尿").any()
        meal_min = day_meal_min(hrow)
        hit = urine_hit and meal_min is not None and meal_min <= threshold1
        if hit:
            detail = f"濃縮尿あり／最小食事摂取率 {meal_min:.0f}%"

    elif kind == "複合:SpO2低下+キーワード":
        spo2 = safe_float(hrow.get("SpO2"), 0) if hrow is not None else 0
        keywords = parse_keywords(rule.get("キーワード")) or ["傾眠", "息苦しい", "呼吸"]
        note_text = ""
        if hrow is not None:
            note_text = f"{clean_text(hrow.get('気になる変化'))} {clean_text(hrow.get('家族共有メモ'))}"
        matched = [kw for kw in keywords if kw and kw in note_text]
        hit = 0 < spo2 <= threshold1 and bool(matched)
        if hit:
            matched_text = "、".join(matched)
            detail = f"SpO2 {spo2:.0f}%／キーワード：{matched_text}"

    if not hit:
        return None

    try:
        message = message.format(
            日数=days,
            閾値1=int(threshold1) if float(threshold1).is_integer() else threshold1,
            閾値2=int(threshold2) if float(threshold2).is_integer() else threshold2,
            該当内容=matched_text or detail,
        )
    except Exception:
        pass

    return {
        "利用者名": user_name,
        "重要度": severity,
        "分類": category,
        "条件名": name,
        "詳細": detail,
        "申し送り文": message,
    }


def build_handover_alerts_by_condition(target_date):
    """条件マスタに基づき、健康・排泄Excelから申し送り候補を抽出する。"""
    target = pd.to_datetime(target_date, errors="coerce")
    if pd.isna(target):
        return pd.DataFrame(columns=["利用者名", "重要度", "分類", "条件名", "詳細", "申し送り文"])
    target_day = target.date()
    try:
        health_df = load_health_data()
    except Exception:
        health_df = pd.DataFrame(columns=HEALTH_COLUMNS)
    try:
        ex_df = load_excretion_data()
    except Exception:
        ex_df = pd.DataFrame(columns=EXCRETION_COLUMNS)
    try:
        rules = load_alert_condition_master()
    except Exception:
        rules = pd.DataFrame(DEFAULT_ALERT_CONDITIONS, columns=ALERT_CONDITION_COLUMNS)

    enabled_rules = rules[rules["使用"].astype(bool)].copy()
    rows = []
    for user in active_users:
        for _, rule in enabled_rules.iterrows():
            hit = check_alert_condition(rule, health_df, ex_df, target_day, user)
            if hit:
                rows.append(hit)
    return pd.DataFrame(rows, columns=["利用者名", "重要度", "分類", "条件名", "詳細", "申し送り文"])


def build_business_handover_auto_extract_text(target_date):
    """条件設定マスタに基づき、健康チェック・排泄チェックExcelから申し送り候補を自動抽出する。"""
    target = pd.to_datetime(target_date, errors="coerce")
    if pd.isna(target):
        return "Excel自動抽出情報：日付を確認できません。"
    target_day = target.date()
    lines = [f"【Excel自動抽出情報】対象日：{target_day.strftime('%Y-%m-%d')} ／ 条件設定マスタに基づく抽出"]

    alert_df = build_handover_alerts_by_condition(target_day)
    if alert_df.empty:
        lines.append("・条件に該当する申し送り候補はありません。")
        return "\n".join(lines)

    severity_order = {"至急": 0, "注意": 1, "観察": 2, "通常": 3}
    alert_df["_order"] = alert_df["重要度"].map(severity_order).fillna(9)
    alert_df = alert_df.sort_values(["_order", "利用者名", "分類", "条件名"]).drop(columns=["_order"])

    for severity in ["至急", "注意", "観察", "通常"]:
        part = alert_df[alert_df["重要度"] == severity]
        if part.empty:
            continue
        mark = "🔴" if severity == "至急" else "🟠" if severity == "注意" else "🟡" if severity == "観察" else "⚪"
        lines.append(f"・{mark}{severity}：")
        for _, r in part.iterrows():
            detail = clean_text(r.get("詳細"))
            detail_text = f"（{detail}）" if detail else ""
            lines.append(
                f"  - {clean_text(r.get('利用者名'))}｜{clean_text(r.get('分類'))}｜{clean_text(r.get('条件名'))}{detail_text}\n"
                f"    → {clean_text(r.get('申し送り文'))}"
            )
    return "\n".join(lines)


def show_alert_condition_master_menu():
    """管理者が申し送り自動抽出条件を編集する画面。"""
    st.subheader("異常検知・申し送り条件設定マスタ")
    st.caption("健康チェック・排泄チェックのExcelデータから、業務全体申し送りに自動表示する条件を設定します。診断ではなく、申し送り候補を拾うための設定です。")

    if not is_admin_user():
        st.warning("条件設定は管理者専用です。")
        return

    c1, c2, c3 = st.columns([1.1, 1.1, 1])
    with c1:
        if st.button("初期おすすめ条件に戻す", use_container_width=True):
            save_alert_condition_master(pd.DataFrame(DEFAULT_ALERT_CONDITIONS, columns=ALERT_CONDITION_COLUMNS))
            st.success("初期おすすめ条件に戻しました。")
            st.rerun()
    with c2:
        if st.button("全条件を使用ONにする", use_container_width=True):
            df_on = load_alert_condition_master()
            df_on["使用"] = True
            save_alert_condition_master(df_on)
            st.success("全条件を使用ONにしました。")
            st.rerun()
    with c3:
        preview_date = st.date_input("抽出プレビュー日", value=today_jst(), key="alert_condition_preview_date")

    df = load_alert_condition_master()
    enabled_count = int(df["使用"].astype(bool).sum()) if "使用" in df.columns else 0
    st.caption(f"現在、使用ONの条件：{enabled_count}件 / {len(df)}件")
    if enabled_count == 0:
        st.warning("使用ONの条件がありません。申し送りには反映されません。左端の『使用』にチェックを入れて保存してください。")

    with st.form("alert_condition_master_form", clear_on_submit=False):
        edited = st.data_editor(
            df,
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "条件ID": st.column_config.TextColumn("条件ID", disabled=True),
                "使用": st.column_config.CheckboxColumn("使用", default=True),
                "重要度": st.column_config.SelectboxColumn("重要度", options=["至急", "注意", "観察", "通常"]),
                "分類": st.column_config.SelectboxColumn("分類", options=["排泄", "食事", "バイタル", "体重", "変化", "複合", "その他"]),
                "条件種別": st.column_config.SelectboxColumn(
                    "条件種別",
                    options=[
                        "未排便", "便性状", "尿性状", "食事低下", "食事低下連続", "発熱", "SpO2低下", "血圧高値", "体重減少", "キーワード",
                        "複合:発熱+食事低下", "複合:濃縮尿+食事低下", "複合:SpO2低下+キーワード",
                    ],
                ),
                "閾値1": st.column_config.TextColumn("閾値1"),
                "閾値2": st.column_config.TextColumn("閾値2"),
                "日数": st.column_config.NumberColumn("日数", min_value=1, max_value=30, step=1),
                "並び順": st.column_config.NumberColumn("並び順", min_value=1, max_value=999, step=1),
            },
            key="alert_condition_master_editor",
        )
        submitted = st.form_submit_button("条件マスタを保存して申し送り表示を更新", type="primary", use_container_width=True)

    if submitted:
        saved_df = save_alert_condition_master(edited)
        st.session_state["alert_condition_master_saved_at"] = format_now_jst("%Y-%m-%d %H:%M:%S")
        st.success(f"条件マスタを保存しました。使用ON：{int(saved_df['使用'].astype(bool).sum())}件。業務全体申し送りの自動抽出に反映されます。")
        st.rerun()

    st.markdown("#### 抽出プレビュー")
    st.caption("ここに表示される内容が、業務全体申し送りの『Excel自動抽出情報』に反映される内容です。")
    alert_df = build_handover_alerts_by_condition(preview_date)
    if alert_df.empty:
        st.info("この日の条件該当者はありません。使用ONの条件、対象日、健康チェック・排泄チェックの記録を確認してください。")
    else:
        st.dataframe(alert_df, use_container_width=True, hide_index=True)
        st.markdown("#### 業務全体申し送りに表示される文章")
        st.info(build_business_handover_auto_extract_text(preview_date))

def show_business_handover_auto_extract_box(target_date):
    """申し送り画面内にExcel自動抽出情報を表示する。"""
    auto_text = build_business_handover_auto_extract_text(target_date)
    st.markdown("#### Excel自動抽出情報")
    st.info(auto_text)
    return auto_text



# =========================
# Ver5予定候補抽出：申し送り → 予定候補 → Excel/CSV出力
# Googleカレンダー等へ直接API連携せず、管理者確認後に取込用データを出力する安全設計。
# =========================
SCHEDULE_KEYWORD_RULES = [
    {"keyword": "受診", "title": "受診", "category": "医療"},
    {"keyword": "病院", "title": "受診・病院", "category": "医療"},
    {"keyword": "通院", "title": "通院", "category": "医療"},
    {"keyword": "往診", "title": "往診", "category": "医療"},
    {"keyword": "訪問診療", "title": "訪問診療", "category": "医療"},
    {"keyword": "訪問看護", "title": "訪問看護", "category": "医療・介護"},
    {"keyword": "訪問", "title": "訪問", "category": "予定"},
    {"keyword": "面談", "title": "面談", "category": "家族・相談"},
    {"keyword": "家族来訪", "title": "家族来訪", "category": "家族"},
    {"keyword": "家族面談", "title": "家族面談", "category": "家族"},
    {"keyword": "来訪", "title": "来訪", "category": "予定"},
    {"keyword": "外出", "title": "外出", "category": "外出"},
    {"keyword": "外泊", "title": "外泊", "category": "外出"},
    {"keyword": "送迎", "title": "送迎", "category": "外出"},
    {"keyword": "美容", "title": "美容", "category": "生活"},
    {"keyword": "理美容", "title": "理美容", "category": "生活"},
    {"keyword": "買い物", "title": "買い物", "category": "生活"},

]

DEFAULT_HANDOVER_KEYWORDS = [
    {"keyword": "受診", "title": "受診", "category": "医療", "is_active": 1, "sort_order": 10},
    {"keyword": "病院", "title": "受診・病院", "category": "医療", "is_active": 1, "sort_order": 20},
    {"keyword": "通院", "title": "通院", "category": "医療", "is_active": 1, "sort_order": 30},
    {"keyword": "往診", "title": "往診", "category": "医療", "is_active": 1, "sort_order": 40},
    {"keyword": "訪問診療", "title": "訪問診療", "category": "医療", "is_active": 1, "sort_order": 50},
    {"keyword": "訪問看護", "title": "訪問看護", "category": "医療・介護", "is_active": 1, "sort_order": 60},
    {"keyword": "訪問", "title": "訪問", "category": "予定", "is_active": 1, "sort_order": 70},
    {"keyword": "家族来訪", "title": "家族来訪", "category": "家族", "is_active": 1, "sort_order": 80},
    {"keyword": "面談", "title": "面談", "category": "家族・相談", "is_active": 1, "sort_order": 90},
    {"keyword": "家族面談", "title": "家族面談", "category": "家族", "is_active": 1, "sort_order": 100},
    {"keyword": "外出", "title": "外出", "category": "外出", "is_active": 1, "sort_order": 110},
    {"keyword": "送迎", "title": "送迎", "category": "外出", "is_active": 1, "sort_order": 120},
]


def normalize_handover_keyword_df(df: pd.DataFrame) -> pd.DataFrame:
    """予定抽出キーワード設定を標準列にそろえる。"""
    if df is None:
        df = pd.DataFrame(columns=HANDOVER_KEYWORD_COLUMNS)
    work = df.copy()
    for col in HANDOVER_KEYWORD_COLUMNS:
        if col not in work.columns:
            work[col] = ""
    work = work[HANDOVER_KEYWORD_COLUMNS].copy()
    work["keyword"] = work["keyword"].map(lambda x: clean_text(x))
    work["title"] = work["title"].map(lambda x: clean_text(x))
    work["category"] = work["category"].map(lambda x: clean_text(x, "予定"))
    work["is_active"] = work["is_active"].map(lambda x: 1 if str(x).lower() in ["1", "true", "yes", "on", "使用", "有効", "対象"] else 0)
    work["sort_order"] = work["sort_order"].map(lambda x: safe_int(x, 100))
    work["created_at"] = work["created_at"].map(lambda x: clean_text(x))
    work["updated_at"] = work["updated_at"].map(lambda x: clean_text(x))
    work = work[work["keyword"] != ""].copy()
    now_value = format_now_jst("%Y-%m-%d %H:%M:%S")
    for idx, row in work.iterrows():
        if not clean_text(row.get("id")):
            work.at[idx, "id"] = "kw_" + hashlib.sha1(clean_text(row.get("keyword")).encode("utf-8")).hexdigest()[:12]
        if not clean_text(row.get("title")):
            work.at[idx, "title"] = clean_text(row.get("keyword"), "予定")
        if not clean_text(row.get("created_at")):
            work.at[idx, "created_at"] = now_value
        work.at[idx, "updated_at"] = clean_text(row.get("updated_at"), now_value)
    work = work.drop_duplicates(subset=["keyword"], keep="last")
    return work.reset_index(drop=True)


def ensure_handover_keyword_table():
    """申し送り予定抽出キーワードのDBテーブルを用意し、初期値を投入する。"""
    try:
        if not sqlite_table_exists(SQLITE_TABLE_HANDOVER_KEYWORDS):
            now_value = format_now_jst("%Y-%m-%d %H:%M:%S")
            rows = []
            for item in DEFAULT_HANDOVER_KEYWORDS:
                rows.append({
                    "id": "kw_" + hashlib.sha1(item["keyword"].encode("utf-8")).hexdigest()[:12],
                    "keyword": item["keyword"],
                    "title": item.get("title", item["keyword"]),
                    "category": item.get("category", "予定"),
                    "is_active": int(item.get("is_active", 1)),
                    "sort_order": int(item.get("sort_order", 100)),
                    "created_at": now_value,
                    "updated_at": now_value,
                })
            save_sqlite_table(
                pd.DataFrame(rows, columns=HANDOVER_KEYWORD_COLUMNS),
                SQLITE_TABLE_HANDOVER_KEYWORDS,
                HANDOVER_KEYWORD_COLUMNS,
                unique_cols=["id"],
                sort_cols=["sort_order", "keyword"],
            )
    except Exception:
        pass


def load_handover_keywords(active_only=False) -> pd.DataFrame:
    ensure_handover_keyword_table()
    try:
        df = load_sqlite_table(SQLITE_TABLE_HANDOVER_KEYWORDS, HANDOVER_KEYWORD_COLUMNS)
        df = normalize_handover_keyword_df(df)
        if active_only:
            df = df[df["is_active"].astype(int) == 1].copy()
        return df.sort_values(["sort_order", "keyword"]).reset_index(drop=True)
    except Exception:
        fallback = pd.DataFrame(DEFAULT_HANDOVER_KEYWORDS)
        fallback["id"] = fallback["keyword"].map(lambda x: "kw_" + hashlib.sha1(str(x).encode("utf-8")).hexdigest()[:12])
        fallback["created_at"] = format_now_jst("%Y-%m-%d %H:%M:%S")
        fallback["updated_at"] = fallback["created_at"]
        fallback = fallback[HANDOVER_KEYWORD_COLUMNS]
        return fallback[fallback["is_active"].astype(int) == 1].copy() if active_only else fallback


def save_handover_keywords(df: pd.DataFrame) -> pd.DataFrame:
    ensure_handover_keyword_table()
    work = normalize_handover_keyword_df(df)
    now_value = format_now_jst("%Y-%m-%d %H:%M:%S")
    work["updated_at"] = now_value
    save_sqlite_table(
        work,
        SQLITE_TABLE_HANDOVER_KEYWORDS,
        HANDOVER_KEYWORD_COLUMNS,
        unique_cols=["id"],
        sort_cols=["sort_order", "keyword"],
    )
    try:
        add_audit_log("予定抽出キーワード設定更新", SQLITE_TABLE_HANDOVER_KEYWORDS, "", f"予定抽出キーワードを{len(work)}件保存")
    except Exception:
        pass
    return work


def load_schedule_keyword_rules() -> list:
    """予定候補抽出に使うキーワードルールをDBから取得する。"""
    df = load_handover_keywords(active_only=True)
    if df.empty:
        return SCHEDULE_KEYWORD_RULES
    rules = []
    for _, row in df.iterrows():
        kw = clean_text(row.get("keyword"))
        if not kw:
            continue
        rules.append({
            "keyword": kw,
            "title": clean_text(row.get("title"), kw),
            "category": clean_text(row.get("category"), "予定"),
        })
    return rules or SCHEDULE_KEYWORD_RULES


SCHEDULE_EXPORT_COLUMNS = [
    "登録する",
    "取込対象",
    "予定ID",
    "元記録ID",
    "元日付",
    "勤務帯",
    "利用者",
    "利用者名",
    "user_id",
    "キーワード",
    "分類",
    "タイトル",
    "件名",
    "日時",
    "開始日",
    "開始時刻",
    "終了日",
    "終了時刻",
    "終日",
    "場所",
    "詳細",
    "内容",
    "元文章",
]

GOOGLE_CALENDAR_CSV_COLUMNS = [
    "Subject",
    "Start Date",
    "Start Time",
    "End Date",
    "End Time",
    "All Day Event",
    "Description",
    "Location",
    "Private",
]

# ひだまり帳内部カレンダー用テーブル
SQLITE_TABLE_HIDAMARI_SCHEDULES = "hidamari_schedules"

HIDAMARI_SCHEDULE_COLUMNS = [
    "予定ID",
    "予定日",
    "開始時刻",
    "終了時刻",
    "終日",
    "user_id",
    "利用者名",
    "分類",
    "タイトル",
    "詳細",
    "場所",
    "元記録ID",
    "元日付",
    "勤務帯",
    "キーワード",
    "元文章",
    "登録日時",
    "登録者",
    "更新日時",
    "更新者",
]


def ensure_hidamari_schedule_table():
    """ひだまり帳内部予定テーブルを作成する。"""
    try:
        if not sqlite_table_exists(SQLITE_TABLE_HIDAMARI_SCHEDULES):
            save_sqlite_table(
                pd.DataFrame(columns=HIDAMARI_SCHEDULE_COLUMNS),
                SQLITE_TABLE_HIDAMARI_SCHEDULES,
                HIDAMARI_SCHEDULE_COLUMNS,
                unique_cols=["予定ID"],
            )
    except Exception:
        pass


def load_hidamari_schedules() -> pd.DataFrame:
    ensure_hidamari_schedule_table()
    return load_sqlite_table(SQLITE_TABLE_HIDAMARI_SCHEDULES, HIDAMARI_SCHEDULE_COLUMNS, date_cols=["予定日", "元日付"])


def save_hidamari_schedules(df: pd.DataFrame):
    ensure_hidamari_schedule_table()
    save_sqlite_table(
        df,
        SQLITE_TABLE_HIDAMARI_SCHEDULES,
        HIDAMARI_SCHEDULE_COLUMNS,
        date_cols=["予定日", "元日付"],
        unique_cols=["予定ID"],
        sort_cols=["予定日", "開始時刻"],
    )


def make_datetime_display(date_text, time_text="") -> str:
    d = clean_text(date_text)
    t = clean_text(time_text)
    return f"{d} {t}".strip() if t else d


def apply_datetime_display_to_schedule_row(row: pd.Series) -> dict:
    """編集された日時欄を開始日・開始時刻へ反映する。"""
    data = dict(row)
    dt_text = clean_text(data.get("日時"))
    if dt_text:
        m = re.search(r"(20\d{2})[-/\.](\d{1,2})[-/\.](\d{1,2})(?:\s+(\d{1,2})[:：](\d{2}))?", dt_text)
        if m:
            try:
                d = date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                data["開始日"] = d.strftime("%Y-%m-%d")
                data["終了日"] = d.strftime("%Y-%m-%d")
                if m.group(4) is not None:
                    h = safe_int(m.group(4), 0)
                    mi = safe_int(m.group(5), 0)
                    data["開始時刻"] = f"{h:02d}:{mi:02d}"
                    if not clean_text(data.get("終了時刻")):
                        data["終了時刻"] = calc_schedule_end_time(data["開始時刻"], 60)
                    data["終日"] = "FALSE"
                else:
                    data["開始時刻"] = clean_text(data.get("開始時刻"))
                    data["終日"] = "FALSE" if clean_text(data.get("開始時刻")) else "TRUE"
            except Exception:
                pass
    # 表示用の編集列を正式列へ同期
    data["取込対象"] = bool(data.get("登録する", data.get("取込対象", True)))
    data["利用者名"] = clean_text(data.get("利用者"), clean_text(data.get("利用者名")))
    if data["利用者名"] == "業務全般":
        data["user_id"] = ""
    elif not clean_text(data.get("user_id")) and data["利用者名"]:
        data["user_id"] = get_user_id_by_name(data["利用者名"])
    data["件名"] = clean_text(data.get("タイトル"), clean_text(data.get("件名"), "予定"))
    data["内容"] = clean_text(data.get("詳細"), clean_text(data.get("内容")))
    return data


def normalize_schedule_editor_df(edited_df: pd.DataFrame) -> pd.DataFrame:
    if edited_df is None or edited_df.empty:
        return pd.DataFrame(columns=SCHEDULE_EXPORT_COLUMNS)
    rows = [apply_datetime_display_to_schedule_row(row) for _, row in edited_df.iterrows()]
    work = pd.DataFrame(rows)
    for col in SCHEDULE_EXPORT_COLUMNS:
        if col not in work.columns:
            work[col] = ""
    # 日時欄も再整形しておく
    work["日時"] = work.apply(lambda r: make_datetime_display(r.get("開始日"), r.get("開始時刻")), axis=1)
    work["登録する"] = work["取込対象"].astype(str).str.lower().isin(["true", "1", "yes", "対象", "取込", "取込対象"])
    return work[SCHEDULE_EXPORT_COLUMNS].copy()


def filter_selected_schedule_rows(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=SCHEDULE_EXPORT_COLUMNS)
    work = normalize_schedule_editor_df(df)
    return work[work["登録する"].astype(str).str.lower().isin(["true", "1", "yes", "対象", "取込", "取込対象"])].copy()


def register_schedules_to_hidamari(candidate_df: pd.DataFrame) -> tuple[int, int]:
    """選択された予定候補を、ひだまり帳内部予定DBへ登録・更新する。戻り値は（登録更新件数, 対象件数）。"""
    selected = filter_selected_schedule_rows(candidate_df)
    if selected.empty:
        return 0, 0

    existing = load_hidamari_schedules()
    if existing is None or existing.empty:
        existing = pd.DataFrame(columns=HIDAMARI_SCHEDULE_COLUMNS)
    else:
        existing = existing.copy()

    # 同じ予定IDは更新扱いにする
    ids = set(selected["予定ID"].map(clean_text).tolist())
    existing = existing[~existing["予定ID"].astype(str).isin(ids)].copy() if "予定ID" in existing.columns else existing

    now_text = format_now_jst("%Y-%m-%d %H:%M:%S")
    user_text = current_login_user()
    rows = []
    for _, row in selected.iterrows():
        schedule_id = clean_text(row.get("予定ID")) or str(uuid.uuid4())
        rows.append({
            "予定ID": schedule_id,
            "予定日": clean_text(row.get("開始日")),
            "開始時刻": clean_text(row.get("開始時刻")),
            "終了時刻": clean_text(row.get("終了時刻")),
            "終日": clean_text(row.get("終日"), "TRUE"),
            "user_id": clean_text(row.get("user_id")),
            "利用者名": clean_text(row.get("利用者名")),
            "分類": clean_text(row.get("分類"), "予定"),
            "タイトル": clean_text(row.get("件名"), "予定"),
            "詳細": clean_text(row.get("内容")),
            "場所": clean_text(row.get("場所")),
            "元記録ID": clean_text(row.get("元記録ID")),
            "元日付": clean_text(row.get("元日付")),
            "勤務帯": clean_text(row.get("勤務帯")),
            "キーワード": clean_text(row.get("キーワード")),
            "元文章": clean_text(row.get("元文章")),
            "登録日時": now_text,
            "登録者": user_text,
            "更新日時": now_text,
            "更新者": user_text,
        })
    add_df = pd.DataFrame(rows, columns=HIDAMARI_SCHEDULE_COLUMNS)
    out = pd.concat([existing, add_df], ignore_index=True)
    save_hidamari_schedules(out)
    try:
        add_audit_log("ひだまり帳予定登録", SQLITE_TABLE_HIDAMARI_SCHEDULES, "", f"申し送り予定候補から{len(add_df)}件を登録・更新")
    except Exception:
        pass
    return len(add_df), len(selected)


def split_handover_lines(text_value: str) -> list:
    """申し送り本文を、予定候補抽出しやすい単位へ分割する。"""
    text_value = clean_text(text_value)
    if not text_value:
        return []
    # 構造化メモの見出しや箇条書きを考慮して分割
    normalized = text_value.replace("。", "。\n").replace("、", "、")
    parts = []
    for line in normalized.splitlines():
        line = clean_text(line)
        if not line:
            continue
        # 長すぎる行は句点で分割した結果を優先
        for part in re.split(r"[。\n\r]+", line):
            part = clean_text(part)
            if part:
                parts.append(part)
    return parts


def get_schedule_keyword_hits(line: str) -> list:
    """1行に含まれる予定化キーワードを返す。長い語を優先して重複を抑える。"""
    line = clean_text(line)
    hits = []
    used = set()
    rules = load_schedule_keyword_rules()
    for rule in sorted(rules, key=lambda x: len(x["keyword"]), reverse=True):
        kw = rule["keyword"]
        if kw in line and kw not in used:
            hits.append(rule)
            used.add(kw)
    return hits


def parse_schedule_date_from_text(line: str, base_date) -> date:
    """申し送り文から日付を推定。見つからなければ元記録日を使う。"""
    if isinstance(base_date, datetime):
        base = base_date.date()
    else:
        try:
            base = pd.to_datetime(base_date, errors="coerce").date()
        except Exception:
            base = today_jst()
    if not base:
        base = today_jst()

    line = clean_text(line)
    if "明後日" in line or "あさって" in line:
        return base + timedelta(days=2)
    if "明日" in line or "翌日" in line:
        return base + timedelta(days=1)
    if "本日" in line or "今日" in line:
        return base

    # 2026/5/20, 2026-05-20, 2026年5月20日
    m = re.search(r"(20\d{2})[年/\-\.](\d{1,2})[月/\-\.](\d{1,2})日?", line)
    if m:
        try:
            return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
        except Exception:
            pass

    # 5/20, 5月20日
    m = re.search(r"(?<!\d)(\d{1,2})[月/\.](\d{1,2})日?", line)
    if m:
        try:
            return date(base.year, int(m.group(1)), int(m.group(2)))
        except Exception:
            pass

    return base


def parse_schedule_time_from_text(line: str):
    """申し送り文から開始時刻を推定。見つからなければ空欄＝終日予定扱い。"""
    line = clean_text(line)
    # 14:30 / 14：30
    m = re.search(r"(\d{1,2})[:：](\d{2})", line)
    if m:
        h = safe_int(m.group(1), 0)
        minute = safe_int(m.group(2), 0)
        if 0 <= h <= 23 and 0 <= minute <= 59:
            return f"{h:02d}:{minute:02d}"

    # 午前10時 / 午後2時30分 / 10時
    m = re.search(r"(午前|午後)?\s*(\d{1,2})時\s*(\d{1,2})?分?", line)
    if m:
        ampm = clean_text(m.group(1))
        h = safe_int(m.group(2), 0)
        minute = safe_int(m.group(3), 0)
        if ampm == "午後" and 1 <= h <= 11:
            h += 12
        if ampm == "午前" and h == 12:
            h = 0
        if 0 <= h <= 23 and 0 <= minute <= 59:
            return f"{h:02d}:{minute:02d}"

    return ""


def calc_schedule_end_time(start_time: str, minutes: int = 60) -> str:
    start_time = clean_text(start_time)
    if not start_time:
        return ""
    try:
        dt = datetime.strptime(start_time, "%H:%M") + timedelta(minutes=minutes)
        return dt.strftime("%H:%M")
    except Exception:
        return ""


def detect_user_name_in_text(line: str) -> str:
    """申し送り文中に利用者名が含まれる場合は拾う。見つからない場合は空欄。"""
    try:
        users = get_active_user_names()
    except Exception:
        users = []
    line = clean_text(line)
    for name in users:
        name_text = clean_text(name)
        if name_text and name_text in line:
            return name_text
        # 「様」抜き表記にも軽く対応
        short_name = name_text.replace("様", "")
        if short_name and short_name in line:
            return name_text
    return ""


def make_schedule_candidate_id(record_id: str, line: str, keyword: str, idx: int) -> str:
    source = f"{clean_text(record_id)}__{clean_text(line)}__{clean_text(keyword)}__{idx}"
    return "sch_" + hashlib.sha1(source.encode("utf-8")).hexdigest()[:12]


def build_schedule_candidate_subject(rule: dict, user_name: str) -> str:
    base = clean_text(rule.get("title"), clean_text(rule.get("keyword"), "予定"))
    user_name = clean_text(user_name)
    return f"{user_name}：{base}" if user_name else base


def extract_schedule_candidates_from_handover_df(df: pd.DataFrame, start_date=None, end_date=None, keyword_filter="") -> pd.DataFrame:
    """業務全体申し送りから、カレンダー取込前の予定候補を抽出する。"""
    if df is None or df.empty:
        return pd.DataFrame(columns=SCHEDULE_EXPORT_COLUMNS)

    work = df.copy()
    if "日付" not in work.columns:
        return pd.DataFrame(columns=SCHEDULE_EXPORT_COLUMNS)
    work["日付_dt"] = pd.to_datetime(work["日付"], errors="coerce")

    if start_date:
        work = work[work["日付_dt"].dt.date >= start_date]
    if end_date:
        work = work[work["日付_dt"].dt.date <= end_date]

    keyword_filter = clean_text(keyword_filter)
    rows = []
    for _, row in work.iterrows():
        record_id = clean_text(row.get("記録ID"))
        record_date = row.get("日付")
        shift = clean_text(row.get("勤務帯"))
        row_target_type = clean_text(row.get("対象区分"))
        row_user_id = clean_text(row.get("user_id"))
        row_user_name = clean_text(row.get("利用者名"))
        if row_user_name == "業務全般":
            row_user_name = ""
            row_user_id = ""
        combined_text = "\n".join([
            clean_text(row.get("全体申し送り")),
            clean_text(row.get("要確認事項")),
        ])
        if keyword_filter and keyword_filter not in combined_text:
            continue

        lines = split_handover_lines(combined_text)
        candidate_idx = 0
        for line in lines:
            hits = get_schedule_keyword_hits(line)
            if not hits:
                continue
            for rule in hits:
                candidate_idx += 1
                schedule_date = parse_schedule_date_from_text(line, record_date)
                start_time = parse_schedule_time_from_text(line)
                end_time = calc_schedule_end_time(start_time, 60)
                detected_user_name = detect_user_name_in_text(line)
                user_name = row_user_name or detected_user_name
                user_id = row_user_id or get_user_id_by_name(user_name)
                subject = build_schedule_candidate_subject(rule, user_name)
                all_day = "TRUE" if not start_time else "FALSE"
                rows.append({
                    "登録する": True,
                    "取込対象": True,
                    "予定ID": make_schedule_candidate_id(record_id, line, rule["keyword"], candidate_idx),
                    "元記録ID": record_id,
                    "元日付": pd.to_datetime(record_date, errors="coerce").strftime("%Y-%m-%d") if not pd.isna(pd.to_datetime(record_date, errors="coerce")) else "",
                    "勤務帯": shift,
                    "利用者": user_name,
                    "利用者名": user_name,
                    "user_id": user_id,
                    "キーワード": rule["keyword"],
                    "分類": rule.get("category", "予定"),
                    "タイトル": subject,
                    "件名": subject,
                    "日時": make_datetime_display(schedule_date.strftime("%Y-%m-%d"), start_time),
                    "開始日": schedule_date.strftime("%Y-%m-%d"),
                    "開始時刻": start_time,
                    "終了日": schedule_date.strftime("%Y-%m-%d"),
                    "終了時刻": end_time,
                    "終日": all_day,
                    "場所": "",
                    "詳細": f"申し送りから抽出：{line}",
                    "内容": f"申し送りから抽出：{line}",
                    "元文章": line,
                })
    if not rows:
        return pd.DataFrame(columns=SCHEDULE_EXPORT_COLUMNS)
    result = pd.DataFrame(rows, columns=SCHEDULE_EXPORT_COLUMNS)
    return result.drop_duplicates(subset=["予定ID"]).reset_index(drop=True)


def convert_to_google_calendar_csv_df(candidate_df: pd.DataFrame) -> pd.DataFrame:
    """GoogleカレンダーCSV取込形式へ変換する。"""
    if candidate_df is None or candidate_df.empty:
        return pd.DataFrame(columns=GOOGLE_CALENDAR_CSV_COLUMNS)

    work = filter_selected_schedule_rows(candidate_df)
    if "取込対象" in work.columns:
        work = work[work["取込対象"].astype(str).str.lower().isin(["true", "1", "yes", "対象", "取込", "取込対象"])]
    if work.empty:
        return pd.DataFrame(columns=GOOGLE_CALENDAR_CSV_COLUMNS)

    out = pd.DataFrame(columns=GOOGLE_CALENDAR_CSV_COLUMNS)
    out["Subject"] = work["件名"].map(lambda x: clean_text(x, "予定"))
    # Google Calendar CSVは環境によって yyyy/mm/dd が安定しやすい
    out["Start Date"] = pd.to_datetime(work["開始日"], errors="coerce").dt.strftime("%Y/%m/%d")
    out["Start Time"] = work["開始時刻"].map(lambda x: clean_text(x))
    out["End Date"] = pd.to_datetime(work["終了日"], errors="coerce").dt.strftime("%Y/%m/%d")
    out["End Time"] = work["終了時刻"].map(lambda x: clean_text(x))
    out["All Day Event"] = work["終日"].map(lambda x: "TRUE" if clean_text(x).upper() in ["TRUE", "1", "YES", "終日"] else "FALSE")
    out["Description"] = work["内容"].map(lambda x: clean_text(x))
    out["Location"] = work["場所"].map(lambda x: clean_text(x))
    out["Private"] = "FALSE"
    return out[GOOGLE_CALENDAR_CSV_COLUMNS]


def dataframe_to_csv_bytes(df: pd.DataFrame) -> bytes:
    """Excelで文字化けしにくいUTF-8 BOM付きCSVを返す。"""
    if df is None:
        df = pd.DataFrame()
    return df.to_csv(index=False, encoding="utf-8-sig").encode("utf-8-sig")



def show_handover_keyword_master_menu():
    """申し送りから予定候補を抽出するためのキーワード設定。"""
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    st.subheader("予定抽出キーワード設定")
    st.caption("ここで使用ONにしたキーワードが、『予定候補抽出・出力』画面で自動使用されます。")

    ensure_handover_keyword_table()
    df = load_handover_keywords(active_only=False)

    if df.empty:
        df = pd.DataFrame(columns=HANDOVER_KEYWORD_COLUMNS)

    view = df.copy()
    view["使用"] = view["is_active"].astype(int).map(lambda x: True if x == 1 else False)
    view = view.rename(columns={
        "keyword": "キーワード",
        "title": "予定タイトル",
        "category": "分類",
        "sort_order": "並び順",
    })
    for col in ["使用", "キーワード", "予定タイトル", "分類", "並び順"]:
        if col not in view.columns:
            view[col] = ""
    editor_cols = ["使用", "キーワード", "予定タイトル", "分類", "並び順"]

    edited = st.data_editor(
        view[editor_cols],
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "使用": st.column_config.CheckboxColumn("使用"),
            "キーワード": st.column_config.TextColumn("キーワード", help="申し送り本文に含まれていたら予定候補にします。例：受診、訪問、外出"),
            "予定タイトル": st.column_config.TextColumn("予定タイトル", help="予定候補のタイトルに使います。空欄ならキーワードを使います。"),
            "分類": st.column_config.SelectboxColumn("分類", options=["医療", "医療・介護", "家族", "家族・相談", "外出", "生活", "予定", "その他"]),
            "並び順": st.column_config.NumberColumn("並び順", min_value=1, max_value=999, step=10),
        },
        key="handover_keyword_master_editor",
    )

    c1, c2 = st.columns(2)
    with c1:
        save_btn = st.button("予定抽出キーワード設定を保存", type="primary", use_container_width=True)
    with c2:
        reset_btn = st.button("初期キーワードを追加", use_container_width=True)

    if reset_btn:
        current = load_handover_keywords(active_only=False)
        current_keywords = set(current["keyword"].map(clean_text).tolist()) if not current.empty else set()
        now_value = format_now_jst("%Y-%m-%d %H:%M:%S")
        rows = current.to_dict("records") if not current.empty else []
        for item in DEFAULT_HANDOVER_KEYWORDS:
            if item["keyword"] in current_keywords:
                continue
            rows.append({
                "id": "kw_" + hashlib.sha1(item["keyword"].encode("utf-8")).hexdigest()[:12],
                "keyword": item["keyword"],
                "title": item.get("title", item["keyword"]),
                "category": item.get("category", "予定"),
                "is_active": int(item.get("is_active", 1)),
                "sort_order": int(item.get("sort_order", 100)),
                "created_at": now_value,
                "updated_at": now_value,
            })
        save_handover_keywords(pd.DataFrame(rows, columns=HANDOVER_KEYWORD_COLUMNS))
        st.success("初期キーワードを追加しました。")
        st.rerun()

    if save_btn:
        rows = []
        now_value = format_now_jst("%Y-%m-%d %H:%M:%S")
        old_df = load_handover_keywords(active_only=False)
        old_map = {clean_text(r.get("keyword")): r for _, r in old_df.iterrows()} if not old_df.empty else {}
        for _, row in edited.iterrows():
            keyword = clean_text(row.get("キーワード"))
            if not keyword:
                continue
            old = old_map.get(keyword, {})
            rows.append({
                "id": clean_text(old.get("id")) or "kw_" + hashlib.sha1(keyword.encode("utf-8")).hexdigest()[:12],
                "keyword": keyword,
                "title": clean_text(row.get("予定タイトル"), keyword),
                "category": clean_text(row.get("分類"), "予定"),
                "is_active": 1 if bool(row.get("使用")) else 0,
                "sort_order": safe_int(row.get("並び順"), 100),
                "created_at": clean_text(old.get("created_at"), now_value),
                "updated_at": now_value,
            })
        saved = save_handover_keywords(pd.DataFrame(rows, columns=HANDOVER_KEYWORD_COLUMNS))
        st.success(f"予定抽出キーワード設定を保存しました。有効：{int(saved['is_active'].astype(int).sum())}件。予定候補抽出に反映されます。")
        st.rerun()

    st.markdown("#### 使用中キーワード")
    active = load_handover_keywords(active_only=True)
    if active.empty:
        st.info("使用ONのキーワードはありません。")
    else:
        st.dataframe(active[["keyword", "title", "category", "sort_order"]].rename(columns={"keyword":"キーワード", "title":"予定タイトル", "category":"分類", "sort_order":"並び順"}), use_container_width=True, hide_index=True)


def show_handover_schedule_export_menu():
    """申し送りから予定候補を抽出し、管理者確認後にExcel/CSVで出力する画面。"""
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    st.subheader("申し送り予定候補の自動抽出・出力")
    st.caption("「予定抽出キーワード設定」に登録されたキーワードを使って、申し送り本文から予定候補を自動抽出します。管理者が確認・修正してから、ひだまり帳登録やカレンダー取込用データを出力します。")

    df = load_business_handover_data()
    if df.empty:
        st.info("業務全体申し送りがまだ登録されていません。")
        return

    work = df.copy()
    work["日付_dt"] = pd.to_datetime(work["日付"], errors="coerce")
    valid_dates = work["日付_dt"].dropna()
    if valid_dates.empty:
        default_start = today_jst() - timedelta(days=7)
        default_end = today_jst()
    else:
        default_start = max(valid_dates.min().date(), today_jst() - timedelta(days=30))
        default_end = valid_dates.max().date()

    active_keywords_df = load_handover_keywords(active_only=True)
    active_keywords = active_keywords_df["keyword"].tolist() if not active_keywords_df.empty else []
    if active_keywords:
        st.caption("現在の抽出キーワード：" + "、".join(active_keywords))
    else:
        st.warning("有効な予定抽出キーワードがありません。『予定抽出キーワード設定』で使用ONにしてください。")

    c1, c2, c3 = st.columns(3)
    with c1:
        start_date = st.date_input("抽出開始日", value=default_start, key="schedule_extract_start")
    with c2:
        end_date = st.date_input("抽出終了日", value=default_end, key="schedule_extract_end")
    with c3:
        keyword_filter = st.text_input("追加の本文絞り込み（任意）", placeholder="例：谷様、家族、午前", key="schedule_extract_keyword")

    candidates = extract_schedule_candidates_from_handover_df(df, start_date=start_date, end_date=end_date, keyword_filter=keyword_filter)

    if candidates.empty:
        st.info("この期間の申し送りから予定候補は見つかりませんでした。キーワードや期間を変えて確認してください。")
        return

    st.success(f"予定候補を {len(candidates)} 件抽出しました。必要に応じて修正してから出力してください。")
    st.caption("時刻が読み取れない予定は『終日予定』として出力されます。時刻が分かる場合は開始時刻・終了時刻を手入力してください。")

    editor_source = candidates.copy()
    editor_source["登録する"] = True
    editor_source["日時"] = editor_source.apply(lambda r: make_datetime_display(r.get("開始日"), r.get("開始時刻")), axis=1)
    editor_source["利用者"] = editor_source["利用者名"]
    editor_source["タイトル"] = editor_source["件名"]
    editor_source["詳細"] = editor_source["内容"]

    edited = st.data_editor(
        editor_source,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_order=["登録する", "日時", "利用者", "user_id", "分類", "タイトル", "詳細", "場所", "予定ID", "元日付", "勤務帯", "キーワード", "元文章"],
        column_config={
            "登録する": st.column_config.CheckboxColumn("登録する"),
            "日時": st.column_config.TextColumn("日時（例：2026-05-27 14:00）"),
            "利用者": st.column_config.TextColumn("利用者"),
            "user_id": st.column_config.TextColumn("利用者ID"),
            "分類": st.column_config.SelectboxColumn("分類", options=["医療", "医療・介護", "家族", "家族・相談", "外出", "生活", "予定", "その他"]),
            "タイトル": st.column_config.TextColumn("タイトル"),
            "詳細": st.column_config.TextColumn("詳細"),
            "場所": st.column_config.TextColumn("場所"),
        },
        disabled=["予定ID", "元日付", "勤務帯", "キーワード", "元文章"],
        key="schedule_candidate_editor",
    )

    normalized_edited = normalize_schedule_editor_df(edited)
    target_df = filter_selected_schedule_rows(normalized_edited)
    google_df = convert_to_google_calendar_csv_df(target_df)

    st.markdown("#### 登録・出力")
    st.caption("まず候補一覧を確認し、必要なら日時・利用者・分類・タイトル・詳細を修正してください。チェックが入った行だけ登録・出力されます。")
    d0, d1, d2, d3 = st.columns(4)
    with d0:
        if st.button("ひだまり帳へ登録", type="primary", use_container_width=True):
            count, total = register_schedules_to_hidamari(normalized_edited)
            if count > 0:
                st.success(f"ひだまり帳の内部予定に {count} 件登録・更新しました。")
            else:
                st.warning("登録対象の予定がありません。『登録する』にチェックを入れてください。")
    with d1:
        st.download_button(
            "予定候補一覧をExcelでダウンロード",
            data=to_excel_download(normalized_edited),
            file_name=f"handover_schedule_candidates_{today_jst().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with d2:
        st.download_button(
            "CSV出力",
            data=dataframe_to_csv_bytes(google_df),
            file_name=f"google_calendar_import_{today_jst().strftime('%Y%m%d')}.csv",
            mime="text/csv",
            use_container_width=True,
        )
    with d3:
        st.download_button(
            "Excel出力",
            data=to_excel_download(google_df),
            file_name=f"google_calendar_import_{today_jst().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    st.markdown("#### ひだまり帳 内部予定一覧")
    schedules = load_hidamari_schedules()
    if schedules.empty:
        st.info("内部予定はまだ登録されていません。")
    else:
        view = schedules.copy()
        view["予定日_dt"] = pd.to_datetime(view["予定日"], errors="coerce")
        view = view.sort_values(["予定日_dt", "開始時刻"], ascending=[False, True]).drop(columns=["予定日_dt"], errors="ignore")
        st.dataframe(view.head(100), use_container_width=True, hide_index=True)

    with st.expander("Googleカレンダー等へ読み込む際の注意", expanded=False):
        st.markdown(
            """
            - まず「予定候補一覧」で内容を確認してください。
            - 時刻が空欄の予定は終日予定として扱います。
            - Googleカレンダーへ取り込む場合は、CSVファイルを使用してください。
            - 施設運用では、いきなり自動登録せず、管理者確認後に取り込む方式が安全です。
            """
        )


def render_business_handover_card(row):
    priority = clean_text(row.get("優先度", "通常"), "通常")
    icon = "【至急】" if priority == "至急" else "【注意】" if priority == "注意" else "【通常】"

    record_date = row.get("日付", "")
    if not isinstance(record_date, str):
        try:
            record_date = pd.to_datetime(record_date).strftime("%Y-%m-%d")
        except Exception:
            record_date = ""

    auto_text = clean_text(row.get("Excel自動抽出情報"), "記載なし").replace(chr(10), "<br>")

    st.markdown(
        f"""
        <div style="
            border:1px solid #e0d6c8;
            border-radius:14px;
            padding:14px 16px;
            margin-bottom:12px;
            background-color:#fffdf7;
        ">
        <b>{icon} {record_date}｜{clean_text(row.get('勤務帯'))}｜{clean_text(row.get('記入者'))}｜対象：{make_handover_target_label(row.get('user_id'), row.get('利用者名'), row.get('対象区分'))}｜{priority}｜{clean_text(row.get('対応状況'))}</b><br><br>
        <b>全体申し送り</b><br>
        {clean_text(row.get('全体申し送り'), '記載なし').replace(chr(10), '<br>')}<br><br>
        <b>要確認事項</b><br>
        {clean_text(row.get('要確認事項'), '記載なし').replace(chr(10), '<br>')}<br><br>
        <b>Excel自動抽出情報</b><br>
        <div style="background-color:#eef7ff;border-left:4px solid #5ca8d8;padding:10px;margin-top:6px;">
        {auto_text}
        </div><br>
        <span style="font-size:12px;color:#666;">記録日時：{clean_text(row.get('記録日時'))}</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    input_excel_path = clean_text(row.get("入力Excelファイル"))
    input_excel_text = clean_text(row.get("入力Excel表示情報"))
    if input_excel_path or input_excel_text:
        with st.expander("入力Excelデータを表示", expanded=True):
            show_business_handover_excel_preview(input_excel_path, input_excel_text)

    photo1 = clean_text(row.get("写真1"))
    photo2 = clean_text(row.get("写真2"))  # 旧データ互換用
    if photo1:
        show_business_handover_photo(photo1, "添付写真")
    elif photo2:
        show_business_handover_photo(photo2, "添付写真（旧写真2）")



def show_business_handover_menu():
    st.header("申し送りを書く・確認する")
    show_observation_perspective("handover")
    st.caption("次の勤務者に伝える出来事・注意点・未対応を残します。")

    # Ver5.0.1 修正：st.tabs だと環境やCSSの影響で「選択しても画面が動かない」ことがあるため、
    # 確実に再描画される radio 方式に変更。
    handover_mode = st.radio(
        "表示する機能",
        ["新規登録", "検索・更新・削除", "予定候補抽出・出力", "予定抽出キーワード設定"],
        horizontal=True,
        key="business_handover_mode_radio",
    )

    if handover_mode == "新規登録":
        df = load_business_handover_data()

        # 先に本日の申し送りを表示し、その下に対応中案件、その下に入力欄を置く
        st.subheader("本日の業務全体申し送り")
        today_df = get_business_handover_by_date(df, today_jst())
        if today_df.empty:
            st.info("本日の業務全体申し送りはまだありません。")
        else:
            for _, row in today_df.iterrows():
                render_business_handover_card(row)

        st.divider()
        show_business_handover_in_progress_section(df)

        st.divider()
        st.subheader("申し送りを入力")

        with st.form("business_handover_form", clear_on_submit=False):
            c1, c2, c3, c4 = st.columns(4)
            with c1:
                record_date = st.date_input("日付", value=today_jst(), key="business_handover_date")
            with c2:
                shift_type = st.selectbox("勤務帯", ["日勤", "夜勤"], index=0, key="business_handover_shift")
            with c3:
                staff_name = st.text_input("記入者", placeholder="例：藤野", key="business_handover_staff")
            with c4:
                target_options = build_handover_target_options()
                target_selection = st.selectbox("対象", target_options, index=0, key="business_handover_target")

            st.markdown("#### 1. 内容")
            st.caption("見たこと、気づき、次に見ることを分けて残します。")
            fact_note = st.text_area(
                "事実（見たこと・起きたこと）",
                height=100,
                placeholder="例：共有スペースの床が濡れていた。15時頃に来客あり。物品残数が少ない。",
                key="business_handover_fact_note",
            )
            insight_note = st.text_area(
                "気づき（普段との違い・気になったこと）",
                height=90,
                placeholder="例：いつもより表情が硬い、動線が混みやすい、確認が必要そう。",
                key="business_handover_insight_note",
            )
            next_note = st.text_area(
                "次に見ること（次勤務者への確認ポイント）",
                height=90,
                placeholder="例：床の状態を再確認、物品補充、家族連絡の有無を確認。",
                key="business_handover_next_note",
            )
            overall_note = format_handover_structured_note(fact_note, insight_note, next_note)

            check_note = st.text_area(
                "要確認事項（未対応・確認中のこと）",
                height=100,
                placeholder="例：明日の往診時間確認、家族連絡の確認、物品残数確認など",
                key="business_handover_check_note",
            )

            c4, c5 = st.columns(2)
            with c4:
                priority = st.selectbox("優先度", ["通常", "注意", "至急"], index=0, key="business_handover_priority")
            with c5:
                status = st.selectbox("対応状況", ["未対応", "対応中", "対応済"], index=0, key="business_handover_status")

            st.markdown("#### 2. 写真添付")
            st.caption("必要なときだけ写真を1枚添付します。")
            photo1_file = st.file_uploader("写真を1枚添付", type=["jpg", "jpeg", "png", "webp"], key="business_handover_photo1")
            photo2_file = None  # 写真1枚運用。旧カラム互換のため変数のみ残す。
            if photo1_file is not None:
                render_photo_compression_preview(photo1_file)

            st.markdown("#### 3. Excel・CSV添付")
            input_excel_file = st.file_uploader(
                "入力済みのExcel・CSVを添付して申し送り内に表示",
                type=["xlsx", "xls", "csv"],
                key="business_handover_input_excel",
            )
            if input_excel_file is not None:
                input_excel_display_text = build_uploaded_excel_display_text(input_excel_file)
                st.info(input_excel_display_text)
                _, input_excel_preview_df = read_uploaded_excel_preview(input_excel_file, max_rows=10)
                if not input_excel_preview_df.empty:
                    st.dataframe(input_excel_preview_df, use_container_width=True, hide_index=True)
            else:
                input_excel_display_text = ""

            auto_extract_text = build_business_handover_auto_extract_text(record_date)
            st.markdown("#### 4. 自動抽出情報")
            st.info(auto_extract_text)

            submitted = st.form_submit_button("申し送りを保存する", use_container_width=True)

        if submitted:
            if not clean_text(staff_name):
                st.warning("記入者を入力してください。")
                st.stop()

            if not clean_text(overall_note) and not clean_text(check_note):
                st.warning("申し送り内容、または要確認事項を入力してください。")
                st.stop()

            target_type, target_user_id, target_user_name = resolve_handover_target(target_selection)

            record_id = make_business_handover_id(record_date, shift_type, staff_name)
            photo1_path = save_business_handover_photo(photo1_file, record_id, 1)
            photo2_path = ""  # Ver5.0 レベル1では写真1枚運用
            input_excel_path = save_business_handover_excel(input_excel_file, record_id)
            input_excel_display_text = build_uploaded_excel_display_text(input_excel_file) if input_excel_file is not None else ""
            auto_extract_text = build_business_handover_auto_extract_text(record_date)

            new_record = {
                "記録ID": record_id,
                "日付": record_date,
                "勤務帯": shift_type,
                "記入者": clean_text(staff_name),
                "対象区分": target_type,
                "user_id": target_user_id,
                "利用者名": target_user_name,
                "全体申し送り": clean_text(overall_note),
                "要確認事項": clean_text(check_note),
                "優先度": priority,
                "対応状況": status,
                "写真1": photo1_path,
                "写真2": photo2_path,
                "Excel自動抽出情報": auto_extract_text,
                "入力Excelファイル": input_excel_path,
                "入力Excel表示情報": input_excel_display_text,
                "記録日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
            }

            df = pd.concat([df, pd.DataFrame([new_record], columns=BUSINESS_HANDOVER_COLUMNS)], ignore_index=True)
            save_business_handover_data(df)
            st.success("申し送りを保存しました。未対応の内容は次の勤務者にも確認してください。")
            st.rerun()

    if handover_mode == "検索・更新・削除":
        st.subheader("申し送りの検索・修正")
        df = load_business_handover_data()

        if df.empty:
            st.info("まだ業務全体申し送りは登録されていません。")
            return

        work = df.copy()
        work["日付"] = pd.to_datetime(work["日付"], errors="coerce")

        valid_dates = work["日付"].dropna()
        if valid_dates.empty:
            default_start = today_jst() - timedelta(days=7)
            default_end = today_jst()
        else:
            default_start = valid_dates.min().date()
            default_end = valid_dates.max().date()

        c1, c2, c3 = st.columns(3)
        with c1:
            start_date = st.date_input("開始日", value=default_start, key="bh_search_start_date")
        with c2:
            end_date = st.date_input("終了日", value=default_end, key="bh_search_end_date")
        with c3:
            keyword = st.text_input("キーワード検索", placeholder="記入者・本文・要確認事項", key="bh_search_keyword")

        c4, c5, c6 = st.columns(3)
        with c4:
            shift_filter = st.selectbox("勤務帯", ["すべて", "日勤", "夜勤"], key="bh_search_shift")
        with c5:
            status_filter = st.selectbox("対応状況", ["すべて", "未対応", "対応中", "対応済"], key="bh_search_status")
        with c6:
            priority_filter = st.selectbox("優先度", ["すべて", "通常", "注意", "至急"], key="bh_search_priority")

        filtered = work[
            (work["日付"].dt.date >= start_date)
            & (work["日付"].dt.date <= end_date)
        ].copy()

        if shift_filter != "すべて":
            filtered = filtered[filtered["勤務帯"] == shift_filter]
        if status_filter != "すべて":
            filtered = filtered[filtered["対応状況"] == status_filter]
        if priority_filter != "すべて":
            filtered = filtered[filtered["優先度"] == priority_filter]

        keyword = clean_text(keyword)
        if keyword:
            search_text = (
                filtered["記入者"].fillna("").astype(str)
                + " " + filtered.get("対象区分", "").fillna("").astype(str)
                + " " + filtered.get("user_id", "").fillna("").astype(str)
                + " " + filtered.get("利用者名", "").fillna("").astype(str)
                + " " + filtered["全体申し送り"].fillna("").astype(str)
                + " " + filtered["要確認事項"].fillna("").astype(str)
            )
            filtered = filtered[search_text.str.contains(keyword, case=False, na=False)]

        if not filtered.empty:
            filtered["_sort_dt"] = pd.to_datetime(filtered["記録日時"], errors="coerce")
            filtered = filtered.sort_values(["日付", "_sort_dt"], ascending=[False, False]).drop(columns=["_sort_dt"])

        st.caption(f"検索結果：{len(filtered)}件")

        if filtered.empty:
            st.info("条件に合う申し送りはありません。")
            return

        display_df = filtered.copy()
        display_df["日付"] = pd.to_datetime(display_df["日付"], errors="coerce").dt.strftime("%Y-%m-%d")
        display_cols = ["日付", "勤務帯", "記入者", "対象区分", "user_id", "利用者名", "優先度", "対応状況", "全体申し送り", "要確認事項", "Excel自動抽出情報", "入力Excel表示情報", "記録日時"]
        st.dataframe(
            display_df[[c for c in display_cols if c in display_df.columns]],
            use_container_width=True,
            hide_index=True,
        )

        st.divider()
        st.subheader("選択した申し送りの更新・削除")

        def make_select_label(row):
            d = pd.to_datetime(row.get("日付"), errors="coerce")
            d_text = d.strftime("%Y-%m-%d") if not pd.isna(d) else "日付不明"
            note = clean_text(row.get("全体申し送り", "")) or clean_text(row.get("要確認事項", ""))
            if len(note) > 24:
                note = note[:24] + "..."
            return f"{d_text}｜{clean_text(row.get('勤務帯'))}｜{clean_text(row.get('記入者'))}｜{clean_text(row.get('優先度'))}｜{clean_text(row.get('対応状況'))}｜{note}"

        select_options = []
        label_to_id = {}
        for _, row in filtered.iterrows():
            label = make_select_label(row)
            rid = clean_text(row.get("記録ID"))
            if not rid:
                continue
            # 同じラベルがあっても選べるように末尾へIDの一部を付ける
            unique_label = f"{label}｜ID:{rid[-6:]}"
            select_options.append(unique_label)
            label_to_id[unique_label] = rid

        selected_label = st.selectbox("編集・削除する申し送りを選択", select_options, key="bh_selected_record")
        selected_id = label_to_id.get(selected_label)

        selected_rows = df[df["記録ID"].astype(str) == str(selected_id)]
        if selected_rows.empty:
            st.warning("選択した記録が見つかりません。")
            return

        selected_row = selected_rows.iloc[0]
        render_business_handover_card(selected_row)

        selected_date = pd.to_datetime(selected_row.get("日付"), errors="coerce")
        if pd.isna(selected_date):
            selected_date_value = today_jst()
        else:
            selected_date_value = selected_date.date()

        shift_options = ["日勤", "夜勤"]
        priority_options = ["通常", "注意", "至急"]
        status_options = ["未対応", "対応中", "対応済"]

        with st.form("business_handover_update_form", clear_on_submit=False):
            u1, u2, u3 = st.columns(3)
            with u1:
                update_date = st.date_input("日付", value=selected_date_value, key="bh_update_date")
            with u2:
                update_shift = st.selectbox(
                    "勤務帯",
                    shift_options,
                    index=shift_options.index(clean_text(selected_row.get("勤務帯"), "日勤")) if clean_text(selected_row.get("勤務帯"), "日勤") in shift_options else 0,
                    key="bh_update_shift",
                )
            with u3:
                update_staff = st.text_input("記入者", value=clean_text(selected_row.get("記入者")), key="bh_update_staff")

            update_target_options = build_handover_target_options()
            current_target_label = make_handover_target_label(selected_row.get("user_id"), selected_row.get("利用者名"), selected_row.get("対象区分"))
            if current_target_label not in update_target_options:
                update_target_options = [current_target_label] + update_target_options
            update_target_selection = st.selectbox(
                "対象",
                update_target_options,
                index=update_target_options.index(current_target_label) if current_target_label in update_target_options else 0,
                key="bh_update_target",
            )

            update_overall = st.text_area(
                "全体申し送り",
                value=clean_text(selected_row.get("全体申し送り")),
                height=150,
                key="bh_update_overall",
            )
            update_check = st.text_area(
                "要確認事項",
                value=clean_text(selected_row.get("要確認事項")),
                height=120,
                key="bh_update_check",
            )

            u4, u5 = st.columns(2)
            with u4:
                update_priority = st.selectbox(
                    "優先度",
                    priority_options,
                    index=priority_options.index(clean_text(selected_row.get("優先度"), "通常")) if clean_text(selected_row.get("優先度"), "通常") in priority_options else 0,
                    key="bh_update_priority",
                )
            with u5:
                update_status = st.selectbox(
                    "対応状況",
                    status_options,
                    index=status_options.index(clean_text(selected_row.get("対応状況"), "未対応")) if clean_text(selected_row.get("対応状況"), "未対応") in status_options else 0,
                    key="bh_update_status",
                )

            st.markdown("#### 写真添付の更新")
            up1, up2 = st.columns(2)
            with up1:
                update_photo1_file = st.file_uploader("写真1を差し替える", type=["jpg", "jpeg", "png", "webp"], key="bh_update_photo1")
                remove_photo1 = st.checkbox("写真1を削除", key="bh_remove_photo1")
            with up2:
                update_photo2_file = st.file_uploader("写真2を差し替える", type=["jpg", "jpeg", "png", "webp"], key="bh_update_photo2")
                remove_photo2 = st.checkbox("写真2を削除", key="bh_remove_photo2")

            st.markdown("#### 入力Excelデータの更新")
            current_excel_path = clean_text(selected_row.get("入力Excelファイル"))
            current_excel_text = clean_text(selected_row.get("入力Excel表示情報"))
            if current_excel_path or current_excel_text:
                show_business_handover_excel_preview(current_excel_path, current_excel_text)
            update_input_excel_file = st.file_uploader(
                "入力Excelを差し替える",
                type=["xlsx", "xls", "csv"],
                key="bh_update_input_excel",
            )
            remove_input_excel = st.checkbox("入力Excelを削除", key="bh_remove_input_excel")
            if update_input_excel_file is not None:
                update_input_excel_display_text = build_uploaded_excel_display_text(update_input_excel_file)
                st.info(update_input_excel_display_text)
                _, update_input_excel_preview_df = read_uploaded_excel_preview(update_input_excel_file, max_rows=10)
                if not update_input_excel_preview_df.empty:
                    st.dataframe(update_input_excel_preview_df, use_container_width=True, hide_index=True)
            else:
                update_input_excel_display_text = current_excel_text

            update_auto_extract_text = build_business_handover_auto_extract_text(update_date)
            st.markdown("#### Excel自動抽出情報（更新時に再作成）")
            st.info(update_auto_extract_text)

            update_submitted = st.form_submit_button("この申し送りを更新する", use_container_width=True)

        if update_submitted:
            if not clean_text(update_staff):
                st.warning("記入者を入力してください。")
                st.stop()
            if not clean_text(update_overall) and not clean_text(update_check):
                st.warning("全体申し送り、または要確認事項を入力してください。")
                st.stop()

            df_update = load_business_handover_data()
            mask = df_update["記録ID"].astype(str) == str(selected_id)
            if not mask.any():
                st.error("更新対象の記録が見つかりません。")
                st.stop()

            update_target_type, update_user_id, update_user_name = resolve_handover_target(update_target_selection)

            df_update.loc[mask, "日付"] = pd.to_datetime(update_date)
            df_update.loc[mask, "勤務帯"] = update_shift
            df_update.loc[mask, "記入者"] = clean_text(update_staff)
            df_update.loc[mask, "対象区分"] = update_target_type
            df_update.loc[mask, "user_id"] = update_user_id
            df_update.loc[mask, "利用者名"] = update_user_name
            df_update.loc[mask, "全体申し送り"] = clean_text(update_overall)
            df_update.loc[mask, "要確認事項"] = clean_text(update_check)
            df_update.loc[mask, "優先度"] = update_priority
            df_update.loc[mask, "対応状況"] = update_status

            current_photo1 = clean_text(selected_row.get("写真1"))
            current_photo2 = clean_text(selected_row.get("写真2"))
            if remove_photo1:
                current_photo1 = ""
            elif update_photo1_file is not None:
                current_photo1 = save_business_handover_photo(update_photo1_file, selected_id, 1)
            if remove_photo2:
                current_photo2 = ""
            elif update_photo2_file is not None:
                current_photo2 = save_business_handover_photo(update_photo2_file, selected_id, 2)

            current_input_excel = clean_text(selected_row.get("入力Excelファイル"))
            current_input_excel_text = clean_text(selected_row.get("入力Excel表示情報"))
            if remove_input_excel:
                current_input_excel = ""
                current_input_excel_text = ""
            elif update_input_excel_file is not None:
                current_input_excel = save_business_handover_excel(update_input_excel_file, selected_id)
                current_input_excel_text = build_uploaded_excel_display_text(update_input_excel_file)

            df_update.loc[mask, "写真1"] = current_photo1
            df_update.loc[mask, "写真2"] = current_photo2
            df_update.loc[mask, "Excel自動抽出情報"] = update_auto_extract_text
            df_update.loc[mask, "入力Excelファイル"] = current_input_excel
            df_update.loc[mask, "入力Excel表示情報"] = current_input_excel_text
            df_update.loc[mask, "記録日時"] = format_now_jst("%Y-%m-%d %H:%M:%S")

            save_business_handover_data(df_update)
            st.success("業務全体申し送りを更新しました。")
            st.rerun()

        st.divider()
        st.subheader("削除")
        st.warning("削除すると、この申し送り記録は一覧と管理者ダッシュボードから消えます。")
        confirm_delete = st.checkbox("この申し送りを削除することを確認しました", key="bh_confirm_delete")
        if st.button("この申し送りを削除する", type="primary", disabled=not confirm_delete, use_container_width=True, key="bh_delete_button"):
            result = delete_business_handover_record(selected_id, source="業務全体申し送り画面から削除")
            if result.get("ok"):
                st.success(
                    f"業務全体申し送りを削除しました。"
                    f" SQLite削除:{result.get('sqlite_deleted', 0)} / Supabase削除:{result.get('supabase_deleted', 0)}"
                )
                st.rerun()
            else:
                if result.get("error"):
                    st.error(f"削除に失敗しました：{result.get('error')}")
                else:
                    st.error("削除対象が見つかりません。SupabaseとSQLiteの記録IDを確認してください。")

    if handover_mode == "予定候補抽出・出力":
        show_handover_schedule_export_menu()

    if handover_mode == "予定抽出キーワード設定":
        show_handover_keyword_master_menu()

def show_admin_business_handover_summary(target_date):
    st.subheader("業務全体申し送り")
    st.caption("出勤時に最初に確認する項目です。確認日の申し送りと、未対応・至急の申し送りを表示します。")

    df = load_business_handover_data()

    if df.empty:
        st.info("業務全体申し送りはまだ登録されていません。")
        return

    target_df = get_business_handover_by_date(df, target_date)
    alert_df = get_business_handover_alerts(df)

    tab1, tab2 = st.tabs(["確認日の申し送り", "未対応・至急"])

    with tab1:
        if target_df.empty:
            st.info("確認日の業務全体申し送りはありません。")
        else:
            for _, row in target_df.iterrows():
                render_business_handover_card(row)

    with tab2:
        if alert_df.empty:
            st.success("未対応・至急の業務全体申し送りはありません。")
        else:
            for _, row in alert_df.iterrows():
                render_business_handover_card(row)




# =========================
# 短期目標・モニタリング機能
# =========================
def ensure_short_goal_files():
    """互換用。実データはSQLiteへ保存します。既存Excelは初回のみ移行します。"""
    migrate_excel_to_sqlite_if_needed(
        SQLITE_TABLE_SHORT_GOAL_MASTER,
        SHORT_GOAL_MASTER_FILE,
        "短期目標マスタ",
        SHORT_GOAL_MASTER_COLUMNS,
        date_cols=["開始日", "終了予定日"],
        unique_cols=["目標ID"],
    )
    migrate_excel_to_sqlite_if_needed(
        SQLITE_TABLE_SHORT_GOAL_CHECKS,
        SHORT_GOAL_CHECK_FILE,
        "実施チェック",
        SHORT_GOAL_CHECK_COLUMNS,
        date_cols=["日付"],
        unique_cols=["記録ID"],
    )
    migrate_excel_to_sqlite_if_needed(
        SQLITE_TABLE_MONITORING_DRAFTS,
        MONITORING_DRAFT_FILE,
        "モニタリング下書き",
        MONITORING_DRAFT_COLUMNS,
        date_cols=["作成日"],
        unique_cols=["下書きID"],
    )


def read_excel_safe(path: Path, columns: list, sheet_name: str | None = None) -> pd.DataFrame:
    """互換用：既存の呼び出しが残っていても安全にExcelを読めるよう残す。"""
    ensure_dirs()
    if not path.exists():
        return pd.DataFrame(columns=columns)
    try:
        if sheet_name:
            df = pd.read_excel(path, sheet_name=sheet_name)
        else:
            df = pd.read_excel(path)
    except Exception:
        return pd.DataFrame(columns=columns)
    for col in columns:
        if col not in df.columns:
            df[col] = ""
    return df[columns].astype("object")



def save_excel_safe(df: pd.DataFrame, path: Path, columns: list, sheet_name: str):
    """
    互換用。
    商品版ではExcelファイルを正データとして保存しません。
    永続保存が必要なデータは各save_*関数でSQLiteへ保存してください。
    """
    return normalize_df_columns(df, columns)

@cache_safe_master_read(ttl=SAFE_READ_CACHE_TTL_SEC)
def load_short_goal_master():
    with perf_timer("load_short_goal_master"):
        ensure_short_goal_files()
        df = load_sqlite_table(
            SQLITE_TABLE_SHORT_GOAL_MASTER,
            SHORT_GOAL_MASTER_COLUMNS,
            date_cols=["開始日", "終了予定日"],
        )
        return attach_user_ids(df)


def save_short_goal_master(df):
    df = normalize_df_columns(df, SHORT_GOAL_MASTER_COLUMNS)
    df = attach_user_ids(df)
    save_sqlite_table(
        df,
        SQLITE_TABLE_SHORT_GOAL_MASTER,
        SHORT_GOAL_MASTER_COLUMNS,
        date_cols=["開始日", "終了予定日"],
        unique_cols=["目標ID"],
    )
    clear_hidamari_read_cache("短期目標マスタ保存")
    add_audit_log("保存", SQLITE_TABLE_SHORT_GOAL_MASTER, "", "短期目標マスタを保存しました")


def load_short_goal_checks(start_date=None, end_date=None, recent_days=None):
    ensure_short_goal_files()
    if recent_days and start_date is None and end_date is None:
        start_date = recent_start_date(recent_days)
        end_date = today_jst()
    if supabase_is_enabled():
        df = supabase_read_table(SQLITE_TABLE_SHORT_GOAL_CHECKS, SHORT_GOAL_CHECK_COLUMNS, date_field="日付", start_date=start_date, end_date=end_date)
    else:
        df = load_sqlite_table(
            SQLITE_TABLE_SHORT_GOAL_CHECKS,
            SHORT_GOAL_CHECK_COLUMNS,
            date_cols=["日付"],
        )
        df = _filter_df_by_date_range(df, "日付", start_date, end_date)
    return attach_user_ids(df)


def save_short_goal_checks(df):
    df = normalize_df_columns(df, SHORT_GOAL_CHECK_COLUMNS)
    df = attach_user_ids(df)
    save_sqlite_table(
        df,
        SQLITE_TABLE_SHORT_GOAL_CHECKS,
        SHORT_GOAL_CHECK_COLUMNS,
        date_cols=["日付"],
        unique_cols=["記録ID"],
    )
    clear_hidamari_read_cache("短期目標実施チェック保存")
    add_audit_log("保存", SQLITE_TABLE_SHORT_GOAL_CHECKS, "", "短期目標実施記録を保存しました")


def load_monitoring_drafts():
    ensure_short_goal_files()
    df = load_sqlite_table(
        SQLITE_TABLE_MONITORING_DRAFTS,
        MONITORING_DRAFT_COLUMNS,
        date_cols=["作成日"],
    )
    return attach_user_ids(df)


def save_monitoring_drafts(df):
    df = normalize_df_columns(df, MONITORING_DRAFT_COLUMNS)
    df = attach_user_ids(df)
    save_sqlite_table(
        df,
        SQLITE_TABLE_MONITORING_DRAFTS,
        MONITORING_DRAFT_COLUMNS,
        date_cols=["作成日"],
        unique_cols=["下書きID"],
    )


def to_excel_download(df: pd.DataFrame) -> bytes:
    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="出力データ")
    return output.getvalue()


def ym_str(d: date):
    return d.strftime("%Y-%m")


def get_active_user_names():
    users = load_active_user_names(include_hidden=False)
    return users if users else DEFAULT_USERS


def _short_goal_text_join(values, limit=6):
    """短期目標サマリー用に、空欄を除いて重複をまとめる。"""
    items = []
    try:
        iterable = list(values)
    except Exception:
        iterable = []
    for value in iterable:
        text = clean_text(value) if "clean_text" in globals() else str(value or "").strip()
        if not text or text.lower() in ["nan", "none", "nat"]:
            continue
        if text not in items:
            items.append(text)
    return "\n".join([f"・{x}" for x in items[:limit]]) if items else "記録なし"


def _build_short_goal_rule_summary(view_df: pd.DataFrame) -> dict:
    """AI未設定時でも止まらない、記録ベースの要約。"""
    if view_df is None or view_df.empty:
        return {
            "理由要約": "対象期間の記録がないため、未実施・一部実施の理由は確認できません。",
            "職員メモ要約": "対象期間の職員メモはありません。",
            "総括コメント": "対象期間の記録がないため、総括コメントは作成できません。",
        }

    reasons = []
    for _, row in view_df.iterrows():
        status = clean_text(row.get("実施状況"))
        reason = clean_text(row.get("未実施理由"))
        if status in ["未実施", "一部実施"] and reason:
            day = clean_text(row.get("日付"))
            reasons.append(f"{day}（{status}）：{reason}")

    memos = []
    for _, row in view_df.iterrows():
        memo = clean_text(row.get("職員メモ"))
        if memo:
            day = clean_text(row.get("日付"))
            status = clean_text(row.get("実施状況"))
            memos.append(f"{day}（{status}）：{memo}")

    total = len(view_df)
    done = int((view_df.get("実施状況", pd.Series(dtype=str)).astype(str) == "実施").sum()) if "実施状況" in view_df.columns else 0
    partial = int((view_df.get("実施状況", pd.Series(dtype=str)).astype(str) == "一部実施").sum()) if "実施状況" in view_df.columns else 0
    not_done = int((view_df.get("実施状況", pd.Series(dtype=str)).astype(str) == "未実施").sum()) if "実施状況" in view_df.columns else 0
    rate = round(((done + partial * 0.5) / total) * 100, 1) if total else 0

    if total:
        general = f"対象期間は{total}件の記録があり、実施{done}件、一部実施{partial}件、未実施{not_done}件、実施状況率は{rate}%です。"
        if not_done or partial:
            general += " 未実施・一部実施の理由と職員メモを確認し、次回の声かけや支援方法の調整に活用してください。"
        else:
            general += " 大きな未実施理由は目立たず、現在の支援の流れを継続しながら観察できます。"
    else:
        general = "対象期間の記録がないため、総括コメントは作成できません。"

    return {
        "理由要約": _short_goal_text_join(reasons, limit=8) if reasons else "未実施・一部実施の理由は記録されていません。",
        "職員メモ要約": _short_goal_text_join(memos, limit=8) if memos else "職員メモは記録されていません。",
        "総括コメント": general,
    }


def generate_ai_short_goal_summary(user_name, goal_text, start_date, end_date, view_df: pd.DataFrame):
    """短期目標の実施記録をAIで短く要約する。API未設定時は通常要約を返す。"""
    fallback = _build_short_goal_rule_summary(view_df)
    api_key = get_openai_api_key("") if "get_openai_api_key" in globals() else ""
    if not api_key:
        return fallback, "OpenAI APIキー未設定のため、通常要約を表示しています。"
    try:
        from openai import OpenAI
    except Exception:
        return fallback, "openaiライブラリが未インストールのため、通常要約を表示しています。"

    try:
        source_cols = ["日付", "利用者名", "短期目標", "実施状況", "本人の様子", "未実施理由", "職員メモ", "入力職員"]
        work = view_df.copy()
        for col in source_cols:
            if col not in work.columns:
                work[col] = ""
        records = work[source_cols].fillna("").astype(str).to_dict(orient="records")
        prompt = f"""
あなたは介護施設の短期目標モニタリング記録の文章整理係です。
医療判断・診断・治療効果の断定は禁止です。
記録に基づき、未実施理由・一部実施理由と職員メモを、管理者が確認しやすい短い文章に整理してください。
推測で事実を追加しないでください。
出力はJSONのみです。

【対象】
利用者：{user_name}
短期目標：{goal_text}
期間：{start_date}〜{end_date}

【記録】
{records}

JSON形式：
{{
  "理由要約": "未実施・一部実施の理由を2〜4文で整理。理由がなければ、記録なしと書く。",
  "職員メモ要約": "職員メモから本人の様子や支援上の注意点を2〜4文で整理。記録がなければ、記録なしと書く。",
  "総括コメント": "実施状況率と記録の傾向を踏まえ、次の確認につながる総括を2〜4文で整理。断定や診断はしない。"
}}
"""
        client = OpenAI(api_key=api_key)
        res = client.chat.completions.create(
            model=get_openai_model("short_goal", "gpt-4o-mini"),
            messages=[
                {"role": "system", "content": "介護記録を断定せず、記録に基づいて短く整理します。"},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
            response_format={"type": "json_object"},
        )
        data = json.loads(res.choices[0].message.content or "{}")
        return {
            "理由要約": clean_text(data.get("理由要約"), fallback["理由要約"]),
            "職員メモ要約": clean_text(data.get("職員メモ要約"), fallback["職員メモ要約"]),
            "総括コメント": clean_text(data.get("総括コメント"), fallback.get("総括コメント", "")),
        }, "AI要約を表示しています。"
    except Exception as e:
        return fallback, f"AI要約中にエラーが出たため、通常要約を表示しています：{e}"


# =========================
# 短期目標AI要約保存（ボタン実行方式）
# =========================
def make_short_goal_ai_summary_setting_key(signature: str) -> str:
    """利用者・目標・期間・件数などの条件から、保存済みAI要約のキーを作る。"""
    raw = clean_text(signature) if "clean_text" in globals() else str(signature or "").strip()
    digest = hashlib.sha256(raw.encode("utf-8")).hexdigest()[:24]
    return f"short_goal_ai_summary::{digest}"


def load_saved_short_goal_ai_summary(signature: str):
    """保存済みの短期目標AI要約をapp_settingsから取得する。"""
    try:
        key = make_short_goal_ai_summary_setting_key(signature)
        data = get_app_setting(key, None) if "get_app_setting" in globals() else None
        if not isinstance(data, dict):
            return None
        summary = data.get("summary")
        if not isinstance(summary, dict):
            return None
        return data
    except Exception:
        return None


def save_short_goal_ai_summary(signature: str, summary: dict, note: str = ""):
    """短期目標AI要約を保存する。次回同条件ではAIを再実行せず保存済みを表示する。"""
    try:
        key = make_short_goal_ai_summary_setting_key(signature)
        data = {
            "signature": signature,
            "summary": summary if isinstance(summary, dict) else {},
            "note": clean_text(note) if "clean_text" in globals() else str(note or ""),
            "saved_at": format_now_jst("%Y-%m-%d %H:%M:%S") if "format_now_jst" in globals() else datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "saved_by": current_login_user() if "current_login_user" in globals() else "",
        }
        if "set_app_setting" in globals():
            set_app_setting(
                key,
                data,
                category="短期目標AI要約",
                description="短期目標モニタリング画面の保存済みAI要約",
            )
        return data
    except Exception:
        return None


def make_short_goal_summary_excel_bytes(
    selected_user,
    selected_goal_text,
    selected_support_text,
    start_date,
    end_date,
    total,
    done,
    partial,
    not_done,
    rate,
    done_only_rate,
    summary: dict,
    work: pd.DataFrame,
):
    """選択した利用者・短期目標の実施状況とAI要約をExcel化する。"""
    output = BytesIO()
    reason_summary = clean_text(summary.get("理由要約"), "記録なし") if isinstance(summary, dict) else "記録なし"
    memo_summary = clean_text(summary.get("職員メモ要約"), "記録なし") if isinstance(summary, dict) else "記録なし"
    general_summary = clean_text(summary.get("総括コメント"), "記録なし") if isinstance(summary, dict) else "記録なし"

    summary_rows = [
        {"項目": "利用者名", "内容": selected_user},
        {"項目": "短期目標", "内容": selected_goal_text},
        {"項目": "支援内容", "内容": selected_support_text},
        {"項目": "集計期間", "内容": f"{start_date}〜{end_date}"},
        {"項目": "記録件数", "内容": total},
        {"項目": "実施状況率", "内容": f"{rate}%"},
        {"項目": "実施のみ率", "内容": f"{done_only_rate}%"},
        {"項目": "実施回数", "内容": done},
        {"項目": "一部実施回数", "内容": partial},
        {"項目": "未実施回数", "内容": not_done},
        {"項目": "未実施理由・一部実施の理由（AI要約）", "内容": reason_summary},
        {"項目": "職員メモ（AI要約）", "内容": memo_summary},
        {"項目": "総括コメント", "内容": general_summary},
        {"項目": "出力日時", "内容": format_now_jst("%Y-%m-%d %H:%M:%S") if "format_now_jst" in globals() else datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
    ]
    summary_df = pd.DataFrame(summary_rows)

    detail = work.copy() if work is not None else pd.DataFrame()
    detail_cols = ["日付", "利用者名", "短期目標", "実施状況", "本人の様子", "未実施理由", "職員メモ", "入力職員", "登録日時"]
    for col in detail_cols:
        if col not in detail.columns:
            detail[col] = ""
    if "日付_dt" in detail.columns:
        detail = detail.sort_values("日付_dt", ascending=True)
    detail_df = detail[detail_cols].fillna("")

    status_df = pd.DataFrame([
        {"実施状況": "実施", "件数": done},
        {"実施状況": "一部実施", "件数": partial},
        {"実施状況": "未実施", "件数": not_done},
    ])

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="要約", index=False)
        status_df.to_excel(writer, sheet_name="実施状況", index=False)
        detail_df.to_excel(writer, sheet_name="詳細記録", index=False)

        try:
            from openpyxl.styles import Alignment, Font, PatternFill
            wb = writer.book
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="EAF4EF")
                for col in ws.columns:
                    max_len = 8
                    col_letter = col[0].column_letter
                    for cell in col:
                        value = str(cell.value or "")
                        max_len = max(max_len, min(len(value) + 2, 60))
                    ws.column_dimensions[col_letter].width = max_len
            wb["要約"].column_dimensions["A"].width = 32
            wb["要約"].column_dimensions["B"].width = 80
        except Exception:
            pass

    output.seek(0)
    return output.getvalue()



def make_short_goal_monitoring_pdf_bytes(
    selected_user,
    selected_goal_text,
    selected_support_text,
    start_date,
    end_date,
    total,
    done,
    partial,
    not_done,
    rate,
    done_only_rate,
    summary: dict,
    work: pd.DataFrame,
    staff_name="",
):
    """介護計画モニタリング票風の1ページPDFを作成する。"""
    if colors is None:
        raise RuntimeError("reportlab が利用できないためPDFを作成できません。")

    output = BytesIO()

    # 日本語フォント。ReportLab標準CIDフォントなので追加フォントファイル不要。
    base_font = register_single_japanese_pdf_font()

    doc = SimpleDocTemplate(
        output,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=12 * mm,
        bottomMargin=12 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "MonitoringTitle",
        parent=styles["Title"],
        fontName=base_font,
        fontSize=15,
        leading=18,
        alignment=1,
        spaceAfter=7,
    )
    normal_style = ParagraphStyle(
        "MonitoringNormal",
        parent=styles["Normal"],
        fontName=base_font,
        fontSize=8.5,
        leading=11,
        wordWrap="CJK",
    )
    small_style = ParagraphStyle(
        "MonitoringSmall",
        parent=styles["Normal"],
        fontName=base_font,
        fontSize=7.5,
        leading=9.5,
        wordWrap="CJK",
        textColor=colors.HexColor("#555555"),
    )

    def P(value, style=normal_style):
        text = short_goal_pdf_text(value, "")
        text = text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
        text = text.replace("\n", "<br/>")
        return Paragraph(text, style)

    detail = work.copy() if work is not None else pd.DataFrame()
    for col in ["本人の様子", "未実施理由", "職員メモ", "入力職員", "実施状況"]:
        if col not in detail.columns:
            detail[col] = ""

    person_notes = short_goal_join_for_pdf(detail.get("本人の様子", []), limit=4)
    reason_summary = short_goal_pdf_text(summary.get("理由要約") if isinstance(summary, dict) else "", "記録なし")
    memo_summary = short_goal_pdf_text(summary.get("職員メモ要約") if isinstance(summary, dict) else "", "記録なし")
    general_summary = short_goal_pdf_text(summary.get("総括コメント") if isinstance(summary, dict) else "", "記録なし")

    achievement = (
        f"実施状況率：{rate}%（実施のみ率：{done_only_rate}%）\n"
        f"対象期間：{total}日／実施：{done}件／一部実施：{partial}件／未実施：{not_done}件"
    )

    if rate >= 80:
        eval_label = "概ね実施できている。"
    elif rate >= 50:
        eval_label = "一部実施があり、支援方法や声かけの継続確認が必要。"
    else:
        eval_label = "未実施日が多く、実施を妨げている要因の確認が必要。"

    evaluation = f"{eval_label}\n{general_summary}"
    future_policy = (
        f"現在の支援内容：{short_goal_pdf_text(selected_support_text, '記載なし')}\n"
        f"今後は、本人の様子と未実施・一部実施の理由を確認しながら、無理のない範囲で支援を継続する。"
    )
    if reason_summary and reason_summary != "記録なし":
        future_policy += f"\n確認事項：{reason_summary}"

    staff_text = short_goal_pdf_text(staff_name, current_login_user() if "current_login_user" in globals() else "管理者")
    created_text = format_now_jst("%Y-%m-%d %H:%M") if "format_now_jst" in globals() else datetime.now().strftime("%Y-%m-%d %H:%M")

    elements = []
    elements.append(Paragraph("介護計画モニタリング票", title_style))
    elements.append(Paragraph("※この帳票は、日々の短期目標実施チェック記録をもとに作成した下書きです。最終確認は管理者が行ってください。", small_style))
    elements.append(Spacer(1, 4))

    meta_data = [
        [P("利用者名"), P(selected_user), P("担当者"), P(staff_text)],
        [P("期間"), P(f"{start_date} 〜 {end_date}"), P("作成日時"), P(created_text)],
    ]
    meta_table = Table(meta_data, colWidths=[25*mm, 65*mm, 25*mm, 60*mm])
    meta_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), base_font),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7CFC4")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EAF4EF")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#EAF4EF")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(meta_table)
    elements.append(Spacer(1, 6))

    body_rows = [
        [P("短期目標"), P(selected_goal_text)],
        [P("実施率"), P(achievement)],
        [P("本人の様子"), P(person_notes)],
        [P("達成状況"), P(achievement)],
        [P("評価"), P(evaluation)],
        [P("今後の支援方針"), P(future_policy)],
        [P("職員メモ要約"), P(memo_summary)],
    ]
    body_table = Table(body_rows, colWidths=[34*mm, 141*mm], repeatRows=0)
    body_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), base_font),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7CFC4")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EAF4EF")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(body_table)
    elements.append(Spacer(1, 6))

    sign_rows = [
        [P("管理者確認"), P("□ 確認済　　□ 修正あり"), P("確認日"), P("　　　年　　　月　　　日")],
    ]
    sign_table = Table(sign_rows, colWidths=[28*mm, 67*mm, 25*mm, 55*mm])
    sign_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), base_font),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#B7CFC4")),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#EAF4EF")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#EAF4EF")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (-1, -1), 5),
        ("RIGHTPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(sign_table)

    doc.build(elements)
    output.seek(0)
    return output.getvalue()


def show_short_goal_selected_summary(goals: pd.DataFrame, checks: pd.DataFrame):
    """利用者・短期目標を選択して、実施状況率と理由・職員メモ要約を表示し、Excel出力する。"""
    st.markdown("### 利用者・短期目標別の実施状況")
    st.caption("利用者と短期目標を選択し、期間内の実施率、未実施理由・一部実施理由、職員メモ要約を確認し、Excel形式で出力できます。")

    if goals is None or goals.empty:
        st.info("短期目標がまだ登録されていません。")
        return

    goals = normalize_df_columns(goals, SHORT_GOAL_MASTER_COLUMNS)
    checks = normalize_df_columns(checks, SHORT_GOAL_CHECK_COLUMNS) if checks is not None else pd.DataFrame(columns=SHORT_GOAL_CHECK_COLUMNS)

    users = sorted([u for u in goals["利用者名"].dropna().astype(str).unique().tolist() if clean_text(u)])
    if not users:
        st.info("短期目標に利用者名が登録されていません。")
        return

    c1, c2, c3 = st.columns([1.2, 1, 1])
    with c1:
        selected_user = st.selectbox("利用者", users, key="short_goal_summary_user")
    with c2:
        start_date = st.date_input("開始日", value=date(today_jst().year, today_jst().month, 1), key="short_goal_summary_start")
    with c3:
        end_date = st.date_input("終了日", value=today_jst(), key="short_goal_summary_end")

    user_goals = goals[goals["利用者名"].astype(str) == str(selected_user)].copy()
    if user_goals.empty:
        st.info("この利用者の短期目標は登録されていません。")
        return

    goal_label_map = {}
    for _, row in user_goals.iterrows():
        gid = clean_text(row.get("目標ID"))
        goal_text = clean_text(row.get("短期目標"))
        status = clean_text(row.get("状態"))
        if not gid or not goal_text:
            continue
        label_goal = goal_text if len(goal_text) <= 55 else goal_text[:55] + "…"
        goal_label_map[f"{label_goal}｜{status}｜{gid[:8]}"] = gid

    if not goal_label_map:
        st.info("選択できる短期目標がありません。")
        return

    selected_goal_label = st.selectbox("短期目標", list(goal_label_map.keys()), key="short_goal_summary_goal")
    selected_goal_id = goal_label_map.get(selected_goal_label, "")
    selected_goal_row = user_goals[user_goals["目標ID"].astype(str) == str(selected_goal_id)].iloc[0]
    selected_goal_text = clean_text(selected_goal_row.get("短期目標"))
    selected_support_text = clean_text(selected_goal_row.get("支援内容"), "支援内容の記載はありません。")

    st.markdown("#### 支援内容")
    st.info(selected_support_text)

    if checks.empty:
        st.info("実施チェック記録がまだありません。")
        return

    work = checks.copy()
    work["日付_dt"] = pd.to_datetime(work["日付"], errors="coerce")
    work = work[
        (work["利用者名"].astype(str) == str(selected_user))
        & (work["目標ID"].astype(str) == str(selected_goal_id))
        & (work["日付_dt"] >= pd.to_datetime(start_date))
        & (work["日付_dt"] <= pd.to_datetime(end_date))
    ].copy()

    # =========================
    # 未入力日を未実施扱いにする集計（Ver4.1+）
    # =========================
    # これまで：入力された実施チェック記録だけを分母にしていた。
    # 変更後：対象期間の全日数を分母にし、記録がない日は「未入力のため未実施扱い」として自動補完する。
    period_dates = pd.date_range(pd.to_datetime(start_date), pd.to_datetime(end_date), freq="D")
    if len(period_dates) == 0:
        st.warning("開始日と終了日を確認してください。")
        return

    for col in SHORT_GOAL_CHECK_COLUMNS:
        if col not in work.columns:
            work[col] = ""

    # 同じ日に複数記録がある場合は、登録日時が新しいものをその日の代表記録にする。
    # 登録日時が空の環境でも止まらないよう、日付で補完する。
    if not work.empty:
        work["日付_only"] = work["日付_dt"].dt.date
        work["_sort_dt"] = pd.to_datetime(work.get("登録日時", ""), errors="coerce")
        work["_sort_dt"] = work["_sort_dt"].fillna(work["日付_dt"])
        work = work.sort_values("_sort_dt").drop_duplicates("日付_only", keep="last").copy()
    else:
        work["日付_only"] = pd.NaT
        work["_sort_dt"] = pd.NaT

    recorded_dates = set([d for d in work["日付_only"].dropna().tolist()]) if "日付_only" in work.columns else set()
    missing_rows = []
    for d in period_dates:
        d_date = d.date()
        if d_date in recorded_dates:
            continue
        missing_rows.append({
            "記録ID": f"AUTO-NOTDONE-{selected_goal_id}-{d.strftime('%Y%m%d')}",
            "日付": d.strftime("%Y-%m-%d"),
            "利用者名": selected_user,
            "user_id": clean_text(selected_goal_row.get("user_id", "")) if "clean_text" in globals() else "",
            "目標ID": selected_goal_id,
            "短期目標": selected_goal_text,
            "実施状況": "未実施",
            "本人の様子": "",
            "未実施理由": "未入力のため未実施扱い",
            "職員メモ": "記録なし。未入力日のため自動的に未実施として集計。",
            "入力職員": "自動判定",
            "登録日時": "",
        })

    if missing_rows:
        work = pd.concat([work, pd.DataFrame(missing_rows)], ignore_index=True)

    work["日付_dt"] = pd.to_datetime(work["日付"], errors="coerce")
    work = work.sort_values("日付_dt", ascending=True).copy()
    if "日付_only" in work.columns:
        work = work.drop(columns=[c for c in ["日付_only", "_sort_dt"] if c in work.columns])

    total = len(period_dates)
    done = int((work["実施状況"].astype(str) == "実施").sum())
    partial = int((work["実施状況"].astype(str) == "一部実施").sum())
    not_done = int((work["実施状況"].astype(str) == "未実施").sum())
    rate = round(((done + partial * 0.5) / total) * 100, 1) if total else 0
    done_only_rate = round((done / total) * 100, 1) if total else 0
    auto_not_done = len(missing_rows)

    m1, m2, m3, m4, m5 = st.columns(5)
    m1.metric("実施状況率", f"{rate}%")
    m2.metric("実施のみ率", f"{done_only_rate}%")
    m3.metric("実施", done)
    m4.metric("一部実施", partial)
    m5.metric("未実施", not_done)
    st.caption("※ 実施状況率は、実施=1点・一部実施=0.5点・未実施=0点として計算しています。")
    st.caption(f"※ 分母は対象期間の全日数（{total}日）です。記録がない日は未実施として自動集計しています（未入力による未実施扱い：{auto_not_done}日）。")

    signature = f"{selected_user}|{selected_goal_id}|{start_date}|{end_date}|{total}|{done}|{partial}|{not_done}|{auto_not_done}"

    # =========================
    # AI要約は自動実行しない（ボタン実行＋保存方式）
    # =========================
    # 画面を開いたり、利用者・期間・短期目標を変更しただけではOpenAI APIを呼ばない。
    # 同じ条件の保存済みAI要約があればそれを表示し、なければ通常要約を表示する。
    saved_ai = load_saved_short_goal_ai_summary(signature)
    if saved_ai:
        summary = saved_ai.get("summary", _build_short_goal_rule_summary(work))
        saved_at = clean_text(saved_ai.get("saved_at", "")) if "clean_text" in globals() else str(saved_ai.get("saved_at", ""))
        saved_by = clean_text(saved_ai.get("saved_by", "")) if "clean_text" in globals() else str(saved_ai.get("saved_by", ""))
        note = f"保存済みAI要約を表示しています。保存日時：{saved_at}" + (f"／作成者：{saved_by}" if saved_by else "")
    else:
        summary = _build_short_goal_rule_summary(work)
        note = "AI要約は未作成です。必要な場合のみ下のボタンで作成・保存してください。"

    if st.button("🤖 AI要約を作成・保存", use_container_width=True, key="short_goal_summary_ai_create_save_button"):
        with st.spinner("AI要約を作成しています..."):
            ai_summary, ai_note = generate_ai_short_goal_summary(selected_user, selected_goal_text, start_date, end_date, work)
            saved = save_short_goal_ai_summary(signature, ai_summary, ai_note)
        if saved:
            st.success("AI要約を作成して保存しました。")
            st.rerun()
        else:
            summary = ai_summary
            note = "AI要約は作成しましたが、保存に失敗しました。画面上には一時表示しています。"

    if note:
        st.caption(note)

    c_reason, c_memo = st.columns(2)
    with c_reason:
        st.markdown("#### 未実施理由・一部実施の理由")
        st.info(summary.get("理由要約", "記録なし"))
    with c_memo:
        st.markdown("#### 職員メモ要約")
        st.info(summary.get("職員メモ要約", "記録なし"))

    st.markdown("#### 総括コメント")
    st.info(summary.get("総括コメント", "記録なし"))

    excel_bytes = make_short_goal_summary_excel_bytes(
        selected_user=selected_user,
        selected_goal_text=selected_goal_text,
        selected_support_text=selected_support_text,
        start_date=start_date,
        end_date=end_date,
        total=total,
        done=done,
        partial=partial,
        not_done=not_done,
        rate=rate,
        done_only_rate=done_only_rate,
        summary=summary,
        work=work,
    )
    safe_user = re.sub(r"[\\/:*?\"<>|\s]+", "_", str(selected_user)).strip("_") or "利用者"
    safe_goal = re.sub(r"[\\/:*?\"<>|\s]+", "_", str(selected_goal_text[:24])).strip("_") or "短期目標"
    st.download_button(
        "📥 実施状況・AI要約をExcelでダウンロード",
        data=excel_bytes,
        file_name=f"{safe_user}_{safe_goal}_短期目標モニタリング_{start_date}_{end_date}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key="short_goal_summary_excel_download",
    )

    # =========================
    # 介護計画モニタリング票 1ページPDF出力
    # =========================
    staff_default = current_login_user() if "current_login_user" in globals() else "管理者"
    staff_name_for_pdf = st.text_input(
        "PDF担当者名",
        value=staff_default,
        key="short_goal_monitoring_pdf_staff",
        help="モニタリング票の担当者欄に表示します。",
    )
    try:
        pdf_bytes = make_short_goal_monitoring_pdf_bytes(
            selected_user=selected_user,
            selected_goal_text=selected_goal_text,
            selected_support_text=selected_support_text,
            start_date=start_date,
            end_date=end_date,
            total=total,
            done=done,
            partial=partial,
            not_done=not_done,
            rate=rate,
            done_only_rate=done_only_rate,
            summary=summary,
            work=work,
            staff_name=staff_name_for_pdf,
        )
        st.download_button(
            "📄 介護計画モニタリング票PDFをダウンロード",
            data=pdf_bytes,
            file_name=f"{safe_user}_{safe_goal}_介護計画モニタリング票_{start_date}_{end_date}.pdf",
            mime="application/pdf",
            use_container_width=True,
            key="short_goal_monitoring_pdf_download",
        )
    except Exception as e:
        st.warning(f"PDF出力を作成できませんでした：{e}")

    with st.expander("対象期間の実施チェック記録を確認", expanded=False):
        show_cols = ["日付", "利用者名", "短期目標", "実施状況", "本人の様子", "未実施理由", "職員メモ", "入力職員", "登録日時"]
        for col in show_cols:
            if col not in work.columns:
                work[col] = ""
        st.dataframe(work.sort_values("日付_dt", ascending=False)[show_cols], use_container_width=True, hide_index=True)


def show_short_goal_top():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return
    st.header("短期目標・モニタリング")
    st.caption("短期目標、実施状況、モニタリング下書きを確認します。")
    goals = load_short_goal_master()
    checks = load_short_goal_checks()
    drafts = load_monitoring_drafts()
    c1, c2, c3 = st.columns(3)
    c1.metric("有効な短期目標", len(goals[goals["状態"].astype(str) == "有効"]) if not goals.empty else 0)
    c2.metric("実施チェック記録", len(checks))
    c3.metric("モニタリング下書き", len(drafts))
    st.markdown(
        """
        <div class="info-box">
        ① 目標を登録<br>
        ② 日々の実施を入力<br>
        ③ 月末や会議前に下書きを確認
        </div>
        """,
        unsafe_allow_html=True,
    )
    st.divider()
    show_short_goal_selected_summary(goals, checks)


def show_short_goal_master():
    st.header("短期目標マスタ")
    st.caption("利用者ごとの短期目標を登録・検索・更新・削除します。日々の実施チェックでは、ここで登録した有効な目標を選択します。")

    users = get_active_user_names()
    df = load_short_goal_master()
    df = normalize_df_columns(df, SHORT_GOAL_MASTER_COLUMNS)
    df = attach_user_ids(df, name_col="利用者名", id_col="user_id")

    if not users:
        st.warning("利用者マスタに表示中の利用者がありません。先に利用者登録を行ってください。")
        return

    def _safe_goal_date_value(value, fallback=None):
        fallback = fallback or today_jst()
        try:
            dt = pd.to_datetime(value, errors="coerce")
            if pd.isna(dt):
                return fallback
            return dt.date()
        except Exception:
            return fallback

    tab_add, tab_search, tab_edit, tab_delete = st.tabs([
        "新規登録",
        "検索・一覧",
        "更新",
        "削除",
    ])

    # -------------------------
    # 新規登録
    # -------------------------
    with tab_add:
        st.subheader("短期目標を新規登録")
        with st.form("goal_master_form", clear_on_submit=True):
            col1, col2, col3 = st.columns(3)
            with col1:
                user_name = st.selectbox("利用者名", users, key="goal_user")
            with col2:
                start_date = st.date_input("開始日", value=today_jst(), key="goal_start")
            with col3:
                end_date = st.date_input("終了予定日", value=today_jst(), key="goal_end")

            short_goal = st.text_area(
                "短期目標",
                placeholder="例：午前中に居室からリビングへ移動し、他利用者と過ごす時間を持つ",
                key="goal_new_short_goal",
            )
            support = st.text_area(
                "支援内容",
                placeholder="例：声かけ、歩行時の見守り、必要時は手を添える",
                key="goal_new_support",
            )

            col4, col5 = st.columns(2)
            with col4:
                status = st.selectbox("状態", ["有効", "終了", "一時停止"], index=0, key="goal_new_status")
            with col5:
                memo = st.text_input("備考", key="goal_new_memo")

            submitted = st.form_submit_button("短期目標を登録", use_container_width=True)

        if submitted:
            if not clean_text(short_goal):
                st.error("短期目標を入力してください。")
            else:
                uid = get_user_id_by_name(user_name) or ensure_user_id_value("", user_name)
                new_row = {
                    "目標ID": str(uuid.uuid4()),
                    "利用者名": user_name,
                    "user_id": uid,
                    "短期目標": clean_text(short_goal),
                    "支援内容": clean_text(support),
                    "開始日": start_date.strftime("%Y-%m-%d"),
                    "終了予定日": end_date.strftime("%Y-%m-%d"),
                    "状態": status,
                    "備考": clean_text(memo),
                    "登録日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
                }
                df2 = pd.concat([df, pd.DataFrame([new_row], columns=SHORT_GOAL_MASTER_COLUMNS)], ignore_index=True)
                save_short_goal_master(df2)
                try:
                    add_audit_log("短期目標マスタ登録", SQLITE_TABLE_SHORT_GOAL_MASTER, new_row["目標ID"], f"{user_name} / {clean_text(short_goal)[:80]}")
                except Exception:
                    pass
                st.success("短期目標を登録しました。")
                st.rerun()

    # -------------------------
    # 検索・一覧
    # -------------------------
    with tab_search:
        st.subheader("登録済み短期目標を検索")
        if df.empty:
            st.info("まだ短期目標が登録されていません。")
        else:
            c1, c2, c3 = st.columns([1.2, 1, 1.6])
            with c1:
                filter_user = st.selectbox("利用者で絞り込み", ["すべて"] + users, key="goal_filter_user")
            with c2:
                filter_status = st.selectbox("状態で絞り込み", ["すべて", "有効", "一時停止", "終了"], key="goal_filter_status")
            with c3:
                keyword = st.text_input("キーワード検索", placeholder="短期目標・支援内容・備考から検索", key="goal_filter_keyword")

            result = df.copy()
            if filter_user != "すべて":
                result = result[result["利用者名"].astype(str) == str(filter_user)].copy()
            if filter_status != "すべて":
                result = result[result["状態"].astype(str) == str(filter_status)].copy()
            if clean_text(keyword):
                kw = clean_text(keyword)
                mask = (
                    result["短期目標"].astype(str).str.contains(kw, case=False, na=False)
                    | result["支援内容"].astype(str).str.contains(kw, case=False, na=False)
                    | result["備考"].astype(str).str.contains(kw, case=False, na=False)
                )
                result = result[mask].copy()

            st.caption(f"検索結果：{len(result)}件")
            display_cols = ["利用者名", "短期目標", "支援内容", "開始日", "終了予定日", "状態", "備考", "登録日時"]
            st.dataframe(result[display_cols], use_container_width=True, hide_index=True)

            if not result.empty:
                download_df = result[SHORT_GOAL_MASTER_COLUMNS].copy()
                buffer = BytesIO()
                with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
                    download_df.to_excel(writer, sheet_name="短期目標マスタ", index=False)
                buffer.seek(0)
                st.download_button(
                    "検索結果をExcelでダウンロード",
                    data=buffer.getvalue(),
                    file_name=f"短期目標マスタ検索結果_{today_jst().strftime('%Y-%m-%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    # -------------------------
    # 更新
    # -------------------------
    with tab_edit:
        st.subheader("短期目標を更新")
        if df.empty:
            st.info("更新できる短期目標がありません。")
        else:
            label_map = {}
            for _, row in df.iterrows():
                goal_id = clean_text(row.get("目標ID"))
                goal_text = clean_text(row.get("短期目標"))
                if len(goal_text) > 45:
                    goal_text = goal_text[:45] + "…"
                label = f"{clean_text(row.get('利用者名'))}｜{clean_text(row.get('状態'))}｜{goal_text}｜{goal_id[:8]}"
                if goal_id:
                    label_map[label] = goal_id

            if not label_map:
                st.warning("更新対象の短期目標IDが見つかりません。登録データを確認してください。")
                return

            selected_label = st.selectbox("更新する短期目標を選択", list(label_map.keys()), key="goal_edit_select")
            selected_id = label_map.get(selected_label, "")
            hit = df[df["目標ID"].astype(str) == str(selected_id)]

            if hit.empty:
                st.warning("選択した短期目標が見つかりません。")
            else:
                # 重要：Streamlitは同じkeyのウィジェット値を保持するため、
                # 更新対象を切り替えても text_area / date_input の中身が前回選択のまま残ることがある。
                # 目標IDごとにフォーム・各入力欄のkeyを分け、選択した短期目標・支援内容を正しく表示する。
                row = hit.iloc[-1]
                edit_key = clean_text(selected_id)[:12] or str(abs(hash(selected_label)))
                current_user = clean_text(row.get("利用者名"))
                user_index = users.index(current_user) if current_user in users else 0
                status_options = ["有効", "終了", "一時停止"]
                current_status = clean_text(row.get("状態"), "有効")
                status_index = status_options.index(current_status) if current_status in status_options else 0

                with st.form(f"goal_master_update_form_{edit_key}"):
                    col1, col2, col3 = st.columns(3)
                    with col1:
                        edit_user = st.selectbox("利用者名", users, index=user_index, key=f"goal_edit_user_{edit_key}")
                    with col2:
                        edit_start = st.date_input("開始日", value=_safe_goal_date_value(row.get("開始日")), key=f"goal_edit_start_{edit_key}")
                    with col3:
                        edit_end = st.date_input("終了予定日", value=_safe_goal_date_value(row.get("終了予定日")), key=f"goal_edit_end_{edit_key}")

                    edit_goal = st.text_area("短期目標", value=clean_text(row.get("短期目標")), key=f"goal_edit_goal_{edit_key}")
                    edit_support = st.text_area("支援内容", value=clean_text(row.get("支援内容")), key=f"goal_edit_support_{edit_key}")

                    col4, col5 = st.columns(2)
                    with col4:
                        edit_status = st.selectbox("状態", status_options, index=status_index, key=f"goal_edit_status_{edit_key}")
                    with col5:
                        edit_memo = st.text_input("備考", value=clean_text(row.get("備考")), key=f"goal_edit_memo_{edit_key}")

                    update_submitted = st.form_submit_button("この内容で更新", use_container_width=True)

                if update_submitted:
                    if not clean_text(edit_goal):
                        st.error("短期目標を入力してください。")
                    else:
                        idx_list = df.index[df["目標ID"].astype(str) == str(selected_id)].tolist()
                        if not idx_list:
                            st.error("更新対象が見つかりません。")
                        else:
                            idx = idx_list[-1]
                            uid = get_user_id_by_name(edit_user) or ensure_user_id_value(clean_text(row.get("user_id")), edit_user)
                            before_summary = f"{clean_text(row.get('利用者名'))} / {clean_text(row.get('短期目標'))[:80]}"
                            df.at[idx, "利用者名"] = edit_user
                            df.at[idx, "user_id"] = uid
                            df.at[idx, "短期目標"] = clean_text(edit_goal)
                            df.at[idx, "支援内容"] = clean_text(edit_support)
                            df.at[idx, "開始日"] = edit_start.strftime("%Y-%m-%d")
                            df.at[idx, "終了予定日"] = edit_end.strftime("%Y-%m-%d")
                            df.at[idx, "状態"] = edit_status
                            df.at[idx, "備考"] = clean_text(edit_memo)
                            if not clean_text(df.at[idx, "登録日時"]):
                                df.at[idx, "登録日時"] = format_now_jst("%Y-%m-%d %H:%M:%S")
                            save_short_goal_master(df)
                            try:
                                add_audit_log(
                                    "短期目標マスタ更新",
                                    SQLITE_TABLE_SHORT_GOAL_MASTER,
                                    selected_id,
                                    f"{before_summary} → {edit_user} / {clean_text(edit_goal)[:80]}",
                                )
                            except Exception:
                                pass
                            st.success("短期目標を更新しました。")
                            st.rerun()

    # -------------------------
    # 削除
    # -------------------------
    with tab_delete:
        st.subheader("短期目標を削除")
        st.warning("削除すると、短期目標マスタから消えます。既に登録済みの日々の実施チェック記録は原則そのまま残ります。")

        if df.empty:
            st.info("削除できる短期目標がありません。")
        else:
            delete_label_map = {}
            for _, row in df.iterrows():
                goal_id = clean_text(row.get("目標ID"))
                goal_text = clean_text(row.get("短期目標"))
                if len(goal_text) > 45:
                    goal_text = goal_text[:45] + "…"
                label = f"{clean_text(row.get('利用者名'))}｜{clean_text(row.get('状態'))}｜{goal_text}｜{goal_id[:8]}"
                if goal_id:
                    delete_label_map[label] = goal_id

            selected_delete_labels = st.multiselect(
                "削除する短期目標を選択",
                list(delete_label_map.keys()),
                key="goal_delete_select",
            )
            selected_delete_ids = [delete_label_map[x] for x in selected_delete_labels if x in delete_label_map]

            if selected_delete_ids:
                preview = df[df["目標ID"].astype(str).isin([str(x) for x in selected_delete_ids])].copy()
                st.dataframe(
                    preview[["利用者名", "短期目標", "支援内容", "開始日", "終了予定日", "状態", "備考"]],
                    use_container_width=True,
                    hide_index=True,
                )

                confirm = st.checkbox("選択した短期目標を削除することを確認しました", key="goal_delete_confirm")
                if st.button("選択した短期目標を削除", type="primary", use_container_width=True, disabled=not confirm):
                    result = delete_short_goal_master_records(selected_delete_ids, source="短期目標マスタ画面から削除")
                    if result.get("error"):
                        st.error(f"削除時に一部エラーがありました：{result.get('error')}")
                    if result.get("ok"):
                        st.success(f"削除しました。SQLite:{result.get('sqlite_deleted', 0)}件 / Supabase:{result.get('supabase_deleted', 0)}件")
                        st.rerun()
                    else:
                        st.error("削除対象が見つかりませんでした。")
            else:
                st.info("削除する短期目標を選択してください。")



def show_daily_goal_check():
    st.header("短期目標の実施チェック")
    st.caption("今日の実施状況を、利用者ごとに記録します。")

    users = get_active_user_names()
    goal_df = load_short_goal_master()
    check_df = load_short_goal_checks()
    check_df = normalize_df_columns(check_df, SHORT_GOAL_CHECK_COLUMNS)
    check_df = attach_user_ids(check_df)

    if not users:
        st.warning("利用者マスタに表示中の利用者がありません。先に利用者登録を行ってください。")
        return

    def _safe_daily_check_date_value(value, fallback=None):
        fallback = fallback or today_jst()
        try:
            dt = pd.to_datetime(value, errors="coerce")
            if pd.isna(dt):
                return fallback
            return dt.date()
        except Exception:
            return fallback

    def _goal_label(row):
        goal_text = clean_text(row.get("短期目標"))
        goal_id = clean_text(row.get("目標ID"))
        if len(goal_text) > 50:
            goal_text = goal_text[:50] + "…"
        return f"{goal_text}｜{goal_id[:8]}" if goal_id else goal_text

    def _build_goal_options(user_name):
        user_goals = goal_df[
            (goal_df["利用者名"].astype(str) == str(user_name))
            & (goal_df["状態"].astype(str) == "有効")
        ].copy()
        goal_label_map = {}
        for _, row in user_goals.iterrows():
            label = _goal_label(row)
            if label:
                goal_label_map[label] = row
        return goal_label_map

    def _build_check_label(row):
        record_id = clean_text(row.get("記録ID"))
        d = pd.to_datetime(row.get("日付"), errors="coerce")
        day_text = d.strftime("%Y-%m-%d") if pd.notna(d) else clean_text(row.get("日付"))
        goal_text = clean_text(row.get("短期目標"))
        if len(goal_text) > 35:
            goal_text = goal_text[:35] + "…"
        return (
            f"{day_text}｜{clean_text(row.get('利用者名'))}｜"
            f"{clean_text(row.get('実施状況'))}｜{goal_text}｜"
            f"入力:{clean_text(row.get('入力職員'))}｜ID:{record_id[:8]}"
        )

    tab_add, tab_search, tab_edit, tab_delete = st.tabs([
        "新規登録",
        "検索・一覧",
        "更新",
        "削除",
    ])

    # -------------------------
    # 新規登録
    # -------------------------
    with tab_add:
        st.subheader("実施チェックを入力")

        col1, col2 = st.columns(2)
        with col1:
            check_date = st.date_input("日付", value=today_jst(), key="daily_goal_date")
        with col2:
            user_name = st.selectbox("利用者名", users, key="daily_goal_user")

        goal_label_map = _build_goal_options(user_name)
        if not goal_label_map:
            st.warning("この利用者の有効な短期目標が登録されていません。先に『短期目標マスタ』で登録してください。")
            return

        selected_goal_label = st.selectbox("短期目標", list(goal_label_map.keys()), key="daily_goal_select")
        selected_goal = goal_label_map[selected_goal_label]

        st.markdown("#### 支援内容")
        st.info(clean_text(selected_goal.get("支援内容"), "支援内容の記載はありません。"))

        with st.form("check_form", clear_on_submit=True):
            col1, col2, col3 = st.columns(3)
            with col1:
                result = st.selectbox("実施状況", ["実施", "一部実施", "未実施"], key="daily_new_result")
            with col2:
                mood = st.selectbox("本人の様子", ["穏やか", "普段通り", "不安あり", "拒否あり", "疲労あり", "痛み訴えあり", "その他"], key="daily_new_mood")
            with col3:
                reflect = st.selectbox("モニタリング反映", ["反映する", "反映しない"], key="daily_new_reflect")
            reason = st.text_input("未実施理由・一部実施の理由", placeholder="例：眠気が強く、声かけのみ実施", key="daily_new_reason")
            staff_memo = st.text_area("職員メモ", placeholder="例：リビングへの移動はできたが、10分ほどで居室へ戻られた", key="daily_new_staff_memo")
            staff_name = st.text_input("入力職員", placeholder="例：藤野", key="daily_new_staff_name")
            submitted = st.form_submit_button("実施チェックを保存する", use_container_width=True)

        if submitted:
            uid = get_user_id_by_name(user_name) or ensure_user_id_value(clean_text(selected_goal.get("user_id")), user_name)
            new_row = {
                "記録ID": str(uuid.uuid4()),
                "日付": check_date.strftime("%Y-%m-%d"),
                "利用者名": user_name,
                "user_id": uid,
                "目標ID": clean_text(selected_goal.get("目標ID")),
                "短期目標": clean_text(selected_goal.get("短期目標")),
                "実施状況": result,
                "本人の様子": mood,
                "未実施理由": clean_text(reason),
                "職員メモ": clean_text(staff_memo),
                "入力職員": clean_text(staff_name),
                "モニタリング反映": reflect,
                "登録日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
            }
            check_df2 = pd.concat([check_df, pd.DataFrame([new_row], columns=SHORT_GOAL_CHECK_COLUMNS)], ignore_index=True)
            save_short_goal_checks(check_df2)
            try:
                add_audit_log("短期目標実施チェック登録", SQLITE_TABLE_SHORT_GOAL_CHECKS, new_row["記録ID"], f"{user_name} / {result} / {clean_text(selected_goal.get('短期目標'))[:80]}")
            except Exception:
                pass
            st.success("実施チェックを保存しました。次回の支援・モニタリングに反映できます。")
            st.rerun()

    # -------------------------
    # 検索・一覧
    # -------------------------
    with tab_search:
        st.subheader("登録済み実施チェックを検索")
        if check_df.empty:
            st.info("まだ実施チェック記録がありません。")
        else:
            c1, c2, c3, c4 = st.columns([1.1, 1, 1, 1.6])
            with c1:
                filter_user = st.selectbox("利用者で絞り込み", ["全員"] + users, key="daily_check_search_user")
            with c2:
                start_date = st.date_input("開始日", value=date(today_jst().year, today_jst().month, 1), key="daily_check_search_start")
            with c3:
                end_date = st.date_input("終了日", value=today_jst(), key="daily_check_search_end")
            with c4:
                keyword = st.text_input("キーワード検索", placeholder="短期目標・理由・職員メモ・入力職員から検索", key="daily_check_search_keyword")

            c5, c6 = st.columns(2)
            with c5:
                filter_result = st.selectbox("実施状況で絞り込み", ["すべて", "実施", "一部実施", "未実施"], key="daily_check_search_result")
            with c6:
                filter_reflect = st.selectbox("モニタリング反映", ["すべて", "反映する", "反映しない"], key="daily_check_search_reflect")

            result_df = check_df.copy()
            result_df["日付_dt"] = pd.to_datetime(result_df["日付"], errors="coerce")
            result_df = result_df[
                (result_df["日付_dt"] >= pd.to_datetime(start_date))
                & (result_df["日付_dt"] <= pd.to_datetime(end_date))
            ].copy()

            if filter_user != "全員":
                result_df = result_df[result_df["利用者名"].astype(str) == str(filter_user)].copy()
            if filter_result != "すべて":
                result_df = result_df[result_df["実施状況"].astype(str) == str(filter_result)].copy()
            if filter_reflect != "すべて":
                result_df = result_df[result_df["モニタリング反映"].astype(str) == str(filter_reflect)].copy()
            if clean_text(keyword):
                kw = clean_text(keyword)
                mask = (
                    result_df["短期目標"].astype(str).str.contains(kw, case=False, na=False)
                    | result_df["未実施理由"].astype(str).str.contains(kw, case=False, na=False)
                    | result_df["職員メモ"].astype(str).str.contains(kw, case=False, na=False)
                    | result_df["入力職員"].astype(str).str.contains(kw, case=False, na=False)
                )
                result_df = result_df[mask].copy()

            st.caption(f"検索結果：{len(result_df)}件")
            show_cols = ["日付", "利用者名", "短期目標", "実施状況", "本人の様子", "未実施理由", "職員メモ", "入力職員", "モニタリング反映", "登録日時"]
            for col in show_cols:
                if col not in result_df.columns:
                    result_df[col] = ""
            if not result_df.empty:
                result_df = result_df.sort_values("日付_dt", ascending=False)
            st.dataframe(result_df[show_cols], use_container_width=True, hide_index=True)

            if not result_df.empty:
                download_df = result_df.drop(columns=["日付_dt"], errors="ignore")
                st.download_button(
                    "検索結果をExcelでダウンロード",
                    data=to_excel_download(download_df),
                    file_name=f"短期目標実施チェック検索結果_{today_jst().strftime('%Y-%m-%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key="daily_check_search_download",
                )

    # -------------------------
    # 更新
    # -------------------------
    with tab_edit:
        st.subheader("登録済み実施チェックを更新")
        if check_df.empty:
            st.info("更新できる実施チェック記録がありません。")
        else:
            c1, c2, c3 = st.columns(3)
            with c1:
                edit_filter_user = st.selectbox("利用者で絞り込み", ["全員"] + users, key="daily_check_edit_filter_user")
            with c2:
                edit_start = st.date_input("開始日", value=date(today_jst().year, today_jst().month, 1), key="daily_check_edit_start")
            with c3:
                edit_end = st.date_input("終了日", value=today_jst(), key="daily_check_edit_end")

            edit_candidates = check_df.copy()
            edit_candidates["記録ID"] = edit_candidates["記録ID"].fillna("").astype(str)
            edit_candidates = edit_candidates[edit_candidates["記録ID"] != ""].copy()
            edit_candidates["日付_dt"] = pd.to_datetime(edit_candidates["日付"], errors="coerce")
            edit_candidates = edit_candidates[
                (edit_candidates["日付_dt"] >= pd.to_datetime(edit_start))
                & (edit_candidates["日付_dt"] <= pd.to_datetime(edit_end))
            ].copy()
            if edit_filter_user != "全員":
                edit_candidates = edit_candidates[edit_candidates["利用者名"].astype(str) == str(edit_filter_user)].copy()

            if edit_candidates.empty:
                st.info("条件に該当する更新対象がありません。")
            else:
                label_to_id = {}
                for _, row in edit_candidates.sort_values("日付_dt", ascending=False).iterrows():
                    label = _build_check_label(row)
                    record_id = clean_text(row.get("記録ID"))
                    if record_id:
                        label_to_id[label] = record_id

                selected_label = st.selectbox("更新する実施チェックを選択", list(label_to_id.keys()), key="daily_check_edit_select")
                selected_record_id = label_to_id.get(selected_label, "")
                hit = check_df[check_df["記録ID"].astype(str) == str(selected_record_id)]

                if hit.empty:
                    st.warning("選択した実施チェック記録が見つかりません。")
                else:
                    row = hit.iloc[-1]
                    edit_key = clean_text(selected_record_id)[:12] or str(abs(hash(selected_label)))

                    current_user = clean_text(row.get("利用者名"))
                    user_index = users.index(current_user) if current_user in users else 0

                    current_goal_id = clean_text(row.get("目標ID"))
                    current_goal_text = clean_text(row.get("短期目標"))

                    selected_user_for_goals = st.selectbox(
                        "利用者名",
                        users,
                        index=user_index,
                        key=f"daily_check_edit_user_outside_{edit_key}",
                    )
                    goal_label_map = _build_goal_options(selected_user_for_goals)

                    if not goal_label_map:
                        st.warning("この利用者には有効な短期目標がありません。短期目標マスタを確認してください。")
                    else:
                        goal_labels = list(goal_label_map.keys())
                        default_goal_index = 0
                        for i, label in enumerate(goal_labels):
                            goal_row = goal_label_map[label]
                            if clean_text(goal_row.get("目標ID")) == current_goal_id or clean_text(goal_row.get("短期目標")) == current_goal_text:
                                default_goal_index = i
                                break

                        result_options = ["実施", "一部実施", "未実施"]
                        mood_options = ["穏やか", "普段通り", "不安あり", "拒否あり", "疲労あり", "痛み訴えあり", "その他"]
                        reflect_options = ["反映する", "反映しない"]

                        current_result = clean_text(row.get("実施状況"), "実施")
                        current_mood = clean_text(row.get("本人の様子"), "普段通り")
                        current_reflect = clean_text(row.get("モニタリング反映"), "反映する")

                        result_index = result_options.index(current_result) if current_result in result_options else 0
                        mood_index = mood_options.index(current_mood) if current_mood in mood_options else 1
                        reflect_index = reflect_options.index(current_reflect) if current_reflect in reflect_options else 0

                        with st.form(f"daily_check_update_form_{edit_key}"):
                            col1, col2 = st.columns(2)
                            with col1:
                                update_date = st.date_input("日付", value=_safe_daily_check_date_value(row.get("日付")), key=f"daily_check_edit_date_{edit_key}")
                            with col2:
                                update_goal_label = st.selectbox("短期目標", goal_labels, index=default_goal_index, key=f"daily_check_edit_goal_{edit_key}")

                            selected_goal_for_update = goal_label_map[update_goal_label]
                            st.markdown("#### 支援内容")
                            st.info(clean_text(selected_goal_for_update.get("支援内容"), "支援内容の記載はありません。"))

                            col3, col4, col5 = st.columns(3)
                            with col3:
                                update_result = st.selectbox("実施状況", result_options, index=result_index, key=f"daily_check_edit_result_{edit_key}")
                            with col4:
                                update_mood = st.selectbox("本人の様子", mood_options, index=mood_index, key=f"daily_check_edit_mood_{edit_key}")
                            with col5:
                                update_reflect = st.selectbox("モニタリング反映", reflect_options, index=reflect_index, key=f"daily_check_edit_reflect_{edit_key}")

                            update_reason = st.text_input(
                                "未実施理由・一部実施の理由",
                                value=clean_text(row.get("未実施理由")),
                                key=f"daily_check_edit_reason_{edit_key}",
                            )
                            update_staff_memo = st.text_area(
                                "職員メモ",
                                value=clean_text(row.get("職員メモ")),
                                key=f"daily_check_edit_staff_memo_{edit_key}",
                            )
                            update_staff_name = st.text_input(
                                "入力職員",
                                value=clean_text(row.get("入力職員")),
                                key=f"daily_check_edit_staff_name_{edit_key}",
                            )

                            update_submitted = st.form_submit_button("この内容で更新", use_container_width=True)

                        if update_submitted:
                            idx_list = check_df.index[check_df["記録ID"].astype(str) == str(selected_record_id)].tolist()
                            if not idx_list:
                                st.error("更新対象が見つかりません。")
                            else:
                                idx = idx_list[-1]
                                uid = get_user_id_by_name(selected_user_for_goals) or ensure_user_id_value(clean_text(row.get("user_id")), selected_user_for_goals)
                                before_summary = f"{clean_text(row.get('日付'))} / {clean_text(row.get('利用者名'))} / {clean_text(row.get('実施状況'))}"
                                check_df.at[idx, "日付"] = update_date.strftime("%Y-%m-%d")
                                check_df.at[idx, "利用者名"] = selected_user_for_goals
                                check_df.at[idx, "user_id"] = uid
                                check_df.at[idx, "目標ID"] = clean_text(selected_goal_for_update.get("目標ID"))
                                check_df.at[idx, "短期目標"] = clean_text(selected_goal_for_update.get("短期目標"))
                                check_df.at[idx, "実施状況"] = update_result
                                check_df.at[idx, "本人の様子"] = update_mood
                                check_df.at[idx, "未実施理由"] = clean_text(update_reason)
                                check_df.at[idx, "職員メモ"] = clean_text(update_staff_memo)
                                check_df.at[idx, "入力職員"] = clean_text(update_staff_name)
                                check_df.at[idx, "モニタリング反映"] = update_reflect
                                if not clean_text(check_df.at[idx, "登録日時"]):
                                    check_df.at[idx, "登録日時"] = format_now_jst("%Y-%m-%d %H:%M:%S")
                                save_short_goal_checks(check_df)
                                try:
                                    add_audit_log(
                                        "短期目標実施チェック更新",
                                        SQLITE_TABLE_SHORT_GOAL_CHECKS,
                                        selected_record_id,
                                        f"{before_summary} → {update_date.strftime('%Y-%m-%d')} / {selected_user_for_goals} / {update_result}",
                                    )
                                except Exception:
                                    pass
                                st.success("実施チェックを更新しました。")
                                st.rerun()

    # -------------------------
    # 削除
    # -------------------------
    with tab_delete:
        st.subheader("登録済み実施チェックを削除")
        st.caption("誤登録した実施チェックを選択して削除できます。削除はSQLiteとSupabaseの両方に反映します。")

        if check_df.empty:
            st.info("削除できる実施チェック記録がありません。")
        else:
            c1, c2, c3 = st.columns(3)
            with c1:
                delete_filter_user = st.selectbox("利用者で絞り込み", ["全員"] + users, key="daily_check_delete_filter_user")
            with c2:
                delete_start = st.date_input("開始日", value=date(today_jst().year, today_jst().month, 1), key="daily_check_delete_start")
            with c3:
                delete_end = st.date_input("終了日", value=today_jst(), key="daily_check_delete_end")

            delete_candidates = check_df.copy()
            delete_candidates["記録ID"] = delete_candidates["記録ID"].fillna("").astype(str)
            delete_candidates = delete_candidates[delete_candidates["記録ID"] != ""].copy()
            delete_candidates["日付_dt"] = pd.to_datetime(delete_candidates["日付"], errors="coerce")
            delete_candidates = delete_candidates[
                (delete_candidates["日付_dt"] >= pd.to_datetime(delete_start))
                & (delete_candidates["日付_dt"] <= pd.to_datetime(delete_end))
            ].copy()
            if delete_filter_user != "全員":
                delete_candidates = delete_candidates[delete_candidates["利用者名"].astype(str) == str(delete_filter_user)].copy()

            if delete_candidates.empty:
                st.info("条件に該当する削除対象がありません。")
            else:
                delete_label_to_id = {}
                for _, row in delete_candidates.sort_values("日付_dt", ascending=False).iterrows():
                    label = _build_check_label(row)
                    record_id = clean_text(row.get("記録ID"))
                    if record_id:
                        delete_label_to_id[label] = record_id

                selected_delete_label = st.selectbox(
                    "削除する実施チェックを選択",
                    list(delete_label_to_id.keys()),
                    key="daily_check_delete_select",
                )
                selected_delete_id = delete_label_to_id.get(selected_delete_label, "")
                delete_hit = check_df[check_df["記録ID"].astype(str) == str(selected_delete_id)].copy()

                if delete_hit.empty:
                    st.warning("選択した実施チェック記録が見つかりません。")
                else:
                    preview_cols = ["日付", "利用者名", "短期目標", "実施状況", "本人の様子", "未実施理由", "職員メモ", "入力職員", "モニタリング反映", "登録日時"]
                    for col in preview_cols:
                        if col not in delete_hit.columns:
                            delete_hit[col] = ""
                    st.markdown("#### 削除対象の確認")
                    st.dataframe(delete_hit[preview_cols], use_container_width=True, hide_index=True)
                    st.warning("この操作は元に戻せません。必要な場合は、削除前にバックアップを作成してください。")

                    confirm_delete = st.checkbox(
                        "この実施チェックを削除することを確認しました",
                        key=f"daily_check_delete_confirm_{clean_text(selected_delete_id)[:12]}",
                    )
                    if st.button(
                        "選択した実施チェックを削除",
                        type="primary",
                        use_container_width=True,
                        disabled=not confirm_delete,
                        key=f"daily_check_delete_button_{clean_text(selected_delete_id)[:12]}",
                    ):
                        result = delete_short_goal_check_records([selected_delete_id], source="日々の実施チェック画面から削除")
                        if result.get("error"):
                            st.error(f"削除時に一部エラーがありました：{result.get('error')}")
                        if result.get("ok"):
                            st.success(f"削除しました。SQLite:{result.get('sqlite_deleted', 0)}件 / Supabase:{result.get('supabase_deleted', 0)}件")
                            st.rerun()
                        else:
                            st.error("削除対象が見つかりませんでした。")

def show_goal_history():
    st.header("実施履歴一覧")
    users = get_active_user_names()
    df = load_short_goal_checks()
    if df.empty:
        st.info("まだ実施チェック記録がありません。")
        return

    col1, col2, col3 = st.columns(3)
    with col1:
        user_name = st.selectbox("利用者名", ["全員"] + users, key="goal_history_user")
    with col2:
        start_date = st.date_input("開始日", value=date(today_jst().year, today_jst().month, 1), key="goal_history_start")
    with col3:
        end_date = st.date_input("終了日", value=today_jst(), key="goal_history_end")

    filtered = df.copy()
    filtered["日付_dt"] = pd.to_datetime(filtered["日付"], errors="coerce")
    filtered = filtered[(filtered["日付_dt"] >= pd.to_datetime(start_date)) & (filtered["日付_dt"] <= pd.to_datetime(end_date))]
    if user_name != "全員":
        filtered = filtered[filtered["利用者名"].astype(str) == str(user_name)]

    if filtered.empty:
        st.info("条件に該当する実施チェック記録はありません。")
        return

    show_df = filtered.drop(columns=["記録ID", "目標ID", "日付_dt"], errors="ignore")
    st.dataframe(show_df, use_container_width=True, hide_index=True)

    st.download_button(
        "表示中データをExcelでダウンロード",
        data=to_excel_download(show_df),
        file_name=f"短期目標実施履歴_{today_jst().strftime('%Y-%m-%d')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

    # 管理者のみ、表示中の履歴から選択して削除できる
    if is_admin_user():
        st.divider()
        st.subheader("実施履歴の削除")
        st.caption("誤入力などを、表示中の履歴から選択して削除できます。削除後は実施チェックExcelに保存し直されます。")

        delete_target_df = filtered.copy()
        delete_target_df["記録ID"] = delete_target_df["記録ID"].fillna("").astype(str)
        delete_target_df = delete_target_df[delete_target_df["記録ID"] != ""].copy()

        if delete_target_df.empty:
            st.info("削除できる履歴がありません。")
        else:
            label_to_id = {}
            for _, row in delete_target_df.sort_values("日付_dt", ascending=False).iterrows():
                record_id = clean_text(row.get("記録ID"))
                day_text = ""
                d = pd.to_datetime(row.get("日付"), errors="coerce")
                if pd.notna(d):
                    day_text = d.strftime("%Y-%m-%d")
                goal_text = clean_text(row.get("短期目標"))
                if len(goal_text) > 35:
                    goal_text = goal_text[:35] + "…"
                label = (
                    f"{day_text}｜{clean_text(row.get('利用者名'))}｜"
                    f"{clean_text(row.get('実施状況'))}｜{goal_text}｜"
                    f"入力:{clean_text(row.get('入力職員'))}｜ID:{record_id[:8]}"
                )
                label_to_id[label] = record_id

            selected_labels = st.multiselect(
                "削除する履歴を選択",
                options=list(label_to_id.keys()),
                key="goal_history_delete_select",
            )

            if selected_labels:
                preview_ids = [label_to_id[label] for label in selected_labels]
                preview = delete_target_df[delete_target_df["記録ID"].isin(preview_ids)].drop(columns=["目標ID", "日付_dt"], errors="ignore")
                st.warning(f"{len(preview_ids)}件を削除対象として選択しています。")
                st.dataframe(preview, use_container_width=True, hide_index=True)

                confirm_delete = st.checkbox(
                    "確認しました。この実施履歴を削除します。",
                    key="goal_history_delete_confirm",
                )
                if st.button("選択した実施履歴を削除", type="secondary", use_container_width=True):
                    if not confirm_delete:
                        st.error("削除する場合は、確認チェックを入れてください。")
                    else:
                        result = delete_short_goal_check_records(preview_ids, source="短期目標実施履歴画面から削除")
                        if result.get("error"):
                            st.error(f"削除中に一部エラーがありました：{result.get('error')}")
                        deleted_count = result.get("sqlite_deleted", 0)
                        if deleted_count > 0:
                            st.success(f"実施履歴を{deleted_count}件削除しました。")
                            st.rerun()
                        else:
                            st.error("削除対象が見つかりません。")
    else:
        st.info("実施履歴の削除は管理者のみ可能です。")


def show_monitoring_draft():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return
    st.header("介護計画モニタリング下書き作成")
    users = get_active_user_names()
    check_df = load_short_goal_checks()
    draft_df = load_monitoring_drafts()
    if check_df.empty:
        st.info("実施チェック記録がまだありません。")
        return

    col1, col2 = st.columns(2)
    with col1:
        user_name = st.selectbox("利用者名", users, key="monitoring_user")
    with col2:
        target_month_date = st.date_input("対象月", value=today_jst(), key="monitoring_month")

    target_month = ym_str(target_month_date)
    tmp = check_df.copy()
    tmp["日付_dt"] = pd.to_datetime(tmp["日付"], errors="coerce")
    tmp["対象月"] = tmp["日付_dt"].dt.strftime("%Y-%m")
    tmp = tmp[(tmp["利用者名"].astype(str) == str(user_name)) & (tmp["対象月"] == target_month) & (tmp["モニタリング反映"].astype(str) == "反映する")]

    if tmp.empty:
        st.warning("この条件に該当する実施チェック記録がありません。")
        return

    goal_list = tmp["短期目標"].dropna().astype(str).unique().tolist()
    selected_goal = st.selectbox("短期目標", goal_list, key="monitoring_goal")
    gdf = tmp[tmp["短期目標"].astype(str) == str(selected_goal)]

    total = len(gdf)
    done = int((gdf["実施状況"] == "実施").sum())
    partial = int((gdf["実施状況"] == "一部実施").sum())
    not_done = int((gdf["実施状況"] == "未実施").sum())
    rate = round(((done + partial * 0.5) / total) * 100, 1) if total else 0
    mood_summary = "／".join(gdf["本人の様子"].dropna().astype(str).value_counts().head(5).index.tolist())
    reasons = "／".join([x for x in gdf["未実施理由"].dropna().astype(str).tolist() if x.strip()][:5])
    memos = [x for x in gdf["職員メモ"].dropna().astype(str).tolist() if x.strip()]
    memo_text = "。".join(memos[:6])

    draft_text = f"{target_month}の記録では、短期目標『{selected_goal}』について、実施{done}回、一部実施{partial}回、未実施{not_done}回でした。実施率の目安は{rate}%です。本人の様子としては『{mood_summary}』が記録されています。"
    if reasons:
        draft_text += f" 未実施・一部実施の理由として『{reasons}』が記録されています。"
    if memo_text:
        draft_text += f" 職員メモでは『{memo_text}』などの記録があります。"
    draft_text += " 今後も本人の様子を確認しながら、無理のない範囲で支援を継続します。"

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("実施率目安", f"{rate}%")
    c2.metric("実施", done)
    c3.metric("一部実施", partial)
    c4.metric("未実施", not_done)

    edited_draft = st.text_area("下書き文", value=draft_text, height=240)
    direction = st.selectbox("今後の方向性", ["継続", "一部見直し", "目標見直し", "終了", "経過観察"])

    col_save, col_download = st.columns(2)
    with col_save:
        if st.button("この下書きを保存", use_container_width=True):
            new_row = {
                "下書きID": str(uuid.uuid4()),
                "作成日": today_jst().strftime("%Y-%m-%d"),
                "対象月": target_month,
                "利用者名": user_name,
                "短期目標": selected_goal,
                "実施率": rate,
                "実施回数": done,
                "一部実施回数": partial,
                "未実施回数": not_done,
                "本人の様子まとめ": mood_summary,
                "未実施理由まとめ": reasons,
                "モニタリング下書き": edited_draft,
                "今後の方向性": direction,
                "作成日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
            }
            draft_df = pd.concat([draft_df, pd.DataFrame([new_row], columns=MONITORING_DRAFT_COLUMNS)], ignore_index=True)
            save_monitoring_drafts(draft_df)
            st.success("モニタリング下書きを保存しました。")
            st.rerun()
    with col_download:
        st.download_button(
            "下書き文をテキストでダウンロード",
            data=edited_draft.encode("utf-8"),
            file_name=f"モニタリング下書き_{user_name}_{target_month}.txt",
            mime="text/plain",
            use_container_width=True,
        )



def _first_non_empty_text(values, default=""):
    for value in values:
        text = clean_text(value)
        if text:
            return text
    return default


def infer_monitoring_item(short_goal, support_text=""):
    """短期目標文からモニタリング表の項目名を推定する。"""
    text = f"{clean_text(short_goal)} {clean_text(support_text)}"
    rules = [
        ("移動", ["移動", "歩行", "車椅子", "車いす", "立位", "移乗", "転倒", "居室", "リビング"]),
        ("排泄", ["排泄", "トイレ", "便", "尿", "誘導", "失禁", "排便", "排尿"]),
        ("食事", ["食事", "摂取", "水分", "嚥下", "むせ", "口腔", "栄養"]),
        ("入浴", ["入浴", "清潔", "更衣", "整容"]),
        ("服薬", ["服薬", "薬", "内服"]),
        ("活動", ["活動", "レク", "交流", "会話", "他利用者", "参加"]),
        ("睡眠", ["睡眠", "夜間", "不眠", "覚醒"]),
    ]
    for label, words in rules:
        if any(word in text for word in words):
            return label
    return "その他"


def build_monthly_health_context(user_name, target_month):
    """健康・排泄データからモニタリング表に添える状況文を作成する。

    注意：ここで作るバイタル等の数値は「ニーズや生活の現状」へは直接入れない。
    バイタル値は、具体的理由・備考やモニタリングまとめの補足として使う。
    """
    try:
        year, month = [int(x) for x in str(target_month).split("-")[:2]]
    except Exception:
        return {"health_text": "", "excretion_text": "", "change_text": ""}

    health_df = get_month_health_data(load_health_data(), user_name, year, month)
    ex_df = get_month_excretion_data(load_excretion_data(), user_name, year, month)

    health_parts = []
    change_parts = []
    if not health_df.empty:
        for col, label, unit in [("体温", "平均体温", "℃"), ("SpO2", "平均SpO2", "%"), ("体重", "平均体重", "kg")]:
            if col in health_df.columns:
                vals = to_number(health_df[col])
                vals = vals[vals > 0]
                if not vals.empty:
                    # 読みやすさのため、項目名と数値の間に半角スペースを入れる
                    health_parts.append(f"{label} {round(float(vals.mean()), 1)}{unit}")
        meal_cols = ["朝食摂取率", "昼食摂取率", "夕食摂取率"]
        meal_vals = []
        for col in meal_cols:
            if col in health_df.columns:
                meal_vals.extend(to_number(health_df[col]).dropna().tolist())
        if meal_vals:
            health_parts.append(f"平均食事摂取率 {round(float(pd.Series(meal_vals).mean()), 1)}%")
        if "気になる変化" in health_df.columns:
            changes = [clean_text(x) for x in health_df["気になる変化"].tolist() if clean_text(x)]
            change_parts.extend(changes[:5])
        if "家族共有メモ" in health_df.columns:
            family_memos = [clean_text(x) for x in health_df["家族共有メモ"].tolist() if clean_text(x)]
            change_parts.extend(family_memos[:3])

    ex_text = ""
    if not ex_df.empty:
        ex_sum = summarize_excretion(ex_df)
        ex_text = f"排尿 {ex_sum.get('排尿回数', 0)}回、排便 {ex_sum.get('排便回数', 0)}回"
        if ex_sum.get("下痢便", 0) or ex_sum.get("水様便", 0):
            ex_text += f"、下痢便・水様便 {ex_sum.get('下痢便', 0) + ex_sum.get('水様便', 0)}回"
        if ex_sum.get("濃縮尿", 0):
            ex_text += f"、濃縮尿 {ex_sum.get('濃縮尿', 0)}回"

    return {
        "health_text": "、".join(health_parts),
        "excretion_text": ex_text,
        "change_text": "。".join(change_parts),
    }


def build_needs_current_status(user_name, item, short_goal, support_text, context):
    """モニタリング表の「ニーズや生活の現状」を作る。

    ここにはバイタル平均・食事摂取率などの数値を直接入れず、
    本人の生活上の困りごと・希望・支援が必要な背景を短く入れる。
    """
    assessment = get_user_assessment(user_name)
    candidates = []
    # 利用者マスタに情報があれば、生活状況・課題・主訴を優先する
    for key in ["主訴", "生活状況", "ADL", "課題"]:
        value = clean_text(assessment.get(key, ""))
        if value:
            candidates.append(value)

    change_text = clean_text(context.get("change_text", ""))
    if change_text:
        candidates.append(change_text[:90])

    if candidates:
        return "。".join(candidates[:2])

    goal_text = f"{clean_text(short_goal)} {clean_text(support_text)}"
    if item == "移動":
        return "移動時のふらつきや転倒リスクに配慮し、安全に移動できるよう見守りや声かけが必要です。"
    if item == "排泄":
        return "排泄リズムやトイレ動作の状態を確認し、必要に応じて声かけや見守りが必要です。"
    if item == "食事":
        return "食事量や水分摂取の様子を確認し、無理なく摂取できるよう支援が必要です。"
    if item == "入浴":
        return "清潔保持や更衣動作の負担に配慮し、安心して生活できるよう支援が必要です。"
    if item == "服薬":
        return "服薬状況を確認し、飲み忘れや不安が出ないよう声かけが必要です。"
    if item == "活動":
        return "生活の中で無理なく参加できる活動や交流の機会を保つことが必要です。"
    if item == "睡眠":
        return "夜間の睡眠状況や日中の様子を確認し、生活リズムを整える支援が必要です。"
    if clean_text(goal_text):
        return "本人の生活状況を確認しながら、設定した短期目標に沿って支援を継続しています。"
    return "日々の生活状況を確認しながら、本人に合った支援を継続しています。"


def build_rule_based_monitoring_rows(user_name, target_month):
    """過去入力データから、介護計画モニタリング表の下書き行を作る。"""
    goals = load_short_goal_master()
    checks = load_short_goal_checks()
    context = build_monthly_health_context(user_name, target_month)

    if goals.empty:
        return pd.DataFrame(columns=["項目", "ニーズや生活の現状", "短期目標", "実施状況", "本人の様子・満足度", "具体的所見", "今後の方向性", "具体的理由・今後の備考"]), "", ""

    active_goals = goals[(goals["利用者名"].astype(str) == str(user_name)) & (goals["状態"].astype(str).isin(["有効", "終了", "一時停止"]))].copy()
    if active_goals.empty:
        return pd.DataFrame(columns=["項目", "ニーズや生活の現状", "短期目標", "実施状況", "本人の様子・満足度", "具体的所見", "今後の方向性", "具体的理由・今後の備考"]), "", ""

    work = checks.copy()
    if not work.empty:
        work["日付_dt"] = pd.to_datetime(work["日付"], errors="coerce")
        work["対象月"] = work["日付_dt"].dt.strftime("%Y-%m")
        work = work[(work["利用者名"].astype(str) == str(user_name)) & (work["対象月"] == str(target_month))]
    else:
        work = pd.DataFrame(columns=SHORT_GOAL_CHECK_COLUMNS + ["日付_dt", "対象月"])

    rows = []
    summary_sentences = []
    issue_sentences = []

    for _, goal in active_goals.iterrows():
        goal_id = clean_text(goal.get("目標ID"))
        short_goal = clean_text(goal.get("短期目標"))
        support = clean_text(goal.get("支援内容"))
        item = infer_monitoring_item(short_goal, support)
        gdf = work[work["目標ID"].astype(str) == goal_id].copy() if goal_id else pd.DataFrame()
        if gdf.empty:
            gdf = work[work["短期目標"].astype(str) == short_goal].copy()

        total = len(gdf)
        done = int((gdf["実施状況"].astype(str) == "実施").sum()) if total else 0
        partial = int((gdf["実施状況"].astype(str) == "一部実施").sum()) if total else 0
        not_done = int((gdf["実施状況"].astype(str) == "未実施").sum()) if total else 0
        rate = round(((done + partial * 0.5) / total) * 100, 1) if total else 0

        mood = ""
        findings = ""
        reasons = ""
        memos = []
        if total:
            mood_values = [clean_text(x) for x in gdf["本人の様子"].tolist() if clean_text(x)]
            mood = _first_non_empty_text(pd.Series(mood_values).value_counts().index.tolist(), "記録上、大きな拒否や不穏は目立ちません。")
            reasons_list = [clean_text(x) for x in gdf["未実施理由"].tolist() if clean_text(x)]
            memo_list = [clean_text(x) for x in gdf["職員メモ"].tolist() if clean_text(x)]
            memos = memo_list[:4]
            reasons = "。".join(reasons_list[:3])
            findings = "。".join(memos) if memos else f"実施{done}回、一部実施{partial}回、未実施{not_done}回。実施率の目安は{rate}%です。"
        else:
            mood = "記録が少ないため、本人の様子は十分に確認できていません。"
            findings = "対象月の実施チェック記録がありません。記録方法または実施状況の確認が必要です。"
            issue_sentences.append(f"{item}について、実施記録が不足しています。")

        if total == 0:
            status_text = "記録未入力"
            direction = "記録を確認して継続する"
        elif rate >= 80:
            status_text = "計画通り実施できた"
            direction = "サービス内容を継続する"
        elif rate >= 50:
            status_text = "一部実施できた"
            direction = "支援方法を一部見直して継続する"
            issue_sentences.append(f"{item}について、一部実施・未実施の理由確認が必要です。")
        else:
            status_text = "実施が少ない"
            direction = "目標または支援方法を見直す"
            issue_sentences.append(f"{item}について、実施率が低めです。")

        # ニーズ欄には支援内容やバイタル平均を混ぜず、生活上の現状・困りごとを出す
        current_status = build_needs_current_status(user_name, item, short_goal, support, context)

        note_parts = []
        if reasons:
            note_parts.append(reasons)
        else:
            note_parts.append("現在の記録上、急な変更を要する内容は目立ちません。本人の様子を見ながら継続して確認します。")
        if support:
            note_parts.append(f"支援内容：{support}")
        if context.get("health_text"):
            note_parts.append(f"健康記録：{context['health_text']}")
        if context.get("excretion_text") and item in ["排泄", "その他"]:
            note_parts.append(f"排泄記録：{context['excretion_text']}")
        note = " ".join(note_parts)
        if item == "移動" and "転倒" in context.get("change_text", ""):
            note += " 転倒リスクに配慮し、移動時の見守りを継続します。"
        if item == "排泄" and context.get("excretion_text"):
            note += f" 排泄状況は、{context['excretion_text']}として記録されています。"

        rows.append({
            "項目": item,
            "ニーズや生活の現状": current_status,
            "短期目標": short_goal,
            "実施状況": status_text,
            "本人の様子・満足度": mood,
            "具体的所見": findings,
            "今後の方向性": direction,
            "具体的理由・今後の備考": note,
        })
        summary_sentences.append(f"{item}の目標は{status_text}（実施率目安{rate}%）でした。")

    summary_text = " ".join(summary_sentences)
    if context.get("health_text"):
        summary_text += f" 健康記録では、{context['health_text']}が確認されています。"
    if context.get("excretion_text"):
        summary_text += f" 排泄記録では、{context['excretion_text']}が確認されています。"
    if context.get("change_text"):
        summary_text += f" 気になる変化として、{context['change_text'][:120]}が記録されています。"
    summary_text += " 今後も本人の様子を確認しながら、無理のない範囲で支援を継続します。"

    if not issue_sentences:
        issue_sentences.append("現在の記録上、新たな生活課題は大きく目立っていません。引き続き、本人の様子と生活リズムを確認します。")
    issue_text = " ".join(dict.fromkeys(issue_sentences))

    return pd.DataFrame(rows), summary_text, issue_text


def generate_ai_monitoring_rows(user_name, target_month, base_rows, summary_text, issue_text):
    """OpenAI APIが設定されている場合のみ、モニタリング文を整える。失敗時は空文字を返す。"""
    api_key = get_openai_api_key("")
    if not api_key:
        return None, "OpenAI APIキーが未設定です。通常下書きを使用してください。"
    try:
        from openai import OpenAI
    except Exception:
        return None, "openaiライブラリが未インストールです。requirements.txtに openai を追加してください。"

    prompt = f"""
あなたは介護施設の介護計画モニタリング表の文章整理係です。
医療判断・診断・治療効果の断定は禁止です。
記録に基づき、以下の表を自然な介護記録文に整えてください。
出力は必ずJSONのみ。推測で事実を追加しないでください。

【対象】
利用者：{user_name}
対象月：{target_month}

【下書き表】
{base_rows.to_dict(orient='records')}

【モニタリングまとめ】
{summary_text}

【新たな生活課題】
{issue_text}

JSON形式：
{{
  "rows": [
    {{"項目":"", "ニーズや生活の現状":"", "短期目標":"", "実施状況":"", "本人の様子・満足度":"", "具体的所見":"", "今後の方向性":"", "具体的理由・今後の備考":""}}
  ],
  "モニタリングまとめ":"",
  "新たな生活課題":""
}}
"""
    try:
        client = OpenAI(api_key=api_key)
        res = client.chat.completions.create(
            model=get_openai_model("monitoring", "gpt-4o-mini"),
            messages=[
                {"role": "system", "content": "介護計画モニタリング表の文章を整理します。断定せず、記録に基づく表現にします。"},
                {"role": "user", "content": prompt},
            ],
            temperature=0.2,
            response_format={"type": "json_object"},
        )
        data = json.loads(res.choices[0].message.content or "{}")
        rows = pd.DataFrame(data.get("rows", []))
        for col in ["項目", "ニーズや生活の現状", "短期目標", "実施状況", "本人の様子・満足度", "具体的所見", "今後の方向性", "具体的理由・今後の備考"]:
            if col not in rows.columns:
                rows[col] = ""
        return (rows[["項目", "ニーズや生活の現状", "短期目標", "実施状況", "本人の様子・満足度", "具体的所見", "今後の方向性", "具体的理由・今後の備考"]], clean_text(data.get("モニタリングまとめ"), summary_text), clean_text(data.get("新たな生活課題"), issue_text)), ""
    except Exception as e:
        return None, f"AI生成中にエラーが出ました：{e}"


def monitoring_table_to_excel(user_name, target_month, rows_df, summary_text, issue_text):
    """画像の様式に近い介護計画モニタリング表をExcelで作成する。"""
    output = BytesIO()
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
    except Exception:
        rows_df.to_excel(output, index=False)
        return output.getvalue()

    wb = Workbook()
    ws = wb.active
    ws.title = "介護計画モニタリング表"

    thin = Side(style="thin", color="555555")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="EDE7D6")
    title_fill = PatternFill("solid", fgColor="F7F1E3")
    accent_fill = PatternFill("solid", fgColor="F4A261")

    ws.merge_cells("A1:H1")
    ws["A1"] = "介護計画モニタリング表"
    ws["A1"].font = Font(size=16, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    info = [
        ("A3", "入居者名", "B3", user_name),
        ("D3", "記入日", "E3", today_jst().strftime("%Y/%m/%d")),
        ("G3", "対象月", "H3", target_month),
        ("A4", "当今確認職員", "B4", current_login_user()),
        ("D4", "総合判定", "E4", ""),
        ("G4", "次回確認予定日", "H4", ""),
    ]
    for label_cell, label, value_cell, value in info:
        ws[label_cell] = label
        ws[label_cell].fill = title_fill
        ws[label_cell].font = Font(bold=True)
        ws[value_cell] = value
        for c in [label_cell, value_cell]:
            ws[c].border = border
            ws[c].alignment = Alignment(vertical="center", wrap_text=True)

    headers = ["項目", "ニーズや生活の現状", "短期目標", "実施状況", "本人の様子・満足度", "具体的所見", "今後の方向性", "具体的理由・今後の備考"]
    start_row = 6
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r_idx, (_, row) in enumerate(rows_df.iterrows(), start_row + 1):
        for c_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=clean_text(row.get(header)))
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        ws.row_dimensions[r_idx].height = 80

    summary_row = start_row + len(rows_df) + 2
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=8)
    ws.cell(summary_row, 1).value = "モニタリングまとめ"
    ws.cell(summary_row, 1).fill = header_fill
    ws.cell(summary_row, 1).font = Font(bold=True)
    ws.cell(summary_row, 1).alignment = Alignment(horizontal="center")
    ws.cell(summary_row, 1).border = border

    ws.merge_cells(start_row=summary_row + 1, start_column=1, end_row=summary_row + 1, end_column=8)
    ws.cell(summary_row + 1, 1).value = summary_text
    ws.cell(summary_row + 1, 1).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(summary_row + 1, 1).border = border
    ws.row_dimensions[summary_row + 1].height = 70

    issue_row = summary_row + 3
    ws.merge_cells(start_row=issue_row, start_column=1, end_row=issue_row, end_column=8)
    ws.cell(issue_row, 1).value = "新たな生活課題"
    ws.cell(issue_row, 1).fill = header_fill
    ws.cell(issue_row, 1).font = Font(bold=True)
    ws.cell(issue_row, 1).alignment = Alignment(horizontal="center")
    ws.cell(issue_row, 1).border = border

    ws.merge_cells(start_row=issue_row + 1, start_column=1, end_row=issue_row + 1, end_column=8)
    ws.cell(issue_row + 1, 1).value = issue_text
    ws.cell(issue_row + 1, 1).alignment = Alignment(vertical="top", wrap_text=True)
    ws.cell(issue_row + 1, 1).border = border
    ws.row_dimensions[issue_row + 1].height = 70

    widths = [12, 26, 24, 18, 20, 26, 22, 42]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.page_setup.orientation = "landscape"
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4
    ws.freeze_panes = "A7"

    wb.save(output)
    return output.getvalue()


def show_ai_monitoring_table_builder():
    st.subheader("AI介護計画モニタリング表 作成")
    st.caption("短期目標・日々の実施チェック・健康記録・排泄記録から、画像のようなモニタリング表の下書きを作成します。")

    users = get_active_user_names()
    if not users:
        st.warning("利用者マスタに表示中の利用者がいません。")
        return

    c1, c2, c3 = st.columns(3)
    with c1:
        user_name = st.selectbox("利用者名", users, key="ai_monitoring_user")
    with c2:
        target_month_date = st.date_input("対象月", value=today_jst(), key="ai_monitoring_month")
    with c3:
        use_ai = st.checkbox("AIで文章を整える", value=False, help=f"OpenAI APIキー設定時のみ使用できます。使用モデル: {get_openai_model('monitoring', 'gpt-4o-mini')}")

    target_month = ym_str(target_month_date)
    if st.button("モニタリング表の下書きを作成", type="primary", use_container_width=True):
        rows_df, summary_text, issue_text = build_rule_based_monitoring_rows(user_name, target_month)
        if rows_df.empty:
            st.warning("対象利用者の短期目標または実施記録が不足しています。先に短期目標マスタ・日々の実施チェックを確認してください。")
        else:
            if use_ai:
                ai_result, ai_error = generate_ai_monitoring_rows(user_name, target_month, rows_df, summary_text, issue_text)
                if ai_error:
                    st.warning(ai_error)
                if ai_result:
                    rows_df, summary_text, issue_text = ai_result
            st.session_state["monitoring_table_rows"] = rows_df
            st.session_state["monitoring_table_summary"] = summary_text
            st.session_state["monitoring_table_issue"] = issue_text
            st.session_state["monitoring_table_user"] = user_name
            st.session_state["monitoring_table_month"] = target_month
            st.success("モニタリング表の下書きを作成しました。内容を確認・修正してからExcel出力してください。")

    rows_df = st.session_state.get("monitoring_table_rows")
    if rows_df is None or len(rows_df) == 0:
        st.info("まだ下書きは作成されていません。")
        return

    st.markdown("#### 表の内容確認・修正")
    edited_rows = st.data_editor(rows_df, use_container_width=True, hide_index=True, num_rows="dynamic", key="monitoring_table_editor")
    summary_text = st.text_area("モニタリングまとめ", value=st.session_state.get("monitoring_table_summary", ""), height=120)
    issue_text = st.text_area("新たな生活課題", value=st.session_state.get("monitoring_table_issue", ""), height=120)

    output_user = st.session_state.get("monitoring_table_user", user_name)
    output_month = st.session_state.get("monitoring_table_month", target_month)
    excel_bytes = monitoring_table_to_excel(output_user, output_month, edited_rows, summary_text, issue_text)
    st.download_button(
        "介護計画モニタリング表をExcelでダウンロード",
        data=excel_bytes,
        file_name=f"介護計画モニタリング表_{output_user}_{output_month}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )

def show_short_goal_data_management():
    st.header("短期目標データ管理")
    tab_ai, tab_data = st.tabs(["AIモニタリング表作成", "登録データ確認"])

    with tab_ai:
        show_ai_monitoring_table_builder()

    with tab_data:
        data_map = {
            "短期目標マスタ": (load_short_goal_master(), SHORT_GOAL_MASTER_FILE),
            "日々の実施チェック": (load_short_goal_checks(), SHORT_GOAL_CHECK_FILE),
            "モニタリング下書き": (load_monitoring_drafts(), MONITORING_DRAFT_FILE),
        }
        for label, (df, path) in data_map.items():
            st.subheader(label)
            st.dataframe(df, use_container_width=True, hide_index=True)
            if not df.empty:
                st.download_button(
                    f"{label}をExcelでダウンロード",
                    data=to_excel_download(df),
                    file_name=f"{label}_{today_jst().strftime('%Y-%m-%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    key=f"download_short_goal_{label}",
                    use_container_width=True,
                )
            else:
                st.info("データはまだありません。")

def show_admin_short_goal_summary(target_date):
    st.subheader("短期目標 実施状況")
    checks = load_short_goal_checks()
    if checks.empty:
        st.info("短期目標の実施チェック記録はまだありません。")
        return
    work = checks.copy()
    work["日付_dt"] = pd.to_datetime(work["日付"], errors="coerce")
    target = pd.to_datetime(target_date, errors="coerce")
    if pd.isna(target):
        st.info("確認日を取得できません。")
        return
    day_df = work[work["日付_dt"].dt.date == target.date()].copy()
    if day_df.empty:
        st.info("確認日の短期目標チェック記録はありません。")
        return
    summary = day_df.groupby("実施状況").size().to_dict()
    c1, c2, c3 = st.columns(3)
    c1.metric("実施", int(summary.get("実施", 0)))
    c2.metric("一部実施", int(summary.get("一部実施", 0)))
    c3.metric("未実施", int(summary.get("未実施", 0)))
    st.dataframe(day_df[["日付", "利用者名", "短期目標", "実施状況", "本人の様子", "未実施理由", "職員メモ", "入力職員"]], use_container_width=True, hide_index=True)

# =========================
# バックアップ
# =========================
# =========================
# Excel出力ユーティリティ
# =========================

def dataframe_to_excel_bytes(sheets: dict):
    """複数のDataFrameを1つのExcelブックにして返す。"""
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        written = False
        for sheet_name, df in sheets.items():
            safe_sheet = str(sheet_name)[:31] if sheet_name else "data"
            work = df.copy() if isinstance(df, pd.DataFrame) else pd.DataFrame()
            # Excelで扱いやすいように日付・日時は文字列化する
            for col in work.columns:
                if pd.api.types.is_datetime64_any_dtype(work[col]):
                    work[col] = work[col].dt.strftime("%Y-%m-%d")
            work.to_excel(writer, index=False, sheet_name=safe_sheet)
            written = True
        if not written:
            pd.DataFrame({"メッセージ": ["出力対象データがありません"]}).to_excel(writer, index=False, sheet_name="データなし")
    buffer.seek(0)
    return buffer.getvalue()



def load_selected_export_data(key):
    """管理者ダウンロード用に、SQLite正データから選択データを読み込む。"""
    if key == "健康チェック":
        return load_health_data()
    if key == "排泄チェック":
        return load_excretion_data()
    if key == "利用者マスタ":
        return load_users(include_hidden=True)
    if key == "業務全体申し送り":
        return load_business_handover_data()
    if key == "短期目標マスタ":
        return load_short_goal_master()
    if key == "短期目標実施チェック":
        return load_short_goal_checks()
    if key == "モニタリング下書き":
        return load_monitoring_drafts()
    if key == "LIFE ADL評価":
        return load_life_adl_data()
    if key == "ログイン履歴":
        return load_login_history()
    if key == "AI分析ログ":
        return load_ai_insight_logs()
    return pd.DataFrame()

def filter_export_dataframe(df, start_date=None, end_date=None, user_name="全員"):
    """記録日・日付・評価日・作成日などの日付列と利用者名で絞り込む。"""
    if df is None or df.empty:
        return pd.DataFrame() if df is None else df
    work = df.copy()

    date_col = None
    for candidate in ["記録日", "日付", "評価日", "作成日", "日時", "登録日時"]:
        if candidate in work.columns:
            date_col = candidate
            break

    if date_col and start_date and end_date:
        work[date_col] = pd.to_datetime(work[date_col], errors="coerce")
        start_dt = pd.to_datetime(start_date)
        end_dt = pd.to_datetime(end_date)
        work = work[(work[date_col] >= start_dt) & (work[date_col] <= end_dt)]

    if user_name != "全員" and "利用者名" in work.columns:
        work = work[work["利用者名"].astype(str) == user_name]

    return work


def show_admin_data_download_menu():
    """管理者が必要なデータを選択してExcelでダウンロードする画面。"""
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("データダウンロード")
    st.caption("健康チェック・排泄チェック・短期目標など、必要なデータを選んでExcel形式でダウンロードできます。")

    export_options = [
        "健康チェック",
        "排泄チェック",
        "利用者マスタ",
        "業務全体申し送り",
        "短期目標マスタ",
        "短期目標実施チェック",
        "モニタリング下書き",
        "LIFE ADL評価",
        "ログイン履歴",
    ]

    selected = st.multiselect(
        "ダウンロードするデータを選択",
        export_options,
        default=["健康チェック", "排泄チェック"],
        help="複数選ぶと、1つのExcelファイル内にシートを分けて出力します。",
    )

    c1, c2, c3 = st.columns(3)
    with c1:
        start_date = st.date_input("開始日", value=today_jst().replace(day=1), key="export_start_date")
    with c2:
        end_date = st.date_input("終了日", value=today_jst(), key="export_end_date")
    with c3:
        user_filter = st.selectbox("利用者で絞り込み", ["全員"] + active_users, key="export_user_filter")

    st.caption("日付列があるデータは期間で絞り込みます。利用者名があるデータは利用者でも絞り込めます。")

    if not selected:
        st.info("ダウンロードするデータを1つ以上選択してください。")
        return

    sheets = {}
    preview_rows = []
    for key in selected:
        df = load_selected_export_data(key)
        filtered = filter_export_dataframe(df, start_date, end_date, user_filter)
        sheets[key] = filtered
        preview_rows.append({"データ種類": key, "出力件数": len(filtered)})

    st.subheader("出力内容の確認")
    st.dataframe(pd.DataFrame(preview_rows), use_container_width=True, hide_index=True)

    with st.expander("選択データのプレビュー"):
        for key, df in sheets.items():
            st.markdown(f"**{key}**")
            if df.empty:
                st.info("該当データはありません。")
            else:
                st.dataframe(df.head(50), use_container_width=True, hide_index=True)

    excel_bytes = dataframe_to_excel_bytes(sheets)
    file_name = f"hidamari_selected_data_{format_now_jst('%Y%m%d_%H%M%S')}.xlsx"
    st.download_button(
        "選択したデータをExcelでダウンロード",
        data=excel_bytes,
        file_name=file_name,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

def show_admin_backup_download():
    """管理者向けバックアップダウンロード欄。"""
    if st.session_state.role != "admin":
        return

    st.divider()
    st.subheader("データバックアップ")
    st.caption("健康チェック・排泄チェック・利用者マスタをまとめて保存できます。定期的にダウンロードしてください。")

    if st.button("バックアップZIPを作成", type="primary", use_container_width=True):
        zip_path, err = create_backup_zip(kind="手動")
        if err:
            st.error(err)
        elif zip_path and Path(zip_path).exists():
            st.session_state["admin_backup_zip_bytes"] = Path(zip_path).read_bytes()
            st.session_state["admin_backup_zip_name"] = Path(zip_path).name
            st.success(f"バックアップを作成しました：{Path(zip_path).name}")

    if st.session_state.get("admin_backup_zip_bytes"):
        st.download_button(
            label="作成済みバックアップZIPをダウンロード",
            data=st.session_state["admin_backup_zip_bytes"],
            file_name=st.session_state.get("admin_backup_zip_name", f"hidamari_backup_{format_now_jst('%Y%m%d_%H%M%S')}.zip"),
            mime="application/zip",
            use_container_width=True,
        )

    with st.expander("個別ファイルでダウンロードする"):
        if HEALTH_FILE.exists():
            with open(HEALTH_FILE, "rb") as f:
                st.download_button(
                    "健康チェックデータをダウンロード",
                    data=f,
                    file_name="health_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        if EXCRETION_FILE.exists():
            with open(EXCRETION_FILE, "rb") as f:
                st.download_button(
                    "排泄チェックデータをダウンロード",
                    data=f,
                    file_name="excretion_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        st.download_button(
            "利用者マスタをダウンロード",
            data=export_user_master_excel_bytes(),
            file_name="user_master.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        if HANDOVER_FILE.exists():
            with open(HANDOVER_FILE, "rb") as f:
                st.download_button(
                    "業務全体申し送りデータをダウンロード",
                    data=f,
                    file_name="business_handover_data.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )

        short_goal_files = [
            (SHORT_GOAL_MASTER_FILE, "短期目標マスタをダウンロード", "short_goal_master.xlsx"),
            (SHORT_GOAL_CHECK_FILE, "短期目標実施チェックをダウンロード", "short_goal_check_data.xlsx"),
            (MONITORING_DRAFT_FILE, "モニタリング下書きをダウンロード", "monitoring_draft_data.xlsx"),
        ]
        for file_path, label, file_name in short_goal_files:
            if file_path.exists():
                with open(file_path, "rb") as f:
                    st.download_button(
                        label,
                        data=f,
                        file_name=file_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )


# =========================
# レポート系
# =========================
def create_family_summary_text(health_df, excretion_df, user_name, year, month):
    """
    家族向け・管理者確認向けの月間文章を作成する。

    Ver4.8.3:
    - 1件だけを拾って終わらないよう、対象月の全健康記録・全排泄記録を確認する。
    - 気になる記録がある場合は日付ごとに列挙する。
    - 気になる記録がない場合も「確認されませんでした」と明記する。
    - 医療判断ではなく、記録整理・継続確認点として表現する。
    """
    target = get_month_health_data(health_df, user_name, year, month)
    ex_target = get_month_excretion_data(excretion_df, user_name, year, month)

    def _fmt_date(value):
        try:
            dt = pd.to_datetime(value, errors="coerce")
            if pd.isna(dt):
                return "日付不明"
            return dt.strftime("%m/%d")
        except Exception:
            return "日付不明"

    def _clean(value):
        try:
            if pd.isna(value):
                return ""
        except Exception:
            pass
        text = str(value).strip()
        if text.lower() in ["nan", "none", "nat"]:
            return ""
        return text

    def _short(value, limit=80):
        text = _clean(value).replace("\n", " ")
        return text[:limit] + ("…" if len(text) > limit else "")

    def _num(value):
        try:
            return float(value)
        except Exception:
            try:
                s = _clean(value).replace("％", "").replace("%", "")
                return float(s) if s else None
            except Exception:
                return None

    def _record_alerts_from_health_row(row):
        """健康チェック1行から、家族向けに強すぎない確認ポイントを作る。"""
        alerts = []
        temp = _num(row.get("体温", ""))
        spo2 = _num(row.get("SpO2", ""))
        bp_high = _num(row.get("血圧上", ""))
        bp_low = _num(row.get("血圧下", ""))
        pulse = _num(row.get("脈拍", ""))
        water = _num(row.get("水分摂取量ml", ""))

        if temp is not None and temp >= 37.5:
            alerts.append(f"体温{temp:g}℃")
        if spo2 is not None and spo2 <= 93:
            alerts.append(f"SpO2 {spo2:g}％")
        if bp_high is not None and bp_high >= 160:
            alerts.append(f"血圧上{bp_high:g}")
        if bp_low is not None and bp_low >= 100:
            alerts.append(f"血圧下{bp_low:g}")
        if pulse is not None and (pulse >= 100 or pulse <= 50):
            alerts.append(f"脈拍{pulse:g}")
        if water is not None and 0 < water < 800:
            alerts.append(f"水分摂取量{water:g}ml")

        for meal_col in ["朝食摂取率", "昼食摂取率", "夕食摂取率"]:
            meal = _num(row.get(meal_col, ""))
            if meal is not None and meal <= 50:
                alerts.append(f"{meal_col.replace('摂取率', '')}{meal:g}％")

        nutrition = _clean(row.get("栄養リスク", ""))
        if nutrition and not nutrition.startswith("0") and nutrition not in ["通常"]:
            alerts.append(f"栄養リスク：{_short(nutrition, 30)}")

        oral = _clean(row.get("口腔状態", ""))
        if oral and not oral.startswith("0") and oral not in ["通常"]:
            alerts.append(f"口腔状態：{_short(oral, 30)}")

        denture = _clean(row.get("義歯使用", ""))
        if denture and ("不具合" in denture or denture.startswith("2")):
            alerts.append(f"義歯：{_short(denture, 30)}")

        family_memo = _clean(row.get("家族共有メモ", ""))
        if family_memo:
            alerts.append(f"家族共有メモ「{_short(family_memo)}」")

        change = _clean(row.get("気になる変化", ""))
        if change:
            alerts.append(f"気になる変化「{_short(change)}」")

        return alerts

    def _record_alerts_from_excretion_row(row):
        alerts = []
        slot = _clean(row.get("時間帯", ""))
        prefix = f"{slot}：" if slot else ""
        urine_type = _clean(row.get("尿性状", ""))
        stool_type = _clean(row.get("便性状", ""))
        stool_amount = _clean(row.get("便量", ""))
        ex_memo = _clean(row.get("排泄メモ", ""))

        if urine_type == "濃縮尿":
            alerts.append(prefix + "濃縮尿")
        if stool_type in ["下痢便", "水様便"]:
            alerts.append(prefix + stool_type)
        if stool_amount and stool_amount != "なし" and stool_type and stool_type != "普通便":
            text = prefix + f"便量{stool_amount}・{stool_type}"
            if text not in alerts:
                alerts.append(text)
        if ex_memo:
            alerts.append(prefix + f"排泄メモ「{_short(ex_memo)}」")
        return alerts

    lines = []
    lines.append(f"【対象期間の全体確認】\n{user_name}の{year}年{month}月の記録全体を確認しました。")
    lines.append("この文章は医療的な判断ではなく、日々の記録をもとにした共有・確認用の整理です。")

    # 健康チェックの概要
    if target.empty:
        lines.append(f"【数値から見た状態】\n{user_name}の{year}年{month}月分の健康チェック記録は、現時点では登録されていません。")
    else:
        try:
            target = target.copy()
            target["記録日"] = pd.to_datetime(target["記録日"], errors="coerce")
            target = target.sort_values("記録日")
        except Exception:
            pass

        temp_mean = to_number(target["体温"]).mean() if "体温" in target.columns else pd.NA
        spo2_mean = to_number(target["SpO2"]).mean() if "SpO2" in target.columns else pd.NA
        weight_mean = to_number(target["体重"]).mean() if "体重" in target.columns else pd.NA

        health_parts = []
        if not pd.isna(temp_mean):
            health_parts.append(f"体温平均{round(float(temp_mean), 1)}℃")
        if not pd.isna(spo2_mean):
            health_parts.append(f"SpO2平均{round(float(spo2_mean), 1)}％")
        if not pd.isna(weight_mean):
            health_parts.append(f"体重平均{round(float(weight_mean), 1)}kg")

        meal_parts = []
        for label in ["朝食摂取率", "昼食摂取率", "夕食摂取率"]:
            if label in target.columns:
                mean = to_number(target[label]).mean()
                if not pd.isna(mean):
                    meal_parts.append(f"{label.replace('摂取率', '')}平均{round(float(mean), 1)}％")

        overview_lines = [f"健康チェック記録は{len(target)}件確認されています。"]
        if health_parts:
            overview_lines.append("記録上、" + "、".join(health_parts) + "として確認されています。")
        if meal_parts:
            overview_lines.append("食事摂取率は、" + "、".join(meal_parts) + "でした。")
        lines.append("【数値から見た状態】\n" + "\n".join(overview_lines))

    # 気になる記録を日付ごとに全件確認
    concern_lines = []
    if not target.empty:
        for _, row in target.iterrows():
            row_alerts = _record_alerts_from_health_row(row)
            if row_alerts:
                concern_lines.append(f"{_fmt_date(row.get('記録日'))}　" + "、".join(row_alerts))

    if not ex_target.empty:
        try:
            ex_target = ex_target.copy()
            ex_target["記録日"] = pd.to_datetime(ex_target["記録日"], errors="coerce")
            ex_target = ex_target.sort_values(["記録日", "時間帯"])
        except Exception:
            pass
        for _, row in ex_target.iterrows():
            row_alerts = _record_alerts_from_excretion_row(row)
            if row_alerts:
                concern_lines.append(f"{_fmt_date(row.get('記録日'))}　" + "、".join(row_alerts))

    if concern_lines:
        # 同じ文章の重複を避ける
        unique_concerns = []
        for item in concern_lines:
            if item not in unique_concerns:
                unique_concerns.append(item)
        lines.append("【気になる記録】\n対象期間中の記録から、次の内容を確認しました。\n" + "\n".join([f"・{x}" for x in unique_concerns[:20]]))
        if len(unique_concerns) > 20:
            lines.append(f"※気になる記録が多いため、上位20件まで表示しています。残り{len(unique_concerns) - 20}件は一覧で確認してください。")
    else:
        lines.append(
            "【気になる記録】\n対象期間中の健康チェック・排泄記録を確認しました。"
            "その他の日について、発熱、SpO2低下、著しい食事低下、濃縮尿、下痢便・水様便、"
            "睡眠不良や生活上の気になる記録は確認されませんでした。"
        )

    # 排泄状況
    if ex_target.empty:
        lines.append("【排泄状況】\n排泄記録は、対象月にはまだ登録されていません。")
    else:
        ex_sum = summarize_excretion(ex_target)
        ex_lines = [
            f"排尿記録{ex_sum['排尿回数']}回、排便記録{ex_sum['排便回数']}回、濃縮尿{ex_sum['濃縮尿']}回、下痢便{ex_sum['下痢便']}回、水様便{ex_sum['水様便']}回として記録されています。"
        ]
        if int(ex_sum.get("排便回数", 0) or 0) == 0:
            ex_lines.append("記録上、対象期間中の排便記録が確認できないため、便秘傾向の有無を継続して確認します。")
        elif int(ex_sum.get("濃縮尿", 0) or 0) > 0 or int(ex_sum.get("下痢便", 0) or 0) > 0 or int(ex_sum.get("水様便", 0) or 0) > 0:
            ex_lines.append("排泄の性状に気になる記録がある日は、水分摂取量、食事量、腹部の様子、普段との差を合わせて確認します。")
        else:
            ex_lines.append("排泄記録上、濃縮尿、下痢便、水様便などの特記は確認されませんでした。")
        lines.append("【排泄状況】\n" + "\n".join(ex_lines))

    assessment = build_assessment_context_text(user_name)
    if assessment:
        lines.append("【アセスメント情報】\nアセスメント情報もふまえ、生活全体の様子を確認しています。\n" + assessment)

    # 継続確認点：気になる記録の有無にかかわらず、現場で次に見る項目を出す
    follow_points = []
    if not target.empty:
        follow_points.extend(["食事摂取量の変化", "SpO2低下時の表情・呼吸状態", "日中の活気や眠気", "本人の普段の様子との差"])
    if not ex_target.empty:
        follow_points.extend(["排便間隔", "尿・便の性状", "水分摂取量との関係"])
    if not follow_points:
        follow_points.append("記録が入力された後、数値と生活の様子を合わせて確認します")

    # 順序を維持して重複削除
    dedup_points = []
    for p in follow_points:
        if p not in dedup_points:
            dedup_points.append(p)
    lines.append("【継続確認が必要な点】\n" + "\n".join([f"・{p}" for p in dedup_points]))

    lines.append("【職員間で共有すること】\n数値だけで判断せず、表情、食事、水分、排泄、睡眠、活動量など、本人の普段の様子との差を職員間で共有しながら見守ります。")

    return "\n\n".join(lines)


def create_hidamari_pdf(health_df, excretion_df, user_name, year, month):
    if colors is None:
        raise RuntimeError("reportlab が利用できません。requirements.txt に reportlab を追加してください。")

    gothic, mincho = register_japanese_pdf_fonts()

    file_path = REPORT_DIR / f"ひだまりレポート_{user_name}_{year}年{month}月.pdf"

    doc = SimpleDocTemplate(
        str(file_path),
        pagesize=A4,
        rightMargin=17 * mm,
        leftMargin=17 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "jp_title",
        parent=styles["Title"],
        fontName=gothic,
        fontSize=22,
        leading=28,
        alignment=1,
        textColor=colors.HexColor("#2F3437"),
    )
    h2_style = ParagraphStyle(
        "jp_h2",
        parent=styles["Heading2"],
        fontName=gothic,
        fontSize=13,
        leading=18,
        textColor=colors.HexColor("#2F3437"),
    )
    body_style = ParagraphStyle(
        "jp_body",
        parent=styles["BodyText"],
        fontName=mincho,
        fontSize=10,
        leading=16,
    )
    small_style = ParagraphStyle(
        "jp_small",
        parent=styles["BodyText"],
        fontName=mincho,
        fontSize=8,
        leading=12,
        textColor=colors.HexColor("#666666"),
    )

    story = []
    story.append(Paragraph(HIDAMARI_REPORT_TITLE, title_style))
    story.append(Spacer(1, 6))
    story.append(Paragraph(f"{user_name}　{year}年{month}月", body_style))
    story.append(Spacer(1, 12))

    summary = create_family_summary_text(health_df, excretion_df, user_name, year, month)
    story.append(Paragraph(HIDAMARI_REPORT_SUMMARY_HEADING, h2_style))
    for para in summary.split("\n\n"):
        story.append(Paragraph(para.replace("\n", "<br/>"), body_style))
        story.append(Spacer(1, 5))

    story.append(PageBreak())
    story.append(Paragraph(HIDAMARI_REPORT_EXCRETION_HEADING, h2_style))

    ex_target = get_month_excretion_data(excretion_df, user_name, year, month)
    if ex_target.empty:
        story.append(Paragraph(HIDAMARI_REPORT_NO_EXCRETION_TEXT, body_style))
    else:
        table_data = [["日付", "時間帯", "尿", "便", "メモ"]]
        for _, row in ex_target.iterrows():
            table_data.append([
                row["記録日"].strftime("%m/%d") if pd.notna(row["記録日"]) else "",
                row["時間帯"],
                f"{row['尿量']}・{row['尿性状']}",
                f"{row['便量']}・{row['便性状']}",
                str(row.get("排泄メモ", ""))[:40],
            ])

        table = Table(table_data, colWidths=[20*mm, 24*mm, 35*mm, 35*mm, 55*mm])
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), mincho),
            ("FONTNAME", (0, 0), (-1, 0), gothic),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F7F4EE")),
            ("BOX", (0, 0), (-1, -1), 0.5, colors.HexColor("#D9D9D9")),
            ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9D9D9")),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ]))
        story.append(table)

    story.append(Spacer(1, 12))
    story.append(Paragraph(HIDAMARI_REPORT_DISCLAIMER, small_style))

    doc.build(story)
    return file_path


def create_handover_text(health_df, excretion_df, target_date):
    lines = [
        f"{target_date.strftime('%Y/%m/%d')}の申し送りまとめです。",
        "医療的な判断ではなく、記録内容をもとにした共有用メモです。",
        "",
    ]

    h = health_df.copy()
    if not h.empty:
        h["記録日"] = pd.to_datetime(h["記録日"], errors="coerce")
        h = h[h["記録日"].dt.date == target_date]

        for _, row in h.iterrows():
            notes = []
            if clean_text(row.get("気になる変化", "")):
                notes.append(f"気になる変化：{row.get('気になる変化')}")
            if clean_text(row.get("家族共有メモ", "")):
                notes.append(f"家族共有メモ：{row.get('家族共有メモ')}")

            vital_alerts = []
            if safe_float(row.get("体温"), 0) >= 37.5:
                vital_alerts.append("体温高め")
            if safe_int(row.get("SpO2"), 100) <= 93:
                vital_alerts.append("SpO2低め")
            if safe_int(row.get("血圧上"), 0) >= 160:
                vital_alerts.append("血圧上高め")

            if vital_alerts:
                notes.append("確認目安：" + "、".join(vital_alerts))

            if notes:
                lines.append(f"■ {row.get('利用者名')}")
                lines.extend([f"・{x}" for x in notes])
                lines.append("")

    e = get_day_excretion_data(excretion_df, target_date, None)
    if not e.empty:
        for user in e["利用者名"].dropna().unique():
            user_ex = e[e["利用者名"] == user]
            alerts = []

            for _, row in user_ex.iterrows():
                if row["尿性状"] == "濃縮尿":
                    alerts.append(f"{row['時間帯']}に濃縮尿")
                if row["便性状"] in ["下痢便", "水様便"]:
                    alerts.append(f"{row['時間帯']}に{row['便性状']}")

            if alerts:
                lines.append(f"■ {user} 排泄確認")
                lines.append("・" + "、".join(alerts))
                lines.append("")

    diff_lines = []
    for user in active_users:
        hdiff = build_health_diff_text(health_df, target_date, user)
        ediff = build_excretion_diff_text(excretion_df, target_date, user)

        if "大きな差分は目立ちません" not in hdiff and "比較できる過去記録はありません" not in hdiff and "本日の健康記録がありません" not in hdiff:
            diff_lines.append(f"■ {user} {hdiff}")

        if "大きな変化は目立ちません" not in ediff and "比較できる過去排泄記録はありません" not in ediff:
            diff_lines.append(f"■ {user} {ediff}")

    if diff_lines:
        lines.append("【前回との差分】")
        lines.extend(diff_lines)
        lines.append("")

    if len(lines) <= 3:
        lines.append("記録上、特に申し送り対象となるメモや注意目安はありません。")

    lines.append("引き続き、普段との違いがないかを確認しながら見守ります。")
    return "\n".join(lines)




# =========================
# 今日のひだまりメッセージ（ランダム表示）
# =========================
HIDAMARI_MESSAGES = [
'小さな記録が、誰かの安心につながります。',
'今日のひと声が、利用者様の一日をやわらかくします。',
'あわてず、ひとつずつ。記録はチームの安心です。',
'いつもの中の小さな違いに、そっと気づけますように。',
'今日も無理なく、やさしいケアでいきましょう。',
'申し送りは、責任を背負うためではなく、チームで分け合うためにあります。',
'気になった時点で、もう大事な気づきです。',
'なんとなく変かも、は立派な観察です。',
'小さな違和感を残せる職場は、利用者様を守る力があります。',
'完璧より、安心できる空気を大切に。',
'急がない声かけが、安心につながる日もあります。',
'いつも通りを守ることも、立派な支援です。',
'今日の記録は、明日の誰かを助けます。',
'一人で抱えず、チームで見守りましょう。',
'深呼吸する時間も、ケアのうちです。',
'利用者様だけでなく、自分の体調も大切に。',
'焦らなくても大丈夫。記録は積み重なります。',
'やさしさは、入力欄の外にもあります。',
'声のトーンひとつで、安心は伝わります。',
'今日の笑顔は、今日の大切なケアです。',
'見る、聞く、待つ。それだけで支援になる時間があります。',
'できたことを一つ見つける日でありますように。',
'記録は評価ではなく、安心をつなぐメモです。',
'昨日との違いに気づけたら、それは大きな一歩です。',
'急がず、でも見逃さず。今日も丁寧にいきましょう。',
'利用者様のペースを尊重できることは、強いケアです。',
'困った時は、早めに共有。それだけで事故は減らせます。',
'小さな確認が、大きな安心につながります。',
'今日も現場を支えるあなたの力があります。',
'声をかける前に一呼吸。それだけで伝わり方が変わります。',
'記録を残すことは、利用者様を一人にしないことです。',
'いつもの表情を知っている職員さんは、現場の宝です。',
'大きなことをしなくても、そばにいることが支援です。',
'今日のケアが、明日の落ち着きにつながることがあります。',
'迷ったら、抱え込まずに共有しましょう。',
'気づきを言葉にすることが、チームケアの第一歩です。',
'静かな見守りも、確かなケアです。',
'今日も、利用者様の生活の一部を支えています。',
'少しの変化を見つける目が、安心を作ります。',
'申し送りは、次の職員さんへの思いやりです。',
'記録は冷たい作業ではなく、あたたかい引き継ぎです。',
'今日の小さなメモが、ご家族の安心にもつながります。',
'ケアは、正解よりも関わり続けることが大切な日があります。',
'利用者様の『その人らしさ』を、今日も少し残していきましょう。',
'ゆっくり話を聞けた時間は、立派な支援です。',
'できないことより、できていることにも目を向けて。',
'焦る日ほど、基本に戻りましょう。',
'今日も一つずつ、確認していけば大丈夫です。',
'体調の変化だけでなく、気持ちの変化にもそっと気づけますように。',
'誰かに相談できることも、専門職の力です。',
'記録があるから、チームで同じ方向を見られます。',
'小さな『ありがとう』を受け取れる日でありますように。',
'いつもの暮らしを守ることは、大きな仕事です。',
'慌ただしい日ほど、やさしい言葉を一つ。',
'今日の安全確認が、安心した夜につながります。',
'気づいたことを残すだけで、次のケアが変わります。',
'職員さんの気づきは、システムより大切な情報です。',
'機械は記録を整理します。安心を作るのは人です。',
'入力は短くても大丈夫。残すことに意味があります。',
'ひとつの記録が、ひとつの見守りになります。',
'利用者様の安心は、毎日の小さな積み重ねから生まれます。',
'声をかける、待つ、見守る。どれも大切なケアです。',
'今日も、無理をしすぎないケアでいきましょう。',
'『いつもと違う』を大切にできる現場は強いです。',
'小さな違いを責めず、そっと共有しましょう。',
'今日の記録は、未来の安心の材料です。',
'手を止めて見る時間も、ケアの時間です。',
'利用者様の表情を思い出しながら記録してみましょう。',
'チームで見れば、気づきはもっとやさしくなります。',
'忙しい中の一言が、利用者様の心を軽くすることがあります。',
'できるだけやさしく、できるだけ正確に。今日もそれで十分です。',
'安全は、細かな確認から育ちます。',
'生活を支える仕事は、目立たなくても深い価値があります。',
'今日の観察が、明日の対応を助けます。',
'記録を残すことは、ケアを見える形にすることです。',
'利用者様の一日は、職員さんの気づきで守られています。',
'『大丈夫かな』と思ったら、残しておきましょう。',
'確認したことも、立派な申し送りです。',
'何もなかった日も、見守った日です。',
'普段通りを確認できることも、安心材料です。',
'今日も、自分を責めすぎず、丁寧に。',
'現場のやさしさは、細かい記録にも表れます。',
'利用者様の安心と、職員さんの安心。どちらも大切です。',
'よく見ることは、よく支えることにつながります。',
'一つの気づきが、転倒予防につながることがあります。',
'一つの声かけが、不安の軽減につながることがあります。',
'食事、排泄、表情。生活の中に大事な情報があります。',
'今日のメモは、明日の会話のきっかけになります。',
'大きな変化でなくても、残してよいのです。',
'職員さんの『気になる』は、現場のセンサーです。',
'急がず、でも流さず。今日もやさしく確認を。',
'穏やかな声は、安心の環境づくりです。',
'ケアは一人で完成しません。つなぐことで強くなります。',
'今日も、利用者様の暮らしをそっと支えています。',
'ボーダーコリーのように、落ち着いて見守る日でありますように。',
'猫のように、少しゆっくりでも大丈夫です。',
'ひなたぼっこのような空気を、今日も少しだけ。',
'尻尾を振るような安心感を、今日の現場にも。',
'やさしい見守りは、ちゃんと伝わっています。',
'忙しさの中にも、ひだまりの時間を一つ。',
]


def get_hidamari_message():
    """ログインごとに1つ選ばれる、介護の心得メッセージ。"""
    if "hidamari_login_message" not in st.session_state:
        st.session_state["hidamari_login_message"] = random.choice(HIDAMARI_MESSAGES)
    return st.session_state["hidamari_login_message"]


# =========================
# あたたかい画面デザイン
# =========================
def show_hidamari_hero(mode="login"):
    """ログイン画面と各ダッシュボード上部の案内表示。"""
    if mode == "login":
        title = "ひだまり 健康チェック管理システム"
        sub = "利用者様に寄り添うために"

        # ログイン画面は、インラインCSSで確実に画面中央へ配置する。
        st.markdown(
            f"""
            <div style="
                width: 100%;
                display: flex;
                justify-content: center;
                align-items: center;
                text-align: center;
                margin: 34px auto 22px auto;
            ">
                <div style="
                    width: min(760px, 92vw);
                    background: linear-gradient(135deg, #F7F2EA 0%, #EEF5EF 58%, #EAF1F5 100%);
                    border: 1px solid rgba(88, 112, 96, 0.16);
                    border-radius: 28px;
                    padding: 28px 26px;
                    box-shadow: 0 10px 28px rgba(55, 64, 58, 0.08);
                ">
                    <div style="font-size: 2.1rem; line-height: 1.25; font-weight: 800; color: #2F6F5E; letter-spacing: 0.02em;">
                        {title}
                    </div>
                    <div style="font-size: 1.15rem; color: #64706A; margin-top: 10px;">
                        {sub}
                    </div>
                </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        return

    elif mode == "staff":
        title = "今日のひだまりメッセージ"
        sub = get_hidamari_message()
    else:
        title = "ひだまり 管理者ダッシュボード"
        sub = get_hidamari_message()

    st.markdown(
        f"""
        <div class="hidamari-hero">
            <div class="hidamari-hero-title">{title}</div>
            <div class="hidamari-hero-sub">{sub}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def show_staff_encouragement():
    # 上部の「今日もお疲れ様です」カードを残すため、
    # ここでは追加メッセージを表示しない。
    return



def show_admin_encouragement():
    st.markdown(
        """
        <div class="admin-welcome">
            <b> 管理者モードです。</b><br>
            入力状況、注意記録、申し送りを確認できます。現場の気づきを、管理の力に変えていきましょう。
        </div>
        """,
        unsafe_allow_html=True,
    )




# =========================
# 現場の気づき構造化・AI管理者支援 Ver1.3追加
# =========================
AI_INSIGHT_LOG_FILE = DATA_DIR / "ai_insight_log.xlsx"
AI_INSIGHT_LOG_COLUMNS = ["作成日時", "分析基準日", "利用者名", "対象期間", "ルール分析", "AI分析結果"]



def ensure_ai_insight_log_file():
    """
    AI分析ログをSQLiteで管理する。
    旧Excelがある場合のみ初回移行し、以後はSQLiteを正とする。
    """
    ensure_dirs()
    if sqlite_table_row_count(SQLITE_TABLE_AI_INSIGHT_LOGS) > 0:
        return

    df = pd.DataFrame(columns=AI_INSIGHT_LOG_COLUMNS)
    if AI_INSIGHT_LOG_FILE.exists():
        try:
            df = pd.read_excel(AI_INSIGHT_LOG_FILE, sheet_name="AI分析ログ")
        except Exception:
            try:
                df = pd.read_excel(AI_INSIGHT_LOG_FILE)
            except Exception:
                df = pd.DataFrame(columns=AI_INSIGHT_LOG_COLUMNS)

    df = normalize_df_columns(df, AI_INSIGHT_LOG_COLUMNS)
    save_sqlite_table(df, SQLITE_TABLE_AI_INSIGHT_LOGS, AI_INSIGHT_LOG_COLUMNS, sort_cols=["作成日時"])


def load_ai_insight_logs():
    ensure_ai_insight_log_file()
    return load_sqlite_table(SQLITE_TABLE_AI_INSIGHT_LOGS, AI_INSIGHT_LOG_COLUMNS).astype("object")


def save_ai_insight_logs(df):
    ensure_dirs()
    df = normalize_df_columns(df, AI_INSIGHT_LOG_COLUMNS)
    save_sqlite_table(df, SQLITE_TABLE_AI_INSIGHT_LOGS, AI_INSIGHT_LOG_COLUMNS, sort_cols=["作成日時"])



def build_ai_insight_pdf_report(row, report_title=AI_INSIGHT_REPORT_TITLE):
    """
    AI分析ログ1件を、管理者確認用PDFとして出力する。
    診断・医療判断ではなく、記録整理と確認ポイントのレポートとして扱う。
    """
    if colors is None:
        raise RuntimeError("ReportLabが利用できません。requirements.txt に reportlab を追加してください。")

    gothic, mincho = register_japanese_pdf_fonts()

    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=14 * mm,
        leftMargin=14 * mm,
        topMargin=14 * mm,
        bottomMargin=14 * mm,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "AiReportTitle",
        parent=styles["Title"],
        fontName=gothic,
        fontSize=16,
        leading=22,
        alignment=1,
        spaceAfter=8 * mm,
    )
    section_style = ParagraphStyle(
        "AiReportSection",
        parent=styles["Heading2"],
        fontName=gothic,
        fontSize=12,
        leading=16,
        textColor=colors.HexColor("#2F6F5E"),
        spaceBefore=5 * mm,
        spaceAfter=3 * mm,
    )
    body_style = ParagraphStyle(
        "AiReportBody",
        parent=styles["BodyText"],
        fontName=mincho,
        fontSize=9.5,
        leading=14,
        spaceAfter=2 * mm,
    )
    small_style = ParagraphStyle(
        "AiReportSmall",
        parent=styles["BodyText"],
        fontName=mincho,
        fontSize=8,
        leading=11,
        textColor=colors.HexColor("#555555"),
    )
    table_style = ParagraphStyle(
        "AiReportTable",
        parent=styles["BodyText"],
        fontName=mincho,
        fontSize=8.5,
        leading=12,
    )

    def cell(v):
        return Paragraph(pdf_safe_text(v).replace("\n", "<br/>"), table_style)

    created_at = row.get("作成日時", "") if isinstance(row, dict) else ""
    base_day = row.get("分析基準日", "") if isinstance(row, dict) else ""
    user_name = row.get("利用者名", "") if isinstance(row, dict) else ""
    period = row.get("対象期間", "") if isinstance(row, dict) else ""
    rule_text = row.get("ルール分析", "") if isinstance(row, dict) else ""
    ai_text = row.get("AI分析結果", "") if isinstance(row, dict) else ""

    story = []
    story.append(Paragraph(report_title, title_style))
    story.append(Paragraph(AI_INSIGHT_REPORT_DISCLAIMER, small_style))
    story.append(Spacer(1, 4 * mm))

    meta_data = [
        [cell("作成日時"), cell(created_at), cell("分析基準日"), cell(base_day)],
        [cell("利用者名"), cell(user_name), cell("対象期間"), cell(period)],
    ]
    meta_table = Table(meta_data, colWidths=[26 * mm, 56 * mm, 28 * mm, 62 * mm])
    meta_table.setStyle(TableStyle([
        ("FONTNAME", (0, 0), (-1, -1), mincho),
        ("BACKGROUND", (0, 0), (0, -1), colors.HexColor("#F3F7F5")),
        ("BACKGROUND", (2, 0), (2, -1), colors.HexColor("#F3F7F5")),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("RIGHTPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story.append(meta_table)

    story.append(Paragraph("1. ルールベース分析", section_style))
    story.extend(paragraph_lines(rule_text, body_style))

    story.append(Paragraph("2. AI管理者アドバイス", section_style))
    story.extend(paragraph_lines(ai_text, body_style))

    story.append(Spacer(1, 5 * mm))
    story.append(Paragraph(AI_INSIGHT_CONFIRMATION_NOTE, small_style))

    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()


def build_ai_insight_log_summary_pdf(logs_df):
    """AI分析ログ一覧をPDFで出力する。"""
    if colors is None:
        raise RuntimeError("ReportLabが利用できません。requirements.txt に reportlab を追加してください。")
    gothic, mincho = register_japanese_pdf_fonts()

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=10 * mm, leftMargin=10 * mm, topMargin=12 * mm, bottomMargin=12 * mm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("AiLogTitle", parent=styles["Title"], fontName=gothic, fontSize=15, leading=20, alignment=1)
    small_style = ParagraphStyle("AiLogSmall", parent=styles["BodyText"], fontName=mincho, fontSize=7.5, leading=10)
    story = [Paragraph(AI_INSIGHT_LOG_SUMMARY_TITLE, title_style), Spacer(1, 5 * mm)]

    if logs_df is None or logs_df.empty:
        story.append(Paragraph(AI_INSIGHT_LOG_EMPTY_TEXT, small_style))
    else:
        show = logs_df.tail(30).copy()
        data = [[Paragraph("作成日時", small_style), Paragraph("利用者名", small_style), Paragraph("対象期間", small_style), Paragraph("AI分析結果（冒頭）", small_style)]]
        for _, r in show.iterrows():
            ai_short = pdf_safe_text(r.get("AI分析結果", ""))[:180]
            data.append([
                Paragraph(pdf_safe_text(r.get("作成日時", "")), small_style),
                Paragraph(pdf_safe_text(r.get("利用者名", "")), small_style),
                Paragraph(pdf_safe_text(r.get("対象期間", "")), small_style),
                Paragraph(ai_short, small_style),
            ])
        table = Table(data, colWidths=[32 * mm, 24 * mm, 42 * mm, 90 * mm], repeatRows=1)
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), mincho),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F3F7F5")),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CCCCCC")),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 4),
            ("RIGHTPADDING", (0, 0), (-1, -1), 4),
            ("TOPPADDING", (0, 0), (-1, -1), 4),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        story.append(table)
    doc.build(story)
    buffer.seek(0)
    return buffer.getvalue()

def generate_ai_structured_advice(context):
    api_key = get_openai_api_key("")
    if not api_key:
        return ""
    try:
        from openai import OpenAI
    except Exception:
        return "OpenAIライブラリが未インストールです。requirements.txt に openai を追加してください。"

    system_prompt = AI_STRUCTURED_ADVICE_SYSTEM_PROMPT
    try:
        client = OpenAI(api_key=api_key)
        res = client.chat.completions.create(
            model=get_openai_model("admin", "gpt-4.1-mini"),
            messages=[
                {"role": "system", "content": system_prompt},
                {"role": "user", "content": context},
            ],
            temperature=0.2,
        )
        return res.choices[0].message.content or ""
    except Exception as e:
        return f"AI分析中にエラーが発生しました：{e}"


def show_structured_insight_menu():
    if not is_admin_user():
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("現場の気づき構造化・AI管理者支援")
    show_observation_perspective("ai")
    st.caption("健康チェック・排泄・申し送り・短期目標を合わせて、管理者が確認しやすい形に整理します。AIは診断せず、記録上の気づきと確認ポイントだけを出します。")

    users = get_active_user_names()
    c1, c2, c3 = st.columns([1.2, 1, 1])
    with c1:
        user_name = st.selectbox("分析する利用者", users, key="structured_insight_user")
    with c2:
        period_mode = st.selectbox(
            "分析期間",
            ["直近7日", "直近14日", "直近30日", "期間指定"],
            index=0,
            key="structured_insight_period_mode",
            help="体調変化は7日、短期目標や申し送り傾向は14〜30日を見ると整理しやすくなります。",
        )
    with c3:
        end_day = st.date_input(
            "分析基準日" if period_mode != "期間指定" else "終了日",
            value=today_jst(),
            key="structured_insight_end_day",
        )

    if period_mode == "期間指定":
        default_start = end_day - timedelta(days=6)
        start_day = st.date_input(
            "開始日",
            value=default_start,
            key="structured_insight_start_day",
        )
        period_label = "期間指定"
    else:
        period_days_map = {"直近7日": 7, "直近14日": 14, "直近30日": 30}
        period_days = period_days_map.get(period_mode, 7)
        start_day = end_day - timedelta(days=period_days - 1)
        period_label = period_mode
        st.caption(f"対象期間：{start_day}〜{end_day}（{period_label}）")

    if start_day > end_day:
        st.warning("開始日が終了日より後になっています。日付を入れ替えて分析します。")
        start_day, end_day = end_day, start_day

    # Ver4.8.3 修正：
    # ここでは target_date という変数は定義されていないため、
    # 画面で選択した分析基準日 end_day を使う。
    # 直近データだけを読み、全件取得による重さを避ける。
    dash_end = end_day
    dash_start = dash_end - timedelta(days=14)
    health_df = load_health_data(start_date=dash_start, end_date=dash_end)
    ex_df = load_excretion_data(start_date=dash_start, end_date=dash_end)
    if 'load_business_handover_data' in globals():
        handover_df = load_business_handover_data(start_date=dash_start, end_date=dash_end)
    else:
        handover_df = read_excel_safe(HANDOVER_FILE, BUSINESS_HANDOVER_COLUMNS)
        handover_df = _filter_df_by_date_range(handover_df, "日付", dash_start, dash_end)
    goal_df = load_short_goal_checks(start_date=dash_start, end_date=dash_end)

    result = analyze_structured_insights(
        health_df,
        ex_df,
        handover_df,
        goal_df,
        user_name,
        end_day,
        start_day=start_day,
        period_label=period_label,
    )

    st.markdown("### ルールベース分析（APIなしでも使用可）")
    left, right = st.columns(2)
    with left:
        st.subheader("記録上の気づき")
        for item in result["findings"]:
            st.write("・" + item)
    with right:
        st.subheader("管理者の確認ポイント")
        for item in result["checks"]:
            st.write("・" + item)

    if result["goal_summary"]:
        st.subheader("短期目標 実施状況")
        st.dataframe(pd.DataFrame({"短期目標分析": result["goal_summary"]}), use_container_width=True, hide_index=True)

    with st.expander("分析対象データを確認", expanded=False):
        st.write(f"対象期間：{start_day}〜{end_day}（{period_label}）")
        st.markdown("#### 健康チェック")
        st.dataframe(filter_records_by_period(health_df, "記録日", start_day, end_day, user_name), use_container_width=True, hide_index=True)
        st.markdown("#### 排泄")
        st.dataframe(filter_records_by_period(ex_df, "記録日", start_day, end_day, user_name), use_container_width=True, hide_index=True)
        st.markdown("#### 短期目標")
        st.dataframe(filter_records_by_period(goal_df, "日付", start_day, end_day, user_name), use_container_width=True, hide_index=True)

    st.markdown("---")
    st.markdown("### AI管理者アドバイス")
    st.caption(f"使用モデル：{get_openai_model('admin', 'gpt-4.1-mini')}")
    if not get_openai_api_key(""):
        st.warning("OpenAI APIキーが未設定です。Streamlit Cloud の Secrets に OPENAI_API_KEY を設定するとAI分析が使えます。")
        st.code('OPENAI_API_KEY = "sk-xxxxxxxxxxxxxxxx"', language="toml")
    else:
        if st.button("AIで管理者向けアドバイスを生成", use_container_width=True):
            context = build_ai_structured_context(
                health_df,
                ex_df,
                handover_df,
                goal_df,
                user_name,
                end_day,
                result,
                start_day=start_day,
                period_label=period_label,
            )
            with st.spinner("AIが記録を整理しています..."):
                ai_text = generate_ai_structured_advice(context)
            st.subheader("AI分析結果")
            st.write(ai_text)

            logs = load_ai_insight_logs()
            rule_text = "\n".join(["【気づき】"] + result["findings"] + ["【確認ポイント】"] + result["checks"] + ["【短期目標】"] + result["goal_summary"])
            new_row = {
                "作成日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
                "分析基準日": end_day,
                "利用者名": user_name,
                "対象期間": f"{start_day}〜{end_day}（{period_label}）",
                "ルール分析": rule_text,
                "AI分析結果": ai_text,
            }
            logs = pd.concat([logs, pd.DataFrame([new_row])], ignore_index=True)
            save_ai_insight_logs(logs)
            st.success("AI分析ログに保存しました。")

    st.markdown("### AI分析ログ")
    logs = load_ai_insight_logs()
    if logs.empty:
        st.info("AI分析ログはまだありません。")
    else:
        show_logs = logs.tail(20).reset_index(drop=True)
        st.dataframe(show_logs, use_container_width=True, hide_index=True)

        st.markdown("#### 管理者向けPDFレポート")
        st.caption("AI分析ログを、会議・確認・保管用のPDFとして出力できます。")
        pdf_col1, pdf_col2 = st.columns([1, 1])
        with pdf_col1:
            selected_idx = st.selectbox(
                "PDFにするログ",
                list(range(len(show_logs))),
                format_func=lambda i: f"{show_logs.iloc[i].get('作成日時', '')}／{show_logs.iloc[i].get('利用者名', '')}／{show_logs.iloc[i].get('対象期間', '')}",
                key="ai_insight_pdf_selected_idx",
            )
            selected_row = show_logs.iloc[selected_idx].to_dict()
            try:
                pdf_bytes = build_ai_insight_pdf_report(selected_row)
                safe_user = re.sub(r"[^0-9A-Za-z一-龥ぁ-んァ-ンー_-]", "_", str(selected_row.get("利用者名", "user")))
                safe_dt = re.sub(r"[^0-9]", "", str(selected_row.get("作成日時", "")))[:14] or format_now_jst("%Y%m%d%H%M%S")
                st.download_button(
                    "選択したAI分析ログをPDFダウンロード",
                    data=pdf_bytes,
                    file_name=f"ai_admin_report_{safe_user}_{safe_dt}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"PDF作成に失敗しました：{e}")
        with pdf_col2:
            try:
                summary_pdf = build_ai_insight_log_summary_pdf(logs)
                st.download_button(
                    "AI分析ログ一覧PDFをダウンロード",
                    data=summary_pdf,
                    file_name=f"ai_insight_log_summary_{format_now_jst('%Y%m%d_%H%M%S')}.pdf",
                    mime="application/pdf",
                    use_container_width=True,
                )
            except Exception as e:
                st.error(f"一覧PDF作成に失敗しました：{e}")


# =========================
# ログイン・デザイン
# =========================
def show_force_password_change_screen():
    """初回ログイン・仮パスワード利用時に通常画面へ進ませず、パスワード変更を求める。"""
    show_hidamari_hero("login")
    login_id = current_login_user()
    accounts = load_accounts()
    hit = accounts[accounts["ログインID"] == login_id]
    if hit.empty:
        st.error("ログイン情報を確認できません。もう一度ログインしてください。")
        if st.button("ログイン画面へ戻る", use_container_width=True):
            st.session_state.logged_in = False
            st.rerun()
        return False

    row = hit.iloc[-1]
    st.warning("安全のため、初回ログイン時はパスワード変更が必要です。")
    st.caption("仮パスワードのまま通常画面へ進むことはできません。新しいパスワードを設定してください。")

    col1, col2, col3 = st.columns([1, 2, 1])
    with col2:
        with st.form("force_password_change_form", clear_on_submit=False):
            st.markdown("### 初回パスワード変更")
            st.text_input("ログインID", value=login_id, disabled=True)
            new_pw = st.text_input("新しいパスワード", type="password", help="8文字以上、英字と数字を両方含めてください。")
            new_pw2 = st.text_input("新しいパスワード（確認）", type="password")
            submitted = st.form_submit_button("パスワードを変更して利用開始", type="primary", use_container_width=True)

        if submitted:
            ok_pw, pw_msg = validate_new_password(login_id, new_pw, new_pw2, clean_text(row.get("パスワードハッシュ")))
            if not ok_pw:
                st.error(pw_msg)
            else:
                ok, msg = update_account_password(login_id, new_pw, force_change="いいえ")
                if ok:
                    add_login_history(login_id, st.session_state.get("user_label", ""), st.session_state.get("role", ""), "成功", "初回パスワード変更完了")
                    st.session_state.force_password_change = False
                    st.session_state["hidamari_login_message"] = "安全設定が完了しました。今日の記録をはじめられます。"
                    st.success("パスワードを変更しました。通常画面へ進みます。")
                    st.rerun()
                else:
                    st.error(msg)

        if st.button("ログアウト", use_container_width=True):
            st.session_state.logged_in = False
            st.session_state.role = None
            st.session_state.user_label = ""
            st.session_state.username = ""
            st.session_state.user_id = ""
            st.session_state.login_user = ""
            st.session_state.login_user_info = {}
            st.session_state.force_password_change = False
            st.rerun()
    return False




def render_auto_backup_download_from_session():
    """終了時バックアップ後、ログイン画面でバックアップZIPをダウンロードできるように表示する。"""
    try:
        backup_file = st.session_state.get("last_backup_file", "")
        backup_name = st.session_state.get("backup_download_name", "")
        if not backup_file:
            return
        path = Path(backup_file)
        if not path.exists():
            st.warning("バックアップファイルが見つかりません。管理者メニューのバックアップ履歴を確認してください。")
            return

        data = path.read_bytes()
        file_name = backup_name or path.name

        st.info("バックアップZIPを作成しました。下のボタンからこの端末に保存してください。")
        st.download_button(
            "バックアップZIPをダウンロード",
            data=data,
            file_name=file_name,
            mime="application/zip",
            type="primary",
            use_container_width=True,
        )

        # ブラウザ設定によって自動保存が止まることがあるため、download_buttonも必ず残す。
        try:
            b64 = base64.b64encode(data).decode("utf-8")
            components.html(
                f"""
                <script>
                (function() {{
                    const a = document.createElement('a');
                    a.href = 'data:application/zip;base64,{b64}';
                    a.download = '{file_name}';
                    document.body.appendChild(a);
                    a.click();
                    document.body.removeChild(a);
                }})();
                </script>
                """,
                height=0,
            )
        except Exception:
            pass

        st.caption("自動ダウンロードが始まらない場合は、上のボタンを押してください。")
    except Exception as e:
        st.warning(f"バックアップダウンロード表示に失敗しました：{e}")


def login_check():
    if "logged_in" not in st.session_state:
        st.session_state.logged_in = False
    if "role" not in st.session_state:
        st.session_state.role = None
    if "user_label" not in st.session_state:
        st.session_state.user_label = ""
    if "force_password_change" not in st.session_state:
        st.session_state.force_password_change = False

    if st.session_state.logged_in:
        if st.session_state.force_password_change:
            return show_force_password_change_screen()
        return True

    show_hidamari_hero("login")

    col1, col2, col3 = st.columns([1, 2, 1])

    with col2:
        st.markdown('<h3 style="text-align:center;">ログイン</h3>', unsafe_allow_html=True)

        # 終了ボタンから戻った直後の安心表示
        end_msg = st.session_state.pop("hidamari_after_logout_message", "")
        if end_msg:
            st.success(end_msg)
        if st.session_state.get("backup_download_pending") or st.session_state.get("last_backup_file"):
            render_auto_backup_download_from_session()
            st.session_state["backup_download_pending"] = False

        input_id = st.text_input("ID")
        input_password = st.text_input("パスワード", type="password")

        if st.button("ログイン", use_container_width=True):
            login_id = clean_text(input_id).lower()
            login_password = clean_text(input_password)
            locked, remaining_seconds = is_login_temporarily_locked(login_id)
            if locked:
                st.error(f"ログイン失敗が続いたため、一時的に制限しています。約{max(remaining_seconds, 1)}秒後に再試行してください。")
                return False

            user, err = authenticate_user(login_id, login_password)

            if user:
                clear_login_failures(login_id)
                st.session_state.logged_in = True
                st.session_state.role = clean_text(user.get("権限", "staff"), "staff")
                st.session_state.user_label = clean_text(user.get("表示名", login_id), login_id)
                st.session_state.username = login_id
                st.session_state.user_id = login_id
                st.session_state.login_user = login_id
                st.session_state.login_user_info = {
                    "id": login_id,
                    "username": login_id,
                    "role": st.session_state.role,
                    "label": st.session_state.user_label,
                }
                st.session_state.force_password_change = account_requires_password_change(user)
                st.session_state["hidamari_login_message"] = random.choice(HIDAMARI_MESSAGES)
                st.rerun()
            else:
                count, remaining, locked_seconds = record_login_failure(login_id)
                if locked_seconds:
                    st.error(f"ログイン失敗が{count}回続いたため、約{locked_seconds}秒間ログインを制限します。")
                elif remaining > 0:
                    st.error((err or "IDまたはパスワードが違います。") + f"（あと{remaining}回失敗すると一時制限されます）")
                else:
                    st.error(err or "IDまたはパスワードが違います。")

    return False


def clear_login_session():
    """ログアウト時にログイン情報だけをクリアする。"""
    st.session_state.logged_in = False
    st.session_state.role = None
    st.session_state.user_label = ""
    st.session_state.username = ""
    st.session_state.user_id = ""
    st.session_state.login_user = ""
    st.session_state.login_user_info = {}
    st.session_state.force_password_change = False
    st.session_state.pop("hidamari_login_message", None)


def render_keep_alive_widget(interval_seconds: int = 240, show_status: bool = False):
    """
    Streamlit Cloudのスリープ対策。
    画面を開いている間、軽い通信を定期実行してセッションを維持する。
    ※ブラウザや端末自体がスリープした場合は通信できないため、完全保証ではありません。

    運用方針：
    - 職員画面では余計な技術表示を出さない
    - 管理者画面のみ、接続維持中の説明表示を出せる
    """
    try:
        interval_ms = max(60, int(interval_seconds)) * 1000
        status_html = ""
        height = 1
        if show_status:
            status_html = (
                f'<div style="font-size:12px;color:#6A5B52;padding:2px 0;">'
                f'接続維持中：{int(interval_seconds)}秒ごとに軽い確認通信を行います。'
                f'</div>'
            )
            height = 28

        components.html(
            f"""
            {status_html}
            <script>
            const hidamariKeepAliveInterval = {interval_ms};
            async function hidamariKeepAlivePing() {{
              try {{
                await fetch(window.location.href, {{
                  method: "GET",
                  cache: "no-store",
                  mode: "no-cors"
                }});
              }} catch (e) {{}}
            }}
            hidamariKeepAlivePing();
            setInterval(hidamariKeepAlivePing, hidamariKeepAliveInterval);
            </script>
            """,
            height=height,
        )
    except Exception:
        pass


def finish_work_backup_and_logout():
    """管理者用：終了時バックアップを作成してからログアウトする。

    職員画面には終了ボタンを表示しない方針のため、
    この処理は管理者の保守操作からのみ呼び出す。
    Supabaseには書き込まず、バックアップZIP作成とセッション終了だけを行う。
    """
    try:
        zip_path, err = create_backup_zip(kind="終了時")
        if err:
            return False, err
        if zip_path:
            try:
                st.session_state["last_backup_file"] = str(zip_path)
                st.session_state["backup_download_pending"] = True
                st.session_state["backup_download_name"] = Path(zip_path).name
                st.session_state["hidamari_after_logout_message"] = (
                    f"終了時バックアップを作成しました：{Path(zip_path).name}"
                )
            except Exception:
                pass
        clear_login_session()
        return True, f"終了時バックアップを作成しました：{Path(zip_path).name if zip_path else ''}"
    except Exception as e:
        return False, f"終了時バックアップに失敗しました：{e}"


def logout_button():
    """サイドバーの終了・接続維持操作。

    職員には終了ボタンも接続維持・手動操作も表示しない。
    管理者だけがバックアップ付き終了、接続確認、通常ログアウトを操作できる。
    """
    if not is_admin_user():
        return

    with st.sidebar:
        st.caption(f"ログイン中：{st.session_state.get('user_label', '')}")

        st.markdown("### 管理者終了操作")
        st.caption("管理者が作業を終えるときだけ、必要に応じてバックアップを作成して終了します。")
        if st.button("バックアップを作成して終了", type="primary", use_container_width=True):
            ok, msg = finish_work_backup_and_logout()
            if ok:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)

        with st.expander("接続維持・手動操作", expanded=False):
            st.caption("管理者用です。職員画面には表示されません。")
            if st.button("今すぐ接続確認", use_container_width=True):
                try:
                    add_audit_log("接続確認", "keep_alive", "", "手動の接続確認を実行")
                except Exception:
                    pass
                st.success(f"接続確認OK：{format_now_jst('%Y-%m-%d %H:%M:%S')}")
            st.caption("通常のログアウトだけ行う場合はこちら。バックアップは作成しません。")
            if st.button("ログアウトのみ", use_container_width=True):
                clear_login_session()
                st.rerun()


def get_standard_menu_groups(role="admin"):
    """標準メニューカテゴリ。自己設定の初期値として使う。"""
    return MENU_GROUPS_ADMIN if role == "admin" else MENU_GROUPS_STAFF


def make_menu_category_rows_from_groups(groups, role="admin"):
    """カテゴリ辞書を編集用DataFrame行へ変換する。"""
    rows = []
    sort_no = 10
    for category, menus in groups.items():
        menu_no = 10
        for menu_name in menus:
            rows.append({
                "表示": True,
                "カテゴリ": clean_text(category, "その他"),
                "メニュー": clean_text(menu_name),
                "並び順": sort_no + menu_no / 100,
                "備考": "標準" if role == "admin" else "職員",
            })
            menu_no += 10
        sort_no += 100
    return rows


def get_standard_menu_category_df(role="admin"):
    groups = get_standard_menu_groups(role)
    return normalize_menu_category_df(pd.DataFrame(make_menu_category_rows_from_groups(groups, role=role)))


def normalize_menu_category_df(df):
    """メニューカテゴリ自己設定の列・型・並びを整える。"""
    columns = ["表示", "カテゴリ", "メニュー", "並び順", "備考"]
    if df is None or df.empty:
        df = pd.DataFrame(columns=columns)
    work = df.copy()
    for col in columns:
        if col not in work.columns:
            work[col] = ""
    work = work[columns].copy()
    work["表示"] = work["表示"].map(lambda x: str(x).lower() in ["true", "1", "yes", "有", "表示", "on"] if not isinstance(x, bool) else x)
    work["カテゴリ"] = work["カテゴリ"].map(lambda x: clean_text(x, "その他"))
    work["メニュー"] = work["メニュー"].map(lambda x: clean_text(x))
    work["並び順"] = pd.to_numeric(work["並び順"], errors="coerce").fillna(9999)
    work["備考"] = work["備考"].map(lambda x: clean_text(x))
    work = work[work["メニュー"] != ""].copy()
    # 非表示メニューは、過去の保存設定に残っていても設定一覧・サイドバーへ出さない。
    try:
        hidden = set(HIDDEN_MENUS)
        work = work[~work["メニュー"].astype(str).isin(hidden)].copy()
    except Exception:
        pass
    work = work.drop_duplicates(subset=["メニュー"], keep="last")
    return work.sort_values(["カテゴリ", "並び順", "メニュー"]).reset_index(drop=True)


def load_menu_category_settings(role="admin"):
    """管理者が編集したメニューカテゴリ設定をSQLiteから読み込む。なければ標準設定を使う。"""
    ensure_dirs()
    role = "admin" if role == "admin" else "staff"
    standard_df = get_standard_menu_category_df(role)

    settings_all = get_app_setting("menu_category_settings_all", None)
    if settings_all is None:
        settings_all = migrate_json_file_setting_to_db(
            "menu_category_settings_all",
            MENU_CATEGORY_SETTINGS_FILE,
            category="メニュー設定",
            default={},
        )
    if not isinstance(settings_all, dict):
        settings_all = {}

    rows = settings_all.get(role, [])
    if rows:
        df = normalize_menu_category_df(pd.DataFrame(rows))
    else:
        df = standard_df.copy()

    # 新機能追加時に自己設定へ自動追記する。既存のカテゴリ変更は維持する。
    existing_menus = set(df["メニュー"].astype(str).tolist())
    missing = standard_df[~standard_df["メニュー"].astype(str).isin(existing_menus)]
    if not missing.empty:
        df = pd.concat([df, missing], ignore_index=True)

    # 管理者が設定画面を非表示にしても復帰できるよう、必ず表示する。
    required_admin_menus = ["メニューカテゴリ設定", "システム設定"]
    if role == "admin":
        for required_menu in required_admin_menus:
            if required_menu in standard_df["メニュー"].tolist():
                if required_menu not in df["メニュー"].tolist():
                    add = standard_df[standard_df["メニュー"] == required_menu]
                    df = pd.concat([df, add], ignore_index=True)
                df.loc[df["メニュー"] == required_menu, "表示"] = True

    return normalize_menu_category_df(df)


def save_menu_category_settings(df, role="admin"):
    """メニューカテゴリ自己設定をSQLiteへ保存する。"""
    ensure_dirs()
    role = "admin" if role == "admin" else "staff"
    clean_df = normalize_menu_category_df(df)
    if role == "admin":
        for required_menu in ["メニューカテゴリ設定", "システム設定"]:
            if required_menu in clean_df["メニュー"].tolist():
                clean_df.loc[clean_df["メニュー"] == required_menu, "表示"] = True

    settings_all = get_app_setting("menu_category_settings_all", {})
    if not isinstance(settings_all, dict):
        settings_all = {}
    settings_all[role] = clean_df.to_dict(orient="records")
    set_app_setting(
        "menu_category_settings_all",
        settings_all,
        category="メニュー設定",
        description="管理者・職員のメニューカテゴリ自己設定",
    )
    try:
        add_audit_log("メニューカテゴリ設定更新", "app_settings", role, "メニューカテゴリ自己設定をSQLiteへ保存")
    except Exception:
        pass


def reset_menu_category_settings(role="admin"):
    """自己設定を標準設定に戻す。"""
    df = get_standard_menu_category_df(role)
    save_menu_category_settings(df, role=role)


def build_menu_groups_from_settings(role="admin"):
    """自己設定からサイドバー用カテゴリ辞書を作る。"""
    df = load_menu_category_settings(role)
    df = df[df["表示"] == True].copy()
    if df.empty:
        return get_standard_menu_groups(role)
    df = df.sort_values(["並び順", "カテゴリ", "メニュー"])
    groups = {}
    for _, row in df.iterrows():
        category = clean_text(row.get("カテゴリ"), "その他")
        menu_name = clean_text(row.get("メニュー"))
        if not menu_name:
            continue
        if menu_name in HIDDEN_MENUS:
            continue
        groups.setdefault(category, [])
        if menu_name not in groups[category]:
            groups[category].append(menu_name)
    return groups


def show_menu_category_settings_menu():
    """管理者がサイドバーのカテゴリ・表示順・表示有無を自己設定する画面。"""
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    ui_section("メニューカテゴリ設定", "標準設定をもとに、管理者自身でサイドバーのカテゴリ名・並び順・表示有無を調整できます。", "🧭")
    ui_card("使い方", "カテゴリ名を書き換えると、左メニューのカテゴリ分けが変わります。表示チェックを外すとメニューを一時的に隠せます。『メニューカテゴリ設定』は復帰用として常に表示されます。", "", soft=True)

    role_target = st.radio("設定対象", ["管理者メニュー", "職員メニュー"], horizontal=True, key="menu_category_role_target")
    role_key = "admin" if role_target == "管理者メニュー" else "staff"

    df = load_menu_category_settings(role_key)
    edited = st.data_editor(
        df,
        use_container_width=True,
        hide_index=True,
        num_rows="fixed",
        column_config={
            "表示": st.column_config.CheckboxColumn("表示"),
            "カテゴリ": st.column_config.TextColumn("カテゴリ", help="例：朝の確認、日々の入力、設定・保守"),
            "メニュー": st.column_config.TextColumn("メニュー", disabled=True, help="機能名は固定です。カテゴリと並び順を変更してください。"),
            "並び順": st.column_config.NumberColumn("並び順", step=1, help="小さい数字ほど上に表示されます。"),
            "備考": st.column_config.TextColumn("備考"),
        },
        key=f"menu_category_editor_{role_key}",
    )

    col1, col2 = st.columns(2)
    with col1:
        if st.button("自分設定を保存", type="primary", use_container_width=True):
            save_menu_category_settings(edited, role=role_key)
            st.success("メニューカテゴリ設定を保存しました。左メニューに反映します。")
            st.rerun()
    with col2:
        if st.button("標準設定に戻す", use_container_width=True):
            reset_menu_category_settings(role=role_key)
            st.success("標準設定に戻しました。")
            st.rerun()

    st.divider()
    st.subheader("現在の表示プレビュー")
    preview_groups = build_menu_groups_from_settings(role_key)
    for cat, menus in preview_groups.items():
        st.markdown(f"**{cat}**")
        st.write(" / ".join(menus))




def build_confirm_points_from_attention(row):
    """注意利用者の表示を責める言葉ではなく確認ポイントへ変換する。"""
    points = []
    item = clean_text(row.get("注意項目"))
    change = clean_text(row.get("気になる変化"))
    family = clean_text(row.get("家族共有メモ"))
    if "体温" in item or "発熱" in item:
        points.append("体温を再確認し、水分・食事・普段との違いを見る")
    if "SpO2" in item:
        points.append("SpO2を再測定し、呼吸・顔色・傾眠の有無を見る")
    if "食事" in item or "摂取" in item:
        points.append("食事量だけでなく、口腔・むせ・好み・疲れを確認する")
    if "排便" in item or "便" in item:
        points.append("排便間隔、水分、腹部症状、下剤等の情報を確認する")
    if change:
        points.append("気になる変化を次勤務者へそのまま共有する")
    if family:
        points.append("家族へ共有してよい内容か、表現を確認する")
    if not points:
        points.append("普段との違いを確認し、必要な共有だけを残す")
    return " / ".join(dict.fromkeys(points))


def add_confirm_points_column(df):
    """注意利用者一覧に確認ポイント列を追加する。"""
    if df is None or df.empty:
        return df
    work = df.copy()
    work["確認すること"] = work.apply(build_confirm_points_from_attention, axis=1)
    preferred = [c for c in ["利用者名", "注意項目", "確認すること", "気になる変化", "家族共有メモ", "記録日"] if c in work.columns]
    rest = [c for c in work.columns if c not in preferred]
    return work[preferred + rest]


def build_process_stop_summary(health_df, ex_df, handover_df, target_date):
    """管理者向けに『どこで止まっているか』を整理する。"""
    rows = []
    try:
        h = health_df.copy()
        if not h.empty:
            h["記録日"] = pd.to_datetime(h["記録日"], errors="coerce")
            h_day = h[h["記録日"].dt.date == target_date]
        else:
            h_day = pd.DataFrame()
        missing_health = max(len(active_users) - len(set(h_day.get("利用者名", []))), 0)
        rows.append({"確認場所": "健康チェック", "止まりやすい点": "未入力・気になる変化の未共有", "件数": missing_health, "次に見ること": "未入力者を責めず、入力できなかった理由と入力導線を見る"})
    except Exception:
        pass
    try:
        ex_day = get_day_excretion_data(ex_df, target_date, None)
        rows.append({"確認場所": "排泄チェック", "止まりやすい点": "時間帯別記録の不足", "件数": 0 if ex_day is None else len(ex_day), "次に見ること": "記録数ではなく、排便なし・濃縮尿など共有すべき点を見る"})
    except Exception:
        pass
    try:
        pending = 0
        if handover_df is not None and not handover_df.empty and "対応状況" in handover_df.columns:
            pending = int(handover_df[handover_df["対応状況"].astype(str).isin(["未対応", "対応中"])].shape[0])
        rows.append({"確認場所": "申し送り", "止まりやすい点": "未対応・対応中のまま残る", "件数": pending, "次に見ること": "誰の責任かではなく、次の一手が書かれているか確認する"})
    except Exception:
        pass
    return pd.DataFrame(rows, columns=["確認場所", "止まりやすい点", "件数", "次に見ること"])


def format_handover_structured_note(fact_text, insight_text, next_text):
    """事実／気づき／次に見ることを既存の申し送り欄へ保存しやすく整形する。"""
    parts = []
    if clean_text(fact_text):
        parts.append("【事実】\n" + clean_text(fact_text))
    if clean_text(insight_text):
        parts.append("【気づき】\n" + clean_text(insight_text))
    if clean_text(next_text):
        parts.append("【次に見ること】\n" + clean_text(next_text))
    return "\n\n".join(parts)



# ============================================================
# Ver4.6：AI管理者アシスタント
# ------------------------------------------------------------
# 管理者が利用者・期間を指定し、健康記録・排泄記録・申し送り・短期目標・
# モニタリングを横断抽出して「見落としを減らす」ための整理レポートを作成する。
# 方針：診断しない／記録を整理する／気になる変化を拾う／最終確認は人が行う。
# ============================================================

def _hidamari_ai_collect_records(user_name, start_date, end_date):
    health_df = hidamari_ai_filter_period(load_health_data(start_date=start_date, end_date=end_date), "記録日", user_name, start_date, end_date)
    ex_df = hidamari_ai_filter_period(load_excretion_data(start_date=start_date, end_date=end_date), "記録日", user_name, start_date, end_date)
    handover_df = hidamari_ai_filter_period(load_business_handover_data(start_date=start_date, end_date=end_date), "日付", user_name, start_date, end_date)
    goal_check_df = hidamari_ai_filter_period(load_sqlite_table(SQLITE_TABLE_SHORT_GOAL_CHECKS, SHORT_GOAL_CHECK_COLUMNS, date_cols=["日付"]), "日付", user_name, start_date, end_date)
    monitoring_df = hidamari_ai_filter_period(load_sqlite_table(SQLITE_TABLE_MONITORING_DRAFTS, MONITORING_DRAFT_COLUMNS, date_cols=["作成日"]), "作成日", user_name, start_date, end_date)

    goal_df = load_sqlite_table(SQLITE_TABLE_SHORT_GOAL_MASTER, SHORT_GOAL_MASTER_COLUMNS, date_cols=["開始日", "終了予定日"])
    if goal_df is not None and not goal_df.empty and "利用者名" in goal_df.columns:
        goal_df = goal_df[goal_df["利用者名"].astype(str).str.strip() == str(user_name).strip()].copy()

    return {
        "health": health_df,
        "excretion": ex_df,
        "handover": handover_df,
        "short_goal_master": goal_df,
        "short_goal_checks": goal_check_df,
        "monitoring": monitoring_df,
    }


def _hidamari_ai_make_report_pdf(report_text, user_name, start_date, end_date):
    if colors is None:
        return None
    ensure_dirs()
    REPORT_DIR.mkdir(parents=True, exist_ok=True)
    pdf_path = REPORT_DIR / ai_admin_report_file_name(user_name, start_date, end_date)
    _, mincho = register_japanese_pdf_fonts()
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=16*mm, leftMargin=16*mm, topMargin=14*mm, bottomMargin=14*mm)
    styles = getSampleStyleSheet()
    base = ParagraphStyle("hidamari_base", parent=styles["Normal"], fontName=mincho, fontSize=9.5, leading=14)
    title = ParagraphStyle("hidamari_title", parent=base, fontSize=15, leading=20, spaceAfter=8)
    story = []
    append_markdown_lines_to_story(report_text, story, base, title)
    doc.build(story)
    pdf_path.write_bytes(buffer.getvalue())
    return pdf_path


def _hidamari_ai_save_insight_log(user_name, start_date, end_date, report_text):
    try:
        ensure_ai_insight_log_file()
        df = load_sqlite_table(SQLITE_TABLE_AI_INSIGHT_LOGS, AI_INSIGHT_LOG_COLUMNS)
        row = {
            "作成日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
            "分析基準日": str(end_date),
            "利用者名": user_name,
            "対象期間": f"{start_date}〜{end_date}",
            "ルール分析": "健康・排泄・申し送り・短期目標・モニタリング横断抽出",
            "AI分析結果": report_text,
        }
        df = pd.concat([df, pd.DataFrame([row], columns=AI_INSIGHT_LOG_COLUMNS)], ignore_index=True)
        save_sqlite_table(df, SQLITE_TABLE_AI_INSIGHT_LOGS, AI_INSIGHT_LOG_COLUMNS, unique_cols=["作成日時", "利用者名", "対象期間"], sort_cols=["作成日時"])
        add_audit_log("AI管理者レポート作成", SQLITE_TABLE_AI_INSIGHT_LOGS, user_name, f"{start_date}〜{end_date}")
        clear_hidamari_read_cache("AI管理者レポート作成")
        return True
    except Exception as e:
        try:
            st.warning(f"AI分析ログ保存に失敗しました：{e}")
        except Exception:
            pass
        return False


def show_ai_admin_assistant_menu():
    """管理者の見落としを減らすAI：利用者・期間指定で横断整理する。"""
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    ui_section("AI管理者アシスタント", "利用者と期間を選び、確認ポイントを整理します。", "🧠")
    ui_card(
        "基本方針",
        "診断はしません。記録を整理し、気になる変化と管理者確認ポイントを拾うための補助機能です。",
        "最終判断は管理者・看護職・主治医等の人が行います。",
        soft=True,
    )

    users_df_local = load_users(include_hidden=False)
    if users_df_local is None or users_df_local.empty or "利用者名" not in users_df_local.columns:
        st.warning("利用者マスタに表示中の利用者がありません。")
        return
    user_options = users_df_local["利用者名"].dropna().astype(str).tolist()

    c1, c2, c3 = st.columns([1.2, 1, 1])
    with c1:
        user_name = st.selectbox("利用者を選ぶ", user_options, key="ai_admin_user_name")
    with c2:
        start_date = st.date_input("開始日", today_jst() - timedelta(days=14), key="ai_admin_start_date")
    with c3:
        end_date = st.date_input("終了日", today_jst(), key="ai_admin_end_date")

    if start_date > end_date:
        st.error("開始日は終了日以前にしてください。")
        return

    run = st.button("記録を一括抽出してAI整理する", type="primary", use_container_width=True)
    if not run and "hidamari_ai_admin_report" not in st.session_state:
        st.info("利用者と期間を選んで、管理者向けレポートを作成してください。")
        return

    if run:
        records = _hidamari_ai_collect_records(user_name, start_date, end_date)
        report = hidamari_ai_build_admin_report(user_name, start_date, end_date, records)
        st.session_state["hidamari_ai_admin_report"] = report
        st.session_state["hidamari_ai_admin_records"] = records
        st.session_state["hidamari_ai_admin_meta"] = {"user_name": user_name, "start_date": start_date, "end_date": end_date}

    report = st.session_state.get("hidamari_ai_admin_report", "")
    records = st.session_state.get("hidamari_ai_admin_records", {})
    meta = st.session_state.get("hidamari_ai_admin_meta", {"user_name": user_name, "start_date": start_date, "end_date": end_date})

    st.subheader("管理者向け整理結果")
    st.markdown(report)

    st.download_button(
        "Markdownでダウンロード",
        data=report.encode("utf-8"),
        file_name=f"hidamari_ai_admin_report_{meta.get('user_name')}_{meta.get('start_date')}_{meta.get('end_date')}.md",
        mime="text/markdown",
        use_container_width=True,
    )

    pdf_path = _hidamari_ai_make_report_pdf(report, meta.get("user_name"), meta.get("start_date"), meta.get("end_date"))
    if pdf_path and Path(pdf_path).exists():
        st.download_button(
            "PDFでダウンロード",
            data=Path(pdf_path).read_bytes(),
            file_name=Path(pdf_path).name,
            mime="application/pdf",
            use_container_width=True,
        )

    if st.button("この結果をAI分析ログに保存する", use_container_width=True):
        if _hidamari_ai_save_insight_log(meta.get("user_name"), meta.get("start_date"), meta.get("end_date"), report):
            st.success("AI分析ログに保存しました。")

    st.divider()
    st.subheader("抽出元データ")
    tabs = st.tabs(["健康", "排泄", "申し送り", "短期目標", "モニタリング"])
    tab_keys = ["health", "excretion", "handover", "short_goal_checks", "monitoring"]
    for tab, key in zip(tabs, tab_keys):
        with tab:
            df = records.get(key, pd.DataFrame())
            if df is None or df.empty:
                st.info("対象期間の記録はありません。")
            else:
                st.dataframe(df, use_container_width=True, hide_index=True)




configure_theme_settings(get_app_setting)
configure_sidebar(build_menu_groups_from_settings, filter_admin_menus)
apply_design()
apply_product_ui_ux()

if not login_check():
    st.stop()

# 画面を開いている間のスリープ対策
# 接続維持の表示・通信・手動操作は管理者だけに限定する。
if is_admin_user():
    render_keep_alive_widget(interval_seconds=240, show_status=True)

logout_button()

# 起動時DB安全チェック（異常時は管理者復元画面で停止）
if not run_startup_database_guard():
    st.stop()

# SQLite・セキュリティテーブル初期化と1日1回自動バックアップ
ensure_hidamari_db()
ensure_security_tables()
run_daily_auto_backup()
run_daily_photo_retention_cleanup()

if st.session_state.role == "admin":
    show_hidamari_hero("admin")
    show_admin_encouragement()
else:
    show_hidamari_hero("staff")
    show_staff_encouragement()

product_ui_notice()


# =========================
# メニュー（Ver3.0：カテゴリ化・iPad最適化）
# =========================
active_users = load_active_user_names(include_hidden=False)
all_users = active_users

menu = render_sidebar_menu(st.session_state.role, APP_VERSION, APP_COPY)


# =========================
# 管理者LIFE管理（月次）
# =========================
LIFE_DB = DATA_DIR / "hidamari_life.db"

def life_conn():
    ensure_dirs()
    conn = sqlite3.connect(LIFE_DB, timeout=DB_BUSY_TIMEOUT_MS / 1000, check_same_thread=False)
    safe_apply_sqlite_pragmas(conn, for_write=True)
    return conn

def init_life_db():
    con = life_conn()
    cur = con.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS life_monthly (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_name TEXT,
        target_month TEXT,
        adl_walk TEXT,
        adl_transfer TEXT,
        adl_eat TEXT,
        adl_toilet TEXT,
        dementia_level TEXT,
        life_check TEXT,
        addition_check TEXT,
        csv_status TEXT,
        manager_memo TEXT
    )
    """)
    con.commit()
    con.close()

def life_read_df(sql):
    con = life_conn()
    df = pd.read_sql(sql, con)
    con.close()
    return df

def life_exec_sql(sql, params):
    con = life_conn()
    cur = con.cursor()
    cur.execute(sql, params)
    con.commit()
    con.close()

def show_manager_life_input():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    init_life_db()
    st.header("管理者LIFE入力（月次）")
    st.caption("ADL評価、認知症自立度、LIFE確認、科学的介護推進体制加算確認を管理者が月次で入力します。")

    if not active_users:
        st.warning("利用者マスタに表示中の利用者がいません。")
        return

    with st.form("life_manager_form"):
        user_name = st.selectbox("利用者名", active_users)
        target_month = st.text_input("対象月", value=format_now_jst("%Y-%m"))

        st.subheader("ADL評価")
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            adl_walk = st.selectbox("歩行", ["自立", "見守り", "一部介助", "全介助"])
        with c2:
            adl_transfer = st.selectbox("移乗", ["自立", "見守り", "一部介助", "全介助"])
        with c3:
            adl_eat = st.selectbox("食事動作", ["自立", "見守り", "一部介助", "全介助"])
        with c4:
            adl_toilet = st.selectbox("排泄動作", ["自立", "見守り", "一部介助", "全介助"])

        st.subheader("認知症・LIFE確認")
        c5, c6, c7, c8 = st.columns(4)
        with c5:
            dementia_level = st.selectbox("認知症自立度", ["自立", "Ⅰ", "Ⅱa", "Ⅱb", "Ⅲa", "Ⅲb", "Ⅳ", "M"])
        with c6:
            life_check = st.selectbox("LIFE確認", ["未確認", "確認済"])
        with c7:
            addition_check = st.selectbox("科学的介護推進体制加算", ["未確認", "対象", "対象外"])
        with c8:
            csv_status = st.selectbox("CSV出力状態", ["未出力", "出力済"])

        manager_memo = st.text_area("管理者メモ")
        submitted = st.form_submit_button("登録する", use_container_width=True)

    if submitted:
        if not clean_text(user_name):
            st.warning("利用者名を選択してください。")
            return

        life_exec_sql("""
        INSERT INTO life_monthly(
            user_name, target_month, adl_walk, adl_transfer, adl_eat, adl_toilet,
            dementia_level, life_check, addition_check, csv_status, manager_memo
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?)
        """, (
            user_name, target_month, adl_walk, adl_transfer, adl_eat, adl_toilet,
            dementia_level, life_check, addition_check, csv_status, manager_memo
        ))
        st.success("管理者LIFEデータを登録しました。")

def judge_life_missing_items(row):
    """LIFE管理データの不足項目を共通判定する。"""
    missing = []
    if row.get("life_check") != "確認済":
        missing.append("LIFE未確認")
    if row.get("csv_status") != "出力済":
        missing.append("CSV未出力")
    if row.get("addition_check") == "未確認":
        missing.append("加算未確認")
    return " / ".join(missing) if missing else "OK"


def show_life_missing_check():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    init_life_db()
    st.header("LIFE不足チェック")
    df = life_read_df("SELECT * FROM life_monthly ORDER BY target_month DESC, user_name ASC")

    if df.empty:
        st.info("LIFE管理データはまだ登録されていません。")
        return

    df["不足項目"] = df.apply(judge_life_missing_items, axis=1)
    st.dataframe(
        df[["user_name", "target_month", "life_check", "addition_check", "csv_status", "不足項目"]],
        use_container_width=True,
        hide_index=True
    )

def show_life_csv_export():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    init_life_db()
    st.header("LIFE CSV出力")
    df = life_read_df("SELECT * FROM life_monthly ORDER BY target_month DESC, user_name ASC")

    if df.empty:
        st.info("出力できるLIFE管理データがありません。")
        return

    st.dataframe(df, use_container_width=True, hide_index=True)
    csv = df.to_csv(index=False).encode("utf-8-sig")
    st.download_button(
        label="LIFE CSVダウンロード",
        data=csv,
        file_name="hidamari_life_export.csv",
        mime="text/csv",
        use_container_width=True
    )
    st.caption("現時点ではLIFE提出補助用の土台CSVです。正式提出形式はLIFE仕様に合わせて最終調整が必要です。")

def show_life_record_list():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    init_life_db()
    st.header("LIFE登録一覧")
    df = life_read_df("SELECT * FROM life_monthly ORDER BY target_month DESC, user_name ASC")
    if df.empty:
        st.info("LIFE管理データはまだ登録されていません。")
    else:
        st.dataframe(df, use_container_width=True, hide_index=True)




# =========================
# システム設定（商品化向け：設定系SQLite一元管理）
# =========================
def show_system_settings_menu():
    """JSON/Excel/コードに散らばる設定を、商品化に向けてSQLite側で管理する画面。"""
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    ui_section("システム設定", "商品化に向けて、UI・色・LIFE・施設情報などの設定をSQLiteに集約して管理します。", "⚙️")
    ui_card(
        "設定DB化の状態",
        "この画面で保存した内容は app_settings テーブルに保存されます。バックアップZIPにはSQLite DBが含まれるため、復元・移行がしやすくなります。",
        "JSON／Excelへ分散しないための土台です。",
        soft=True,
    )

    initialize_default_app_settings()

    tab1, tab2, tab3, tab4, tab5 = st.tabs(["施設設定", "UI設定", "色設定", "LIFE設定", "設定一覧"])

    with tab1:
        st.subheader("施設設定")
        facility = get_app_setting("facility_settings", {})
        if not isinstance(facility, dict):
            facility = {}
        with st.form("facility_settings_form"):
            c1, c2 = st.columns(2)
            with c1:
                facility_name = st.text_input("施設名", value=clean_text(facility.get("施設名"), "ひだまり"))
                service_type = st.text_input("事業種別", value=clean_text(facility.get("事業種別"), "小規模介護施設"))
                capacity = st.text_input("定員", value=clean_text(facility.get("定員")))
            with c2:
                manager = st.text_input("管理者名", value=clean_text(facility.get("管理者名")))
                tel = st.text_input("連絡先", value=clean_text(facility.get("連絡先")))
                address = st.text_input("所在地", value=clean_text(facility.get("所在地")))
            if st.form_submit_button("施設設定を保存", type="primary", use_container_width=True):
                set_app_setting(
                    "facility_settings",
                    {
                        "施設名": facility_name,
                        "事業種別": service_type,
                        "定員": capacity,
                        "所在地": address,
                        "管理者名": manager,
                        "連絡先": tel,
                    },
                    category="施設設定",
                    description="施設名・管理者・帳票表示用の基本情報",
                )
                add_audit_log("施設設定更新", "app_settings", "facility_settings", "施設設定をSQLiteへ保存")
                st.success("施設設定を保存しました。")
                st.rerun()

    with tab2:
        st.subheader("UI設定")
        ui = get_app_setting("ui_settings", {})
        if not isinstance(ui, dict):
            ui = {}
        with st.form("ui_settings_form"):
            theme_name = st.text_input("テーマ名", value=clean_text(ui.get("テーマ"), "ひだまり標準"))
            ipad_opt = st.checkbox("iPad最適化", value=bool(ui.get("iPad最適化", True)))
            large_button = st.checkbox("ボタン大型化", value=bool(ui.get("ボタン大型化", True)))
            card_view = st.checkbox("カード表示", value=bool(ui.get("カード表示", True)))
            font_scale = st.number_input("フォント倍率", min_value=0.8, max_value=1.4, value=float(ui.get("フォント倍率", 1.0)), step=0.05)
            if st.form_submit_button("UI設定を保存", type="primary", use_container_width=True):
                set_app_setting(
                    "ui_settings",
                    {
                        "テーマ": theme_name,
                        "iPad最適化": ipad_opt,
                        "ボタン大型化": large_button,
                        "カード表示": card_view,
                        "フォント倍率": font_scale,
                    },
                    category="UI設定",
                    description="画面表示・タブレット対応設定",
                )
                add_audit_log("UI設定更新", "app_settings", "ui_settings", "UI設定をSQLiteへ保存")
                st.success("UI設定を保存しました。")
                st.rerun()

    with tab3:
        st.subheader("色設定")
        colors_setting = get_color_settings()
        with st.form("color_settings_form"):
            c1, c2 = st.columns(2)
            with c1:
                staff_bg = st.text_input("職員画面 背景色", value=clean_text(colors_setting.get("staff_bg"), "#FFFDF7"))
                staff_accent = st.text_input("職員画面 アクセント色", value=clean_text(colors_setting.get("staff_accent"), "#C9705C"))
                alert_color = st.text_input("注意色", value=clean_text(colors_setting.get("alert"), "#C9705C"))
            with c2:
                admin_bg = st.text_input("管理者画面 背景色", value=clean_text(colors_setting.get("admin_bg"), "#F6F8F7"))
                admin_accent = st.text_input("管理者画面 アクセント色", value=clean_text(colors_setting.get("admin_accent"), "#2F6F5E"))
                success_color = st.text_input("確認済み色", value=clean_text(colors_setting.get("success"), "#2F6F5E"))
            if st.form_submit_button("色設定を保存", type="primary", use_container_width=True):
                set_app_setting(
                    "color_settings",
                    {
                        "staff_bg": staff_bg,
                        "staff_accent": staff_accent,
                        "admin_bg": admin_bg,
                        "admin_accent": admin_accent,
                        "alert": alert_color,
                        "success": success_color,
                    },
                    category="色設定",
                    description="UIカラー設定",
                )
                add_audit_log("色設定更新", "app_settings", "color_settings", "色設定をSQLiteへ保存")
                st.success("色設定を保存しました。再読み込み後に反映されます。")
                st.rerun()

    with tab4:
        st.subheader("LIFE設定")
        life = get_app_setting("life_settings", {})
        if not isinstance(life, dict):
            life = {}
        with st.form("life_settings_form"):
            c1, c2 = st.columns(2)
            with c1:
                month_default = st.selectbox("対象月初期値", ["当月", "前月"], index=0 if clean_text(life.get("対象月初期値"), "当月") == "当月" else 1)
                missing_view = st.checkbox("LIFE不足表示", value=bool(life.get("LIFE不足表示", True)))
                csv_confirm = st.checkbox("CSV出力前確認", value=bool(life.get("CSV出力前確認", True)))
            with c2:
                avoid_diag = st.checkbox("診断表現を避ける", value=bool(life.get("診断表現を避ける", True)))
                ai整理 = st.checkbox("AIは整理係", value=bool(life.get("AIは整理係", True)))
            if st.form_submit_button("LIFE設定を保存", type="primary", use_container_width=True):
                set_app_setting(
                    "life_settings",
                    {
                        "対象月初期値": month_default,
                        "LIFE不足表示": missing_view,
                        "CSV出力前確認": csv_confirm,
                        "診断表現を避ける": avoid_diag,
                        "AIは整理係": ai整理,
                    },
                    category="LIFE設定",
                    description="LIFE管理・AI整理・CSV出力設定",
                )
                add_audit_log("LIFE設定更新", "app_settings", "life_settings", "LIFE設定をSQLiteへ保存")
                st.success("LIFE設定を保存しました。")
                st.rerun()

    with tab5:
        st.subheader("app_settings 一覧")
        settings_df = get_all_app_settings_df()
        if settings_df.empty:
            st.info("設定はまだ登録されていません。")
        else:
            st.dataframe(settings_df.sort_values(["分類", "設定キー"]), use_container_width=True, hide_index=True)
            st.download_button(
                "設定一覧をExcelでダウンロード",
                data=to_excel_download(settings_df),
                file_name=f"app_settings_{today_jst().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )


# =========================
# 自分専用ダッシュボード設定
# =========================
DASHBOARD_SETTINGS_FILE = DATA_DIR / "dashboard_settings.json"

DASHBOARD_ITEMS = {
    "前日の申し送り": "前日の業務全体申し送りを表示",
    "前日の気になる変化（全員）": "前日の健康チェックに入力された『気になる変化』を全利用者分表示",
    "未対応・至急申し送り": "未対応・至急の申し送りを表示",
    "排便3日なし": "確認日までに排便が3日間ない利用者を表示",
    "注意利用者": "発熱・SpO2低下・食事低下・気になる変化などを表示",
    "確認日の排泄状況": "確認日の排泄記録を表示",
    "最新体重・未測定確認": "最新体重と14日以上未測定を表示",
    "LIFE不足チェック": "LIFE確認・加算確認・CSV出力の不足を表示",
    "短期目標 実施状況": "短期目標の実施状況を表示",
}

DEFAULT_DASHBOARD_ITEMS = [
    "前日の申し送り",
    "未対応・至急申し送り",
    "排便3日なし",
    "注意利用者",
    "最新体重・未測定確認",
]

def load_dashboard_settings(username=None):
    ensure_dirs()
    username = username or current_login_user()
    if "dashboard_enabled_items" in st.session_state:
        return set([x for x in st.session_state["dashboard_enabled_items"] if x in DASHBOARD_ITEMS])

    data = get_app_setting("dashboard_settings_all", None)
    if data is None:
        data = migrate_json_file_setting_to_db(
            "dashboard_settings_all",
            DASHBOARD_SETTINGS_FILE,
            category="ダッシュボード設定",
            default={},
        )
    if not isinstance(data, dict):
        data = {}

    items = data.get(username)
    if items is None:
        # kanriで保存したものを他キーでも拾えるようにする
        items = data.get("kanri", DEFAULT_DASHBOARD_ITEMS)

    return set([x for x in items if x in DASHBOARD_ITEMS])

def save_dashboard_settings(username, enabled_items):
    ensure_dirs()
    data = get_app_setting("dashboard_settings_all", {})
    if not isinstance(data, dict):
        data = {}

    # ログインキーの揺れで反映されないのを防ぐため、管理者はkanriにも保存
    clean_items = [x for x in enabled_items if x in DASHBOARD_ITEMS]
    data[username] = clean_items
    if username == "kanri" or is_admin_user():
        data["kanri"] = clean_items

    set_app_setting(
        "dashboard_settings_all",
        data,
        category="ダッシュボード設定",
        description="自分専用ダッシュボードの表示項目設定",
    )


def show_custom_dashboard_page():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    st.header("自分専用ダッシュボード")
    st.caption("『自分専用ダッシュボード設定』で選択した項目だけを表示します。")

    check_date = st.date_input("確認日", value=today_jst(), key="custom_dashboard_check_date")
    show_my_dashboard_blocks(check_date)

def show_custom_dashboard_settings():
    if not is_admin_user():
        st.warning("このメニューは管理者専用です。")
        return

    st.header("自分専用ダッシュボード設定")
    st.caption("管理者ダッシュボードに表示する項目を選べます。保存後、管理者ダッシュボードへ戻ると反映されます。")

    username = current_login_user()
    current = load_dashboard_settings(username)

    enabled_items = []
    st.markdown("#### 表示する項目を選択")
    for item, desc in DASHBOARD_ITEMS.items():
        checked = st.checkbox(
            item,
            value=item in current,
            help=desc,
            key=f"dashboard_setting_{item}"
        )
        if checked:
            enabled_items.append(item)

    c1, c2 = st.columns(2)

    with c1:
        if st.button("自分設定を保存", type="primary", use_container_width=True):
            save_dashboard_settings(username, enabled_items)
            st.session_state["dashboard_settings_saved"] = True
            st.session_state["dashboard_enabled_items"] = enabled_items
            st.success("設定を保存しました。管理者ダッシュボードに戻ると反映されます。")
            st.rerun()

    with c2:
        if st.button("標準設定に戻す", use_container_width=True):
            save_dashboard_settings(username, DEFAULT_DASHBOARD_ITEMS)
            st.success("標準設定に戻しました。")
            st.rerun()

    st.divider()
    st.subheader("現在保存されている表示項目")
    saved = load_dashboard_settings(username)
    if saved:
        st.write(" / ".join(saved))
    else:
        st.info("表示項目は選択されていません。")

def show_my_dashboard_blocks(target_date=None):
    if not is_admin_user():
        return

    username = current_login_user()
    enabled = load_dashboard_settings(username)

    if target_date is None:
        target_date = today_jst()


    if not enabled:
        st.info("表示項目が選択されていません。『自分専用ダッシュボード設定』で表示項目を選んでください。")
        return

    health_cache = {}
    ex_cache = {}
    handover_cache = {}

    def get_dashboard_health_df(start_date=None, end_date=None, recent_days=None):
        key = (str(start_date or ""), str(end_date or ""), str(recent_days or ""))
        if key not in health_cache:
            health_cache[key] = load_health_data(start_date=start_date, end_date=end_date, recent_days=recent_days)
        return health_cache[key]

    def get_dashboard_ex_df(start_date=None, end_date=None, recent_days=None):
        key = (str(start_date or ""), str(end_date or ""), str(recent_days or ""))
        if key not in ex_cache:
            ex_cache[key] = load_excretion_data(start_date=start_date, end_date=end_date, recent_days=recent_days)
        return ex_cache[key]

    def get_dashboard_handover_df(start_date=None, end_date=None, recent_days=None):
        key = (str(start_date or ""), str(end_date or ""), str(recent_days or ""))
        if key not in handover_cache:
            handover_cache[key] = load_business_handover_data(start_date=start_date, end_date=end_date, recent_days=recent_days)
        return handover_cache[key]

    if "前日の申し送り" in enabled:
        st.subheader("前日の申し送り")
        try:
            prev_day = target_date - timedelta(days=1)
            df = get_dashboard_handover_df(start_date=prev_day, end_date=prev_day)
            prev_df = get_business_handover_by_date(df, prev_day)
            if prev_df.empty:
                st.info("前日の申し送りはありません。")
            else:
                for _, row in prev_df.iterrows():
                    render_business_handover_card(row)
        except Exception as e:
            st.warning(f"前日の申し送りを表示できませんでした: {e}")

    if "前日の気になる変化（全員）" in enabled:
        st.subheader("前日の気になる変化（全員）")
        st.caption("確認日の前日に、健康チェックへ入力された『気になる変化』を全利用者分まとめて表示します。")
        try:
            prev_day = target_date - timedelta(days=1)
            change_df = get_dashboard_health_df(start_date=prev_day, end_date=prev_day).copy()

            if change_df.empty:
                st.info("健康チェックデータがありません。")
            elif "記録日" not in change_df.columns:
                st.warning("健康チェックデータに『記録日』列がありません。")
            elif "気になる変化" not in change_df.columns:
                st.warning("健康チェックデータに『気になる変化』列がありません。")
            else:
                change_df["記録日_dt"] = pd.to_datetime(change_df["記録日"], errors="coerce")
                change_df["気になる変化"] = change_df["気になる変化"].fillna("").astype(str).str.strip()
                prev_changes = change_df[
                    (change_df["記録日_dt"].dt.date == prev_day)
                    & (change_df["気になる変化"] != "")
                ].copy()

                if prev_changes.empty:
                    st.success(f"{prev_day.strftime('%Y/%m/%d')} の気になる変化はありません。")
                else:
                    prev_changes = prev_changes.sort_values(
                        ["利用者名", "登録日時"] if "登録日時" in prev_changes.columns else ["利用者名", "記録日_dt"]
                    )

                    for col in ["記録日", "利用者名", "気になる変化", "家族共有メモ", "入力者", "登録日時"]:
                        if col not in prev_changes.columns:
                            prev_changes[col] = ""

                    display_df = prev_changes[["記録日", "利用者名", "気になる変化", "家族共有メモ", "入力者", "登録日時"]].copy()
                    display_df["日付"] = pd.to_datetime(display_df["記録日"], errors="coerce").dt.strftime("%Y/%m/%d")
                    display_df = display_df[["日付", "利用者名", "気になる変化", "家族共有メモ", "入力者", "登録日時"]]

                    st.warning(f"{prev_day.strftime('%Y/%m/%d')} に、気になる変化が {len(display_df)} 件あります。")
                    st.dataframe(display_df, use_container_width=True, hide_index=True)

                    memo_lines = []
                    st.markdown("#### 確認メモ")
                    for _, row in display_df.iterrows():
                        date_label_raw = clean_text(row.get("日付", ""))
                        row_user_raw = clean_text(row.get("利用者名", ""))
                        change_text_raw = clean_text(row.get("気になる変化", ""))
                        family_text_raw = clean_text(row.get("家族共有メモ", ""))
                        staff_text_raw = clean_text(row.get("入力者", ""))
                        date_label = html_escape_text(date_label_raw)
                        row_user = html_escape_text(row_user_raw)
                        change_text = html_escape_text(change_text_raw)
                        family_text = html_escape_text(family_text_raw)
                        staff_text = html_escape_text(staff_text_raw)
                        family_display = family_text if family_text else "記録なし"
                        staff_display = staff_text if staff_text else "未入力"

                        st.markdown(
                            f"""
                            <div style='background:#FFF8E8; border:1px solid #E5C782; border-radius:14px; padding:12px 14px; margin:8px 0;'>
                                <b>{date_label}　{row_user}</b><br>
                                <span style='color:#7A4A00;'>気になる変化：</span>{change_text}<br>
                                <span style='color:#666;'>家族共有メモ：</span>{family_display}<br>
                                <span style='color:#888; font-size:0.9rem;'>入力者：{staff_display}</span>
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )

                        memo_lines.append(
                            f"{date_label_raw}　{row_user_raw}\n"
                            f"気になる変化：{change_text_raw}\n"
                            f"家族共有メモ：{family_text_raw if family_text_raw else '記録なし'}\n"
                            f"入力者：{staff_text_raw if staff_text_raw else '未入力'}"
                        )

                    export_text = f"前日の気になる変化（全員）　{prev_day.strftime('%Y/%m/%d')}\n\n" + "\n\n".join(memo_lines)
                    st.text_area(
                        "コピー用テキスト",
                        value=export_text,
                        height=220,
                        key=f"prev_day_changes_all_text_{prev_day.strftime('%Y%m%d')}",
                    )
                    st.download_button(
                        "前日の気になる変化（全員）をテキストでダウンロード",
                        data=export_text.encode("utf-8-sig"),
                        file_name=f"前日の気になる変化_全員_{prev_day.strftime('%Y%m%d')}.txt",
                        mime="text/plain",
                        use_container_width=True,
                        key=f"prev_day_changes_all_download_{prev_day.strftime('%Y%m%d')}",
                    )
        except Exception as e:
            st.warning(f"前日の気になる変化（全員）を表示できませんでした: {e}")

    if "未対応・至急申し送り" in enabled:
        st.subheader("未対応・至急申し送り")
        try:
            df = get_dashboard_handover_df()
            alert_df = get_business_handover_alerts(df)
            if alert_df.empty:
                st.success("未対応・至急の申し送りはありません。")
            else:
                for _, row in alert_df.iterrows():
                    render_business_handover_card(row)
        except Exception as e:
            st.warning(f"未対応・至急申し送りを表示できませんでした: {e}")

    if "排便3日なし" in enabled:
        st.subheader("排便3日なし")
        try:
            no_stool_df = build_no_stool_3days_users(get_dashboard_ex_df(start_date=target_date - timedelta(days=4), end_date=target_date), target_date)
            if no_stool_df.empty:
                st.success("3日以上の未排便該当者はいません。")
            else:
                st.dataframe(no_stool_df, use_container_width=True, hide_index=True)
        except Exception as e:
            st.warning(f"排便3日なしを表示できませんでした: {e}")

    if "注意利用者" in enabled:
        st.subheader("注意利用者")
        try:
            attention_df = build_attention_users(get_dashboard_health_df(start_date=target_date, end_date=target_date), get_dashboard_ex_df(start_date=target_date - timedelta(days=4), end_date=target_date), target_date)
            if attention_df.empty:
                st.success("注意表示の対象者はいません。")
            else:
                st.dataframe(add_confirm_points_column(attention_df), use_container_width=True, hide_index=True)
        except Exception as e:
            st.warning(f"注意利用者を表示できませんでした: {e}")

    if "最新体重・未測定確認" in enabled:
        try:
            show_latest_weight_block(get_dashboard_health_df(), active_users if "active_users" in globals() else None, target_date)
            show_weight_overdue_block(get_dashboard_health_df(), active_users if "active_users" in globals() else None, target_date, threshold_days=14)
        except Exception as e:
            st.warning(f"体重確認を表示できませんでした: {e}")

    if "確認日の排泄状況" in enabled:
        st.subheader("確認日の排泄状況")
        try:
            day_ex = get_day_excretion_data(get_dashboard_ex_df(start_date=target_date, end_date=target_date), target_date, None)
            if day_ex.empty:
                st.info("確認日の排泄記録はありません。")
            else:
                st.dataframe(day_ex, use_container_width=True, hide_index=True)
        except Exception as e:
            st.warning(f"排泄状況を表示できませんでした: {e}")

    if "LIFE不足チェック" in enabled:
        st.subheader("LIFE不足チェック")
        try:
            if "life_read_df" in globals():
                init_life_db()
                life_df = life_read_df("SELECT * FROM life_monthly ORDER BY target_month DESC, user_name ASC")
                if life_df.empty:
                    st.info("LIFE管理データはまだありません。")
                else:
                    life_df["不足項目"] = life_df.apply(judge_life_missing_items, axis=1)
                    st.dataframe(
                        life_df[["user_name", "target_month", "life_check", "addition_check", "csv_status", "不足項目"]],
                        use_container_width=True,
                        hide_index=True
                    )
            else:
                st.info("LIFE管理機能がまだ読み込まれていません。")
        except Exception as e:
            st.warning(f"LIFE不足チェックを表示できませんでした: {e}")

    if "短期目標 実施状況" in enabled:
        st.subheader("短期目標 実施状況")
        try:
            checks = load_short_goal_checks()
            if checks.empty:
                st.info("短期目標の実施チェック記録はまだありません。")
            else:
                work = checks.copy()
                work["日付_dt"] = pd.to_datetime(work["日付"], errors="coerce")
                month_start = date(target_date.year, target_date.month, 1)
                month_df = work[work["日付_dt"] >= pd.to_datetime(month_start)].copy()
                if month_df.empty:
                    st.info("今月の短期目標実施チェックはまだありません。")
                else:
                    summary = month_df.groupby(["利用者名", "実施状況"]).size().reset_index(name="件数")
                    st.dataframe(summary, use_container_width=True, hide_index=True)
        except Exception as e:
            st.warning(f"短期目標実施状況を表示できませんでした: {e}")


def _filter_records_by_date(df, date_col, target_date):
    if df is None or df.empty or date_col not in df.columns:
        return pd.DataFrame()
    work = df.copy()
    work[date_col] = pd.to_datetime(work[date_col], errors="coerce")
    return work[work[date_col].dt.date == target_date].copy()


def show_dashboard_today_tasks(today, health_df, ex_df):
    st.subheader("今日やること")
    st.caption("未入力と未対応を先に確認します。")

    today_health = _filter_records_by_date(health_df, "記録日", today)
    today_ex = get_day_excretion_data(ex_df, today, None)
    health_users = set(today_health.get("利用者名", pd.Series(dtype=str)).dropna().astype(str)) if not today_health.empty else set()
    ex_users = set(today_ex.get("利用者名", pd.Series(dtype=str)).dropna().astype(str)) if not today_ex.empty else set()
    user_count = len(active_users) if "active_users" in globals() else 0

    try:
        handover_df = load_business_handover_data(start_date=today, end_date=today)
        if handover_df.empty:
            open_handover = 0
        else:
            status = handover_df.get("対応状況", pd.Series(dtype=str)).fillna("").astype(str)
            priority = handover_df.get("優先度", pd.Series(dtype=str)).fillna("").astype(str)
            check_note = handover_df.get("要確認事項", pd.Series(dtype=str)).fillna("").astype(str).str.strip()
            open_handover = int(((status != "対応済") & ((priority.isin(["注意", "至急"])) | (check_note != ""))).sum())
    except Exception:
        open_handover = 0

    c1, c2, c3, c4 = st.columns(4)
    c1.metric("健康チェック未入力", max(user_count - len(health_users), 0))
    c2.metric("排泄チェック未入力", max(user_count - len(ex_users), 0))
    c3.metric("未対応の申し送り", open_handover)
    c4.metric("今日の利用者数", user_count)


def show_dashboard_recent_attention(target_date, health_df, ex_df):
    st.subheader("最近の注意点")
    st.caption("確認日を含む直近の体調・排泄・申し送りを見ます。")

    try:
        attention_df = build_attention_users(health_df, ex_df, target_date)
    except Exception:
        attention_df = pd.DataFrame()

    recent_start = target_date - timedelta(days=2)
    try:
        recent_handover = load_business_handover_data(start_date=recent_start, end_date=target_date)
    except Exception:
        recent_handover = pd.DataFrame()

    if attention_df.empty and recent_handover.empty:
        st.success("直近の注意表示はありません。")
        return

    if not attention_df.empty:
        show_cols = [c for c in ["利用者名", "確認ポイント", "内容"] if c in add_confirm_points_column(attention_df).columns]
        st.dataframe(add_confirm_points_column(attention_df)[show_cols].head(8), use_container_width=True, hide_index=True)

    if not recent_handover.empty:
        for col in ["日付", "対象", "利用者名", "優先度", "対応状況", "要確認事項"]:
            if col not in recent_handover.columns:
                recent_handover[col] = ""
        handover_alerts = recent_handover[
            (recent_handover["対応状況"].astype(str) != "対応済")
            | (recent_handover["優先度"].astype(str).isin(["注意", "至急"]))
        ].copy()
        if not handover_alerts.empty:
            st.caption("未対応・注意の申し送り")
            st.dataframe(handover_alerts[["日付", "対象", "利用者名", "優先度", "対応状況", "要確認事項"]].head(8), use_container_width=True, hide_index=True)

# =========================
# 管理者ダッシュボード
# =========================
if menu == "管理者ダッシュボード":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("管理者ダッシュボード")
    show_observation_perspective("admin")

    health_df = load_health_data()
    ex_df = load_excretion_data()
    today = today_jst()
    yesterday = today - timedelta(days=1)

    st.markdown(
        """
        <div class="info-box">
            <b>朝の確認画面です。</b><br>
            今日の未入力・未対応と、確認日の注意点を見ます。
        </div>
        """,
        unsafe_allow_html=True,
    )

    target_date = st.date_input(
        "確認する日付（初期表示：昨日）",
        value=yesterday,
        key="admin_dashboard_target_date",
        help="朝の確認では昨日の日付を基本にします。必要に応じて別日も確認できます。",
    )

    target_excretion = get_day_excretion_data(ex_df, target_date, None)

    h_target = health_df.copy()
    if not h_target.empty:
        h_target["記録日"] = pd.to_datetime(h_target["記録日"], errors="coerce")
        h_target = h_target[h_target["記録日"].dt.date == target_date]

    col1, col2, col3, col4 = st.columns(4)
    col1.metric("確認日の健康記録", len(h_target))
    col2.metric("確認日の排泄記録", len(target_excretion))
    col3.metric("利用者数", len(active_users))

    ex_sum = summarize_excretion(target_excretion)
    col4.metric("確認日の排便記録", ex_sum["排便回数"])

    st.markdown("---")
    show_dashboard_today_tasks(today, health_df, ex_df)
    st.markdown("---")
    show_dashboard_recent_attention(target_date, health_df, ex_df)

    st.markdown("---")
    show_latest_weight_block(health_df, active_users, target_date)
    show_weight_overdue_block(health_df, active_users, target_date, threshold_days=14)

    # 出勤時に最初に確認したい項目として、業務全体申し送りを上部に表示
    st.markdown("---")
    show_admin_business_handover_summary(target_date)
    st.markdown("---")
    show_admin_short_goal_summary(target_date)
    st.markdown("---")

    st.subheader("確認日の注意利用者")
    attention_df = build_attention_users(health_df, ex_df, target_date)
    if attention_df.empty:
        st.success("確認日の注意利用者はありません。")
    else:
        st.warning("確認したい利用者がいます。")
        st.dataframe(add_confirm_points_column(attention_df), use_container_width=True, hide_index=True)

    st.subheader("どこで止まっているか")
    st.caption("職員個人ではなく、記録・共有・対応の流れを確認します。")
    try:
        st.dataframe(build_process_stop_summary(health_df, ex_df, load_business_handover_data(), target_date), use_container_width=True, hide_index=True)
    except Exception as e:
        st.info(f"流れの確認表を作成できませんでした: {e}")

    st.subheader("直近3日間、排便記録がない利用者")
    st.caption("確認する日付を含めた直近3日間で、排便記録がない利用者を表示します。")
    no_stool_3days_df = build_no_stool_3days_users(ex_df, target_date)
    if no_stool_3days_df.empty:
        st.success("確認日時点で、3日以上排便がない利用者はありません。")
    else:
        st.warning("排便状況を確認したい利用者がいます。")
        st.dataframe(no_stool_3days_df, use_container_width=True, hide_index=True)

    st.subheader("前日の申し送り確認")
    st.caption("初期表示は昨日です。出勤時に、前日の気になる変化・家族共有メモ・注意項目を確認できます。")
    st.text_area(
        "申し送りメモ",
        value=create_handover_text(health_df, ex_df, target_date),
        height=320,
    )

    st.subheader("前日の排泄状況確認")
    st.caption("初期表示は昨日です。出勤時に、前日の排尿・排便・濃縮尿・下痢便・水様便の有無を確認できます。")
    if target_excretion.empty:
        st.info("前日（確認日）の排泄記録はまだありません。")
    else:
        st.dataframe(target_excretion, use_container_width=True, hide_index=True)

        if ex_sum["濃縮尿"] or ex_sum["下痢便"] or ex_sum["水様便"]:
            st.warning(
                f"確認項目：濃縮尿 {ex_sum['濃縮尿']}件、"
                f"下痢便 {ex_sum['下痢便']}件、水様便 {ex_sum['水様便']}件"
            )
        else:
            st.success("前日（確認日）の排泄状況で大きな注意記録はありません。")

    show_admin_backup_download()


elif menu == "現場の気づき構造化・AI管理者支援":
    show_structured_insight_menu()

elif menu == "AI管理者アシスタント":
    show_ai_admin_assistant_menu()

# =========================
# 業務全体申し送り
# =========================
elif menu == "業務全体申し送り":
    show_business_handover_menu()

elif menu == "短期目標・モニタリング" and is_admin_user():
    show_short_goal_top()

elif menu == "短期目標マスタ":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()
    show_short_goal_master()

elif menu == "日々の実施チェック":
    show_daily_goal_check()

elif menu == "実施履歴一覧":
    show_goal_history()

elif menu == "モニタリング下書き作成" and is_admin_user():
    show_monitoring_draft()

elif menu == "短期目標データ管理":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()
    show_short_goal_data_management()


# =========================
# LIFE入力標準化
# =========================
elif menu == "LIFE入力標準化" and is_admin_user():
    show_life_standardization_menu()

elif menu == "加算シミュレーション" and is_admin_user():
    show_addon_simulation_menu()




elif menu == "自分専用ダッシュボード":
    show_custom_dashboard_page()

elif menu == "自分専用ダッシュボード設定":
    show_custom_dashboard_settings()

elif menu == "管理者LIFE入力":
    show_manager_life_input()

elif menu == "LIFE不足チェック":
    show_life_missing_check()

elif menu == "LIFE CSV出力":
    show_life_csv_export()

elif menu == "LIFE登録一覧":
    show_life_record_list()



elif menu == "写真から半自動入力":
    show_photo_import_menu()


# =========================
# 健康チェック入力
# =========================
elif menu == "健康チェック入力":
    st.header("健康チェックを書く")
    show_observation_perspective("health")

    if st.session_state.role == "staff":
        st.write("記録日・利用者・入力者を確認してから、バイタル、食事、気になる変化の順に入力します。")

    if not active_users:
        st.warning("利用者マスタに表示中の利用者がいません。")
        st.stop()

    col1, col2, col3 = st.columns(3)
    with col1:
        record_date = st.date_input("記録日", value=today_jst(), key="health_date")
    with col2:
        user_name = st.selectbox("利用者名", active_users, key="health_user")
    with col3:
        input_staff = st.text_input("入力者", placeholder="例：藤野", key="health_staff")

    # 入力画面では選択日の前後だけ読む。既存データ確認のために全件取得しない。
    health_df = load_health_data(start_date=record_date, end_date=record_date)
    idx = find_health_index(health_df, record_date, user_name)

    if idx is None:
        existing_row = None
        st.markdown(
            """
            <div style='background:#EAF4FF; border:1px solid #9CC7F0; color:#174A7C; padding:12px 14px; border-radius:10px; margin:8px 0 12px 0;'>
                <b>この記録日・利用者名の健康チェックデータはありません。</b><br>
                保存すると新しい記録として登録されます。
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        existing_row = health_df.loc[idx]
        st.markdown(
            """
            <div style='background:#FFF3E0; border:1px solid #F0B36A; color:#8A4B00; padding:12px 14px; border-radius:10px; margin:8px 0 12px 0;'>
                <b>この記録日・利用者名の健康チェックデータは既にあります。</b><br>
                保存するとこの内容で更新されます。前回の内容を表示しています。
            </div>
            """,
            unsafe_allow_html=True,
        )

    def row_float(col, default):
        if existing_row is None:
            return default
        return safe_float(existing_row.get(col), default)

    def row_int(col, default):
        if existing_row is None:
            return default
        return safe_int(existing_row.get(col), default)

    def row_text(col, default=""):
        if existing_row is None:
            return default
        return clean_text(existing_row.get(col), default)

    with st.form("health_form", clear_on_submit=False):
        st.subheader("1. バイタル")

        c1, c2, c3 = st.columns(3)
        with c1:
            temp = st.number_input("体温", min_value=0.0, max_value=45.0, value=row_float("体温", 0.0), step=0.1)
        with c2:
            bp_high = st.number_input("血圧上", min_value=0, max_value=250, value=row_int("血圧上", 0), step=1)
        with c3:
            bp_low = st.number_input("血圧下", min_value=0, max_value=150, value=row_int("血圧下", 0), step=1)

        c4, c5, c6 = st.columns(3)
        with c4:
            pulse = st.number_input("脈拍", min_value=0, max_value=200, value=row_int("脈拍", 0), step=1)
        with c5:
            spo2 = st.number_input("SpO2", min_value=0, max_value=100, value=row_int("SpO2", 0), step=1)
        with c6:
            existing_weight_text = format_weight_value(row_text("体重", ""))
            weight_raw = st.text_input(
                "体重（任意）",
                value=existing_weight_text,
                placeholder="例：56.2",
                help="週1回など、測定した日だけ入力します。未測定の場合は空欄でOKです。",
            )
            st.caption("※未測定の場合は空欄でOK")

        st.divider()
        st.subheader("2. 食事量")
        st.caption("朝・昼・夕の食事量を選びます。")

        m1, m2, m3 = st.columns(3)
        with m1:
            breakfast_default = row_text("朝食摂取区分", meal_option_from_percent(row_int("朝食摂取率", 80)))
            breakfast_code = st.selectbox("朝食", MEAL_INTAKE_OPTIONS, index=get_life_option_index(MEAL_INTAKE_OPTIONS, breakfast_default, 1))
        with m2:
            lunch_default = row_text("昼食摂取区分", meal_option_from_percent(row_int("昼食摂取率", 80)))
            lunch_code = st.selectbox("昼食", MEAL_INTAKE_OPTIONS, index=get_life_option_index(MEAL_INTAKE_OPTIONS, lunch_default, 1))
        with m3:
            dinner_default = row_text("夕食摂取区分", meal_option_from_percent(row_int("夕食摂取率", 80)))
            dinner_code = st.selectbox("夕食", MEAL_INTAKE_OPTIONS, index=get_life_option_index(MEAL_INTAKE_OPTIONS, dinner_default, 1))

        breakfast = MEAL_INTAKE_PERCENT[breakfast_code]
        lunch = MEAL_INTAKE_PERCENT[lunch_code]
        dinner = MEAL_INTAKE_PERCENT[dinner_code]

        st.divider()
        st.subheader("3. 水分・口腔・メモ")
        l1, l2, l3, l4 = st.columns(4)
        with l1:
            water_ml = st.number_input("水分摂取量ml", min_value=0, max_value=5000, value=row_int("水分摂取量ml", 0), step=50)
        with l2:
            nutrition_risk = st.selectbox("栄養リスク", NUTRITION_RISK_OPTIONS, index=get_life_option_index(NUTRITION_RISK_OPTIONS, row_text("栄養リスク", "0: 通常")))
        with l3:
            oral_status = st.selectbox("口腔状態", ORAL_STATUS_OPTIONS, index=get_life_option_index(ORAL_STATUS_OPTIONS, row_text("口腔状態", "9: 未確認"), 4))
        with l4:
            denture_status = st.selectbox("義歯使用", DENTURE_OPTIONS, index=get_life_option_index(DENTURE_OPTIONS, row_text("義歯使用", "9: 未確認"), 3))

        life_memo = st.text_area("LIFE補助メモ", value=row_text("LIFE補助メモ"), placeholder="食事・水分・口腔・栄養面で気になる点")
        family_memo = st.text_area("家族共有メモ", value=row_text("家族共有メモ"), placeholder="ご家族へ共有してよい内容を入力")
        changes = st.text_area("気になる変化", value=row_text("気になる変化"), placeholder="食事、睡眠、歩行、表情、体調など")

        submitted = st.form_submit_button("健康チェックを保存する")

    if submitted:
        weight, weight_error = parse_optional_weight(weight_raw)
        if weight_error:
            st.error(weight_error)
            st.stop()

        record = {
            "記録日": record_date,
            "利用者名": user_name,
            "体温": temp,
            "血圧上": bp_high,
            "血圧下": bp_low,
            "脈拍": pulse,
            "SpO2": spo2,
            "体重": weight,
            "朝食摂取率": breakfast,
            "昼食摂取率": lunch,
            "夕食摂取率": dinner,
            "朝食摂取区分": breakfast_code,
            "昼食摂取区分": lunch_code,
            "夕食摂取区分": dinner_code,
            "水分摂取量ml": water_ml,
            "栄養リスク": nutrition_risk,
            "口腔状態": oral_status,
            "義歯使用": denture_status,
            "LIFE補助メモ": life_memo,
            "家族共有メモ": family_memo,
            "気になる変化": changes,
            "登録日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
            "入力者": input_staff,
        }

        errors, warnings = validate_health_record(record)

        if errors:
            st.error("保存できません。赤い表示の項目を確認してください。")
            for msg in errors:
                st.error(msg)
        else:
            if warnings:
                st.warning("保存前に確認してください。")
                for msg in warnings:
                    st.warning(msg)

            diff_text = build_health_diff_text(health_df, record_date, user_name, record)
            st.info(diff_text)

            action = upsert_health_record(record)
            st.success(f"健康チェックを{action}しました。申し送りが必要な内容は共有してください。")
            st.rerun()


# =========================
# 排泄チェック入力
# 未入力チェック一覧＋スマホ用ワンタップ風UI
# =========================
elif menu == "排泄チェック入力":
    st.header("排泄チェックを書く")
    show_observation_perspective("excretion")
    st.caption("記録日・利用者・入力者を確認して、時間帯ごとに尿・便を入力します。")

    if st.session_state.role == "staff":
        st.markdown("### 排泄チェックを書く")
        st.write("日中帯、夜間帯の順に確認します。")

    if not active_users:
        st.warning("利用者マスタに表示中の利用者がいません。")
        st.stop()


    # まず日付を選ぶ。日付が変わると未入力一覧も切り替わる。
    top1, top2 = st.columns([1, 2])
    with top1:
        record_date = st.date_input("記録日", value=today_jst(), key="ex_input_date")

    # 利用者・入力者選択
    col1, col2 = st.columns(2)
    with col1:
        # 未入力一覧はページ下部へ移動したため、ここでは通常の利用者選択にする
        user_name = st.selectbox("利用者名", active_users, key="ex_input_user")
    with col2:
        input_staff = st.text_input("入力者", placeholder="例：藤野", key="ex_input_staff")

    ex_df = load_excretion_data(start_date=record_date, end_date=record_date)
    day_data = get_day_excretion_data(ex_df, record_date, user_name)

    if day_data.empty:
        st.markdown(
            """
            <div style='background:#EAF4FF; border:1px solid #9CC7F0; color:#174A7C; padding:12px 14px; border-radius:10px; margin:8px 0 12px 0;'>
                <b>この記録日・利用者名の排泄データはありません。</b><br>
                保存すると新しい排泄記録として登録されます。
            </div>
            """,
            unsafe_allow_html=True,
        )
    else:
        st.markdown(
            """
            <div style='background:#FFF3E0; border:1px solid #F0B36A; color:#8A4B00; padding:12px 14px; border-radius:10px; margin:8px 0 12px 0;'>
                <b>この記録日・利用者名の排泄データは既にあります。</b><br>
                保存すると時間帯ごとに更新されます。
            </div>
            """,
            unsafe_allow_html=True,
        )

    # スマホ用ワンタップ風UI
    st.caption("尿量・便量を選び、必要な時間帯だけメモを残します。")

    with st.form("excretion_form", clear_on_submit=False):
        records_to_save = []

        def one_tap_radio(label, options, index, key):
            return st.radio(
                label,
                options,
                index=index,
                horizontal=True,
                key=key,
            )

        def render_slot(slot, time_label, card_color, border_color):
            existing = get_excretion_row(ex_df, record_date, user_name, slot)
            sig = "new" if existing is None else hashlib.md5(str(existing.to_dict()).encode("utf-8")).hexdigest()[:8]
            key_base = f"ex_{record_date}_{user_name}_{slot}_{sig}"

            st.markdown(
                f"""
                <div style='background:{card_color}; padding:12px; border-radius:14px; border:1px solid {border_color}; margin-bottom:10px;'>
                    <b style='font-size:16px;'>{slot}</b><br>
                    <span style='font-size:12px; color:#666;'>{time_label}</span>
                </div>
                """,
                unsafe_allow_html=True,
            )

            # Ver4.6：排泄チェックの初期値を現場入力向けに最適化
            # 新規入力時は、頻度の高い「尿量：中」「尿性状：普通尿」を最初から選択する。
            # 既存データがある場合は、保存済みの値をそのまま表示する。
            urine_amount_default = existing.get("尿量", "中") if existing is not None else "中"
            urine_type_default = existing.get("尿性状", "普通尿") if existing is not None else "普通尿"
            stool_amount_default = existing.get("便量", "なし") if existing is not None else "なし"
            stool_type_default = existing.get("便性状", "なし") if existing is not None else "なし"

            urine_amount = one_tap_radio(
                f"{slot} 尿量",
                URINE_AMOUNT_OPTIONS,
                get_option_index(URINE_AMOUNT_OPTIONS, urine_amount_default),
                f"{key_base}_urine_amount",
            )

            urine_type = one_tap_radio(
                f"{slot} 尿性状",
                URINE_TYPE_OPTIONS,
                get_option_index(URINE_TYPE_OPTIONS, urine_type_default),
                f"{key_base}_urine_type",
            )

            stool_amount = one_tap_radio(
                f"{slot} 便量",
                STOOL_AMOUNT_OPTIONS,
                get_option_index(STOOL_AMOUNT_OPTIONS, stool_amount_default),
                f"{key_base}_stool_amount",
            )

            stool_type = one_tap_radio(
                f"{slot} 便性状",
                STOOL_TYPE_OPTIONS,
                get_option_index(STOOL_TYPE_OPTIONS, stool_type_default),
                f"{key_base}_stool_type",
            )

            memo = st.text_area(
                f"{slot} メモ",
                value=existing.get("排泄メモ", "") if existing is not None else "",
                key=f"{key_base}_memo",
                height=70,
                placeholder="必要時のみ入力",
            )

            if urine_amount == "なし":
                urine_type = "なし"
            if stool_amount == "なし":
                stool_type = "なし"

            records_to_save.append({
                "記録日": record_date,
                "利用者名": user_name,
                "時間帯": slot,
                "時間帯目安": time_label,
                "尿量": urine_amount,
                "尿性状": urine_type,
                "便量": stool_amount,
                "便性状": stool_type,
                "排泄メモ": memo,
                "入力者": input_staff,
                "登録日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
            })

        st.markdown("#### 1. 日中帯（9時〜17時）")
        day_cols = st.columns(3)

        for col, (slot, time_label) in zip(day_cols, EXCRETION_SLOTS[:3]):
            with col:
                render_slot(slot, time_label, "#FFF7EC", "#E5D5BF")

        st.markdown("#### 2. 夜間帯（18時〜翌8時）")
        night_cols = st.columns(3)

        for col, (slot, time_label) in zip(night_cols, EXCRETION_SLOTS[3:]):
            with col:
                render_slot(slot, time_label, "#EEF4FA", "#C9D8E6")

        submitted = st.form_submit_button("排泄チェックを保存する")

    if submitted:
        all_errors = []
        all_warnings = []

        for record in records_to_save:
            errors, warnings = validate_excretion_record(record)
            for msg in errors:
                all_errors.append(f"{record['時間帯']}：{msg}")
            for msg in warnings:
                all_warnings.append(f"{record['時間帯']}：{msg}")

        if all_errors:
            st.error("保存できません。赤い表示の時間帯を確認してください。")
            for msg in all_errors:
                st.error(msg)
        else:
            if all_warnings:
                st.warning("保存前に確認してください。")
                for msg in all_warnings:
                    st.warning(msg)

            for record in records_to_save:
                upsert_excretion_record(record)

            st.info(build_excretion_diff_text(load_excretion_data(), record_date, user_name))
            st.success("排泄チェックを保存しました。時間帯ごとの記録を更新しました。")
            st.rerun()

    st.subheader("この日の排泄記録")
    day_data = get_day_excretion_data(load_excretion_data(), record_date, user_name)
    if day_data.empty:
        st.info("この日の排泄記録はまだありません。")
    else:
        st.dataframe(day_data, use_container_width=True, hide_index=True)

    st.divider()

    # 未入力チェック一覧（ページ下部へ移動）
    st.subheader("未入力チェック一覧")

    missing_rows = []
    completed_rows = []

    for user in active_users:
        for slot, time_label in EXCRETION_SLOTS:
            row = get_excretion_row(load_excretion_data(), record_date, user, slot)

            if row is None:
                missing_rows.append(
                    {
                        "利用者名": user,
                        "時間帯": slot,
                        "時間帯目安": time_label,
                        "状態": "未入力",
                    }
                )
            else:
                completed_rows.append(
                    {
                        "利用者名": user,
                        "時間帯": slot,
                        "時間帯目安": time_label,
                        "状態": "入力済み",
                    }
                )

    total_count = len(active_users) * len(EXCRETION_SLOTS)
    completed_count = len(completed_rows)
    missing_count = len(missing_rows)

    m1, m2, m3 = st.columns(3)
    m1.metric("必要記録数", total_count)
    m2.metric("入力済み", completed_count)
    m3.metric("未入力", missing_count)

    if missing_rows:
        st.warning("未入力の排泄記録があります。")

        with st.expander("未入力一覧を表示する", expanded=True):
            st.dataframe(
                pd.DataFrame(missing_rows),
                use_container_width=True,
                hide_index=True,
            )
    else:
        st.success("この日の排泄記録はすべて入力済みです。")

    with st.expander("入力済み一覧を表示する", expanded=False):
        if completed_rows:
            st.dataframe(
                pd.DataFrame(completed_rows),
                use_container_width=True,
                hide_index=True,
            )
        else:
            st.info("この日の入力済み記録はまだありません。")


# =========================
# 過去データ管理
# =========================
elif menu == "過去データ管理":
    st.header("過去データ管理")
    st.caption("健康チェックだけでなく、入力状況・注意記録・業務全体申し送りを切り替えて確認できます。")

    data_mode = st.selectbox(
        "確認・管理するデータ種別",
        ["健康チェック", "入力状況", "注意記録", "業務全体申し送り"],
        key="past_data_mode",
    )

    # ---------------------------------
    # 健康チェック：従来の検索・更新・削除
    # ---------------------------------
    if data_mode == "健康チェック":
        st.subheader("健康チェックデータ")
        st.caption("健康チェックデータを、記録日＋利用者名で検索・更新・削除します。")

        # 過去データ管理の初期表示は直近7日だけ読む。一覧検索では選択月だけ読み直す。
        health_df = load_health_data(recent_days=7)

        if health_df.empty:
            st.info("まだ健康チェックデータがありません。")
        else:
            col1, col2 = st.columns(2)
            with col1:
                key_date = st.date_input("記録日", value=today_jst(), key="past_health_date")
            with col2:
                key_user = st.selectbox("利用者名", all_users, key="past_health_user")

            idx = find_health_index(health_df, key_date, key_user)

            if idx is None:
                st.info("この記録日・利用者名の健康チェックデータはありません。")
            else:
                st.success("該当データが見つかりました。")
                row = health_df.loc[idx]

                with st.form("health_update_form"):
                    c1, c2, c3 = st.columns(3)
                    with c1:
                        temp = st.number_input("体温", value=safe_float(row.get("体温"), 0.0), step=0.1)
                    with c2:
                        bp_high = st.number_input("血圧上", value=safe_int(row.get("血圧上"), 0), step=1)
                    with c3:
                        bp_low = st.number_input("血圧下", value=safe_int(row.get("血圧下"), 0), step=1)

                    c4, c5, c6 = st.columns(3)
                    with c4:
                        pulse = st.number_input("脈拍", value=safe_int(row.get("脈拍"), 0), step=1)
                    with c5:
                        spo2 = st.number_input("SpO2", value=safe_int(row.get("SpO2"), 0), step=1)
                    with c6:
                        weight_text = st.text_input("体重（任意）", value=format_weight_value(row.get("体重", "")), help="未測定の場合は空欄でOKです。")
                        weight, weight_error = parse_optional_weight(weight_text)
                        if weight_error:
                            st.warning(weight_error)

                    m1, m2, m3 = st.columns(3)
                    with m1:
                        breakfast = st.slider("朝食", 0, 100, safe_int(row.get("朝食摂取率"), 80), step=10)
                    with m2:
                        lunch = st.slider("昼食", 0, 100, safe_int(row.get("昼食摂取率"), 80), step=10)
                    with m3:
                        dinner = st.slider("夕食", 0, 100, safe_int(row.get("夕食摂取率"), 80), step=10)

                    family_memo = st.text_area("家族共有メモ", value=clean_text(row.get("家族共有メモ", "")))
                    changes = st.text_area("気になる変化", value=clean_text(row.get("気になる変化", "")))
                    staff = st.text_input("入力者", value=clean_text(row.get("入力者", "")))

                    update_submit = st.form_submit_button("更新する")

                if update_submit:
                    record = {
                        "記録日": key_date,
                        "利用者名": key_user,
                        "体温": temp,
                        "血圧上": bp_high,
                        "血圧下": bp_low,
                        "脈拍": pulse,
                        "SpO2": spo2,
                        "体重": weight,
                        "朝食摂取率": breakfast,
                        "昼食摂取率": lunch,
                        "夕食摂取率": dinner,
                        "朝食摂取区分": meal_option_from_percent(breakfast),
                        "昼食摂取区分": meal_option_from_percent(lunch),
                        "夕食摂取区分": meal_option_from_percent(dinner),
                        "水分摂取量ml": clean_text(row.get("水分摂取量ml", "")),
                        "栄養リスク": clean_text(row.get("栄養リスク", "")),
                        "口腔状態": clean_text(row.get("口腔状態", "")),
                        "義歯使用": clean_text(row.get("義歯使用", "")),
                        "LIFE補助メモ": clean_text(row.get("LIFE補助メモ", "")),
                        "家族共有メモ": family_memo,
                        "気になる変化": changes,
                        "登録日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
                        "入力者": staff,
                    }
                    action = upsert_health_record(record)
                    try:
                        add_audit_log("健康チェック更新", SQLITE_TABLE_HEALTH, f"{key_date}_{key_user}", "過去データ管理から更新")
                    except Exception:
                        pass
                    st.success(f"{action}しました。")
                    st.rerun()

                st.warning("削除すると元に戻せません。")
                delete_check = st.checkbox("この健康チェックデータを削除する")

                if st.button("削除する"):
                    if not delete_check:
                        st.error("削除する場合は確認チェックを入れてください。")
                    else:
                        # 健康チェック画面では該当レコードを row に保持しているため、
                        # 未定義の current ではなく row から user_id を取得する。
                        target_user_id = row.get("user_id", "") if hasattr(row, "get") else ""
                        result = delete_health_record(key_date, key_user, user_id=target_user_id, source="過去データ管理から削除")
                        show_delete_result_and_rerun(result, "削除しました。")

            st.divider()
            st.subheader("一覧検索")

            if not health_df.empty:
                year = st.number_input("年", min_value=2024, max_value=2035, value=today_jst().year, step=1, key="past_health_year")
                month = st.number_input("月", min_value=1, max_value=12, value=today_jst().month, step=1, key="past_health_month")
                user_filter = st.selectbox("利用者で絞り込み", ["全員"] + all_users, key="past_health_filter_user")
                month_start = date(int(year), int(month), 1)
                month_end = (month_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
                health_df = load_health_data(start_date=month_start, end_date=month_end)
                health_df["記録日"] = pd.to_datetime(health_df["記録日"], errors="coerce")

                result = health_df[
                    (health_df["記録日"].dt.year == int(year))
                    & (health_df["記録日"].dt.month == int(month))
                ]
                if user_filter != "全員":
                    result = result[result["利用者名"] == user_filter]

                st.dataframe(result.sort_values(["記録日", "利用者名"]), use_container_width=True, hide_index=True)
                st.download_button(
                    "この一覧をExcelでダウンロード",
                    data=to_excel_download(result.sort_values(["記録日", "利用者名"])),
                    file_name=f"health_records_{int(year)}_{int(month):02d}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )

    # ---------------------------------
    # 入力状況：健康・排泄・短期目標・申し送りの入力有無
    # ---------------------------------
    elif data_mode == "入力状況":
        st.subheader("入力状況")
        st.caption("指定日の健康チェック、排泄チェック、日々の実施チェック、業務全体申し送りの入力状況を一覧で確認します。")

        target_day = st.date_input("確認日", value=today_jst(), key="past_input_status_date")
        user_filter = st.selectbox("利用者で絞り込み", ["全員"] + all_users, key="past_input_status_user")

        health_df = load_health_data(start_date=target_day, end_date=target_day)
        ex_df = load_excretion_data(start_date=target_day, end_date=target_day)
        goal_check_df = load_short_goal_check_data() if "load_short_goal_check_data" in globals() else pd.DataFrame(columns=SHORT_GOAL_CHECK_COLUMNS)
        handover_df = load_business_handover_data(start_date=target_day, end_date=target_day)

        if not health_df.empty:
            health_df["記録日"] = pd.to_datetime(health_df["記録日"], errors="coerce")
        if not ex_df.empty:
            ex_df["記録日"] = pd.to_datetime(ex_df["記録日"], errors="coerce")
        if not goal_check_df.empty:
            goal_check_df["日付"] = pd.to_datetime(goal_check_df["日付"], errors="coerce")
        if not handover_df.empty:
            handover_df["日付"] = pd.to_datetime(handover_df["日付"], errors="coerce")

        target_users = all_users if user_filter == "全員" else [user_filter]
        status_rows = []
        for user_name in target_users:
            h_hit = pd.DataFrame()
            e_hit = pd.DataFrame()
            g_hit = pd.DataFrame()

            if not health_df.empty:
                h_hit = health_df[
                    (health_df["記録日"].dt.date == target_day)
                    & (health_df["利用者名"].astype(str) == str(user_name))
                ]

            if not ex_df.empty:
                e_hit = ex_df[
                    (ex_df["記録日"].dt.date == target_day)
                    & (ex_df["利用者名"].astype(str) == str(user_name))
                ]

            if not goal_check_df.empty:
                g_hit = goal_check_df[
                    (goal_check_df["日付"].dt.date == target_day)
                    & (goal_check_df["利用者名"].astype(str) == str(user_name))
                ]

            status_rows.append({
                "確認日": target_day.strftime("%Y-%m-%d"),
                "利用者名": user_name,
                "健康チェック": "入力済" if not h_hit.empty else "未入力",
                "排泄チェック": f"{len(e_hit)}件" if not e_hit.empty else "未入力",
                "日々の実施チェック": f"{len(g_hit)}件" if not g_hit.empty else "未入力",
                "健康メモ": "あり" if (not h_hit.empty and h_hit.get("気になる変化", pd.Series(dtype=str)).astype(str).str.strip().ne("").any()) else "",
            })

        status_df = pd.DataFrame(status_rows)
        handover_count = 0
        if not handover_df.empty:
            handover_count = len(handover_df[handover_df["日付"].dt.date == target_day])

        c1, c2, c3, c4 = st.columns(4)
        with c1:
            st.metric("健康チェック入力済", int((status_df["健康チェック"] == "入力済").sum()))
        with c2:
            st.metric("健康チェック未入力", int((status_df["健康チェック"] == "未入力").sum()))
        with c3:
            st.metric("排泄記録あり", int((status_df["排泄チェック"] != "未入力").sum()))
        with c4:
            st.metric("当日の申し送り", handover_count)

        st.dataframe(status_df, use_container_width=True, hide_index=True)
        st.download_button(
            "入力状況をExcelでダウンロード",
            data=to_excel_download(status_df),
            file_name=f"input_status_{target_day.strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

        with st.expander("当日の業務全体申し送りを確認", expanded=False):
            if handover_df.empty:
                st.info("申し送りデータはありません。")
            else:
                view_handover = handover_df[handover_df["日付"].dt.date == target_day].copy()
                if view_handover.empty:
                    st.info("この日の申し送りはありません。")
                else:
                    st.dataframe(view_handover, use_container_width=True, hide_index=True)

    # ---------------------------------
    # 注意記録：条件マスタに基づく抽出結果
    # ---------------------------------
    elif data_mode == "注意記録":
        st.subheader("注意記録")
        st.caption("条件設定マスタに基づいて、健康チェック・排泄チェックから注意候補を抽出します。診断ではなく、申し送り候補の確認です。")

        target_day = st.date_input("抽出日", value=today_jst(), key="past_alert_date")
        alert_df = build_handover_alerts_by_condition(target_day)

        if alert_df.empty:
            st.info("この日の条件該当者はありません。")
        else:
            f1, f2 = st.columns(2)
            with f1:
                severity_filter = st.selectbox("重要度で絞り込み", ["すべて"] + sorted(alert_df["重要度"].dropna().astype(str).unique().tolist()), key="past_alert_severity")
            with f2:
                user_filter = st.selectbox("利用者で絞り込み", ["全員"] + sorted(alert_df["利用者名"].dropna().astype(str).unique().tolist()), key="past_alert_user")

            view = alert_df.copy()
            if severity_filter != "すべて":
                view = view[view["重要度"].astype(str) == severity_filter]
            if user_filter != "全員":
                view = view[view["利用者名"].astype(str) == user_filter]

            severity_order = {"至急": 0, "注意": 1, "観察": 2, "通常": 3}
            view["_order"] = view["重要度"].map(severity_order).fillna(9)
            view = view.sort_values(["_order", "利用者名", "分類", "条件名"]).drop(columns=["_order"])

            st.dataframe(view, use_container_width=True, hide_index=True)

            st.markdown("#### 申し送り文プレビュー")
            st.info(build_business_handover_auto_extract_text(target_day))

            st.download_button(
                "注意記録をExcelでダウンロード",
                data=to_excel_download(view),
                file_name=f"alert_records_{pd.to_datetime(target_day).strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    # ---------------------------------
    # 業務全体申し送り：検索・更新・削除
    # ---------------------------------
    elif data_mode == "業務全体申し送り":
        st.subheader("業務全体申し送り")
        st.caption("業務全体申し送りを、日付・勤務帯・優先度・対応状況で検索し、必要に応じて更新・削除します。")

        df = load_business_handover_data()

        if df.empty:
            st.info("まだ業務全体申し送りは登録されていません。")
        else:
            work = df.copy()
            work["日付"] = pd.to_datetime(work["日付"], errors="coerce")

            c1, c2, c3 = st.columns(3)
            with c1:
                from_day = st.date_input("開始日", value=today_jst() - timedelta(days=30), key="past_handover_from")
            with c2:
                to_day = st.date_input("終了日", value=today_jst(), key="past_handover_to")
            with c3:
                keyword = st.text_input("キーワード", key="past_handover_keyword", placeholder="申し送り・要確認事項など")

            c4, c5, c6 = st.columns(3)
            with c4:
                shift_filter = st.selectbox("勤務帯", ["すべて"] + sorted([x for x in work["勤務帯"].dropna().astype(str).unique().tolist() if x]), key="past_handover_shift")
            with c5:
                priority_filter = st.selectbox("優先度", ["すべて"] + sorted([x for x in work["優先度"].dropna().astype(str).unique().tolist() if x]), key="past_handover_priority")
            with c6:
                status_filter = st.selectbox("対応状況", ["すべて"] + sorted([x for x in work["対応状況"].dropna().astype(str).unique().tolist() if x]), key="past_handover_status")

            result = work[
                (work["日付"].dt.date >= from_day)
                & (work["日付"].dt.date <= to_day)
            ].copy()

            if shift_filter != "すべて":
                result = result[result["勤務帯"].astype(str) == shift_filter]
            if priority_filter != "すべて":
                result = result[result["優先度"].astype(str) == priority_filter]
            if status_filter != "すべて":
                result = result[result["対応状況"].astype(str) == status_filter]
            if clean_text(keyword):
                kw = clean_text(keyword)
                search_cols = ["全体申し送り", "要確認事項", "Excel自動抽出情報", "入力Excel表示情報", "記入者"]
                mask = pd.Series(False, index=result.index)
                for col in search_cols:
                    if col in result.columns:
                        mask = mask | result[col].astype(str).str.contains(kw, case=False, na=False)
                result = result[mask]

            st.dataframe(result.sort_values(["日付", "記録日時"], ascending=[False, False]), use_container_width=True, hide_index=True)

            st.download_button(
                "申し送り一覧をExcelでダウンロード",
                data=to_excel_download(result.sort_values(["日付", "記録日時"], ascending=[False, False])),
                file_name=f"business_handover_{from_day.strftime('%Y%m%d')}_{to_day.strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

            st.divider()
            st.markdown("#### 選択した申し送りを更新・削除")

            if result.empty:
                st.info("更新・削除できる対象がありません。")
            else:
                result_ids = result["記録ID"].astype(str).tolist()
                selected_id = st.selectbox("対象の記録ID", result_ids, key="past_handover_selected_id")
                selected = df[df["記録ID"].astype(str) == str(selected_id)]

                if selected.empty:
                    st.error("対象データが見つかりません。")
                else:
                    row = selected.iloc[-1]
                    with st.form("past_handover_update_form"):
                        uc1, uc2, uc3 = st.columns(3)
                        with uc1:
                            edit_date = st.date_input("日付", value=pd.to_datetime(row.get("日付"), errors="coerce").date() if not pd.isna(pd.to_datetime(row.get("日付"), errors="coerce")) else today_jst(), key="past_handover_edit_date")
                        with uc2:
                            edit_shift = st.selectbox("勤務帯", ["日勤", "夜勤"], index=0 if clean_text(row.get("勤務帯"), "日勤") == "日勤" else 1, key="past_handover_edit_shift")
                        with uc3:
                            edit_priority = st.selectbox("優先度", ["通常", "注意", "至急"], index=["通常", "注意", "至急"].index(clean_text(row.get("優先度"), "通常")) if clean_text(row.get("優先度"), "通常") in ["通常", "注意", "至急"] else 0, key="past_handover_edit_priority")

                        edit_writer = st.text_input("記入者", value=clean_text(row.get("記入者", "")), key="past_handover_edit_writer")
                        edit_status = st.selectbox(
                            "対応状況",
                            ["未対応", "対応中", "完了", "共有のみ"],
                            index=["未対応", "対応中", "完了", "共有のみ"].index(clean_text(row.get("対応状況"), "未対応")) if clean_text(row.get("対応状況"), "未対応") in ["未対応", "対応中", "完了", "共有のみ"] else 0,
                            key="past_handover_edit_status",
                        )
                        edit_main = st.text_area("全体申し送り", value=clean_text(row.get("全体申し送り", "")), height=120, key="past_handover_edit_main")
                        edit_confirm = st.text_area("要確認事項", value=clean_text(row.get("要確認事項", "")), height=120, key="past_handover_edit_confirm")
                        edit_auto = st.text_area("Excel自動抽出情報", value=clean_text(row.get("Excel自動抽出情報", "")), height=150, key="past_handover_edit_auto")
                        update_handover_submit = st.form_submit_button("申し送りを更新する", type="primary", use_container_width=True)

                    if update_handover_submit:
                        update_df = load_business_handover_data()
                        mask = update_df["記録ID"].astype(str) == str(selected_id)
                        if not mask.any():
                            st.error("更新対象の記録が見つかりません。")
                        else:
                            update_df.loc[mask, "日付"] = pd.to_datetime(edit_date)
                            update_df.loc[mask, "勤務帯"] = edit_shift
                            update_df.loc[mask, "記入者"] = edit_writer
                            update_df.loc[mask, "優先度"] = edit_priority
                            update_df.loc[mask, "対応状況"] = edit_status
                            update_df.loc[mask, "全体申し送り"] = edit_main
                            update_df.loc[mask, "要確認事項"] = edit_confirm
                            update_df.loc[mask, "Excel自動抽出情報"] = edit_auto
                            update_df.loc[mask, "記録日時"] = format_now_jst("%Y-%m-%d %H:%M:%S")
                            save_business_handover_data(update_df)
                            try:
                                add_audit_log("申し送り更新", SQLITE_TABLE_HANDOVER, selected_id, "過去データ管理から更新")
                            except Exception:
                                pass
                            st.success("申し送りを更新しました。")
                            st.rerun()

                    st.warning("削除すると元に戻せません。")
                    delete_handover_check = st.checkbox("この申し送りを削除する", key="past_handover_delete_check")
                    if st.button("選択した申し送りを削除する", disabled=not delete_handover_check, key="past_handover_delete_button"):
                        delete_df = load_business_handover_data()
                        before_count = len(delete_df)
                        delete_df = delete_df[delete_df["記録ID"].astype(str) != str(selected_id)].copy()
                        if len(delete_df) == before_count:
                            st.error("削除対象が見つかりません。")
                        else:
                            result = delete_business_handover_record(selected_id, source="過去データ管理から削除")
                            show_delete_result_and_rerun(result, "申し送りを削除しました。")


# =========================
# 排泄詳細管理
# =========================
elif menu == "排泄詳細管理":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("排泄詳細管理")
    st.caption("排泄チェックデータを、記録日＋利用者名＋時間帯で管理します。")

    ex_df = load_excretion_data()

    if ex_df.empty:
        st.info("まだ排泄チェックデータがありません。")
        st.stop()

    col1, col2, col3 = st.columns(3)
    with col1:
        ex_user = st.selectbox("利用者", ["全員"] + all_users, key="ex_admin_user")
    with col2:
        start_date = st.date_input("開始日", value=today_jst(), key="ex_admin_start")
    with col3:
        end_date = st.date_input("終了日", value=today_jst(), key="ex_admin_end")

    work = ex_df.copy()
    work["記録日"] = pd.to_datetime(work["記録日"], errors="coerce")
    work = work[
        (work["記録日"].dt.date >= start_date)
        & (work["記録日"].dt.date <= end_date)
    ]

    if ex_user != "全員":
        work = work[work["利用者名"] == ex_user]

    st.subheader("排泄サマリー")
    if work.empty:
        st.warning("該当する排泄データがありません。")
    else:
        summary_rows = []
        for user in work["利用者名"].dropna().unique():
            user_df = work[work["利用者名"] == user]
            s = summarize_excretion(user_df)
            summary_rows.append({
                "利用者名": user,
                "記録数": len(user_df),
                "排尿回数": s["排尿回数"],
                "排便回数": s["排便回数"],
                "濃縮尿": s["濃縮尿"],
                "下痢便": s["下痢便"],
                "水様便": s["水様便"],
                "排便なし枠": s["排便なし枠"],
            })
        st.dataframe(pd.DataFrame(summary_rows), use_container_width=True, hide_index=True)

        st.subheader("注意して確認したい排泄記録")
        alert = work[
            (work["尿性状"] == "濃縮尿")
            | (work["便性状"].isin(["下痢便", "水様便"]))
        ]

        if alert.empty:
            st.success("指定期間内に、濃縮尿・下痢便・水様便の記録はありません。")
        else:
            st.warning("確認したい排泄記録があります。")
            st.dataframe(alert, use_container_width=True, hide_index=True)

        st.subheader("時系列の排泄詳細")
        slot_order = {slot: i for i, (slot, _) in enumerate(EXCRETION_SLOTS)}
        work["_slot_order"] = work["時間帯"].map(slot_order).fillna(99)
        work = work.sort_values(["記録日", "利用者名", "_slot_order"]).drop(columns=["_slot_order"])
        st.dataframe(work, use_container_width=True, hide_index=True)

        csv = work.to_csv(index=False).encode("utf-8-sig")
        st.download_button(
            "排泄詳細CSVをダウンロード",
            data=csv,
            file_name="排泄詳細データ.csv",
            mime="text/csv",
        )

        st.subheader("管理者向け確認メモ")
        memo_lines = [
            "排泄詳細データをもとにした管理者確認メモです。",
            "医療判断ではなく、職員間の共有と見守り方針の整理に使用してください。",
            "",
        ]

        for _, row in pd.DataFrame(summary_rows).iterrows():
            memo_lines.append(
                f"■ {row['利用者名']}：排尿{row['排尿回数']}回、排便{row['排便回数']}回、"
                f"濃縮尿{row['濃縮尿']}回、下痢便{row['下痢便']}回、水様便{row['水様便']}回。"
            )

        st.text_area("確認メモ", value="\n".join(memo_lines), height=260)

    st.divider()
    st.subheader("排泄データの更新・削除")

    c1, c2, c3 = st.columns(3)
    with c1:
        key_date = st.date_input("更新対象日", value=today_jst(), key="ex_edit_date")
    with c2:
        key_user = st.selectbox("更新対象利用者", all_users, key="ex_edit_user")
    with c3:
        key_slot = st.selectbox("時間帯", [slot for slot, _ in EXCRETION_SLOTS], key="ex_edit_slot")

    current = get_excretion_row(load_excretion_data(), key_date, key_user, key_slot)

    if current is None:
        st.info("このキーの排泄データはありません。")
    else:
        st.success("該当する排泄データがあります。")

    with st.form("ex_edit_form"):
        time_label = dict(EXCRETION_SLOTS).get(key_slot, "")

        urine_amount = st.selectbox(
            "尿量",
            URINE_AMOUNT_OPTIONS,
            index=get_option_index(URINE_AMOUNT_OPTIONS, current.get("尿量", "中") if current is not None else "中"),
        )
        urine_type = st.selectbox(
            "尿性状",
            URINE_TYPE_OPTIONS,
            index=get_option_index(URINE_TYPE_OPTIONS, current.get("尿性状", "普通尿") if current is not None else "普通尿"),
        )
        stool_amount = st.selectbox(
            "便量",
            STOOL_AMOUNT_OPTIONS,
            index=get_option_index(STOOL_AMOUNT_OPTIONS, current.get("便量", "なし") if current is not None else "なし"),
        )
        stool_type = st.selectbox(
            "便性状",
            STOOL_TYPE_OPTIONS,
            index=get_option_index(STOOL_TYPE_OPTIONS, current.get("便性状", "なし") if current is not None else "なし"),
        )
        memo = st.text_area("排泄メモ", value=current.get("排泄メモ", "") if current is not None else "")
        staff = st.text_input("入力者", value=current.get("入力者", "") if current is not None else "")

        submit = st.form_submit_button("登録・更新する")

    if submit:
        if urine_amount == "なし":
            urine_type = "なし"
        if stool_amount == "なし":
            stool_type = "なし"

        record = {
            "記録日": key_date,
            "利用者名": key_user,
            "時間帯": key_slot,
            "時間帯目安": time_label,
            "尿量": urine_amount,
            "尿性状": urine_type,
            "便量": stool_amount,
            "便性状": stool_type,
            "排泄メモ": memo,
            "入力者": staff,
            "登録日時": format_now_jst("%Y-%m-%d %H:%M:%S"),
        }
        action = upsert_excretion_record(record)
        st.success(f"排泄データを{action}しました。")
        st.rerun()

    if current is not None:
        delete_check = st.checkbox("この排泄データを削除する")
        if st.button("排泄データを削除する"):
            if not delete_check:
                st.error("削除する場合は確認チェックを入れてください。")
            else:
                ex_df = load_excretion_data()
                idx = find_excretion_index(ex_df, key_date, key_user, key_slot)
                if idx is not None:
                    target_user_id = current.get("user_id", "") if isinstance(current, dict) else ""
                    result = delete_excretion_record(key_date, key_user, key_slot, user_id=target_user_id, source="排泄詳細管理から削除")
                    show_delete_result_and_rerun(result, "削除しました。")


# =========================
# 家族向けレポート作成
# =========================
elif menu == "家族向けレポート作成":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("家族向けレポート作成")
    st.caption("利用者と対象月を選ぶと、家族共有用の文章を作成します。")

    if not all_users:
        st.warning("利用者が登録されていません。")
        st.stop()

    col1, col2, col3 = st.columns(3)
    with col1:
        report_user = st.selectbox("利用者", all_users, key="family_report_user")
    with col2:
        report_year = st.number_input("対象年", min_value=2024, max_value=2035, value=today_jst().year, step=1)
    with col3:
        report_month = st.number_input("対象月", min_value=1, max_value=12, value=today_jst().month, step=1)

    report_start = date(int(report_year), int(report_month), 1)
    report_end = (report_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
    health_df = load_health_data(start_date=report_start, end_date=report_end)
    ex_df = load_excretion_data(start_date=report_start, end_date=report_end)

    report_text = create_family_summary_text(health_df, ex_df, report_user, report_year, report_month)
    st.text_area("家族向け文章", value=report_text, height=420)

    ex_target = get_month_excretion_data(ex_df, report_user, report_year, report_month)
    with st.expander("排泄記録を確認する"):
        if ex_target.empty:
            st.info("対象月の排泄記録はありません。")
        else:
            st.dataframe(ex_target, use_container_width=True, hide_index=True)


# =========================
# ひだまりレポートPDF
# =========================
elif menu == "ひだまりレポートPDF":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("ひだまりレポートPDF")
    st.caption("利用者と対象月を選び、PDFを作成します。")

    if not all_users:
        st.warning("利用者が登録されていません。")
        st.stop()

    col1, col2, col3 = st.columns(3)
    with col1:
        pdf_user = st.selectbox("利用者", all_users, key="pdf_user")
    with col2:
        pdf_year = st.number_input("対象年", min_value=2024, max_value=2035, value=today_jst().year, step=1, key="pdf_year")
    with col3:
        pdf_month = st.number_input("対象月", min_value=1, max_value=12, value=today_jst().month, step=1, key="pdf_month")

    if st.button("ひだまりレポートPDFを作成する"):
        try:
            pdf_start = date(int(pdf_year), int(pdf_month), 1)
            pdf_end = (pdf_start.replace(day=28) + timedelta(days=4)).replace(day=1) - timedelta(days=1)
            path = create_hidamari_pdf(load_health_data(start_date=pdf_start, end_date=pdf_end), load_excretion_data(start_date=pdf_start, end_date=pdf_end), pdf_user, pdf_year, pdf_month)
            with open(path, "rb") as f:
                st.download_button(
                    "PDFをダウンロード",
                    data=f,
                    file_name=path.name,
                    mime="application/pdf",
                )
            st.success("PDFを作成しました。")
        except Exception as e:
            st.error(f"PDF作成中にエラーが発生しました：{e}")


# =========================
# 管理者支援
# =========================
elif menu == "管理者支援":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("管理者支援")
    health_df = load_health_data()
    ex_df = load_excretion_data()

    tab1, tab2, tab3, tab4, tab5, tab6, tab7, tab8 = st.tabs([
        "AI家族レポート",
        "バイタル推移グラフ",
        "気になる変化",
        "ChatGPT連携",
        "申し送り支援",
        "注意通知",
        "条件設定マスタ変更",
        "体重未測定確認",
    ])

    with tab1:
        st.subheader("AI家族レポート自動文章")
        col1, col2, col3 = st.columns(3)
        with col1:
            ai_user = st.selectbox("利用者", all_users, key="ai_user")
        with col2:
            ai_year = st.number_input("対象年", min_value=2024, max_value=2035, value=today_jst().year, step=1, key="ai_year")
        with col3:
            ai_month = st.number_input("対象月", min_value=1, max_value=12, value=today_jst().month, step=1, key="ai_month")

        summary = create_family_summary_text(health_df, ex_df, ai_user, ai_year, ai_month)
        st.text_area("家族向け文章", value=summary, height=360)

    with tab2:
        st.subheader("バイタル推移グラフ")
        if health_df.empty:
            st.info("データがありません。")
        else:
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                graph_user = st.selectbox("利用者", all_users, key="graph_user")
            with col2:
                graph_item = st.selectbox("項目", ["体温", "血圧上", "血圧下", "脈拍", "SpO2", "体重", "朝食摂取率", "昼食摂取率", "夕食摂取率"], key="graph_item")
            with col3:
                graph_year = st.number_input("年", min_value=2024, max_value=2035, value=today_jst().year, step=1, key="graph_year")
            with col4:
                graph_month = st.number_input("月", min_value=1, max_value=12, value=today_jst().month, step=1, key="graph_month")

            target = get_month_health_data(health_df, graph_user, graph_year, graph_month)
            if target.empty:
                st.warning("対象データがありません。")
            else:
                chart_df = target[["記録日", graph_item]].copy()
                chart_df[graph_item] = pd.to_numeric(chart_df[graph_item], errors="coerce")
                chart_df = chart_df.dropna()
                chart_df = chart_df.set_index("記録日")
                st.line_chart(chart_df)


    with tab3:
        st.subheader("気になる変化（日付別一覧）")
        st.caption("健康チェック入力の『気になる変化』を、利用者・日付で絞り込み、日付ごとに確認できます。")

        if health_df.empty:
            st.info("健康チェックデータがありません。")
        else:
            # Ver4.6追加：利用者に「全員」を追加し、年月だけではなく日付範囲で絞り込めるようにする。
            user_options_for_change = ["全員"] + [u for u in all_users if clean_text(u) and clean_text(u) != "全員"]

            default_end_date = today_jst()
            default_start_date = default_end_date.replace(day=1)

            col1, col2, col3 = st.columns(3)
            with col1:
                change_user = st.selectbox("利用者", user_options_for_change, key="change_user")
            with col2:
                change_start_date = st.date_input("開始日", value=default_start_date, key="change_start_date")
            with col3:
                change_end_date = st.date_input("終了日", value=default_end_date, key="change_end_date")

            if change_start_date > change_end_date:
                st.error("開始日は終了日以前にしてください。")
                st.stop()

            change_target = health_df.copy()
            if "記録日" not in change_target.columns:
                st.error("健康チェックデータに『記録日』列がありません。")
                st.stop()

            change_target["記録日_dt"] = pd.to_datetime(change_target["記録日"], errors="coerce")
            start_dt = pd.to_datetime(change_start_date)
            end_dt = pd.to_datetime(change_end_date)
            change_target = change_target[
                (change_target["記録日_dt"] >= start_dt)
                & (change_target["記録日_dt"] <= end_dt)
            ].copy()

            if change_user != "全員" and "利用者名" in change_target.columns:
                change_target = change_target[
                    change_target["利用者名"].fillna("").astype(str).str.strip() == clean_text(change_user)
                ].copy()

            period_label = f"{change_start_date.strftime('%Y/%m/%d')}〜{change_end_date.strftime('%Y/%m/%d')}"
            user_label = "全利用者" if change_user == "全員" else change_user

            if change_target.empty:
                st.warning(f"{user_label} の {period_label} に、健康チェックデータがありません。")
            else:
                change_rows = change_target.copy()
                if "気になる変化" not in change_rows.columns:
                    st.error("健康チェックデータに『気になる変化』列がありません。")
                    st.stop()

                change_rows["気になる変化"] = change_rows["気になる変化"].fillna("").astype(str).str.strip()
                change_rows = change_rows[change_rows["気になる変化"] != ""]

                if change_rows.empty:
                    st.success(f"{user_label} の {period_label} に『気になる変化』の記録はありません。")
                else:
                    change_rows = change_rows.sort_values(["記録日_dt", "利用者名"] if "利用者名" in change_rows.columns else ["記録日_dt"])

                    display_cols = []
                    for col in ["記録日", "利用者名", "気になる変化", "家族共有メモ", "入力者", "登録日時"]:
                        if col not in change_rows.columns:
                            change_rows[col] = ""
                        display_cols.append(col)

                    display_df = change_rows[display_cols].copy()
                    display_df["日付"] = pd.to_datetime(display_df["記録日"], errors="coerce").dt.strftime("%Y/%m/%d")
                    display_df = display_df[["日付", "利用者名", "気になる変化", "家族共有メモ", "入力者", "登録日時"]]

                    st.warning(f"{user_label} の {period_label} に、気になる変化が {len(display_df)} 件あります。")
                    st.dataframe(display_df, use_container_width=True, hide_index=True)

                    st.markdown("#### 日付ごとの確認メモ")
                    memo_lines = []
                    for _, row in display_df.iterrows():
                        date_label_raw = clean_text(row.get("日付", ""))
                        row_user_raw = clean_text(row.get("利用者名", ""))
                        change_text_raw = clean_text(row.get("気になる変化", ""))
                        family_text_raw = clean_text(row.get("家族共有メモ", ""))
                        staff_text_raw = clean_text(row.get("入力者", ""))
                        date_label = html_escape_text(date_label_raw)
                        row_user = html_escape_text(row_user_raw)
                        change_text = html_escape_text(change_text_raw)
                        family_text = html_escape_text(family_text_raw)
                        staff_text = html_escape_text(staff_text_raw)
                        family_display = family_text if family_text else "記録なし"
                        staff_display = staff_text if staff_text else "未入力"

                        st.markdown(
                            f"""
                            <div style='background:#FFF8E8; border:1px solid #E5C782; border-radius:14px; padding:12px 14px; margin:8px 0;'>
                                <b>{date_label}　{row_user}</b><br>
                                <span style='color:#7A4A00;'>気になる変化：</span>{change_text}<br>
                                <span style='color:#666;'>家族共有メモ：</span>{family_display}<br>
                                <span style='color:#888; font-size:0.9rem;'>入力者：{staff_display}</span>
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )

                        memo_lines.append(
                            f"{date_label_raw}　{row_user_raw}\n"
                            f"気になる変化：{change_text_raw}\n"
                            f"家族共有メモ：{family_text_raw if family_text_raw else '記録なし'}\n"
                            f"入力者：{staff_text_raw if staff_text_raw else '未入力'}"
                        )

                    safe_user_label = re.sub(r"[\\/:*?\"<>|\s]+", "_", user_label)
                    export_text = f"{user_label}　{period_label}　気になる変化一覧\n\n" + "\n\n".join(memo_lines)
                    st.text_area("コピー用テキスト", value=export_text, height=260)
                    st.download_button(
                        "気になる変化一覧をテキストでダウンロード",
                        data=export_text.encode("utf-8-sig"),
                        file_name=f"気になる変化一覧_{safe_user_label}_{change_start_date.strftime('%Y%m%d')}_{change_end_date.strftime('%Y%m%d')}.txt",
                        mime="text/plain",
                        use_container_width=True,
                    )

    with tab4:
        st.subheader("ChatGPT連携用プロンプト")
        col1, col2, col3 = st.columns(3)
        with col1:
            prompt_user = st.selectbox("利用者", all_users, key="prompt_user")
        with col2:
            prompt_year = st.number_input("年", min_value=2024, max_value=2035, value=today_jst().year, step=1, key="prompt_year")
        with col3:
            prompt_month = st.number_input("月", min_value=1, max_value=12, value=today_jst().month, step=1, key="prompt_month")

        target_h = get_month_health_data(health_df, prompt_user, prompt_year, prompt_month)
        target_e = get_month_excretion_data(ex_df, prompt_user, prompt_year, prompt_month)

        prompt = f"""あなたは介護施設の家族向けレポートを整える文章整理係です。
以下の健康チェック記録・排泄記録・アセスメント情報をもとに、ご家族へ渡す月間レポート文を作成してください。

【重要ルール】
・医療判断、診断、治療効果の断定はしない。
・「問題ありません」「改善しました」「安心です」と断定しない。
・記録に基づく表現にする。
・不安を煽らず、やわらかく丁寧な文章にする。
・対象期間中の全記録を確認し、一部の日付だけを拾って終わらない。
・気になる記録がある場合は日付ごとに列挙する。
・気になる記録がない場合も「その他の日について、特記すべき体調変化や生活上の気になる記録は確認されませんでした」と明記する。
・発熱、SpO2低下、食事摂取量低下、水分摂取量低下、排便なし、不穏、転倒、睡眠不良、表情や活気の変化、申し送りの注意事項を確認する。
・単に平均値をまとめるだけでなく、期間全体を見たうえで、気になる点、継続確認点、職員間で共有すべき点を整理する。

【出力形式】
【対象期間の全体確認】
【気になる記録】
【数値から見た状態】
【排泄状況】
【継続確認が必要な点】
【職員間で共有すること】

【利用者】
{prompt_user}

【アセスメント情報】
{build_assessment_context_text(prompt_user)}

【健康チェック記録】
{target_h.to_string(index=False)}

【排泄記録】
{target_e.to_string(index=False)}
"""
        st.text_area("プロンプト", value=prompt, height=520)

    with tab5:
        st.subheader("申し送り支援")
        target_date = st.date_input("対象日", value=today_jst(), key="handover_date")
        handover = create_handover_text(health_df, ex_df, target_date)
        st.text_area("申し送り案", value=handover, height=360)

    with tab6:
        st.subheader("注意通知")
        alert_date = st.date_input("注意通知の対象日", value=today_jst(), key="alert_date")
        alert_df = build_handover_alerts_by_condition(alert_date)

        if alert_df.empty:
            st.success("対象日の注意通知はありません。")
        else:
            st.warning("条件設定マスタに該当する注意通知があります。")
            st.dataframe(alert_df, use_container_width=True, hide_index=True)

        st.caption("注意通知は診断ではなく、条件設定マスタと記録に基づく確認支援です。")


    with tab7:
        st.subheader("条件設定マスタ変更")
        st.caption("未排便・発熱・SpO2低下・食事量低下など、業務全体申し送りや注意通知で使う抽出条件を管理します。")
        show_alert_condition_master_menu()

    with tab8:
        st.subheader("体重未測定確認")
        st.caption("体重は測定した日だけ入力します。14日以上測定が空いている場合だけ確認対象にします。")
        check_day = st.date_input("確認日", value=today_jst(), key="weight_overdue_check_day")
        show_latest_weight_block(health_df, all_users, check_day)
        show_weight_overdue_block(health_df, all_users, check_day, threshold_days=14)


# =========================
# 利用者マスタ管理
# =========================
elif menu == "メニューカテゴリ設定":
    show_menu_category_settings_menu()

elif menu == "システム設定":
    show_system_settings_menu()

elif menu == "データダウンロード":
    show_admin_data_download_menu()

elif menu == "利用者マスタ管理":
    if st.session_state.role != "admin":
        st.error("この画面は管理者専用です。")
        st.stop()

    st.header("利用者マスタ管理")
    st.caption("利用者の追加・非表示・アセスメント情報の登録管理ができます。")

    df_users = load_users(include_hidden=True)

    st.subheader("現在の利用者一覧")
    st.dataframe(df_users, use_container_width=True, hide_index=True)

    st.divider()
    st.subheader("利用者を追加")

    with st.form("add_user_form", clear_on_submit=True):
        new_user = st.text_input("追加する利用者名", placeholder="例：田中様")
        add_submit = st.form_submit_button("追加する")

    if add_submit:
        ok, msg = add_user(new_user)
        if ok:
            st.success(msg)
            st.rerun()
        else:
            st.error(msg)

    st.divider()
    st.subheader("アセスメント情報の登録・更新")

    if df_users.empty:
        st.info("利用者が登録されていません。")
    else:
        selected_user = st.selectbox("アセスメントを編集する利用者", df_users["利用者名"].tolist(), key="assessment_user")
        selected = df_users[df_users["利用者名"] == selected_user].iloc[0]

        with st.form("assessment_form"):
            values = {}
            values["基本情報"] = st.text_area("基本情報（氏名・住所など）", value=clean_text(selected.get("基本情報", "")), height=80)
            values["主訴"] = st.text_area("主訴（本人・家族の希望や困りごと）", value=clean_text(selected.get("主訴", "")), height=100)
            values["生活状況"] = st.text_area("生活状況（1日の流れ）", value=clean_text(selected.get("生活状況", "")), height=120)
            values["ADL"] = st.text_area("ADL（日常生活動作）", value=clean_text(selected.get("ADL", "")), height=100)
            values["IADL"] = st.text_area("IADL（生活関連動作）", value=clean_text(selected.get("IADL", "")), height=100)
            values["認知機能"] = st.text_area("認知機能（判断・記憶）", value=clean_text(selected.get("認知機能", "")), height=100)
            values["健康状態"] = st.text_area("健康状態（疾患・服薬）", value=clean_text(selected.get("健康状態", "")), height=100)
            values["課題"] = st.text_area("課題（支援が必要な問題点）", value=clean_text(selected.get("課題", "")), height=100)
            values["支援内容"] = st.text_area("支援内容（具体的な対応）", value=clean_text(selected.get("支援内容", "")), height=100)
            assessment_submit = st.form_submit_button("アセスメント情報を保存する")

        if assessment_submit:
            df_save = load_users(include_hidden=True)
            mask = df_save["利用者名"] == selected_user
            for col, value in values.items():
                df_save.loc[mask, col] = value
            save_users(df_save)
            st.success("アセスメント情報を保存しました。")
            st.rerun()

    st.divider()
    st.subheader("利用者を入力候補から外す")

    visible_users = load_active_user_names(include_hidden=False)

    if visible_users:
        target_user = st.selectbox("対象利用者", visible_users, key="hide_user")
        st.warning("この操作は、入力画面の候補から外すだけです。過去データとアセスメント情報は削除されません。")

        if st.button("入力候補から外す"):
            ok, msg = hide_user(target_user)
            if ok:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)

    hidden_df = load_users(include_hidden=True)
    hidden_df = hidden_df[hidden_df["表示"] == "非表示"]

    st.subheader("非表示の利用者を戻す")

    if not hidden_df.empty:
        restore_user = st.selectbox("表示に戻す利用者", hidden_df["利用者名"].tolist(), key="restore_user")
        if st.button("表示に戻す"):
            ok, msg = add_user(restore_user)
            if ok:
                st.success(msg)
                st.rerun()
            else:
                st.error(msg)
    else:
        st.info("非表示の利用者はいません。")

    st.download_button(
        "利用者マスタExcelをダウンロード",
        data=export_user_master_excel_bytes(),
        file_name="利用者マスタ.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
# =========================
# 利用者ID移行チェック
# =========================
elif menu == "利用者ID移行チェック" and is_admin_user():
    show_user_id_migration_check()
elif menu == "利用者名ゆれ紐づけマスタ" and is_admin_user():
    show_user_name_alias_master_menu()

# =========================
# ログイン・職員ID管理
# =========================
elif menu == "ログイン・職員ID管理" and is_admin_user():
    show_login_user_management_menu()
elif menu == "セキュリティ・保守管理" and is_admin_user():
    # Streamlit magic が戻り値（DeltaGenerator）を画面表示しないよう、明示的に代入して呼び出す
    _security_menu_result = show_security_maintenance_menu()
    _security_menu_result = None
